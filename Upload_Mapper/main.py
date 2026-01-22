"""
상품 등록 맵퍼 메인 GUI
"""

import os
import sys
from pathlib import Path

# PyInstaller로 빌드된 환경에서 리소스 경로 찾기
def get_base_path():
    """실행 파일의 기본 경로를 반환 (PyInstaller 환경 고려)"""
    if getattr(sys, 'frozen', False):
        # PyInstaller로 빌드된 실행 파일
        # 실행 파일이 있는 디렉토리 사용
        return Path(sys.executable).parent
    else:
        # 개발 환경
        return Path(__file__).parent

# 모듈 경로 설정 (상대 import를 위한 경로 추가)
current_dir = get_base_path()
if str(current_dir) not in sys.path:
    sys.path.insert(0, str(current_dir))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime
from typing import Optional, Dict, Tuple, List

# xlrd 1.2.0을 사용하여 .xls 파일 읽기 지원
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

def read_excel_with_fallback(file_path, sheet_name=None, **kwargs):
    """엑셀 파일 읽기 (xls 파일 지원 포함)
    
    .xls 파일(97-03 워크시트)은 xlrd 1.2.0으로 직접 읽고,
    .xlsx 파일은 openpyxl을 사용합니다.
    """
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xls':
        # .xls 파일은 xlrd 1.2.0으로 직접 읽기
        if not HAS_XLRD:
            raise ImportError("xlrd가 설치되지 않았습니다. pip install xlrd==1.2.0 을 실행하세요.")
        
        try:
            workbook = xlrd.open_workbook(file_path)
            if sheet_name:
                try:
                    sheet = workbook.sheet_by_name(sheet_name)
                except xlrd.XLRDError:
                    # 시트가 없으면 첫 번째 시트 사용
                    sheet = workbook.sheet_by_index(0)
            else:
                sheet = workbook.sheet_by_index(0)
            
            # 헤더 읽기 (첫 번째 행)
            header = []
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col)
                # xlrd는 숫자 타입을 float로 반환하므로 문자열로 변환
                if isinstance(cell_value, float) and cell_value == int(cell_value):
                    header.append(str(int(cell_value)))
                else:
                    header.append(str(cell_value) if cell_value else f"Unnamed: {col}")
            
            # 데이터 읽기
            data = []
            start_row = 1
            if sheet_name and '기본정보' in str(sheet_name):
                # 이셀러스는 2행이 설명탭이므로 3행부터 (인덱스 2부터)
                start_row = 2
            
            for row_idx in range(start_row, sheet.nrows):
                row_data = []
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col)
                    # xlrd 타입 처리
                    if isinstance(cell_value, float):
                        # 날짜 타입 체크
                        if sheet.cell_type(row_idx, col) == xlrd.XL_CELL_DATE:
                            # 날짜를 문자열로 변환
                            date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                            if date_tuple[0] != 0:  # 유효한 날짜인 경우
                                from datetime import datetime
                                cell_value = datetime(*date_tuple).strftime('%Y-%m-%d')
                            else:
                                cell_value = ""
                        elif cell_value == int(cell_value):
                            cell_value = int(cell_value)
                    row_data.append(cell_value)
                data.append(row_data)
            
            # DataFrame 생성
            df = pd.DataFrame(data, columns=header)
            return df
        except Exception as e:
            raise Exception(f".xls 파일 읽기 실패: {str(e)}")
    else:
        # .xlsx 파일은 openpyxl 사용
        engine = kwargs.pop('engine', 'openpyxl')
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, **kwargs)
        else:
            df = pd.read_excel(file_path, engine=engine, **kwargs)
        
        # 이셀러스 기본정보/확장정보 시트는 2행 설명탭 처리 (3행부터 데이터)
        if sheet_name and ('기본정보' in str(sheet_name) or '확장정보' in str(sheet_name)) and len(df) > 0:
            # 헤더는 유지하고 데이터는 3행부터 (인덱스 2부터)
            df = df.iloc[2:].reset_index(drop=True)
        
        return df

class ToolTip:
    """간단한 툴팁 클래스"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.after_id = None
        self.widget.bind('<Enter>', self.on_enter)
        self.widget.bind('<Leave>', self.on_leave)
        # 클릭 시 툴팁 즉시 숨기기
        self.widget.bind('<Button-1>', self.on_click)
        self.widget.bind('<ButtonRelease-1>', self.on_click)
    
    def on_enter(self, event=None):
        # 약간의 지연 후 툴팁 표시 (실수로 마우스를 올렸을 때 방지)
        if self.after_id:
            self.widget.after_cancel(self.after_id)
        self.after_id = self.widget.after(500, self.show_tooltip)  # 0.5초 지연
    
    def on_leave(self, event=None):
        # 지연된 툴팁 표시 취소
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None
        self.hide_tooltip()
    
    def on_click(self, event=None):
        # 클릭 시 툴팁 즉시 숨기기
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None
        self.hide_tooltip()
        # 클릭 이벤트는 원래 위젯으로 전달되도록 함
    
    def show_tooltip(self):
        if self.tip_window:
            return  # 이미 표시 중이면 무시
        
        x, y, _, _ = self.widget.bbox("insert") if hasattr(self.widget, 'bbox') else (0, 0, 0, 0)
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        
        # 툴팁 창을 최상위로 설정하되, 클릭 이벤트는 통과하도록 처리
        try:
            self.tip_window.wm_attributes('-topmost', True)
            # 툴팁 창의 모든 마우스 이벤트를 원래 위젯으로 전달
            self.tip_window.bind('<Button-1>', lambda e: self._pass_click_event(e))
            self.tip_window.bind('<ButtonRelease-1>', lambda e: self._pass_click_event(e))
        except:
            pass
        
        label = tk.Label(self.tip_window, text=self.text, 
                        background="#ffffe0", relief="solid", borderwidth=1,
                        font=("맑은 고딕", 9), justify="left", wraplength=300)
        label.pack()
        # 라벨에도 클릭 이벤트 전달
        label.bind('<Button-1>', lambda e: self._pass_click_event(e))
        label.bind('<ButtonRelease-1>', lambda e: self._pass_click_event(e))
    
    def _pass_click_event(self, event):
        """클릭 이벤트를 원래 위젯으로 전달"""
        self.hide_tooltip()
        # 이벤트를 원래 위젯으로 시뮬레이션
        # 실제로는 툴팁이 숨겨지면 자동으로 원래 위젯이 클릭됩니다
    
    def hide_tooltip(self):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None

# 모듈 import (절대 import로 변경)
from config_manager import MapperConfig
from solutions import get_solution, list_solutions

# 런처에서 JobManager import 시도
try:
    import sys
    from pathlib import Path
    # 상위 디렉토리를 경로에 추가
    parent_dir = Path(__file__).parent.parent
    if str(parent_dir) not in sys.path:
        sys.path.insert(0, str(parent_dir))
    from main_launcher_v8_Casche import JobManager
    HAS_LAUNCHER = True
except ImportError:
    HAS_LAUNCHER = False
    class JobManager:
        @classmethod
        def update_status(cls, filename, **kwargs):
            pass

class UploadMapperGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("상품 등록 맵퍼 (Upload Mapper)")
        self.geometry("1400x900")
        self.configure(bg="#F0F2F5")
        
        # 기본 디렉토리 설정
        self.base_dir = get_base_path()
        self.config_manager = MapperConfig(self.base_dir)
        
        # 데이터 저장
        self.solution_name: Optional[str] = None
        self.solution_instance = None
        self.solution_df: Optional[pd.DataFrame] = None
        self.processed_df: Optional[pd.DataFrame] = None
        self.solution_file_path: Optional[str] = None
        self.processed_file_path: Optional[str] = None
        self.detected_market: Optional[str] = None  # 감지된 마켓 (스스, 쿠팡, 11번가, 옥션, 지마켓, 토스, 톡스토어)
        
        # 설정 로드
        self.config_data = self.config_manager.load_config()
        
        self._init_ui()
        self._load_last_settings()
    
    def _init_ui(self):
        """UI 초기화"""
        # 헤더
        header = tk.Frame(self, bg="#2C3E50", height=60)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="📋 상품 등록 맵퍼", 
                font=("맑은 고딕", 16, "bold"), 
                bg="#2C3E50", fg="white").pack(expand=True, pady=15)
        
        # 메인 컨테이너
        main_container = tk.Frame(self, bg="#F0F2F5")
        main_container.pack(fill="both", expand=True, padx=15, pady=15)
        
        # =====================================================================
        # 1단계: 솔루션 선택 및 파일 선택
        # =====================================================================
        step1_frame = tk.LabelFrame(main_container, 
                                    text=" [STEP 1] 솔루션 선택 및 파일 불러오기 ",
                                    font=("맑은 고딕", 11, "bold"),
                                    bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        step1_frame.pack(fill="x", pady=(0, 10), ipady=10, ipadx=10)
        
        # 솔루션 선택
        solution_frame = tk.Frame(step1_frame, bg="#FFFFFF")
        solution_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(solution_frame, text="등록 솔루션 선택:", 
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        
        self.solution_var = tk.StringVar()
        solution_combo = ttk.Combobox(solution_frame, 
                                      textvariable=self.solution_var,
                                      values=list_solutions(),
                                      state="readonly",
                                      width=20,
                                      font=("맑은 고딕", 10))
        solution_combo.pack(side="left", padx=5)
        solution_combo.bind("<<ComboboxSelected>>", self._on_solution_selected)
        
        # 다팔자 CSV 생성 버튼 (다팔자 선택 시에만 표시)
        self.csv_button_frame = tk.Frame(solution_frame, bg="#FFFFFF")
        self.csv_button_frame.pack(side="left", padx=10)
        self.csv_button = None  # 초기에는 None
        self.ownerclan_button = None  # 오너클랜 원가/배송비 버튼
        self.batch_mapping_button = None  # 일괄 매핑 버튼
        
        # 파일 선택 영역
        file_frame = tk.Frame(step1_frame, bg="#FFFFFF")
        file_frame.pack(fill="x", padx=10, pady=5)
        
        # 솔루션 엑셀 파일
        solution_file_frame = tk.Frame(file_frame, bg="#FFFFFF")
        solution_file_frame.pack(fill="x", pady=5)
        
        solution_label = tk.Label(solution_file_frame, text="① 등록 솔루션 엑셀:", 
                font=("맑은 고딕", 10),
                bg="#FFFFFF", fg="#333", width=18, anchor="w")
        solution_label.pack(side="left", padx=5)
        ToolTip(solution_label, "다팔자 등록 솔루션 엑셀 파일을 선택하세요.\n파일명 접두사로 마켓이 자동 감지됩니다.")
        
        self.solution_file_label = tk.Label(solution_file_frame, 
                                           text="파일을 선택하세요",
                                           font=("맑은 고딕", 9),
                                           bg="#FFFFFF", fg="#666",
                                           width=50, anchor="w", relief="sunken", bd=1)
        self.solution_file_label.pack(side="left", padx=5, fill="x", expand=True)
        
        solution_load_btn = tk.Button(solution_file_frame, text="📂 불러오기",
                 command=self._load_solution_file,
                 bg="#546E7A", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=10, pady=3)
        solution_load_btn.pack(side="left", padx=5)
        ToolTip(solution_load_btn, "등록 솔루션 엑셀 파일을 불러옵니다.")
        
        # 가공 엑셀 파일
        processed_file_frame = tk.Frame(file_frame, bg="#FFFFFF")
        processed_file_frame.pack(fill="x", pady=5)
        
        processed_label = tk.Label(processed_file_frame, text="② 가공된 엑셀:", 
                font=("맑은 고딕", 10),
                bg="#FFFFFF", fg="#333", width=18, anchor="w")
        processed_label.pack(side="left", padx=5)
        ToolTip(processed_label, "가공된 엑셀 파일을 선택하세요.\n파일명 접두사(스스, 쿠팡, 11번가 등)로 마켓이 자동 감지됩니다.\n필수 컬럼: 마켓판매가격, 판매자 부담 할인")
        
        self.processed_file_label = tk.Label(processed_file_frame, 
                                             text="파일을 선택하세요",
                                             font=("맑은 고딕", 9),
                                             bg="#FFFFFF", fg="#666",
                                             width=50, anchor="w", relief="sunken", bd=1)
        self.processed_file_label.pack(side="left", padx=5, fill="x", expand=True)
        
        processed_load_btn = tk.Button(processed_file_frame, text="📂 불러오기",
                 command=self._load_processed_file,
                 bg="#27ae60", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=10, pady=3)
        processed_load_btn.pack(side="left", padx=5)
        ToolTip(processed_load_btn, "가공된 엑셀 파일을 불러옵니다.\n파일명에서 마켓이 자동 감지됩니다.")
        
        # =====================================================================
        # 2단계: 컬럼 매핑 설정
        # =====================================================================
        step2_frame = tk.LabelFrame(main_container,
                                    text=" [STEP 2] 컬럼 매핑 설정 ",
                                    font=("맑은 고딕", 11, "bold"),
                                    bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        step2_frame.pack(fill="both", expand=True, pady=(0, 10), ipady=10, ipadx=10)
        
        # 매핑 테이블
        mapping_container = tk.Frame(step2_frame, bg="#FFFFFF")
        mapping_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(mapping_container, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(mapping_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.mapping_frame = scrollable_frame
        
        # 매핑 설정 버튼
        btn_frame = tk.Frame(step2_frame, bg="#FFFFFF")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        # 마켓 정보 표시 (다팔자, 이셀러스 솔루션인 경우)
        self.market_info_label = tk.Label(btn_frame,
                                          text="",
                                          font=("맑은 고딕", 9, "bold"),
                                          bg="#FFFFFF", fg="#e74c3c")
        self.market_info_label.pack(side="left", padx=5, pady=5)
        
        # 설명 라벨
        info_label = tk.Label(btn_frame, 
                             text="💡 매핑 안함 = 해당 컬럼은 매핑하지 않음 | 기본값 = 매핑된 컬럼에 값이 없을 때 사용할 기본값",
                             font=("맑은 고딕", 8),
                             bg="#FFFFFF", fg="#666",
                             justify="left")
        info_label.pack(side="left", padx=5, pady=5)
        ToolTip(info_label, "• 매핑 안함: 해당 컬럼은 매핑하지 않습니다.\n• 기본값: 매핑된 컬럼에 값이 없을 때 사용할 기본값을 입력합니다.\n• 마켓별 기본 매핑이 자동으로 적용됩니다.")
        
        save_btn = tk.Button(btn_frame, text="💾 매핑 설정 저장",
                 command=self._save_mapping_config,
                 bg="#3498db", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=15, pady=5)
        save_btn.pack(side="right", padx=5)
        ToolTip(save_btn, "현재 설정한 매핑을 저장합니다.\n다팔자 솔루션의 경우 마켓별로 별도 저장됩니다.")
        
        load_btn = tk.Button(btn_frame, text="📥 매핑 설정 불러오기",
                 command=self._load_mapping_config,
                 bg="#95a5a6", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=15, pady=5)
        load_btn.pack(side="right", padx=5)
        ToolTip(load_btn, "저장된 매핑 설정을 불러옵니다.\n가공 엑셀을 불러올 때 자동으로 불러옵니다.")
        
        # 상세설명 문구 설정 버튼
        detail_btn = tk.Button(btn_frame, text="📝 상세설명 문구 설정",
                 command=self._edit_detail_texts,
                 bg="#9b59b6", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=15, pady=5)
        detail_btn.pack(side="left", padx=5)
        ToolTip(detail_btn, "상세정보 상단/하단에 표시될 문구를 설정합니다.\nHTML 형식 사용 가능하며, 템플릿을 선택할 수 있습니다.")
        
        # 솔루션별 고급 설정 버튼
        advanced_btn = tk.Button(btn_frame, text="⚙️ 솔루션별 고급 설정",
                 command=self._open_advanced_settings,
                 bg="#e67e22", fg="white",
                 font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2",
                 padx=15, pady=5)
        advanced_btn.pack(side="left", padx=5)
        ToolTip(advanced_btn, "배송비 계산 방식, 옵션금액 규칙,\n상세정보 상단/하단 이미지 설정 등을 변경할 수 있습니다.")
        
        # 마켓 설정 버튼 (다팔자 솔루션인 경우에만 표시)
        self.market_settings_button = tk.Button(btn_frame,
                                                text="🔧 마켓 설정",
                                                command=self._open_market_settings,
                                                bg="#16a085", fg="white",
                                                font=("맑은 고딕", 9, "bold"),
                                                relief="raised", cursor="hand2",
                                                padx=15, pady=5)
        # 초기에는 숨김
        self.market_settings_button.pack_forget()
        ToolTip(self.market_settings_button, "마켓 접두사를 추가/수정/삭제할 수 있습니다.\n새로운 마켓이 추가될 때 사용합니다.\n예: '네이버' 접두사 추가")
        
        # =====================================================================
        # 3단계: 실행 및 결과
        # =====================================================================
        step3_frame = tk.LabelFrame(main_container,
                                    text=" [STEP 3] 매핑 실행 및 결과 저장 ",
                                    font=("맑은 고딕", 11, "bold"),
                                    bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        step3_frame.pack(fill="x", pady=(0, 10), ipady=10, ipadx=10)
        
        action_frame = tk.Frame(step3_frame, bg="#FFFFFF")
        action_frame.pack(fill="x", padx=10, pady=5)
        
        execute_btn = tk.Button(action_frame, text="🚀 매핑 실행",
                 command=self._execute_mapping,
                 bg="#e74c3c", fg="white",
                 font=("맑은 고딕", 12, "bold"),
                 relief="raised", cursor="hand2",
                 padx=20, pady=8)
        execute_btn.pack(side="left", padx=5)
        execute_tooltip = ToolTip(execute_btn, "설정한 매핑 규칙을 적용하여 결과를 생성합니다.\n다팔자 솔루션의 경우 원본 파일에 저장됩니다.\n(백업 파일 자동 생성)")
        # 클릭 시 툴팁 숨기기 보장
        execute_btn.bind('<Button-1>', lambda e: execute_tooltip.hide_tooltip())
        execute_btn.bind('<ButtonRelease-1>', lambda e: execute_tooltip.hide_tooltip())
        
        self.status_label = tk.Label(action_frame,
                                     text="준비됨",
                                     font=("맑은 고딕", 10),
                                     bg="#FFFFFF", fg="#27ae60")
        self.status_label.pack(side="left", padx=20)
        
        # 로그 영역
        log_frame = tk.Frame(main_container, bg="#FFFFFF", bd=1, relief="sunken")
        log_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        tk.Label(log_frame, text="📋 작업 로그",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=5, pady=5)
        
        self.log_text = tk.Text(log_frame, height=8, font=("Consolas", 9),
                               bg="#F8F9FA", fg="#333", relief="flat",
                               wrap=tk.WORD)
        self.log_text.pack(fill="both", expand=True, padx=5, pady=(0, 5))
        
        scrollbar_log = ttk.Scrollbar(self.log_text, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar_log.set)
        scrollbar_log.pack(side="right", fill="y")
    
    def _log(self, message: str, level: str = "INFO"):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_msg = f"[{timestamp}] [{level}] {message}\n"
        self.log_text.insert(tk.END, log_msg)
        self.log_text.see(tk.END)
        self.update_idletasks()
    
    def _on_solution_selected(self, event=None):
        """솔루션 선택 시 호출"""
        self.solution_name = self.solution_var.get()
        if self.solution_name:
            try:
                self.solution_instance = get_solution(self.solution_name)
                self._log(f"솔루션 선택: {self.solution_name}")
                
                # 다팔자 선택 시 CSV 생성 버튼 및 마켓 설정 버튼 표시
                if self.solution_name == "다팔자":
                    if self.csv_button is None:
                        self.csv_button = tk.Button(self.csv_button_frame,
                                                   text="📄 다팔자 등록용 CSV 생성",
                                                   command=self._create_dafalza_csv,
                                                   bg="#e74c3c", fg="white",
                                                   font=("맑은 고딕", 9, "bold"),
                                                   relief="raised", cursor="hand2",
                                                   padx=10, pady=3)
                        self.csv_button.pack(side="left", padx=5)
                        ToolTip(self.csv_button, "여러 엑셀 파일을 선택하여\n다팔자 등록용 CSV 파일을 일괄 생성합니다.\n각 엑셀의 첫 번째 열만 추출하여 CSV로 저장합니다.")
                    
                    # 마켓 설정 버튼 표시
                    if hasattr(self, 'market_settings_button'):
                        self.market_settings_button.pack(side="left", padx=5, before=self.market_info_label)
                elif self.solution_name == "이셀러스":
                    # 이셀러스 선택 시 CSV 생성 버튼 표시
                    if self.csv_button is None:
                        self.csv_button = tk.Button(self.csv_button_frame,
                                                   text="📄 이셀러스 등록용 CSV 생성",
                                                   command=self._create_esellers_csv,
                                                   bg="#e74c3c", fg="white",
                                                   font=("맑은 고딕", 9, "bold"),
                                                   relief="raised", cursor="hand2",
                                                   padx=10, pady=3)
                        self.csv_button.pack(side="left", padx=5)
                        ToolTip(self.csv_button, "여러 엑셀 파일을 선택하여\n이셀러스 등록용 CSV 파일을 생성합니다.\n각 엑셀의 상품코드 열을 추출하여 CSV로 저장합니다.")
                    
                    # 오너클랜 원가/배송비 버튼 표시
                    if self.ownerclan_button is None:
                        self.ownerclan_button = tk.Button(self.csv_button_frame,
                                                         text="💰 오너클랜 원가/배송비",
                                                         command=self._create_ownerclan_excel,
                                                         bg="#27ae60", fg="white",
                                                         font=("맑은 고딕", 9, "bold"),
                                                         relief="raised", cursor="hand2",
                                                         padx=10, pady=3)
                        self.ownerclan_button.pack(side="left", padx=5)
                        ToolTip(self.ownerclan_button, "엑셀 파일을 선택하여\n오너클랜 원가/배송비 정보를 추출합니다.\n상품코드, 오너클랜판매가, 배송비, 배송유형, 최대구매수량, 반품배송비 컬럼만 추출합니다.")
                    
                    # 일괄 매핑 버튼 표시
                    if not hasattr(self, 'batch_mapping_button') or self.batch_mapping_button is None:
                        self.batch_mapping_button = tk.Button(self.csv_button_frame,
                                                             text="🔄 일괄 매핑",
                                                             command=self._batch_mapping_esellers,
                                                             bg="#9b59b6", fg="white",
                                                             font=("맑은 고딕", 9, "bold"),
                                                             relief="raised", cursor="hand2",
                                                             padx=10, pady=3)
                        self.batch_mapping_button.pack(side="left", padx=5)
                        ToolTip(self.batch_mapping_button, "여러 등록 솔루션 엑셀과 가공 엑셀을\n파일명 기준으로 자동 매칭하여 일괄 매핑합니다.\n날짜_스토어명_스토어사업자번호 패턴으로 매칭됩니다.")
                else:
                    # 다른 솔루션 선택 시 버튼 숨김
                    if self.csv_button is not None:
                        self.csv_button.destroy()
                        self.csv_button = None
                    if self.ownerclan_button is not None:
                        self.ownerclan_button.destroy()
                        self.ownerclan_button = None
                    if hasattr(self, 'batch_mapping_button') and self.batch_mapping_button is not None:
                        self.batch_mapping_button.destroy()
                        self.batch_mapping_button = None
                    if hasattr(self, 'market_settings_button'):
                        self.market_settings_button.pack_forget()
                
                # 기존 매핑 설정 불러오기
                if self.processed_df is not None:
                    self._update_mapping_ui()
                    self._load_mapping_config()
            except Exception as e:
                messagebox.showerror("오류", f"솔루션 로드 실패:\n{str(e)}")
                self._log(f"솔루션 로드 실패: {str(e)}", "ERROR")
    
    def _load_solution_file(self):
        """등록 솔루션 엑셀 파일 불러오기"""
        if not self.solution_name or not self.solution_instance:
            messagebox.showwarning("경고", "먼저 등록 솔루션을 선택해주세요.")
            return
        
        # 솔루션별 기본 경로 제안
        initial_dir = None
        if self.solution_instance.temp_path_template:
            username = os.getenv("USERNAME") or os.getenv("USER")
            if username:
                default_path = self.solution_instance.temp_path_template.format(username=username)
                if os.path.exists(default_path):
                    initial_dir = default_path
        
        file_path = filedialog.askopenfilename(
            title=f"{self.solution_name} 엑셀 파일 선택",
            initialdir=initial_dir,
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            # 이셀러스는 기본정보 시트를 읽어야 함
            if self.solution_name == "이셀러스":
                # read_excel_with_fallback에서 2행 설명탭 처리 (3행부터 데이터)
                self.solution_df = read_excel_with_fallback(file_path, sheet_name="기본정보")
                # 모든 행이 유지되어야 하므로 추가 처리 없음
            else:
                self.solution_df = read_excel_with_fallback(file_path)
            
            self.solution_file_path = file_path
            filename = os.path.basename(file_path)
            self.solution_file_label.config(text=filename, fg="#27ae60")
            self._log(f"솔루션 엑셀 로드 완료: {filename} ({len(self.solution_df)}행, {len(self.solution_df.columns)}컬럼)")
            self._update_mapping_ui()
        except Exception as e:
            messagebox.showerror("오류", f"파일 로드 실패:\n{str(e)}")
            self._log(f"파일 로드 실패: {str(e)}", "ERROR")
    
    def _detect_market_from_filename(self, filename: str) -> Optional[str]:
        """파일명에서 마켓 접두사 감지
        
        Args:
            filename: 파일명
            
        Returns:
            감지된 마켓명 또는 None
        """
        filename_lower = filename.lower()
        
        # 설정에서 마켓 접두사 목록 가져오기 (동적으로 확장 가능)
        market_prefixes = self.config_manager.get_market_prefixes()
        
        # 접두사 길이 순으로 정렬 (긴 접두사 우선, 예: "스마트스토어"가 "스스"보다 먼저 체크)
        sorted_prefixes = sorted(market_prefixes.items(), key=lambda x: len(x[0]), reverse=True)
        
        # 파일명에 접두사가 포함되어 있는지 확인 (시작 부분 우선, 그 다음 전체 검색)
        for prefix, market in sorted_prefixes:
            prefix_lower = prefix.lower()
            # 1. 파일명이 접두사로 시작하는 경우 (우선)
            if filename_lower.startswith(prefix_lower):
                return market
            # 2. 파일명에 접두사가 포함되어 있는 경우
            # 단, 접두사 앞뒤로 구분자(_, -, 공백 등)가 있거나 파일명 시작/끝인 경우만
            if prefix_lower in filename_lower:
                # 접두사 위치 찾기
                idx = filename_lower.find(prefix_lower)
                # 접두사 앞이 파일명 시작이거나 구분자인지 확인
                if idx == 0 or filename_lower[idx-1] in ['_', '-', ' ', '.']:
                    # 접두사 뒤 확인: 파일명 끝이거나 구분자이거나 영문자/숫자인 경우 허용
                    next_idx = idx + len(prefix_lower)
                    if next_idx >= len(filename_lower):
                        # 파일명 끝
                        return market
                    else:
                        next_char = filename_lower[next_idx]
                        # 구분자 또는 영문자/숫자인 경우 허용
                        if next_char in ['_', '-', ' ', '.']:
                            return market
                        # 영문자나 숫자가 바로 오는 경우도 허용 (예: "스스A1", "쿠팡2024")
                        if next_char.isalnum():
                            return market
        
        return None
    
    def _load_processed_file(self):
        """가공된 엑셀 파일 불러오기"""
        file_path = filedialog.askopenfilename(
            title="가공된 엑셀 파일 선택",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            # 엑셀 파일 읽기 (#N/A, 빈 값 처리)
            self.processed_df = read_excel_with_fallback(
                file_path, 
                na_values=['#N/A', '#N/A!', '#NA', '-1.#IND', '-1.#QNAN', '-NaN', '-nan', '1.#IND', '1.#QNAN', '<NA>', 'N/A', 'NA', 'NULL', 'NaN', 'n/a', 'nan', 'null']
            )
            self.processed_file_path = file_path
            filename = os.path.basename(file_path)
            
            # 마켓 감지 (다팔자, 이셀러스 솔루션인 경우)
            self.detected_market = None
            if self.solution_name in ["다팔자", "이셀러스"]:
                self.detected_market = self._detect_market_from_filename(filename)
                if self.detected_market:
                    self._log(f"마켓 감지: {self.detected_market}")
                    # 마켓 정보 표시
                    if hasattr(self, 'market_info_label'):
                        self.market_info_label.config(text=f"📌 감지된 마켓: {self.detected_market}", fg="#e74c3c")
                else:
                    self._log("마켓 감지 실패: 파일명 접두사를 확인할 수 없습니다.", "WARNING")
                    if hasattr(self, 'market_info_label'):
                        self.market_info_label.config(text="⚠️ 마켓 미감지", fg="#f39c12")
            else:
                if hasattr(self, 'market_info_label'):
                    self.market_info_label.config(text="")
            
            # 필수 컬럼 검증 (다팔자 솔루션인 경우)
            if self.solution_name == "다팔자":
                missing_cols = []
                if "마켓판매가격" not in self.processed_df.columns:
                    missing_cols.append("마켓판매가격")
                if "판매자 부담 할인" not in self.processed_df.columns:
                    missing_cols.append("판매자 부담 할인")
                
                if missing_cols:
                    messagebox.showwarning(
                        "경고",
                        f"가공된 엑셀에 필수 컬럼이 없습니다:\n\n" +
                        "\n".join(f"- {col}" for col in missing_cols) +
                        "\n\n매핑을 진행할 수 없습니다."
                    )
                    self.processed_df = None
                    self.processed_file_path = None
                    self.processed_file_label.config(text="파일을 선택하세요", fg="#666")
                    return
            
            self.processed_file_label.config(text=filename, fg="#27ae60")
            self._log(f"가공 엑셀 로드 완료: {filename} ({len(self.processed_df)}행, {len(self.processed_df.columns)}컬럼)")
            self._update_mapping_ui()
            # 매핑 설정 자동 불러오기
            if self.solution_name:
                self._load_mapping_config()
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("오류", f"파일 로드 실패:\n{str(e)}")
            self._log(f"파일 로드 실패: {str(e)}", "ERROR")
            self._log(f"상세 오류:\n{error_detail}", "ERROR")
    
    def _update_mapping_ui(self):
        """매핑 UI 업데이트"""
        # 기존 위젯 제거
        for widget in self.mapping_frame.winfo_children():
            widget.destroy()
        
        if self.processed_df is None:
            return
        
        # 헤더
        header_frame = tk.Frame(self.mapping_frame, bg="#E8EAF6")
        header_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(header_frame, text="가공 엑셀 컬럼", 
                font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=25).pack(side="left", padx=5)
        tk.Label(header_frame, text="→", 
                font=("맑은 고딕", 10, "bold"),
                bg="#E8EAF6", fg="#546E7A", width=5).pack(side="left")
        tk.Label(header_frame, text="솔루션 엑셀 컬럼", 
                font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=25).pack(side="left", padx=5)
        tk.Label(header_frame, text="기본값 (매핑 없을 시)", 
                font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=20).pack(side="left", padx=5)
        
        # 매핑 행들
        self.mapping_vars = {}
        processed_columns = list(self.processed_df.columns)
        
        solution_columns = []
        if self.solution_df is not None:
            solution_columns = list(self.solution_df.columns)
        
        for idx, proc_col in enumerate(processed_columns):
            row_frame = tk.Frame(self.mapping_frame, bg="#FFFFFF")
            row_frame.pack(fill="x", padx=5, pady=2)
            
            # 가공 엑셀 컬럼명
            tk.Label(row_frame, text=proc_col,
                    font=("맑은 고딕", 9),
                    bg="#FFFFFF", fg="#333", width=25, anchor="w").pack(side="left", padx=5)
            
            tk.Label(row_frame, text="→",
                    font=("맑은 고딕", 10),
                    bg="#FFFFFF", fg="#999", width=5).pack(side="left")
            
            # 솔루션 컬럼 선택
            sol_var = tk.StringVar()
            sol_combo = ttk.Combobox(row_frame,
                                    textvariable=sol_var,
                                    values=["(매핑 안함)"] + solution_columns,
                                    state="readonly",
                                    width=25,
                                    font=("맑은 고딕", 9))
            sol_combo.pack(side="left", padx=5)
            
            # 기본값 입력
            default_var = tk.StringVar()
            default_entry = tk.Entry(row_frame,
                                    textvariable=default_var,
                                    width=20,
                                    font=("맑은 고딕", 9))
            default_entry.pack(side="left", padx=5)
            
            self.mapping_vars[proc_col] = {
                "solution_col": sol_var,
                "default_value": default_var
            }
    
    def _edit_detail_texts(self):
        """상세설명 상단/하단 문구 편집"""
        if not self.solution_name:
            messagebox.showwarning("경고", "솔루션을 선택해주세요.")
            return
        
        config_data = self.config_manager.get_solution_config(self.solution_name)
        current_top = config_data.get("detail_top_text", "")
        current_bottom = config_data.get("detail_bottom_text", "")
        current_top_template = config_data.get("detail_top_template", "")
        current_bottom_template = config_data.get("detail_bottom_template", "")
        
        # 템플릿 목록 가져오기
        top_templates = self.config_manager.list_templates("top")
        bottom_templates = self.config_manager.list_templates("bottom")
        
        # 팝업 창 생성
        popup = tk.Toplevel(self)
        popup.title("상세설명 문구 설정")
        popup.geometry("900x700")
        popup.configure(bg="#FFFFFF")
        
        # 템플릿 선택 영역
        template_frame = tk.Frame(popup, bg="#FFFFFF")
        template_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(template_frame, text="상단 템플릿:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        
        top_template_var = tk.StringVar(value=current_top_template)
        top_template_combo = ttk.Combobox(template_frame,
                                         textvariable=top_template_var,
                                         values=["(직접 입력)"] + top_templates,
                                         state="readonly",
                                         width=20)
        top_template_combo.pack(side="left", padx=5)
        
        tk.Label(template_frame, text="하단 템플릿:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333").pack(side="left", padx=10)
        
        bottom_template_var = tk.StringVar(value=current_bottom_template)
        bottom_template_combo = ttk.Combobox(template_frame,
                                            textvariable=bottom_template_var,
                                            values=["(직접 입력)"] + bottom_templates,
                                            state="readonly",
                                            width=20)
        bottom_template_combo.pack(side="left", padx=5)
        
        tk.Label(popup, text="상세설명 상단 문구 (HTML 형식 가능):",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=(10, 5))
        
        top_text = tk.Text(popup, height=8, font=("맑은 고딕", 10),
                          bg="#F8F9FA", fg="#333", wrap=tk.WORD)
        top_text.pack(fill="both", expand=True, padx=10, pady=5)
        top_text.insert("1.0", current_top)
        
        def on_top_template_change(event=None):
            template_name = top_template_var.get()
            if template_name and template_name != "(직접 입력)":
                template_content = self.config_manager.get_template("top", template_name)
                top_text.delete("1.0", tk.END)
                top_text.insert("1.0", template_content)
        
        top_template_combo.bind("<<ComboboxSelected>>", on_top_template_change)
        
        tk.Label(popup, text="상세설명 하단 문구 (HTML 형식 가능):",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=(10, 5))
        
        bottom_text = tk.Text(popup, height=8, font=("맑은 고딕", 10),
                             bg="#F8F9FA", fg="#333", wrap=tk.WORD)
        bottom_text.pack(fill="both", expand=True, padx=10, pady=5)
        bottom_text.insert("1.0", current_bottom)
        
        def on_bottom_template_change(event=None):
            template_name = bottom_template_var.get()
            if template_name and template_name != "(직접 입력)":
                template_content = self.config_manager.get_template("bottom", template_name)
                bottom_text.delete("1.0", tk.END)
                bottom_text.insert("1.0", template_content)
        
        bottom_template_combo.bind("<<ComboboxSelected>>", on_bottom_template_change)
        
        def save_texts():
            top_content = top_text.get("1.0", tk.END).strip()
            bottom_content = bottom_text.get("1.0", tk.END).strip()
            top_template = top_template_var.get() if top_template_var.get() != "(직접 입력)" else ""
            bottom_template = bottom_template_var.get() if bottom_template_var.get() != "(직접 입력)" else ""
            
            config_data = self.config_manager.get_solution_config(self.solution_name)
            config_data["detail_top_text"] = top_content
            config_data["detail_bottom_text"] = bottom_content
            config_data["detail_top_template"] = top_template
            config_data["detail_bottom_template"] = bottom_template
            self.config_manager.save_solution_config(self.solution_name, config_data)
            
            self._log(f"상세설명 문구 저장 완료: {self.solution_name}")
            messagebox.showinfo("완료", "상세설명 문구가 저장되었습니다.")
            popup.destroy()
        
        btn_frame = tk.Frame(popup, bg="#FFFFFF")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(btn_frame, text="저장", command=save_texts,
                 bg="#3498db", fg="white",
                 font=("맑은 고딕", 10, "bold"),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)
        
        tk.Button(btn_frame, text="취소", command=popup.destroy,
                 bg="#95a5a6", fg="white",
                 font=("맑은 고딕", 10),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)
    
    def _save_mapping_config(self):
        """매핑 설정 저장"""
        if not self.solution_name:
            messagebox.showwarning("경고", "솔루션을 선택해주세요.")
            return
        
        if not self.mapping_vars:
            messagebox.showwarning("경고", "매핑할 컬럼이 없습니다.")
            return
        
        column_mapping = {}
        default_values = {}
        
        for proc_col, vars_dict in self.mapping_vars.items():
            sol_col = vars_dict["solution_col"].get()
            default_val = vars_dict["default_value"].get().strip()
            
            if sol_col and sol_col != "(매핑 안함)":
                column_mapping[proc_col] = sol_col
                if default_val:
                    default_values[sol_col] = default_val
        
        config_data = self.config_manager.get_solution_config(self.solution_name, self.detected_market)
        config_data["column_mapping"] = column_mapping
        config_data["default_values"] = default_values
        
        # 마켓 정보와 함께 저장
        self.config_manager.save_solution_config(self.solution_name, config_data, self.detected_market)
        
        if self.solution_name == "다팔자" and self.detected_market:
            self._log(f"매핑 설정 저장 완료: {self.solution_name} ({self.detected_market})")
            messagebox.showinfo("완료", f"매핑 설정이 저장되었습니다.\n\n마켓: {self.detected_market}")
        else:
            self._log(f"매핑 설정 저장 완료: {self.solution_name}")
            messagebox.showinfo("완료", "매핑 설정이 저장되었습니다.")
    
    def _load_mapping_config(self):
        """저장된 매핑 설정 불러오기"""
        if not self.solution_name:
            return
        
        if not self.mapping_vars:
            return
        
        # 마켓 정보와 함께 설정 불러오기
        config_data = self.config_manager.get_solution_config(self.solution_name, self.detected_market)
        column_mapping = config_data.get("column_mapping", {})
        default_values = config_data.get("default_values", {})
        
        # 기본 매핑 규칙 적용
        if self.solution_instance:
            default_mapping = self.solution_instance.get_default_mapping()
            for proc_col, sol_col in default_mapping.items():
                if proc_col in self.mapping_vars:
                    self.mapping_vars[proc_col]["solution_col"].set(sol_col)
        
        # 다팔자 솔루션의 경우 마켓별 기본 매핑 추가
        if self.solution_name == "다팔자" and self.detected_market and self.processed_df is not None:
            market_mapping = self._get_market_specific_mapping()
            for proc_col, sol_col in market_mapping.items():
                if proc_col in self.mapping_vars:
                    # 이미 기본 매핑이 설정되지 않은 경우에만 적용
                    current_value = self.mapping_vars[proc_col]["solution_col"].get()
                    if not current_value or current_value == "(매핑 안함)":
                        self.mapping_vars[proc_col]["solution_col"].set(sol_col)
        
        # 저장된 매핑 복원
        for proc_col, vars_dict in self.mapping_vars.items():
            if proc_col in column_mapping:
                sol_col = column_mapping[proc_col]
                vars_dict["solution_col"].set(sol_col)
            elif proc_col not in (self.solution_instance.get_default_mapping() if self.solution_instance else {}):
                # 마켓별 매핑에도 없는 경우에만 "(매핑 안함)" 설정
                if not (self.solution_name == "다팔자" and self.detected_market and 
                        proc_col in self._get_market_specific_mapping()):
                    vars_dict["solution_col"].set("(매핑 안함)")
            
            # 기본값 복원
            sol_col = vars_dict["solution_col"].get()
            if sol_col in default_values:
                vars_dict["default_value"].set(default_values[sol_col])
            else:
                vars_dict["default_value"].set("")
        
        if self.solution_name == "다팔자" and self.detected_market:
            self._log(f"매핑 설정 불러오기 완료: {self.solution_name} ({self.detected_market})")
        else:
            self._log(f"매핑 설정 불러오기 완료: {self.solution_name}")
    
    def _get_market_specific_mapping(self) -> Dict[str, str]:
        """마켓별 기본 매핑 규칙 반환
        
        Returns:
            {가공엑셀컬럼: 솔루션엑셀컬럼} 딕셔너리
        """
        if not self.detected_market or self.processed_df is None:
            return {}
        
        # solution_df가 None인지 안전하게 확인
        solution_columns = []
        if self.solution_df is not None:
            try:
                solution_columns = list(self.solution_df.columns)
            except Exception:
                solution_columns = []
        
        mapping = {}
        
        # 공통: 마켓판매가격 -> 가격
        if "마켓판매가격" in self.processed_df.columns and "가격" in solution_columns:
            mapping["마켓판매가격"] = "가격"
        
        # 공통: 판매자 부담 할인
        if "판매자 부담 할인" in self.processed_df.columns and "판매자 부담 할인" in solution_columns:
            mapping["판매자 부담 할인"] = "판매자 부담 할인"
        
        # 마켓별 이미지 컬럼 매핑
        if self.detected_market == "11번가":
            # 11번가: 대표 이미지
            if "대표 이미지" in self.processed_df.columns and "대표 이미지" in solution_columns:
                mapping["대표 이미지"] = "대표 이미지"
        elif self.detected_market == "쿠팡":
            # 쿠팡: 목록 이미지 -> 등록 솔루션의 '목록 이미지' 또는 '대표 이미지'
            if "목록 이미지" in self.processed_df.columns and self.solution_df is not None:
                if "목록 이미지" in solution_columns:
                    mapping["목록 이미지"] = "목록 이미지"
                elif "대표 이미지" in solution_columns:
                    mapping["목록 이미지"] = "대표 이미지"
        elif self.detected_market in ["옥션", "지마켓"]:
            # 옥션, 지마켓: 사용URL 또는 목록이미지 (띄어쓰기 없음) 또는 목록 이미지
            if self.solution_df is not None:
                # 가공 엑셀의 가능한 컬럼명 확인 (사용URL 우선)
                processed_image_col = None
                if "사용URL" in self.processed_df.columns:
                    processed_image_col = "사용URL"
                elif "목록이미지" in self.processed_df.columns:
                    processed_image_col = "목록이미지"
                elif "목록 이미지" in self.processed_df.columns:
                    processed_image_col = "목록 이미지"
                
                if processed_image_col:
                    # 등록 솔루션의 가능한 컬럼명 확인
                    if "목록 이미지" in solution_columns:
                        mapping[processed_image_col] = "목록 이미지"
                    elif "대표 이미지" in solution_columns:
                        mapping[processed_image_col] = "대표 이미지"
        
        # 마켓별 배송비 매핑
        if self.detected_market in ["옥션", "지마켓"]:
            # 옥션, 지마켓: 등록 솔루션에 배송비, 교환배송비 컬럼이 없음
            # 가공 엑셀의 반품배송비를 등록 솔루션의 반품배송비로 매핑
            if "반품배송비" in self.processed_df.columns and "반품배송비" in solution_columns:
                mapping["반품배송비"] = "반품배송비"
        
        return mapping
    
    def _open_advanced_settings(self):
        """솔루션별 고급 설정 창 열기"""
        if not self.solution_name:
            messagebox.showwarning("경고", "솔루션을 선택해주세요.")
            return
        
        config_data = self.config_manager.get_solution_config(self.solution_name)
        
        # 팝업 창 생성
        popup = tk.Toplevel(self)
        popup.title(f"{self.solution_name} 고급 설정")
        popup.geometry("1000x800")
        popup.configure(bg="#FFFFFF")
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(popup, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # =====================================================================
        # 1. 배송비 계산 방식 설정
        # =====================================================================
        shipping_frame = tk.LabelFrame(scrollable_frame,
                                      text="🚚 배송비 계산 방식",
                                      font=("맑은 고딕", 11, "bold"),
                                      bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        shipping_frame.pack(fill="x", padx=10, pady=10)
        
        shipping_method_var = tk.StringVar(value=config_data.get("shipping_method", "standard"))
        
        tk.Label(shipping_frame, text="배송비 계산 방식 선택:",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=10)
        
        method_frame = tk.Frame(shipping_frame, bg="#FFFFFF")
        method_frame.pack(fill="x", padx=20, pady=5)
        
        tk.Radiobutton(method_frame, text="형식 1: 기본 배송비 변환 로직 (금액대별 변환)",
                      variable=shipping_method_var, value="standard",
                      font=("맑은 고딕", 9),
                      bg="#FFFFFF", fg="#333",
                      activebackground="#FFFFFF").pack(anchor="w", pady=5)
        
        tk.Label(method_frame, 
                text="  • 금액대별 변환 규칙 적용\n  • 반품배송비 = 변경된 배송비 + 1000\n  • 교환배송비 = 변경된 반품배송비 × 2",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        tk.Radiobutton(method_frame, text="형식 2: 무료배송으로 전환",
                      variable=shipping_method_var, value="free",
                      font=("맑은 고딕", 9),
                      bg="#FFFFFF", fg="#333",
                      activebackground="#FFFFFF").pack(anchor="w", pady=5)
        
        tk.Label(method_frame, 
                text="  • 배송비 = 0\n  • 반품배송비 = 기존 배송비 + 1000\n  • 교환배송비 = 변경된 반품배송비 × 2",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        # =====================================================================
        # 2. 옵션금액 규칙 설정
        # =====================================================================
        option_price_frame = tk.LabelFrame(scrollable_frame,
                                          text="💰 옵션금액 규칙",
                                          font=("맑은 고딕", 11, "bold"),
                                          bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        option_price_frame.pack(fill="x", padx=10, pady=10)
        
        option_price_rule_var = tk.StringVar(value=config_data.get("option_price_rule", "smartstore"))
        
        tk.Label(option_price_frame, text="옵션금액 규칙 선택:",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=10)
        
        option_rule_frame = tk.Frame(option_price_frame, bg="#FFFFFF")
        option_rule_frame.pack(fill="x", padx=20, pady=5)
        
        tk.Radiobutton(option_rule_frame, text="스마트스토어 기준 (기본값)",
                      variable=option_price_rule_var, value="smartstore",
                      font=("맑은 고딕", 9),
                      bg="#FFFFFF", fg="#333",
                      activebackground="#FFFFFF").pack(anchor="w", pady=5)
        
        tk.Label(option_rule_frame, 
                text="  • P < 2,000: max_delta = P (10원 단위 내림)\n  • 2,000 ≤ P < 10,000: max_delta = P (10원 단위 내림)\n  • P ≥ 10,000: max_delta = P × 0.5\n  • 단위 내림 규칙:\n    - 1만원 이하: 10원 단위\n    - 1만원 초과 3만원 이하: 100원 단위\n    - 3만원 초과 6만원 이하: 500원 단위\n    - 6만원 초과: 1000원 단위\n  • 양수 값이 2개 이상이면 비율 유지 스케일링",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        tk.Radiobutton(option_rule_frame, text="사용 안함",
                      variable=option_price_rule_var, value="none",
                      font=("맑은 고딕", 9),
                      bg="#FFFFFF", fg="#333",
                      activebackground="#FFFFFF").pack(anchor="w", pady=5)
        
        tk.Label(option_rule_frame, 
                text="  • 옵션금액 자동 보정 기능을 사용하지 않습니다.",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        # =====================================================================
        # 3. 상세정보 상단 설정
        # =====================================================================
        detail_top_frame = tk.LabelFrame(scrollable_frame,
                                         text="📝 상세정보 상단 설정",
                                         font=("맑은 고딕", 11, "bold"),
                                         bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        detail_top_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(detail_top_frame, 
                text="※ 상단에는 '등록 솔루션 엑셀'의 '대표 이미지' 컬럼 값이 이미지로 자동 삽입됩니다.",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=10, pady=(10, 5))
        
        # 이미지 사이즈 설정
        image_size_frame = tk.Frame(detail_top_frame, bg="#FFFFFF")
        image_size_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(image_size_frame, text="이미지 사이즈:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333", width=15).pack(side="left", padx=5)
        
        image_width_var = tk.StringVar(value=str(config_data.get("detail_top_image_width", 500)))
        image_height_var = tk.StringVar(value=str(config_data.get("detail_top_image_height", 500)))
        
        tk.Label(image_size_frame, text="가로:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(image_size_frame, textvariable=image_width_var, font=("맑은 고딕", 9), width=8).pack(side="left", padx=2)
        tk.Label(image_size_frame, text="px", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=2)
        
        tk.Label(image_size_frame, text="세로:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(image_size_frame, textvariable=image_height_var, font=("맑은 고딕", 9), width=8).pack(side="left", padx=2)
        tk.Label(image_size_frame, text="px", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=2)
        
        # 상품명 텍스트
        tk.Label(detail_top_frame, text="상품명 표시 텍스트:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=(10, 5))
        
        product_name_text_var = tk.StringVar(value=config_data.get("detail_top_product_name_text", "[상품명: {상품명}]"))
        product_name_entry = tk.Entry(detail_top_frame, textvariable=product_name_text_var,
                                     font=("맑은 고딕", 9),
                                     width=80)
        product_name_entry.pack(fill="x", padx=10, pady=5)
        
        # 상품명 스타일 설정
        product_name_style_frame = tk.Frame(detail_top_frame, bg="#FFFFFF")
        product_name_style_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(product_name_style_frame, text="상품명 스타일:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333", width=15).pack(side="left", padx=5)
        
        product_name_color_var = tk.StringVar(value=config_data.get("detail_top_product_name_color", "blue"))
        product_name_font_size_var = tk.StringVar(value=str(config_data.get("detail_top_product_name_font_size", 10)))
        
        tk.Label(product_name_style_frame, text="컬러:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(product_name_style_frame, textvariable=product_name_color_var, font=("맑은 고딕", 9), width=10).pack(side="left", padx=2)
        
        tk.Label(product_name_style_frame, text="폰트 사이즈:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(product_name_style_frame, textvariable=product_name_font_size_var, font=("맑은 고딕", 9), width=8).pack(side="left", padx=2)
        tk.Label(product_name_style_frame, text="px", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=2)
        
        tk.Label(detail_top_frame, 
                text="※ {상품명} 부분이 '등록 솔루션 엑셀'의 '상품명' 컬럼 값(원본상품명)으로 치환됩니다.",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#999",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        # 필독 문구
        tk.Label(detail_top_frame, text="필독 문구:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=(10, 5))
        
        notice_text_var = tk.StringVar(value=config_data.get("detail_top_notice_text", "[필독] 제품명 및 상세설명에 기재된 '본품'만 발송됩니다."))
        notice_entry = tk.Entry(detail_top_frame, textvariable=notice_text_var,
                               font=("맑은 고딕", 9),
                               width=80)
        notice_entry.pack(fill="x", padx=10, pady=5)
        
        # 필독 문구 스타일 설정
        notice_style_frame = tk.Frame(detail_top_frame, bg="#FFFFFF")
        notice_style_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(notice_style_frame, text="필독 문구 스타일:",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333", width=15).pack(side="left", padx=5)
        
        notice_bg_color_var = tk.StringVar(value=config_data.get("detail_top_notice_bg_color", "yellow"))
        notice_padding_var = tk.StringVar(value=config_data.get("detail_top_notice_padding", "2px 5px"))
        
        tk.Label(notice_style_frame, text="배경 컬러:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(notice_style_frame, textvariable=notice_bg_color_var, font=("맑은 고딕", 9), width=10).pack(side="left", padx=2)
        
        tk.Label(notice_style_frame, text="패딩:", font=("맑은 고딕", 9), bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        tk.Entry(notice_style_frame, textvariable=notice_padding_var, font=("맑은 고딕", 9), width=12).pack(side="left", padx=2)
        tk.Label(notice_style_frame, text="(예: 2px 5px)", font=("맑은 고딕", 8), bg="#FFFFFF", fg="#999").pack(side="left", padx=2)
        
        # =====================================================================
        # 4. 상세정보 하단 설정
        # =====================================================================
        detail_bottom_frame = tk.LabelFrame(scrollable_frame,
                                           text="🖼️ 상세정보 하단 이미지 설정",
                                           font=("맑은 고딕", 11, "bold"),
                                           bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        detail_bottom_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(detail_bottom_frame, 
                text="하단 이미지 URL 목록 (랜덤으로 1개 선택됩니다):",
                font=("맑은 고딕", 9, "bold"),
                bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10, pady=(10, 5))
        
        # 이미지 URL 목록 (Text 위젯)
        url_list_frame = tk.Frame(detail_bottom_frame, bg="#FFFFFF")
        url_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        url_list_text = tk.Text(url_list_frame, height=10, font=("Consolas", 9),
                               bg="#F8F9FA", fg="#333", wrap=tk.WORD)
        url_list_text.pack(side="left", fill="both", expand=True)
        
        url_scrollbar = ttk.Scrollbar(url_list_frame, orient="vertical", command=url_list_text.yview)
        url_list_text.configure(yscrollcommand=url_scrollbar.set)
        url_scrollbar.pack(side="right", fill="y")
        
        # 기존 URL 목록 불러오기
        bottom_image_urls = config_data.get("detail_bottom_image_urls", [])
        if bottom_image_urls:
            url_list_text.insert("1.0", "\n".join(bottom_image_urls))
        else:
            # 기본값
            default_urls = [
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_0.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_1.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_2.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_3.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_4.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_5.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_6.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_7.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_8.jpg',
                'https://ai.esmplus.com/kohaz94/detailfolder/exchange_9.jpg'
            ]
            url_list_text.insert("1.0", "\n".join(default_urls))
        
        tk.Label(detail_bottom_frame, 
                text="※ 한 줄에 하나씩 URL을 입력하세요. (날짜별/마켓별로 변경 가능)",
                font=("맑은 고딕", 8),
                bg="#FFFFFF", fg="#999",
                justify="left").pack(anchor="w", padx=20, pady=2)
        
        # 저장 버튼
        def save_advanced_settings():
            # 배송비 계산 방식
            config_data["shipping_method"] = shipping_method_var.get()
            
            # 옵션금액 규칙
            config_data["option_price_rule"] = option_price_rule_var.get()
            
            # 상단 설정
            config_data["detail_top_image_width"] = int(image_width_var.get()) if image_width_var.get().strip().isdigit() else 500
            config_data["detail_top_image_height"] = int(image_height_var.get()) if image_height_var.get().strip().isdigit() else 500
            config_data["detail_top_product_name_text"] = product_name_text_var.get().strip()
            config_data["detail_top_product_name_color"] = product_name_color_var.get().strip()
            config_data["detail_top_product_name_font_size"] = int(product_name_font_size_var.get()) if product_name_font_size_var.get().strip().isdigit() else 10
            config_data["detail_top_notice_text"] = notice_text_var.get().strip()
            config_data["detail_top_notice_bg_color"] = notice_bg_color_var.get().strip()
            config_data["detail_top_notice_padding"] = notice_padding_var.get().strip()
            
            # 하단 이미지 URL 목록
            url_content = url_list_text.get("1.0", tk.END).strip()
            url_list = [url.strip() for url in url_content.split("\n") if url.strip()]
            config_data["detail_bottom_image_urls"] = url_list
            
            self.config_manager.save_solution_config(self.solution_name, config_data)
            
            self._log(f"고급 설정 저장 완료: {self.solution_name}")
            messagebox.showinfo("완료", "고급 설정이 저장되었습니다.")
            popup.destroy()
        
        btn_frame = tk.Frame(popup, bg="#FFFFFF")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(btn_frame, text="💾 저장", command=save_advanced_settings,
                 bg="#3498db", fg="white",
                 font=("맑은 고딕", 10, "bold"),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)
        
        tk.Button(btn_frame, text="취소", command=popup.destroy,
                 bg="#95a5a6", fg="white",
                 font=("맑은 고딕", 10),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)
    
    def _load_last_settings(self):
        """마지막 설정 불러오기"""
        config = self.config_data
        if "last_solution" in config:
            self.solution_var.set(config["last_solution"])
            self._on_solution_selected()
    
    def _execute_mapping(self):
        """매핑 실행"""
        if self.solution_df is None:
            messagebox.showwarning("경고", "등록 솔루션 엑셀을 불러오세요.")
            return
        
        if self.processed_df is None:
            messagebox.showwarning("경고", "가공된 엑셀을 불러오세요.")
            return
        
        if not self.solution_instance:
            messagebox.showwarning("경고", "솔루션을 선택해주세요.")
            return
        
        try:
            self._log("매핑 실행 시작...")
            self.status_label.config(text="매핑 중...", fg="#f39c12")
            self.update_idletasks()
            
            # 매핑 정보 수집
            column_mapping = {}
            default_values = {}
            
            for proc_col, vars_dict in self.mapping_vars.items():
                sol_col = vars_dict["solution_col"].get()
                default_val = vars_dict["default_value"].get().strip()
                
                if sol_col and sol_col != "(매핑 안함)":
                    column_mapping[proc_col] = sol_col
                    if default_val:
                        default_values[sol_col] = default_val
            
            # 기본 매핑 규칙 추가
            default_mapping = self.solution_instance.get_default_mapping()
            for proc_col, sol_col in default_mapping.items():
                if proc_col not in column_mapping:
                    column_mapping[proc_col] = sol_col
            
            if not column_mapping:
                messagebox.showwarning("경고", "매핑할 컬럼이 없습니다.")
                return
            
            # 매핑 정보 로그 출력
            self._log("=" * 60)
            self._log("📋 매핑 정보")
            self._log("=" * 60)
            self._log(f"등록 솔루션 엑셀: {len(self.solution_df)}행")
            self._log(f"가공 엑셀: {len(self.processed_df)}행")
            self._log(f"매핑된 컬럼 수: {len(column_mapping)}개")
            self._log("")
            self._log("📌 컬럼 매핑:")
            for proc_col, sol_col in sorted(column_mapping.items()):
                default_val_str = ""
                if sol_col in default_values:
                    default_val_str = f" (기본값: {default_values[sol_col]})"
                self._log(f"  • [{proc_col}] → [{sol_col}]{default_val_str}")
            
            if default_values:
                self._log("")
                self._log("💡 기본값 설정:")
                for sol_col, default_val in default_values.items():
                    self._log(f"  • [{sol_col}]: {default_val}")
            
            self._log("=" * 60)
            
            # 마켓판매가격 및 판매자 부담 할인 매핑 확인 (다팔자 솔루션의 경우)
            if self.solution_name == "다팔자":
                # '가격' 컬럼에 매핑된 컬럼 확인
                price_mapped = False
                for proc_col, sol_col in column_mapping.items():
                    if sol_col == "가격":
                        price_mapped = True
                        # 가공 엑셀에 해당 컬럼이 있는지 확인
                        if proc_col in self.processed_df.columns:
                            # 값이 있는지 확인
                            has_values = self.processed_df[proc_col].notna().any()
                            if not has_values:
                                response = messagebox.askyesno(
                                    "경고",
                                    f"가공된 엑셀의 '{proc_col}' 컬럼에 값이 없습니다.\n\n"
                                    f"마켓판매가격을 미리 계산하여 가공된 엑셀에 설정해주세요.\n\n"
                                    f"계속 진행하시겠습니까?"
                                )
                                if not response:
                                    return
                        break
                
                if not price_mapped:
                    response = messagebox.askyesno(
                        "경고",
                        "다팔자 솔루션의 '가격' 컬럼에 매핑이 설정되지 않았습니다.\n\n"
                        "가공된 엑셀의 '마켓판매가격' 컬럼을 '가격'에 매핑하거나,\n"
                        "가공된 엑셀에 판매가격을 미리 계산하여 설정해주세요.\n\n"
                        "계속 진행하시겠습니까?"
                    )
                    if not response:
                        return
                
                # '판매자 부담 할인' 매핑 확인
                discount_mapped = False
                for proc_col, sol_col in column_mapping.items():
                    if sol_col == "판매자 부담 할인":
                        discount_mapped = True
                        break
                
                if not discount_mapped:
                    self._log("'판매자 부담 할인' 컬럼이 매핑되지 않았습니다. (선택사항)", "WARNING")
            
            # 결과 데이터프레임 생성 (솔루션 엑셀 구조 복사)
            result_df = self.solution_df.copy()
            
            # 설정 가져오기 (마켓 정보 포함)
            config = self.config_manager.get_solution_config(self.solution_name, self.detected_market)
            config["default_values"] = default_values
            
            # 다팔자, 이셀러스 솔루션의 경우 감지된 마켓 정보를 config에 추가
            if self.solution_name in ["다팔자", "이셀러스"] and self.detected_market:
                config["detected_market"] = self.detected_market
            
            # 이셀러스 솔루션의 경우 가공 파일 경로를 config에 추가 (폴더명 추출용)
            if self.solution_name == "이셀러스" and self.processed_file_path:
                config["processed_file_path"] = self.processed_file_path
            
            # 솔루션별 매핑 규칙 적용
            self._log("")
            self._log("🔄 매핑 적용 중...")
            result_df = self.solution_instance.apply_mapping(
                result_df, self.processed_df, column_mapping, config
            )
            
            # 매핑 결과 통계 수집 (팝업 창 표시용)
            matched_codes = None
            unmatched_result = None
            unmatched_processed = None
            
            # 식별자 기준 매핑 통계 (다팔자: 상품코드, 이셀러스: 판매자 관리코드 = 상품코드)
            identifier_col = None
            processed_identifier_col = None
            if self.solution_name == "이셀러스":
                identifier_col = "판매자 관리코드"
                processed_identifier_col = "상품코드"
            else:
                identifier_col = "상품코드"
                processed_identifier_col = "상품코드"
            
            if processed_identifier_col in self.processed_df.columns and identifier_col in result_df.columns:
                processed_codes = set(self.processed_df[processed_identifier_col].dropna().astype(str))
                result_codes = set(result_df[identifier_col].dropna().astype(str))
                matched_codes = processed_codes & result_codes
                unmatched_result = result_codes - processed_codes
                unmatched_processed = processed_codes - result_codes
                
                identifier_name = "판매자 관리코드" if self.solution_name == "이셀러스" else "상품코드"
                self._log(f"매칭된 {identifier_name}: {len(matched_codes)}개")
                if unmatched_result:
                    self._log(f"등록 솔루션에만 있는 {identifier_name}: {len(unmatched_result)}개", "WARNING")
                if unmatched_processed:
                    self._log(f"가공 엑셀에만 있는 {identifier_name}: {len(unmatched_processed)}개", "WARNING")
            
            # 컬럼별 매핑 통계
            self._log("")
            self._log("📈 컬럼별 매핑 결과:")
            for proc_col, sol_col in sorted(column_mapping.items()):
                if sol_col in result_df.columns:
                    # 매핑된 값의 개수 (빈 값 제외)
                    mapped_count = result_df[sol_col].notna().sum()
                    total_count = len(result_df)
                    mapped_pct = (mapped_count / total_count * 100) if total_count > 0 else 0
                    
                    # 기본값으로 채워진 경우 확인
                    default_filled_count = 0
                    if sol_col in default_values:
                        default_val = default_values[sol_col]
                        default_filled_count = (result_df[sol_col] == default_val).sum()
                    
                    status_icon = "✅" if mapped_count > 0 else "❌"
                    log_msg = f"  {status_icon} [{proc_col}] → [{sol_col}]: {mapped_count}/{total_count}행 매핑됨 ({mapped_pct:.1f}%)"
                    if default_filled_count > 0:
                        log_msg += f" (기본값 사용: {default_filled_count}행)"
                    self._log(log_msg)
            
            # 솔루션별 특화 규칙 적용
            # 원본 solution_df를 전달하여 원본 상품명 등을 가져올 수 있도록 함
            self._log("")
            self._log("🔄 솔루션별 특화 규칙 적용 중...")
            result_df = self.solution_instance.apply_solution_specific_rules(
                result_df, self.processed_df, config, original_solution_df=self.solution_df
            )
            
            # 빈 값에 기본값 채우기
            if default_values:
                self._log("")
                self._log("💾 기본값 적용 중...")
                for sol_col, default_val in default_values.items():
                    if sol_col in result_df.columns:
                        before_count = result_df[sol_col].isna().sum()
                        result_df[sol_col] = result_df[sol_col].fillna(default_val)
                        after_count = result_df[sol_col].isna().sum()
                        filled_count = before_count - after_count
                        if filled_count > 0:
                            self._log(f"  • [{sol_col}]: {filled_count}행에 기본값 '{default_val}' 적용")
            
            self._log("=" * 60)
            
            # 결과 저장
            # 다팔자 솔루션의 경우 원본 파일에 저장
            if self.solution_name == "다팔자" and self.solution_file_path:
                output_path = self.solution_file_path
                # 백업 생성 (원본 파일이 존재하는 경우)
                backup_path = output_path.replace(".xlsx", "_backup.xlsx")
                if os.path.exists(output_path):
                    import shutil
                    try:
                        shutil.copy2(output_path, backup_path)
                        self._log(f"원본 파일 백업 생성: {backup_path}")
                    except Exception as backup_error:
                        self._log(f"백업 생성 실패: {str(backup_error)}", "ERROR")
                if output_path:
                    result_df.to_excel(output_path, index=False)
            elif self.solution_name == "이셀러스":
                # 이셀러스: 파일명 변경 및 확장정보 시트 포함
                if self.processed_file_path and self.solution_file_path:
                    # 가공 엑셀 파일명에서 확장자 제거
                    processed_filename = os.path.splitext(os.path.basename(self.processed_file_path))[0]
                    # 등록 솔루션 엑셀의 확장자 가져오기
                    solution_ext = os.path.splitext(self.solution_file_path)[1]
                    # 파일명: '이셀완료' + 가공 엑셀 파일명
                    output_filename = f"이셀완료{processed_filename}{solution_ext}"
                    output_dir = os.path.dirname(self.processed_file_path)
                    
                    # 솔루션 엑셀 파일명에 '_기본카테고리'가 있는지 확인
                    solution_filename = os.path.basename(self.solution_file_path)
                    has_basic_category = '_기본카테고리' in solution_filename
                    
                    # 날짜 추출 (파일명 앞부분에서 YYYYMMDD 형식 추출)
                    import re
                    date_match = re.match(r'^(\d{8})', processed_filename)
                    if date_match:
                        date_str = date_match.group(1)
                        # 기본카테고리인 경우 별도 폴더 생성
                        if has_basic_category:
                            subfolder_name = f"{date_str}이셀완료_기본카테고리"
                        else:
                            subfolder_name = f"{date_str}이셀완료"
                        output_dir = os.path.join(output_dir, subfolder_name)
                        # 폴더가 없으면 생성
                        if not os.path.exists(output_dir):
                            os.makedirs(output_dir)
                    
                    output_path = os.path.join(output_dir, output_filename)
                else:
                    output_path = filedialog.asksaveasfilename(
                        title="매핑 결과 저장",
                        defaultextension=".xls",
                        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
                    )
                
                if output_path:
                    # 확장정보 시트 읽기 및 마켓별 처리
                    extension_df = None
                    if self.solution_file_path:
                        try:
                            extension_df = read_excel_with_fallback(self.solution_file_path, sheet_name="확장정보")
                            
                            # 마켓별 확장정보 처리
                            if extension_df is not None and self.detected_market:
                                extension_df = self._process_extension_sheet_by_market(extension_df, self.detected_market)
                        except Exception as e:
                            self._log(f"확장정보 시트 읽기 실패: {str(e)}", "WARNING")
                            extension_df = None
                    
                    # 기본정보 시트 전처리
                    # 1. BY열(Unnamed: 76) 제거 또는 빈 값으로 설정
                    if 'Unnamed: 76' in result_df.columns:
                        result_df['Unnamed: 76'] = ''
                        self._log("BY열(Unnamed: 76)을 빈 열로 설정했습니다.")
                    
                    # 2. 2행 빈 행 추가 (이셀러스 형식: 1행=헤더, 2행=빈행, 3행부터=데이터)
                    # 빈 행을 DataFrame으로 생성하여 앞에 추가
                    empty_row = pd.DataFrame([[''] * len(result_df.columns)], columns=result_df.columns)
                    result_df_with_empty_row = pd.concat([empty_row, result_df], ignore_index=True)
                    
                    # 확장정보: 2행 빈 행 추가 (기본정보와 동일한 형식)
                    extension_df_with_empty_row = None
                    if extension_df is not None and len(extension_df) > 0:
                        extension_empty_row = pd.DataFrame([[''] * len(extension_df.columns)], columns=extension_df.columns)
                        extension_df_with_empty_row = pd.concat([extension_empty_row, extension_df], ignore_index=True)
                    
                    # ExcelWriter로 여러 시트 저장
                    file_ext = os.path.splitext(output_path)[1].lower()
                    if file_ext == '.xls':
                        # .xls 파일은 xlwt 사용 (또는 .xlsx로 저장)
                        # xlwt는 복잡하므로 .xlsx로 저장 후 사용자에게 알림
                        output_path_xlsx = output_path.replace('.xls', '.xlsx')
                        with pd.ExcelWriter(output_path_xlsx, engine='openpyxl') as writer:
                            result_df_with_empty_row.to_excel(writer, sheet_name='기본정보', index=False)
                            if extension_df_with_empty_row is not None:
                                extension_df_with_empty_row.to_excel(writer, sheet_name='확장정보', index=False)
                        self._log(f"⚠️ .xls 형식은 .xlsx로 저장되었습니다: {output_path_xlsx}", "WARNING")
                        output_path = output_path_xlsx
                    else:
                        # .xlsx 파일
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            result_df_with_empty_row.to_excel(writer, sheet_name='기본정보', index=False)
                            if extension_df_with_empty_row is not None:
                                extension_df_with_empty_row.to_excel(writer, sheet_name='확장정보', index=False)
                    
                    # 이셀러스 저장 후 로그 및 팝업 표시
                    if output_path:
                        self._log("")
                        self._log("=" * 60)
                        self._log("✅ 매핑 완료!")
                        self._log(f"저장 위치: {output_path}")
                        self._log(f"결과 파일: {len(result_df)}행, {len(result_df.columns)}컬럼")
                        if extension_df is not None:
                            self._log(f"확장정보 시트: {len(extension_df)}행, {len(extension_df.columns)}컬럼")
                        self._log("=" * 60)
                        self.status_label.config(text="완료", fg="#27ae60")
                        
                        # 매핑 결과 팝업 창 표시 (저장 후)
                        self._show_mapping_result_popup(
                            result_df, column_mapping, default_values,
                            matched_codes, unmatched_result, unmatched_processed, output_path
                        )
            else:
                output_path = filedialog.asksaveasfilename(
                    title="매핑 결과 저장",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                if output_path:
                    result_df.to_excel(output_path, index=False)
                self._log("")
                self._log("=" * 60)
                self._log("✅ 매핑 완료!")
                self._log(f"저장 위치: {output_path}")
                self._log(f"결과 파일: {len(result_df)}행, {len(result_df.columns)}컬럼")
                self._log("=" * 60)
                self.status_label.config(text="완료", fg="#27ae60")
                
                # 매핑 결과 팝업 창 표시 (저장 후)
                self._show_mapping_result_popup(
                    result_df, column_mapping, default_values,
                    matched_codes, unmatched_result, unmatched_processed, output_path
                )
            
        except Exception as e:
            error_msg = f"매핑 실행 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            self.status_label.config(text="오류", fg="#e74c3c")
            messagebox.showerror("오류", error_msg)
    
    def _create_dafalza_csv(self):
        """다팔자 등록용 CSV 생성 (여러 엑셀 파일 일괄 처리)"""
        try:
            # 여러 엑셀 파일 선택
            file_paths = filedialog.askopenfilenames(
                title="다팔자 등록용 CSV 생성을 위한 엑셀 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            
            if not file_paths:
                return
            
            self._log(f"선택된 파일 수: {len(file_paths)}개")
            self.status_label.config(text="CSV 생성 중...", fg="#f39c12")
            self.update_idletasks()
            
            success_count = 0
            error_count = 0
            error_files = []
            
            # 통합 CSV 생성을 위한 데이터 수집
            all_first_columns = []  # 모든 파일의 첫 번째 열 데이터
            all_excel_data = []  # 전체 엑셀 병합을 위한 데이터
            
            for file_path in file_paths:
                try:
                    # 엑셀 파일 읽기
                    df = read_excel_with_fallback(file_path, header=None)
                    
                    if df.empty:
                        self._log(f"⚠️ 빈 파일: {os.path.basename(file_path)}", "WARNING")
                        error_count += 1
                        error_files.append(os.path.basename(file_path))
                        continue
                    
                    # 1행 헤더 삭제
                    df = df.iloc[1:]
                    
                    # 1열만 남기고 2열부터 모두 삭제
                    if df.shape[1] > 0:
                        df_first_col = df.iloc[:, [0]]  # 첫 번째 열만 선택
                    else:
                        self._log(f"⚠️ 열이 없는 파일: {os.path.basename(file_path)}", "WARNING")
                        error_count += 1
                        error_files.append(os.path.basename(file_path))
                        continue
                    
                    # CSV 파일 경로 생성 (원본 파일명과 동일, 확장자만 .csv)
                    base_path = Path(file_path)
                    csv_path = base_path.with_suffix('.csv')
                    
                    # CSV로 저장 (인덱스 없이, 헤더 없이)
                    df_first_col.to_csv(csv_path, index=False, header=False, encoding='utf-8-sig')
                    
                    success_count += 1
                    self._log(f"✅ 생성 완료: {csv_path.name}")
                    
                    # 통합 CSV 생성을 위한 데이터 수집 (첫 번째 열)
                    first_col_values = df_first_col.iloc[:, 0].dropna().astype(str).tolist()
                    all_first_columns.extend(first_col_values)
                    
                    # 전체 엑셀 병합을 위한 데이터 수집
                    all_excel_data.append(df)
                    
                except Exception as e:
                    error_count += 1
                    error_files.append(os.path.basename(file_path))
                    self._log(f"❌ 오류 ({os.path.basename(file_path)}): {str(e)}", "ERROR")
            
            # 통합 CSV 생성 (상품코드 기준 중복 제거)
            if all_first_columns and success_count > 0:
                try:
                    self._log("")
                    self._log("통합 CSV 생성 중...")
                    
                    # 상품코드 기준 중복 제거 (순서 유지)
                    seen_codes = set()
                    unique_codes = []
                    for code in all_first_columns:
                        code_str = str(code).strip()
                        if code_str and code_str not in seen_codes:
                            seen_codes.add(code_str)
                            unique_codes.append(code_str)
                    
                    # 통합 CSV 저장
                    # 첫 번째 파일과 같은 디렉토리에 저장
                    if file_paths:
                        first_file_dir = Path(file_paths[0]).parent
                        merged_csv_path = first_file_dir / "통합_상품코드_중복제거.csv"
                        
                        # DataFrame으로 변환하여 저장
                        merged_df = pd.DataFrame(unique_codes, columns=[0])
                        merged_df.to_csv(merged_csv_path, index=False, header=False, encoding='utf-8-sig')
                        
                        self._log(f"✅ 통합 CSV 생성 완료: {merged_csv_path.name} ({len(unique_codes)}개 상품코드, 중복 제거: {len(all_first_columns) - len(unique_codes)}개)")
                    
                except Exception as e:
                    self._log(f"⚠️ 통합 CSV 생성 실패: {str(e)}", "WARNING")
            
            # 전체 엑셀 병합 파일 생성
            if all_excel_data and success_count > 0:
                try:
                    self._log("")
                    self._log("전체 엑셀 병합 파일 생성 중...")
                    
                    # 모든 엑셀 데이터 병합
                    merged_excel_df = pd.concat(all_excel_data, ignore_index=True)
                    
                    # 상품코드 기준 중복 제거 (첫 번째 열이 상품코드라고 가정)
                    if merged_excel_df.shape[1] > 0:
                        # 첫 번째 열 기준으로 중복 제거
                        merged_excel_df = merged_excel_df.drop_duplicates(subset=[0], keep='first')
                        
                        # 병합 파일 저장
                        if file_paths:
                            first_file_dir = Path(file_paths[0]).parent
                            merged_excel_path = first_file_dir / "통합_전체엑셀_병합.xlsx"
                            
                            merged_excel_df.to_excel(merged_excel_path, index=False, header=False, engine='openpyxl')
                            
                            original_count = sum(len(df) for df in all_excel_data)
                            merged_count = len(merged_excel_df)
                            removed_count = original_count - merged_count
                            
                            self._log(f"✅ 전체 엑셀 병합 파일 생성 완료: {merged_excel_path.name}")
                            self._log(f"   원본 행 수: {original_count}개, 병합 후: {merged_count}개, 중복 제거: {removed_count}개")
                    
                except Exception as e:
                    self._log(f"⚠️ 전체 엑셀 병합 파일 생성 실패: {str(e)}", "WARNING")
                    import traceback
                    self._log(traceback.format_exc(), "ERROR")
            
            # 결과 메시지
            result_msg = f"CSV 생성 완료!\n\n"
            result_msg += f"✅ 성공: {success_count}개\n"
            if error_count > 0:
                result_msg += f"❌ 실패: {error_count}개\n"
                result_msg += f"실패 파일: {', '.join(error_files[:5])}"
                if len(error_files) > 5:
                    result_msg += f" 외 {len(error_files) - 5}개"
            
            if error_count == 0:
                messagebox.showinfo("완료", result_msg)
            else:
                messagebox.showwarning("완료 (일부 실패)", result_msg)
            
            self.status_label.config(text="CSV 생성 완료", fg="#27ae60")
            self._log(f"총 {len(file_paths)}개 파일 처리 완료 (성공: {success_count}, 실패: {error_count})")
            
        except Exception as e:
            error_msg = f"CSV 생성 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            self.status_label.config(text="오류", fg="#e74c3c")
            messagebox.showerror("오류", error_msg)
            import traceback
            self._log(traceback.format_exc(), "ERROR")
    
    def _create_esellers_csv(self):
        """이셀러스 등록용 CSV 생성 (여러 엑셀 파일 일괄 처리)
        
        용도: 각 도매처에서 상품코드 기반으로 이셀러스 양식 엑셀을 받기 위한 CSV 생성
        
        CSV 구조:
        - 각 엑셀 파일마다 1개의 열을 차지 (10개 파일 → 10열)
        - 1행: 각 엑셀 파일명에서 추출한 날짜_마켓코드 값들 (헤더, 예: "20260117_스스A1-0")
        - 2행부터: 각 엑셀의 '상품코드' 열 내용을 세로로 채움
        
        예시 (3개 파일 선택 시):
        파일1: 20260117_스스A1-0, 파일2: 20260118_옥션A1-0, 파일3: 20260119_지마켓A1-0
        
        CSV 결과:
        20260117_스스A1-0,20260118_옥션A1-0,20260119_지마켓A1-0
        상품코드1,상품코드1,상품코드1
        상품코드2,상품코드2,상품코드2
        ...
        """
        # #region agent log
        import json
        log_path = r"c:\Users\kohaz\Desktop\Python\.cursor\debug.log"
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "pre-fix",
                    "hypothesisId": "A",
                    "location": "main.py:_create_esellers_csv:entry",
                    "message": "이셀러스 CSV 생성 시작",
                    "data": {},
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + "\n")
        except: pass
        # #endregion
        
        try:
            # 여러 엑셀 파일 선택
            file_paths = filedialog.askopenfilenames(
                title="이셀러스 등록용 CSV 생성을 위한 엑셀 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "A",
                        "location": "main.py:_create_esellers_csv:files_selected",
                        "message": "선택된 파일 수",
                        "data": {"file_count": len(file_paths) if file_paths else 0},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            if not file_paths:
                return
            
            self._log(f"선택된 파일 수: {len(file_paths)}개")
            self.status_label.config(text="CSV 생성 중...", fg="#f39c12")
            self.update_idletasks()
            
            # 파일명에서 날짜_마켓코드 추출 함수 (이셀러스의 _extract_date_market_code와 동일한 로직)
            def extract_date_market_code(filename):
                """파일명에서 날짜_마켓코드 추출 (예: "20260117_스스A1-0")"""
                import re
                # 확장자 제거
                base_name = os.path.splitext(os.path.basename(filename))[0]
                # 언더스코어로 분리하여 첫 두 부분만 추출
                parts = base_name.split('_')
                if len(parts) >= 2:
                    # 날짜 형식 검증 (8자리 숫자)
                    if re.match(r'^\d{8}$', parts[0]):
                        return f"{parts[0]}_{parts[1]}"
                    # 날짜 형식이 아니어도 첫 두 부분 반환
                    return f"{parts[0]}_{parts[1]}"
                # 언더스코어가 없거나 1개만 있으면 첫 부분만 반환
                return parts[0] if parts else base_name[:20]
            
            # 각 파일의 상품코드 데이터 수집
            file_headers = []  # 헤더 (날짜_마켓코드)
            file_data = {}  # {날짜_마켓코드: [상품코드 리스트]}
            
            success_count = 0
            error_count = 0
            error_files = []
            
            for file_path in file_paths:
                try:
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "B",
                                "location": "main.py:_create_esellers_csv:processing_file",
                                "message": "파일 처리 시작",
                                "data": {"file_path": os.path.basename(file_path)},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    # 파일명에서 날짜_마켓코드 추출
                    header_name = extract_date_market_code(file_path)
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "B",
                                "location": "main.py:_create_esellers_csv:extracted_header",
                                "message": "추출된 헤더명",
                                "data": {"header_name": header_name},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    # 엑셀 파일 읽기
                    df = read_excel_with_fallback(file_path)
                    
                    if df.empty:
                        self._log(f"⚠️ 빈 파일: {os.path.basename(file_path)}", "WARNING")
                        error_count += 1
                        error_files.append(os.path.basename(file_path))
                        continue
                    
                    # '상품코드' 열 확인
                    if '상품코드' not in df.columns:
                        self._log(f"⚠️ '상품코드' 열이 없는 파일: {os.path.basename(file_path)}", "WARNING")
                        error_count += 1
                        error_files.append(os.path.basename(file_path))
                        continue
                    
                    # 상품코드 열 추출 (NaN 제거)
                    product_codes = df['상품코드'].dropna().astype(str).tolist()
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "C",
                                "location": "main.py:_create_esellers_csv:product_codes_extracted",
                                "message": "상품코드 추출 완료",
                                "data": {"code_count": len(product_codes), "first_5": product_codes[:5]},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    if not product_codes:
                        self._log(f"⚠️ 상품코드가 없는 파일: {os.path.basename(file_path)}", "WARNING")
                        error_count += 1
                        error_files.append(os.path.basename(file_path))
                        continue
                    
                    # 헤더와 데이터 저장
                    file_headers.append(header_name)
                    file_data[header_name] = product_codes
                    
                    success_count += 1
                    self._log(f"✅ 처리 완료: {os.path.basename(file_path)} (헤더: {header_name}, 상품코드: {len(product_codes)}개)")
                    
                except Exception as e:
                    error_count += 1
                    error_files.append(os.path.basename(file_path))
                    self._log(f"❌ 오류 ({os.path.basename(file_path)}): {str(e)}", "ERROR")
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "D",
                                "location": "main.py:_create_esellers_csv:error",
                                "message": "파일 처리 오류",
                                "data": {"file": os.path.basename(file_path), "error": str(e)},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
            
            # CSV 생성
            if file_headers and success_count > 0:
                try:
                    self._log("")
                    self._log("CSV 생성 중...")
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "E",
                                "location": "main.py:_create_esellers_csv:csv_creation_start",
                                "message": "CSV 생성 시작",
                                "data": {"header_count": len(file_headers), "headers": file_headers},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    # 최대 행 수 계산 (가장 긴 상품코드 리스트의 길이)
                    max_rows = max(len(codes) for codes in file_data.values()) if file_data else 0
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "E",
                                "location": "main.py:_create_esellers_csv:max_rows_calculated",
                                "message": "최대 행 수 계산",
                                "data": {"max_rows": max_rows},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    # CSV 데이터 구성
                    csv_rows = []
                    # 1행: 헤더 (날짜_마켓코드)
                    csv_rows.append(file_headers)
                    
                    # 2행부터: 각 열의 상품코드
                    for row_idx in range(max_rows):
                        row_data = []
                        for header in file_headers:
                            codes = file_data[header]
                            if row_idx < len(codes):
                                row_data.append(codes[row_idx])
                            else:
                                row_data.append("")  # 빈 셀
                        csv_rows.append(row_data)
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "E",
                                "location": "main.py:_create_esellers_csv:csv_rows_prepared",
                                "message": "CSV 행 데이터 준비 완료",
                                "data": {"total_rows": len(csv_rows), "first_row_sample": csv_rows[0][:3] if csv_rows else []},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                    
                    # DataFrame으로 변환
                    csv_df = pd.DataFrame(csv_rows)
                    
                    # 엑셀 파일 저장 경로
                    if file_paths:
                        first_file_dir = Path(file_paths[0]).parent
                        excel_path = first_file_dir / "이셀러스_등록용_상품코드.xlsx"
                        
                        # 엑셀로 저장 (인덱스 없이, 헤더 없이)
                        csv_df.to_excel(excel_path, index=False, header=False, engine='openpyxl')
                        
                        # #region agent log
                        try:
                            with open(log_path, 'a', encoding='utf-8') as f:
                                f.write(json.dumps({
                                    "sessionId": "debug-session",
                                    "runId": "pre-fix",
                                    "hypothesisId": "E",
                                    "location": "main.py:_create_esellers_csv:excel_saved",
                                    "message": "엑셀 파일 저장 완료",
                                    "data": {"excel_path": str(excel_path), "rows": len(csv_rows), "cols": len(file_headers)},
                                    "timestamp": int(__import__('time').time() * 1000)
                                }, ensure_ascii=False) + "\n")
                        except: pass
                        # #endregion
                        
                        # openpyxl을 사용하여 포맷팅 적용 (헤더 색상, 열 너비)
                        try:
                            from openpyxl import load_workbook
                            from openpyxl.styles import PatternFill, Font
                            
                            workbook = load_workbook(excel_path)
                            sheet = workbook.active
                            
                            # 헤더 배경색 설정 (1행)
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 진한 파란색
                            header_font = Font(bold=True, color="FFFFFF")  # 흰색 굵은 글씨
                            
                            for col_idx, header in enumerate(file_headers, start=1):
                                cell = sheet.cell(row=1, column=col_idx)
                                cell.fill = header_fill
                                cell.font = header_font
                            
                            # 열 너비 자동 조절 (1행의 각 셀 내용 길이에 따라)
                            for col_idx, header in enumerate(file_headers, start=1):
                                column_letter = sheet.cell(row=1, column=col_idx).column_letter
                                # 헤더 길이를 기준으로 너비 설정 (한글 고려하여 약간 여유 공간 추가)
                                header_length = len(str(header))
                                # 한글은 약 1.5배 정도 더 넓게 계산
                                estimated_width = max(header_length * 1.5, 15)  # 최소 15
                                sheet.column_dimensions[column_letter].width = min(estimated_width, 50)  # 최대 50
                            
                            # 추가 시트: 전체 상품코드 (중복 제거, a-z 정렬)
                            unique_codes = []  # 변수 초기화
                            try:
                                # 모든 상품코드 수집
                                all_product_codes = []
                                for codes in file_data.values():
                                    all_product_codes.extend(codes)
                                
                                # 중복 제거 및 a-z 정렬
                                unique_codes = sorted(set(all_product_codes), key=str.lower)  # 대소문자 구분 없이 정렬
                                
                                # #region agent log
                                try:
                                    with open(log_path, 'a', encoding='utf-8') as f:
                                        f.write(json.dumps({
                                            "sessionId": "debug-session",
                                            "runId": "pre-fix",
                                            "hypothesisId": "G",
                                            "location": "main.py:_create_esellers_csv:all_codes_collected",
                                            "message": "전체 상품코드 수집 및 정렬",
                                            "data": {"total_count": len(all_product_codes), "unique_count": len(unique_codes)},
                                            "timestamp": int(__import__('time').time() * 1000)
                                        }, ensure_ascii=False) + "\n")
                                except: pass
                                # #endregion
                                
                                # 새 시트 생성
                                new_sheet = workbook.create_sheet(title="전체상품코드")
                                
                                # 1행: 헤더 '전체상품코드'
                                header_cell = new_sheet.cell(row=1, column=1, value="전체상품코드")
                                header_cell.fill = header_fill
                                header_cell.font = header_font
                                
                                # 2행부터: 상품코드 리스트
                                for row_idx, code in enumerate(unique_codes, start=2):
                                    new_sheet.cell(row=row_idx, column=1, value=code)
                                
                                # 열 너비 조절
                                new_sheet.column_dimensions['A'].width = 30
                                
                                # #region agent log
                                try:
                                    with open(log_path, 'a', encoding='utf-8') as f:
                                        f.write(json.dumps({
                                            "sessionId": "debug-session",
                                            "runId": "pre-fix",
                                            "hypothesisId": "G",
                                            "location": "main.py:_create_esellers_csv:all_codes_sheet_created",
                                            "message": "전체상품코드 시트 생성 완료",
                                            "data": {"row_count": len(unique_codes) + 1},
                                            "timestamp": int(__import__('time').time() * 1000)
                                        }, ensure_ascii=False) + "\n")
                                except: pass
                                # #endregion
                                
                            except Exception as e:
                                self._log(f"⚠️ 전체상품코드 시트 생성 실패: {str(e)}", "WARNING")
                            
                            workbook.save(excel_path)
                            
                            # #region agent log
                            try:
                                with open(log_path, 'a', encoding='utf-8') as f:
                                    f.write(json.dumps({
                                        "sessionId": "debug-session",
                                        "runId": "pre-fix",
                                        "hypothesisId": "F",
                                        "location": "main.py:_create_esellers_csv:formatting_applied",
                                        "message": "포맷팅 적용 완료",
                                        "data": {"headers_count": len(file_headers)},
                                        "timestamp": int(__import__('time').time() * 1000)
                                    }, ensure_ascii=False) + "\n")
                            except: pass
                            # #endregion
                            
                        except ImportError:
                            self._log("⚠️ openpyxl이 설치되지 않아 포맷팅을 적용할 수 없습니다.", "WARNING")
                        except Exception as e:
                            self._log(f"⚠️ 포맷팅 적용 실패: {str(e)}", "WARNING")
                        
                        self._log(f"✅ 엑셀 파일 생성 완료: {excel_path.name}")
                        self._log(f"   헤더: {len(file_headers)}개, 최대 행 수: {max_rows}개")
                        
                        # 전체 상품코드 통계 출력
                        if unique_codes:
                            self._log(f"   전체상품코드 시트: {len(unique_codes)}개 (중복 제거 후)")
                    
                except Exception as e:
                    self._log(f"⚠️ CSV 생성 실패: {str(e)}", "WARNING")
                    import traceback
                    self._log(traceback.format_exc(), "ERROR")
                    
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "F",
                                "location": "main.py:_create_esellers_csv:csv_creation_error",
                                "message": "CSV 생성 오류",
                                "data": {"error": str(e)},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
            
            # 결과 메시지
            result_msg = f"CSV 생성 완료!\n\n"
            result_msg += f"✅ 성공: {success_count}개\n"
            if error_count > 0:
                result_msg += f"❌ 실패: {error_count}개\n"
                result_msg += f"실패 파일: {', '.join(error_files[:5])}"
                if len(error_files) > 5:
                    result_msg += f" 외 {len(error_files) - 5}개"
            
            if error_count == 0:
                messagebox.showinfo("완료", result_msg)
            else:
                messagebox.showwarning("완료 (일부 실패)", result_msg)
            
            self.status_label.config(text="CSV 생성 완료", fg="#27ae60")
            self._log(f"총 {len(file_paths)}개 파일 처리 완료 (성공: {success_count}, 실패: {error_count})")
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "G",
                        "location": "main.py:_create_esellers_csv:completion",
                        "message": "CSV 생성 완료",
                        "data": {"success_count": success_count, "error_count": error_count},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
        except Exception as e:
            error_msg = f"CSV 생성 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            self.status_label.config(text="오류", fg="#e74c3c")
            messagebox.showerror("오류", error_msg)
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "H",
                        "location": "main.py:_create_esellers_csv:exception",
                        "message": "전체 예외 발생",
                        "data": {"error": str(e)},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
    
    def _create_ownerclan_excel(self):
        """오너클랜 원가/배송비 엑셀 생성
        
        엑셀 파일에서 다음 컬럼만 추출:
        - 상품코드
        - 오너클랜판매가
        - 배송비
        - 배송유형
        - 최대구매수량
        - 반품배송비
        
        첫 시트의 2행이 헤더 (1행은 병합 헤더), 3행부터 실제 상품정보
        """
        # #region agent log
        import json
        log_path = r"c:\Users\kohaz\Desktop\Python\.cursor\debug.log"
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "pre-fix",
                    "hypothesisId": "A",
                    "location": "main.py:_create_ownerclan_excel:entry",
                    "message": "오너클랜 원가/배송비 엑셀 생성 시작",
                    "data": {},
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + "\n")
        except: pass
        # #endregion
        
        try:
            # 엑셀 파일 선택
            file_path = filedialog.askopenfilename(
                title="오너클랜 원가/배송비 추출을 위한 엑셀 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "A",
                        "location": "main.py:_create_ownerclan_excel:file_selected",
                        "message": "파일 선택 완료",
                        "data": {"file_path": file_path if file_path else "None"},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            if not file_path:
                return
            
            self._log(f"선택된 파일: {os.path.basename(file_path)}")
            self.status_label.config(text="엑셀 생성 중...", fg="#f39c12")
            self.update_idletasks()
            
            # 엑셀 파일 읽기 (첫 번째 시트, 헤더 없이 읽기)
            df = read_excel_with_fallback(file_path, sheet_name=0, header=None)
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "B",
                        "location": "main.py:_create_ownerclan_excel:file_loaded",
                        "message": "엑셀 파일 로드 완료",
                        "data": {"row_count": len(df), "col_count": len(df.columns), "first_row_sample": df.iloc[0].tolist()[:5] if len(df) > 0 else [], "second_row_sample": df.iloc[1].tolist()[:5] if len(df) > 1 else []},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            if df.empty:
                messagebox.showerror("오류", "엑셀 파일이 비어있습니다.")
                return
            
            # 2행이 헤더이므로, 1행(인덱스 0)은 병합 헤더로 무시하고 2행(인덱스 1)을 헤더로 사용
            if len(df) < 2:
                messagebox.showerror("오류", "엑셀 파일에 헤더 행(2행)이 없습니다.")
                return
            
            # 2행(인덱스 1)을 헤더로 설정하고, 1행(인덱스 0)과 2행(인덱스 1) 제거
            # 컬럼명을 문자열로 변환하고 공백 제거
            df.columns = [str(col).strip() if pd.notna(col) else f"Unnamed_{i}" for i, col in enumerate(df.iloc[1])]
            df = df.iloc[2:].reset_index(drop=True)  # 1행과 2행 제거하고 인덱스 리셋 (3행부터 데이터)
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "B",
                        "location": "main.py:_create_ownerclan_excel:header_set",
                        "message": "헤더 설정 완료",
                        "data": {"columns": list(df.columns)[:10]},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            # 필요한 컬럼만 추출
            required_columns = ['상품코드', '오너클랜판매가', '배송비', '배송유형', '최대구매수량', '반품배송비']
            
            # 컬럼명 확인 (대소문자, 공백 등 고려)
            available_columns = [str(col).strip() for col in df.columns]  # 모든 컬럼명을 문자열로 변환하고 공백 제거
            selected_columns = []
            missing_columns = []
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "C1",
                        "location": "main.py:_create_ownerclan_excel:columns_normalized",
                        "message": "컬럼명 정규화 완료",
                        "data": {"available_columns": available_columns[:20], "required_columns": required_columns},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            for req_col in required_columns:
                req_col_normalized = req_col.strip()
                found = False
                matched_col = None
                
                # 1. 정확한 매칭 시도 (대소문자 구분)
                for avail_col in available_columns:
                    if avail_col == req_col_normalized:
                        matched_col = avail_col
                        found = True
                        break
                
                # 2. 공백 제거 후 매칭 시도
                if not found:
                    req_col_no_space = req_col_normalized.replace(' ', '').replace('\t', '')
                    for avail_col in available_columns:
                        avail_col_no_space = avail_col.replace(' ', '').replace('\t', '')
                        if avail_col_no_space == req_col_no_space:
                            matched_col = avail_col
                            found = True
                            break
                
                # 3. 부분 일치 시도 (공백 무시, 대소문자 무시)
                if not found:
                    req_col_lower = req_col_normalized.lower().replace(' ', '').replace('\t', '')
                    for avail_col in available_columns:
                        avail_col_lower = str(avail_col).lower().replace(' ', '').replace('\t', '')
                        if req_col_lower in avail_col_lower or avail_col_lower in req_col_lower:
                            matched_col = avail_col
                            found = True
                            break
                
                if found and matched_col:
                    selected_columns.append(matched_col)
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "C2",
                                "location": "main.py:_create_ownerclan_excel:column_matched",
                                "message": "컬럼 매칭 성공",
                                "data": {"required": req_col, "matched": matched_col},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
                else:
                    missing_columns.append(req_col)
                    # #region agent log
                    try:
                        with open(log_path, 'a', encoding='utf-8') as f:
                            f.write(json.dumps({
                                "sessionId": "debug-session",
                                "runId": "pre-fix",
                                "hypothesisId": "C3",
                                "location": "main.py:_create_ownerclan_excel:column_not_found",
                                "message": "컬럼 찾기 실패",
                                "data": {"required": req_col, "available_sample": available_columns[:5]},
                                "timestamp": int(__import__('time').time() * 1000)
                            }, ensure_ascii=False) + "\n")
                    except: pass
                    # #endregion
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "C",
                        "location": "main.py:_create_ownerclan_excel:columns_found",
                        "message": "컬럼 찾기 완료",
                        "data": {"selected": selected_columns, "missing": missing_columns},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            if missing_columns:
                # 모든 컬럼명을 문자열로 변환 (정수형 컬럼명 처리)
                missing_str = ', '.join(str(col) for col in missing_columns)
                available_str = ', '.join(str(col) for col in available_columns[:20])
                error_msg = f"다음 컬럼을 찾을 수 없습니다:\n{missing_str}\n\n사용 가능한 컬럼:\n{available_str}"
                messagebox.showerror("오류", error_msg)
                return
            
            # 필요한 컬럼만 추출
            result_df = df[selected_columns].copy()
            
            # 컬럼명을 원래 이름으로 변경 (정확한 매칭을 위해)
            column_mapping = {}
            for i, req_col in enumerate(required_columns):
                if selected_columns[i] != req_col:
                    column_mapping[selected_columns[i]] = req_col
            
            if column_mapping:
                result_df = result_df.rename(columns=column_mapping)
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "D",
                        "location": "main.py:_create_ownerclan_excel:columns_extracted",
                        "message": "컬럼 추출 완료",
                        "data": {"row_count": len(result_df), "columns": list(result_df.columns)},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            # 4개의 추가 컬럼 생성 및 계산
            # 배송비를 숫자로 변환 (NaN 처리)
            result_df['배송비_숫자'] = pd.to_numeric(result_df['배송비'], errors='coerce').fillna(0)
            result_df['오너클랜판매가_숫자'] = pd.to_numeric(result_df['오너클랜판매가'], errors='coerce').fillna(0)
            
            # 3000원가, 3000배송비 계산
            def calc_3000_price(row):
                shipping = row['배송비_숫자']
                price = row['오너클랜판매가_숫자']
                
                if shipping >= 3000:
                    return price + shipping - 3000
                else:
                    return price
            
            def calc_3000_shipping(row):
                shipping = row['배송비_숫자']
                if shipping >= 3000:
                    return 3000
                elif shipping > 0:
                    return 3000
                else:
                    return 0
            
            result_df['3000원가'] = result_df.apply(calc_3000_price, axis=1)
            result_df['3000배송비'] = result_df.apply(calc_3000_shipping, axis=1)
            
            # 무배원가, 무배배송비 계산
            def calc_free_price(row):
                shipping = row['배송비_숫자']
                price = row['오너클랜판매가_숫자']
                
                if shipping > 0:
                    return price + shipping
                else:
                    return price
            
            def calc_free_shipping(row):
                shipping = row['배송비_숫자']
                
                if shipping > 3000:
                    return shipping
                elif shipping > 0:  # 3000 > 배송비 > 0
                    return 3000
                else:  # 배송비 = 0
                    return shipping  # 0
            
            result_df['무배원가'] = result_df.apply(calc_free_price, axis=1)
            result_df['무배배송비'] = result_df.apply(calc_free_shipping, axis=1)
            
            # 임시 컬럼 삭제
            result_df = result_df.drop(columns=['배송비_숫자', '오너클랜판매가_숫자'])
            
            # 컬럼 순서 재정렬: 기존 컬럼 + 새 컬럼 4개
            base_columns = ['상품코드', '오너클랜판매가', '배송비', '배송유형', '최대구매수량', '반품배송비']
            new_columns = ['3000원가', '3000배송비', '무배원가', '무배배송비']
            result_df = result_df[base_columns + new_columns]
            
            # 엑셀 파일 저장
            file_dir = Path(file_path).parent
            base_name = Path(file_path).stem
            output_path = file_dir / f"{base_name}_오너클랜원가배송비.xlsx"
            
            # 엑셀로 저장
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                # 포맷팅 적용
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                
                # 헤더 스타일 적용 (1행)
                from openpyxl.styles import PatternFill, Font
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # 3000원가/3000배송비 컬럼 색상 (다른 색상)
                color1_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")  # 파란색
                # 무배원가/무배배송비 컬럼 색상 (또 다른 색상)
                color2_fill = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid")  # 녹색
                
                # 컬럼 인덱스 찾기
                base_col_count = len(['상품코드', '오너클랜판매가', '배송비', '배송유형', '최대구매수량', '반품배송비'])
                col_3000_price_idx = base_col_count + 1  # 3000원가 컬럼 인덱스 (1-based)
                col_3000_shipping_idx = base_col_count + 2  # 3000배송비 컬럼 인덱스
                col_free_price_idx = base_col_count + 3  # 무배원가 컬럼 인덱스
                col_free_shipping_idx = base_col_count + 4  # 무배배송비 컬럼 인덱스
                
                for col_idx in range(1, len(result_df.columns) + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    
                    # 3000원가, 3000배송비 컬럼
                    if col_idx == col_3000_price_idx or col_idx == col_3000_shipping_idx:
                        cell.fill = color1_fill
                    # 무배원가, 무배배송비 컬럼
                    elif col_idx == col_free_price_idx or col_idx == col_free_shipping_idx:
                        cell.fill = color2_fill
                    # 기본 컬럼
                    else:
                        cell.fill = header_fill
                
                # 열 너비 자동 조절
                for col_idx, col_name in enumerate(result_df.columns, start=1):
                    column_letter = sheet.cell(row=1, column=col_idx).column_letter
                    col_width = max(len(str(col_name)) * 1.5, 15)
                    sheet.column_dimensions[column_letter].width = min(col_width, 50)
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "E",
                        "location": "main.py:_create_ownerclan_excel:excel_saved",
                        "message": "엑셀 파일 저장 완료",
                        "data": {"output_path": str(output_path), "row_count": len(result_df)},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
            
            self._log(f"✅ 엑셀 파일 생성 완료: {output_path.name}")
            self._log(f"   추출된 행 수: {len(result_df)}개")
            self._log(f"   추출된 컬럼: {', '.join(result_df.columns)}")
            
            self.status_label.config(text="엑셀 생성 완료", fg="#27ae60")
            messagebox.showinfo("완료", f"엑셀 파일 생성 완료!\n\n파일명: {output_path.name}\n행 수: {len(result_df)}개")
            
        except Exception as e:
            error_msg = f"엑셀 생성 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            self.status_label.config(text="오류", fg="#e74c3c")
            messagebox.showerror("오류", error_msg)
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "pre-fix",
                        "hypothesisId": "F",
                        "location": "main.py:_create_ownerclan_excel:exception",
                        "message": "예외 발생",
                        "data": {"error": str(e)},
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + "\n")
            except: pass
            # #endregion
    
    def _extract_date_store_pattern(self, filename: str):
        """파일명에서 날짜_스토어명_스토어사업자번호 패턴 추출
        
        예시:
        - 등록 솔루션: 'shiningmall_20260117_스스A1-0_ESELLERS_2200408_1_1' → '20260117_스스A1-0'
        - 가공 엑셀: '20260117_스스A1-0_패션의류...' → '20260117_스스A1-0'
        
        Returns:
            패턴 문자열 (예: '20260117_스스A1-0') 또는 None
        """
        import re
        # 확장자 제거
        base_name = os.path.splitext(os.path.basename(filename))[0]
        
        # 언더스코어로 분리
        parts = base_name.split('_')
        
        # 날짜 패턴 찾기 (8자리 숫자)
        for i, part in enumerate(parts):
            if re.match(r'^\d{8}$', part):  # 8자리 숫자 (날짜)
                # 날짜 다음 부분이 있으면 스토어명_스토어사업자번호로 간주
                if i + 1 < len(parts):
                    # 날짜_스토어명_스토어사업자번호 형식
                    store_part = parts[i + 1]
                    # 스토어사업자번호는 보통 숫자나 알파벳+숫자 조합
                    # 예: '스스A1-0', '스스B2-0' 등
                    if store_part:
                        return f"{part}_{store_part}"
        
        return None
    
    def _batch_mapping_esellers(self):
        """이셀러스 일괄 매핑 기능 - 파일 선택 UI 제공"""
        try:
            # 일괄 매핑 창 생성
            batch_window = tk.Toplevel(self)
            batch_window.title("🔄 이셀러스 일괄 매핑")
            
            # 메인 창의 위치 가져오기
            try:
                main_x = self.winfo_x()
                main_y = self.winfo_y()
                main_width = self.winfo_width()
                main_height = self.winfo_height()
            except:
                # 메인 창 위치를 가져올 수 없는 경우 기본값 사용
                main_x = 100
                main_y = 100
                main_width = 800
                main_height = 600
            
            # 일괄 매핑 창 크기 (세로 크기 증가 - 버튼이 보이도록 충분히 크게)
            batch_width = 1100
            batch_height = 950
            
            # 메인 창과 같은 위치에 배치 (중앙 정렬)
            batch_x = max(0, main_x + (main_width - batch_width) // 2)
            batch_y = max(0, main_y + (main_height - batch_height) // 2)
            
            # 창 위치 및 크기 설정
            batch_window.geometry(f"{batch_width}x{batch_height}+{batch_x}+{batch_y}")
            batch_window.configure(bg="#F0F2F5")
            
            # 부모 창 설정 및 항상 위에 고정
            batch_window.transient(self)  # 부모 창 설정
            batch_window.attributes('-topmost', True)  # 항상 위에 고정
            batch_window.lift()  # 창을 맨 위로 올리기
            batch_window.focus_force()  # 포커스 강제 설정
            batch_window.update()  # 창 업데이트 강제
        except Exception as e:
            error_msg = f"일괄 매핑 창 생성 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            messagebox.showerror("오류", error_msg)
            return
        
        # 선택된 파일 저장 변수
        solution_files = []  # [(파일경로, 파일명)]
        processed_files = []  # [(파일경로, 파일명)]
        
        # 매핑 모드 (1:1 모드 또는 1:N 모드)
        mapping_mode = tk.StringVar(value="1:N")  # 기본값: 1:N 모드
        
        # 제목
        title_frame = tk.Frame(batch_window, bg="#FFFFFF", relief="solid", bd=1)
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        tk.Label(title_frame, text="🔄 이셀러스 일괄 매핑", 
                font=("맑은 고딕", 16, "bold"), bg="#FFFFFF", fg="#2c3e50").pack(pady=15)
        
        # 모드 선택 프레임 (눈에 띄게 강조)
        mode_frame = tk.Frame(title_frame, bg="#E8F4F8", relief="solid", bd=2)
        mode_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # 모드 선택 제목 (강조)
        mode_title_frame = tk.Frame(mode_frame, bg="#E8F4F8")
        mode_title_frame.pack(fill="x", padx=15, pady=(10, 5))
        tk.Label(mode_title_frame, text="📌 매핑 모드 선택", 
                font=("맑은 고딕", 12, "bold"), bg="#E8F4F8", fg="#2980b9").pack(side="left")
        
        # 라디오 버튼 프레임
        radio_frame = tk.Frame(mode_frame, bg="#E8F4F8")
        radio_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        # update_preview 함수는 나중에 정의되므로, 임시로 빈 함수를 사용하고 나중에 재설정
        def update_preview_placeholder():
            pass
        
        mode_radio_1 = tk.Radiobutton(radio_frame, text="1:1 매칭 (파일명 패턴 기준 자동 매칭)", 
                      variable=mapping_mode, value="1:1",
                      font=("맑은 고딕", 11), bg="#E8F4F8", fg="#34495e",
                      activebackground="#D5E8F0", activeforeground="#2c3e50",
                      selectcolor="#FFFFFF", cursor="hand2")
        mode_radio_1.pack(side="left", padx=10, pady=5)
        
        mode_radio_2 = tk.Radiobutton(radio_frame, text="1:N 매칭 (등록솔루션 1개 + 가공엑셀 여러 개) ⭐ 기본", 
                      variable=mapping_mode, value="1:N",
                      font=("맑은 고딕", 11, "bold"), bg="#E8F4F8", fg="#27ae60",
                      activebackground="#D5E8F0", activeforeground="#27ae60",
                      selectcolor="#FFFFFF", cursor="hand2")
        mode_radio_2.pack(side="left", padx=10, pady=5)
        
        # 설명 레이블 (더 명확하게)
        desc_frame = tk.Frame(mode_frame, bg="#E8F4F8")
        desc_frame.pack(fill="x", padx=15, pady=(0, 10))
        tk.Label(desc_frame, 
                text="• 1:1 모드: 파일명 패턴(날짜_스토어명_스토어사업자번호) 기준으로 자동 매칭\n• 1:N 모드: 등록솔루션 1개에 여러 가공엑셀의 상품코드가 매핑 (각 가공엑셀마다 개별 결과 파일 생성)", 
                font=("맑은 고딕", 9), bg="#E8F4F8", fg="#555555", justify="left").pack(anchor="w")
        
        tk.Label(title_frame, text="", bg="#FFFFFF").pack(pady=(0, 5))  # 간격
        
        # 메인 컨텐츠 프레임 (하단 버튼 공간 확보)
        main_frame = tk.Frame(batch_window, bg="#F0F2F5")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(10, 0))
        
        # 왼쪽: 파일 선택 영역
        left_frame = tk.Frame(main_frame, bg="#FFFFFF", relief="solid", bd=1)
        left_frame.pack(side="left", fill="both", padx=(0, 10))
        
        # 등록 솔루션 엑셀 영역
        solution_frame = tk.LabelFrame(left_frame, text="① 등록 솔루션 엑셀", 
                                       font=("맑은 고딕", 11, "bold"), bg="#FFFFFF", fg="#2c3e50", bd=1)
        solution_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        solution_btn_frame = tk.Frame(solution_frame, bg="#FFFFFF")
        solution_btn_frame.pack(fill="x", padx=10, pady=10)
        
        solution_list_frame = tk.Frame(solution_frame, bg="#FFFFFF")
        solution_list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # 등록 솔루션 엑셀 리스트박스
        solution_listbox = tk.Listbox(solution_list_frame, font=("맑은 고딕", 9), 
                                      bg="#F8F9FA", fg="#2c3e50", selectmode=tk.EXTENDED)
        solution_listbox.pack(side="left", fill="both", expand=True)
        solution_scrollbar = ttk.Scrollbar(solution_list_frame, orient="vertical", command=solution_listbox.yview)
        solution_listbox.configure(yscrollcommand=solution_scrollbar.set)
        solution_scrollbar.pack(side="right", fill="y")
        
        def load_solution_files():
            # 파일 선택 다이얼로그를 열기 전에 topmost 해제
            batch_window.attributes('-topmost', False)
            batch_window.update()
            
            # 1:N 모드일 때는 1개만 선택 가능
            if mapping_mode.get() == "1:N":
                file = filedialog.askopenfilename(
                    title="등록 솔루션 엑셀 파일 선택 (1개만)",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                )
                files = [file] if file else []
            else:
                files = filedialog.askopenfilenames(
                    title="등록 솔루션 엑셀 파일 선택 (여러 개 선택 가능)",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                )
            
            # 파일 선택 후 다시 topmost 설정
            batch_window.attributes('-topmost', True)
            batch_window.lift()
            batch_window.focus_force()
            
            if files:
                solution_files.clear()
                solution_listbox.delete(0, tk.END)
                for file_path in files:
                    file_name = os.path.basename(file_path)
                    solution_files.append((file_path, file_name))
                    solution_listbox.insert(tk.END, file_name)
                update_preview()
        
        def remove_selected_solution():
            selected_indices = solution_listbox.curselection()
            for idx in reversed(selected_indices):
                solution_files.pop(idx)
                solution_listbox.delete(idx)
            update_preview()
        
        tk.Button(solution_btn_frame, text="📂 불러오기", command=load_solution_files,
                 bg="#3498db", fg="white", font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2", padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(solution_btn_frame, text="🗑️ 선택 삭제", command=remove_selected_solution,
                 bg="#e74c3c", fg="white", font=("맑은 고딕", 9),
                 relief="raised", cursor="hand2", padx=15, pady=5).pack(side="left", padx=5)
        
        # 가공 엑셀 영역
        processed_frame = tk.LabelFrame(left_frame, text="② 가공된 엑셀", 
                                        font=("맑은 고딕", 11, "bold"), bg="#FFFFFF", fg="#2c3e50", bd=1)
        processed_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        processed_btn_frame = tk.Frame(processed_frame, bg="#FFFFFF")
        processed_btn_frame.pack(fill="x", padx=10, pady=10)
        
        processed_list_frame = tk.Frame(processed_frame, bg="#FFFFFF")
        processed_list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # 가공 엑셀 리스트박스
        processed_listbox = tk.Listbox(processed_list_frame, font=("맑은 고딕", 9), 
                                       bg="#F8F9FA", fg="#2c3e50", selectmode=tk.EXTENDED)
        processed_listbox.pack(side="left", fill="both", expand=True)
        processed_scrollbar = ttk.Scrollbar(processed_list_frame, orient="vertical", command=processed_listbox.yview)
        processed_listbox.configure(yscrollcommand=processed_scrollbar.set)
        processed_scrollbar.pack(side="right", fill="y")
        
        def load_processed_files():
            # 파일 선택 다이얼로그를 열기 전에 topmost 해제
            batch_window.attributes('-topmost', False)
            batch_window.update()
            
            files = filedialog.askopenfilenames(
                title="가공된 엑셀 파일 선택 (여러 개 선택 가능)",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            # 파일 선택 후 다시 topmost 설정
            batch_window.attributes('-topmost', True)
            batch_window.lift()
            batch_window.focus_force()
            
            if files:
                processed_files.clear()
                processed_listbox.delete(0, tk.END)
                for file_path in files:
                    file_name = os.path.basename(file_path)
                    processed_files.append((file_path, file_name))
                    processed_listbox.insert(tk.END, file_name)
                update_preview()
        
        def remove_selected_processed():
            selected_indices = processed_listbox.curselection()
            for idx in reversed(selected_indices):
                processed_files.pop(idx)
                processed_listbox.delete(idx)
            update_preview()
        
        tk.Button(processed_btn_frame, text="📂 불러오기", command=load_processed_files,
                 bg="#3498db", fg="white", font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2", padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(processed_btn_frame, text="🗑️ 선택 삭제", command=remove_selected_processed,
                 bg="#e74c3c", fg="white", font=("맑은 고딕", 9),
                 relief="raised", cursor="hand2", padx=15, pady=5).pack(side="left", padx=5)
        
        # 오른쪽: 미리보기 영역
        preview_frame = tk.Frame(main_frame, bg="#FFFFFF", relief="solid", bd=1)
        preview_frame.pack(side="right", fill="both", expand=True)
        
        preview_title = tk.Label(preview_frame, text="📋 매칭 미리보기", 
                                font=("맑은 고딕", 12, "bold"), bg="#FFFFFF", fg="#2c3e50")
        preview_title.pack(pady=15)
        
        # 스크롤 가능한 미리보기 프레임
        preview_canvas = tk.Canvas(preview_frame, bg="#FFFFFF")
        preview_scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=preview_canvas.yview)
        preview_scrollable = tk.Frame(preview_canvas, bg="#FFFFFF")
        
        preview_scrollable.bind(
            "<Configure>",
            lambda e: preview_canvas.configure(scrollregion=preview_canvas.bbox("all"))
        )
        
        preview_canvas.create_window((0, 0), window=preview_scrollable, anchor="nw")
        preview_canvas.configure(yscrollcommand=preview_scrollbar.set)
        
        preview_canvas.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        preview_scrollbar.pack(side="right", fill="y", pady=10)
        
        # update_preview 함수 정의 (라디오 버튼에서 참조하기 위해)
        def update_preview():
            # 기존 미리보기 내용 삭제
            for widget in preview_scrollable.winfo_children():
                widget.destroy()
            
            if not solution_files and not processed_files:
                tk.Label(preview_scrollable, text="파일을 불러오면 매칭 결과가 표시됩니다", 
                        font=("맑은 고딕", 10), bg="#FFFFFF", fg="#7f8c8d").pack(pady=50)
                return
            
            mode = mapping_mode.get()
            
            # 1:N 모드 처리
            if mode == "1:N":
                if not solution_files:
                    tk.Label(preview_scrollable, text="⚠️ 등록 솔루션 엑셀을 1개 선택해주세요", 
                            font=("맑은 고딕", 10), bg="#FFFFFF", fg="#e67e22").pack(pady=50)
                    batch_window.matched_pairs = []
                    batch_window.unmatched_solution = []
                    batch_window.unmatched_processed = []
                    return
                
                if len(solution_files) > 1:
                    tk.Label(preview_scrollable, text="⚠️ 1:N 모드에서는 등록 솔루션 엑셀을 1개만 선택할 수 있습니다", 
                            font=("맑은 고딕", 10), bg="#FFFFFF", fg="#e67e22").pack(pady=50)
                    batch_window.matched_pairs = []
                    batch_window.unmatched_solution = []
                    batch_window.unmatched_processed = []
                    return
                
                if not processed_files:
                    tk.Label(preview_scrollable, text="⚠️ 가공된 엑셀을 1개 이상 선택해주세요", 
                            font=("맑은 고딕", 10), bg="#FFFFFF", fg="#e67e22").pack(pady=50)
                    batch_window.matched_pairs = []
                    batch_window.unmatched_solution = []
                    batch_window.unmatched_processed = []
                    return
                
                # 1:N 모드 미리보기
                sol_file, sol_name = solution_files[0]
                
                # 등록 솔루션 엑셀 정보 추출
                solution_product_count = 0
                try:
                    sol_df = read_excel_with_fallback(sol_file, sheet_name="기본정보")
                    if len(sol_df) > 0:
                        sol_df_data = sol_df.iloc[2:].reset_index(drop=True)  # 2행 설명탭 제거
                        if '판매자 관리코드' in sol_df_data.columns:
                            solution_product_count = sol_df_data['판매자 관리코드'].dropna().astype(str).nunique()
                except Exception as e:
                    pass
                
                info_lf = tk.LabelFrame(preview_scrollable, text="✅ 1:N 매핑 정보", 
                                       font=("맑은 고딕", 10, "bold"), bg="#FFFFFF", fg="#27ae60", bd=1)
                info_lf.pack(fill="x", padx=10, pady=10)
                
                tk.Label(info_lf, text=f"등록 솔루션 엑셀: {sol_name}", 
                        font=("맑은 고딕", 9, "bold"), bg="#FFFFFF", fg="#2c3e50").pack(anchor="w", padx=8, pady=(5, 2))
                tk.Label(info_lf, text=f"  상품코드 (판매자 관리코드): {solution_product_count}개", 
                        font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=15, pady=1)
                tk.Label(info_lf, text=f"가공된 엑셀: {len(processed_files)}개", 
                        font=("맑은 고딕", 9), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=15, pady=(1, 5))
                
                files_lf = tk.LabelFrame(preview_scrollable, text=f"📋 가공된 엑셀 목록 ({len(processed_files)}개)", 
                                        font=("맑은 고딕", 10, "bold"), bg="#FFFFFF", fg="#2c3e50", bd=1)
                files_lf.pack(fill="x", padx=10, pady=10)
                
                for idx, (proc_file, proc_name) in enumerate(processed_files, 1):
                    item_frame = tk.Frame(files_lf, bg="#FFFFFF", relief="solid", bd=1)
                    item_frame.pack(fill="x", padx=5, pady=2)
                    
                    tk.Label(item_frame, text=f"{idx}. {proc_name}", 
                            font=("맑은 고딕", 8, "bold"), bg="#FFFFFF", fg="#2c3e50").pack(anchor="w", padx=8, pady=(3, 1))
                    
                    # 가공 엑셀 정보 추출
                    try:
                        proc_df = read_excel_with_fallback(proc_file)
                        
                        # 상품코드 개수
                        product_code_count = 0
                        if '상품코드' in proc_df.columns:
                            product_code_count = proc_df['상품코드'].dropna().astype(str).nunique()
                        
                        tk.Label(item_frame, text=f"   상품코드: {product_code_count}개", 
                                font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=15, pady=1)
                        
                        # 배송비 컬럼 정보
                        has_shipping = '배송비' in proc_df.columns
                        if has_shipping:
                            shipping_values = proc_df['배송비'].dropna().astype(str).value_counts().head(5)
                            shipping_info = ", ".join([f"{val}({cnt}개)" for val, cnt in shipping_values.items()])
                            if len(proc_df['배송비'].dropna().unique()) > 5:
                                shipping_info += f" 외 {len(proc_df['배송비'].dropna().unique()) - 5}개"
                            tk.Label(item_frame, text=f"   배송비: ✅ 있음 - {shipping_info}", 
                                    font=("맑은 고딕", 8), bg="#FFFFFF", fg="#27ae60").pack(anchor="w", padx=15, pady=1)
                        else:
                            tk.Label(item_frame, text=f"   배송비: ❌ 없음", 
                                    font=("맑은 고딕", 8), bg="#FFFFFF", fg="#e74c3c").pack(anchor="w", padx=15, pady=1)
                    except Exception as e:
                        tk.Label(item_frame, text=f"   ⚠️ 파일 읽기 실패: {str(e)[:30]}", 
                                font=("맑은 고딕", 8), bg="#FFFFFF", fg="#e67e22").pack(anchor="w", padx=15, pady=1)
                
                # 1:N 모드 데이터 저장
                batch_window.matched_pairs = [(sol_file, processed_files, "1:N", sol_name, None)]
                batch_window.unmatched_solution = []
                batch_window.unmatched_processed = []
                return
            
            # 1:1 모드 처리 (기존 로직)
            # 파일명 패턴 추출 및 매칭
            solution_patterns = {}
            processed_patterns = {}
            
            for sol_file, sol_name in solution_files:
                pattern = self._extract_date_store_pattern(sol_file)
                if pattern:
                    solution_patterns[pattern] = (sol_file, sol_name)
            
            for proc_file, proc_name in processed_files:
                pattern = self._extract_date_store_pattern(proc_file)
                if pattern:
                    processed_patterns[pattern] = (proc_file, proc_name)
            
            # 매칭 결과 계산
            matched_pairs = []
            unmatched_solution = []
            unmatched_processed = []
            
            for pattern, (sol_file, sol_name) in solution_patterns.items():
                if pattern in processed_patterns:
                    proc_file, proc_name = processed_patterns[pattern]
                    matched_pairs.append((sol_file, proc_file, pattern, sol_name, proc_name))
                else:
                    unmatched_solution.append((sol_file, sol_name, pattern))
            
            for pattern, (proc_file, proc_name) in processed_patterns.items():
                if pattern not in solution_patterns:
                    unmatched_processed.append((proc_file, proc_name, pattern))
            
            # 매칭된 파일 쌍 표시
            if matched_pairs:
                matched_lf = tk.LabelFrame(preview_scrollable, text=f"✅ 매칭된 파일 쌍 ({len(matched_pairs)}개)", 
                                          font=("맑은 고딕", 10, "bold"), bg="#FFFFFF", fg="#27ae60", bd=1)
                matched_lf.pack(fill="x", padx=10, pady=10)
                
                for idx, (sol_file, proc_file, pattern, sol_name, proc_name) in enumerate(matched_pairs, 1):
                    pair_frame = tk.Frame(matched_lf, bg="#FFFFFF", relief="solid", bd=1)
                    pair_frame.pack(fill="x", padx=5, pady=3)
                    
                    tk.Label(pair_frame, text=f"{idx}. 패턴: {pattern}", 
                            font=("맑은 고딕", 9, "bold"), bg="#FFFFFF", fg="#2c3e50").pack(anchor="w", padx=8, pady=(5, 2))
                    tk.Label(pair_frame, text=f"   등록 솔루션: {sol_name}", 
                            font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=15, pady=1)
                    
                    # 등록 솔루션 엑셀 상품코드 개수
                    sol_product_count = 0
                    try:
                        sol_df = read_excel_with_fallback(sol_file, sheet_name="기본정보")
                        if len(sol_df) > 0:
                            sol_df_data = sol_df.iloc[2:].reset_index(drop=True)
                            if '판매자 관리코드' in sol_df_data.columns:
                                sol_product_count = sol_df_data['판매자 관리코드'].dropna().astype(str).nunique()
                        tk.Label(pair_frame, text=f"     상품코드 (판매자 관리코드): {sol_product_count}개", 
                                font=("맑은 고딕", 7), bg="#FFFFFF", fg="#7f8c8d").pack(anchor="w", padx=23, pady=0)
                    except:
                        pass
                    
                    tk.Label(pair_frame, text=f"   가공 엑셀: {proc_name}", 
                            font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=15, pady=1)
                    
                    # 가공 엑셀 정보
                    try:
                        proc_df = read_excel_with_fallback(proc_file)
                        
                        # 상품코드 개수
                        proc_product_count = 0
                        if '상품코드' in proc_df.columns:
                            proc_product_count = proc_df['상품코드'].dropna().astype(str).nunique()
                        tk.Label(pair_frame, text=f"     상품코드: {proc_product_count}개", 
                                font=("맑은 고딕", 7), bg="#FFFFFF", fg="#7f8c8d").pack(anchor="w", padx=23, pady=0)
                        
                        # 배송비 컬럼 정보
                        has_shipping = '배송비' in proc_df.columns
                        if has_shipping:
                            shipping_values = proc_df['배송비'].dropna().astype(str).value_counts().head(3)
                            shipping_info = ", ".join([f"{val}({cnt}개)" for val, cnt in shipping_values.items()])
                            if len(proc_df['배송비'].dropna().unique()) > 3:
                                shipping_info += f" 외 {len(proc_df['배송비'].dropna().unique()) - 3}개"
                            tk.Label(pair_frame, text=f"     배송비: ✅ {shipping_info}", 
                                    font=("맑은 고딕", 7), bg="#FFFFFF", fg="#27ae60").pack(anchor="w", padx=23, pady=0)
                        else:
                            tk.Label(pair_frame, text=f"     배송비: ❌ 없음", 
                                    font=("맑은 고딕", 7), bg="#FFFFFF", fg="#e74c3c").pack(anchor="w", padx=23, pady=0)
                    except:
                        pass
                    
                    tk.Label(pair_frame, text="", bg="#FFFFFF").pack(anchor="w", padx=8, pady=1)  # 간격
            
            # 매칭 안 된 등록 솔루션
            if unmatched_solution:
                unmatched_sol_lf = tk.LabelFrame(preview_scrollable, text=f"⚠️ 매칭 안 된 등록 솔루션 ({len(unmatched_solution)}개)", 
                                                 font=("맑은 고딕", 10, "bold"), bg="#FFFFFF", fg="#e67e22", bd=1)
                unmatched_sol_lf.pack(fill="x", padx=10, pady=10)
                
                for sol_file, sol_name, pattern in unmatched_solution:
                    item_frame = tk.Frame(unmatched_sol_lf, bg="#FFFFFF", relief="solid", bd=1)
                    item_frame.pack(fill="x", padx=5, pady=3)
                    
                    tk.Label(item_frame, text=f"파일: {sol_name}", 
                            font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=8, pady=3)
                    if pattern:
                        tk.Label(item_frame, text=f"패턴: {pattern}", 
                                font=("맑은 고딕", 8), bg="#FFFFFF", fg="#7f8c8d").pack(anchor="w", padx=8, pady=(0, 3))
            
            # 매칭 안 된 가공 엑셀
            if unmatched_processed:
                unmatched_proc_lf = tk.LabelFrame(preview_scrollable, text=f"⚠️ 매칭 안 된 가공 엑셀 ({len(unmatched_processed)}개)", 
                                                  font=("맑은 고딕", 10, "bold"), bg="#FFFFFF", fg="#e67e22", bd=1)
                unmatched_proc_lf.pack(fill="x", padx=10, pady=10)
                
                for proc_file, proc_name, pattern in unmatched_processed:
                    item_frame = tk.Frame(unmatched_proc_lf, bg="#FFFFFF", relief="solid", bd=1)
                    item_frame.pack(fill="x", padx=5, pady=3)
                    
                    tk.Label(item_frame, text=f"파일: {proc_name}", 
                            font=("맑은 고딕", 8), bg="#FFFFFF", fg="#34495e").pack(anchor="w", padx=8, pady=3)
                    if pattern:
                        tk.Label(item_frame, text=f"패턴: {pattern}", 
                                font=("맑은 고딕", 8), bg="#FFFFFF", fg="#7f8c8d").pack(anchor="w", padx=8, pady=(0, 3))
            
            # 매칭 결과 저장 (실행 시 사용)
            batch_window.matched_pairs = matched_pairs
            batch_window.unmatched_solution = unmatched_solution
            batch_window.unmatched_processed = unmatched_processed
        
        # update_preview 함수 정의 후 라디오 버튼의 command 설정
        mode_radio_1.config(command=update_preview)
        mode_radio_2.config(command=update_preview)
        
        # 초기 미리보기 메시지
        tk.Label(preview_scrollable, text="파일을 불러오면 매칭 결과가 표시됩니다", 
                font=("맑은 고딕", 10), bg="#FFFFFF", fg="#7f8c8d").pack(pady=50)
        
        # 하단 버튼 프레임 (고정 높이로 확보)
        bottom_frame = tk.Frame(batch_window, bg="#FFFFFF", relief="solid", bd=1, height=80)
        bottom_frame.pack(fill="x", padx=20, pady=(10, 20))
        bottom_frame.pack_propagate(False)  # 높이 고정
        
        def execute_batch_mapping():
            if not hasattr(batch_window, 'matched_pairs') or not batch_window.matched_pairs:
                messagebox.showwarning("경고", "매칭된 파일 쌍이 없습니다.\n파일을 불러온 후 매칭 결과를 확인해주세요.")
                return
            
            mode = mapping_mode.get()
            batch_window.destroy()
            self._execute_batch_mapping(
                batch_window.matched_pairs,
                batch_window.unmatched_solution,
                batch_window.unmatched_processed,
                mapping_mode=mode
            )
        
        button_inner_frame = tk.Frame(bottom_frame, bg="#FFFFFF")
        button_inner_frame.pack(expand=True, fill="both")
        
        tk.Button(button_inner_frame, text="✅ 매핑 실행", command=execute_batch_mapping,
                 bg="#27ae60", fg="white", font=("맑은 고딕", 12, "bold"),
                 relief="raised", cursor="hand2", padx=40, pady=10).pack(side="left", padx=10)
        
        tk.Button(button_inner_frame, text="취소", command=batch_window.destroy,
                 bg="#95a5a6", fg="white", font=("맑은 고딕", 11),
                 relief="raised", cursor="hand2", padx=40, pady=10).pack(side="left", padx=5)
    
    def _execute_batch_mapping(self, matched_pairs, unmatched_solution, unmatched_processed, mapping_mode="1:1"):
        """일괄 매핑 실행
        
        Args:
            matched_pairs: 매칭된 파일 쌍 리스트
            unmatched_solution: 매칭 안 된 등록 솔루션 리스트
            unmatched_processed: 매칭 안 된 가공 엑셀 리스트
            mapping_mode: 매핑 모드 ("1:1" 또는 "1:N")
        """
        try:
            self._log("=" * 60)
            self._log(f"🔄 일괄 매핑 시작 ({mapping_mode} 모드)")
            self._log("=" * 60)
            
            # 1:N 모드 처리: 등록 솔루션 엑셀 1개를 기준으로 각 가공 엑셀마다 개별 결과 파일 생성
            if mapping_mode == "1:N":
                if not matched_pairs or len(matched_pairs) != 1:
                    messagebox.showerror("오류", "1:N 모드에서는 등록 솔루션 엑셀 1개와 가공 엑셀 여러 개가 필요합니다.")
                    return
                
                sol_file, processed_files_list, pattern, sol_name, _ = matched_pairs[0]
                
                self._log(f"등록 솔루션: {sol_name}")
                self._log(f"가공 엑셀: {len(processed_files_list)}개")
                self._log("")
                
                self.status_label.config(text="1:N 매핑 중...", fg="#f39c12")
                self.update_idletasks()
                
                # 프로그레스 바 창 생성
                progress_window = tk.Toplevel(self)
                progress_window.title("매핑 진행 중...")
                progress_window.geometry("500x150")
                progress_window.attributes('-topmost', True)
                progress_window.transient(self)
                
                # 메인 창 위치 가져오기
                try:
                    main_x = self.winfo_x()
                    main_y = self.winfo_y()
                    progress_window.geometry(f"500x150+{main_x}+{main_y}")
                except:
                    pass
                
                progress_frame = tk.Frame(progress_window, bg="#FFFFFF")
                progress_frame.pack(fill="both", expand=True, padx=20, pady=20)
                
                progress_label = tk.Label(progress_frame, text="매핑 진행 중...", 
                                         font=("맑은 고딕", 11, "bold"), bg="#FFFFFF", fg="#2c3e50")
                progress_label.pack(pady=(0, 10))
                
                progress_bar = ttk.Progressbar(progress_frame, length=460, mode='determinate')
                progress_bar.pack(fill="x", pady=5)
                
                progress_status = tk.Label(progress_frame, text="", 
                                          font=("맑은 고딕", 9), bg="#FFFFFF", fg="#7f8c8d")
                progress_status.pack(pady=5)
                
                try:
                    # 등록 솔루션 엑셀 전체 로드 (기본정보 + 확장정보)
                    self.solution_file_path = sol_file
                    solution_df_full = read_excel_with_fallback(sol_file, sheet_name="기본정보")
                    extension_df_full = None
                    try:
                        extension_df_full = read_excel_with_fallback(sol_file, sheet_name="확장정보")
                    except Exception as e:
                        self._log(f"확장정보 시트 읽기 실패 (무시): {str(e)}", "WARNING")
                    
                    if len(solution_df_full) == 0:
                        messagebox.showerror("오류", "등록 솔루션 엑셀 파일을 읽을 수 없습니다.")
                        return
                    
                    # 2행 설명탭 처리 (3행부터 데이터)
                    solution_df_full = solution_df_full.iloc[2:].reset_index(drop=True)
                    
                    results = []  # [(성공여부, 가공엑셀명, 상품코드수, 오류메시지)]
                    
                    # 각 가공 엑셀마다 개별 처리
                    total_files = len(processed_files_list)
                    for idx, (proc_file, proc_name) in enumerate(processed_files_list, 1):
                        try:
                            # 프로그레스 바 업데이트 (파일 로드 시작)
                            progress = (idx - 1) * 100 / total_files
                            progress_bar['value'] = progress
                            progress_status.config(text=f"파일 로드 중: {idx}/{total_files} - {proc_name}")
                            progress_window.update()  # update_idletasks() 대신 update() 사용
                            
                            self._log("")
                            self._log("=" * 60)
                            self._log(f"📌 처리 {idx}/{len(processed_files_list)}: {proc_name}")
                            self._log("=" * 60)
                            
                            # 가공 엑셀 로드
                            proc_df = read_excel_with_fallback(proc_file)
                            
                            # 프로그레스 바 업데이트 (파일 로드 완료)
                            progress = (idx - 1) * 100 / total_files + 5
                            progress_bar['value'] = progress
                            progress_status.config(text=f"데이터 처리 중: {idx}/{total_files} - {proc_name}")
                            progress_window.update()
                            
                            # 상품코드 추출
                            if '상품코드' not in proc_df.columns:
                                self._log(f"⚠️ '상품코드' 컬럼이 없어 스킵됨: {proc_name}", "WARNING")
                                results.append((False, proc_name, 0, "'상품코드' 컬럼이 없음"))
                                continue
                            
                            # 가공 엑셀의 상품코드 목록 추출
                            product_codes = proc_df['상품코드'].dropna().astype(str).unique().tolist()
                            product_codes_set = set(product_codes)
                            
                            self._log(f"  상품코드 {len(product_codes)}개 발견: {', '.join(product_codes[:5])}{'...' if len(product_codes) > 5 else ''}")
                            
                            # 등록 솔루션 엑셀에서 해당 상품코드에 해당하는 행만 필터링 (기본정보 시트)
                            # '판매자 관리코드' 컬럼과 가공 엑셀의 '상품코드' 매칭
                            if '판매자 관리코드' not in solution_df_full.columns:
                                self._log(f"⚠️ '판매자 관리코드' 컬럼이 없어 스킵됨: {proc_name}", "WARNING")
                                results.append((False, proc_name, 0, "'판매자 관리코드' 컬럼이 없음"))
                                continue
                            
                            # 필터링: 판매자 관리코드가 가공 엑셀의 상품코드 목록에 있는 행만
                            filtered_solution_df = solution_df_full[
                                solution_df_full['판매자 관리코드'].astype(str).isin(product_codes_set)
                            ].copy()
                            
                            if len(filtered_solution_df) == 0:
                                self._log(f"⚠️ 매칭되는 행이 없어 스킵됨: {proc_name}", "WARNING")
                                results.append((False, proc_name, len(product_codes), "매칭되는 행이 없음"))
                                continue
                            
                            self._log(f"  매칭된 행: {len(filtered_solution_df)}행")
                            
                            # 확장정보 시트 필터링 (원본번호* 기준)
                            filtered_extension_df = None
                            if extension_df_full is not None and len(extension_df_full) > 0:
                                # 기본정보의 '원본번호*' 값 추출 (필터링된 행만)
                                if '원본번호*' in filtered_solution_df.columns:
                                    original_numbers = filtered_solution_df['원본번호*'].dropna().astype(str).unique().tolist()
                                    original_numbers_set = set(original_numbers)
                                    
                                    # 확장정보 시트에서 '원본번호*' 컬럼이 있는지 확인
                                    if '원본번호*' in extension_df_full.columns:
                                        filtered_extension_df = extension_df_full[
                                            extension_df_full['원본번호*'].astype(str).isin(original_numbers_set)
                                        ].copy()
                                        self._log(f"  확장정보 매칭된 행: {len(filtered_extension_df)}행")
                                else:
                                    self._log(f"⚠️ '원본번호*' 컬럼이 없어 확장정보 필터링 불가", "WARNING")
                            
                            # 필터링된 데이터로 매핑 실행
                            self.solution_df = filtered_solution_df.reset_index(drop=True)
                            self.processed_df = proc_df
                            self.processed_file_path = proc_file
                            
                            # 마켓 감지
                            self.detected_market = self._detect_market_from_filename(proc_file)
                            
                            # 매핑 실행 전 프로그레스 바 업데이트
                            progress = (idx - 1) * 100 / total_files
                            progress_bar['value'] = progress
                            progress_status.config(text=f"매핑 실행 중: {idx}/{total_files} - {proc_name}")
                            progress_window.update()
                            
                            # 매핑 실행 (필터링된 확장정보 전달)
                            success = self._execute_mapping_internal(skip_popup=True, filtered_extension_df=filtered_extension_df, 
                                                                    progress_window=progress_window, progress_bar=progress_bar, 
                                                                    progress_status=progress_status, current_file=idx, total_files=total_files)
                            
                            # 매핑 완료 후 프로그레스 바 업데이트
                            progress = idx * 100 / total_files
                            progress_bar['value'] = progress
                            progress_status.config(text=f"완료: {idx}/{total_files} - {proc_name}")
                            progress_window.update()
                            
                            if success:
                                results.append((True, proc_name, len(product_codes), None))
                                self._log(f"✅ 처리 완료: {proc_name}")
                            else:
                                results.append((False, proc_name, len(product_codes), "매핑 실행 실패"))
                                self._log(f"❌ 처리 실패: {proc_name}", "ERROR")
                                
                        except Exception as e:
                            error_msg = str(e)
                            results.append((False, proc_name, 0, error_msg))
                            self._log(f"❌ 오류 발생: {proc_name} - {error_msg}", "ERROR")
                            import traceback
                            self._log(traceback.format_exc(), "ERROR")
                    
                    # 결과 요약
                    self._log("")
                    self._log("=" * 60)
                    self._log("📊 1:N 매핑 결과 요약")
                    self._log("=" * 60)
                    
                    success_count = sum(1 for r in results if r[0])
                    fail_count = len(results) - success_count
                    
                    self._log(f"✅ 성공: {success_count}개")
                    self._log(f"❌ 실패: {fail_count}개")
                    self._log("")
                    
                    # 결과 팝업
                    result_msg = f"1:N 매핑 완료!\n\n"
                    result_msg += f"등록 솔루션: {sol_name}\n"
                    result_msg += f"가공 엑셀: {len(processed_files_list)}개\n\n"
                    result_msg += f"✅ 성공: {success_count}개\n"
                    result_msg += f"❌ 실패: {fail_count}개"
                    
                    if fail_count > 0:
                        result_msg += "\n\n실패한 파일:\n"
                        for success, proc_name, code_count, error in results:
                            if not success:
                                result_msg += f"• {proc_name}"
                                if error:
                                    result_msg += f" ({error[:30]}...)"
                                result_msg += "\n"
                    
                    if fail_count == 0:
                        messagebox.showinfo("완료", result_msg)
                    else:
                        messagebox.showwarning("완료 (일부 실패)", result_msg)
                    
                except Exception as e:
                    error_msg = str(e)
                    self._log(f"❌ 1:N 매핑 오류: {error_msg}", "ERROR")
                    import traceback
                    self._log(traceback.format_exc(), "ERROR")
                    messagebox.showerror("오류", f"1:N 매핑 실패:\n{error_msg}")
                finally:
                    # 프로그레스 바 창 닫기
                    try:
                        progress_bar['value'] = 100
                        progress_status.config(text="완료!")
                        progress_window.update_idletasks()
                        progress_window.after(500, progress_window.destroy)  # 0.5초 후 자동 닫기
                    except:
                        pass
                
                self.status_label.config(text="1:N 매핑 완료", fg="#27ae60")
                return
            
            # 1:1 모드 처리 (기존 로직)
            self._log(f"매칭된 파일 쌍: {len(matched_pairs)}개")
            self._log("")
            
            self.status_label.config(text="일괄 매핑 중...", fg="#f39c12")
            self.update_idletasks()
            
            # 프로그레스 바 창 생성
            progress_window = tk.Toplevel(self)
            progress_window.title("매핑 진행 중...")
            progress_window.geometry("500x150")
            progress_window.attributes('-topmost', True)
            progress_window.transient(self)
            
            # 메인 창 위치 가져오기
            try:
                main_x = self.winfo_x()
                main_y = self.winfo_y()
                progress_window.geometry(f"500x150+{main_x}+{main_y}")
            except:
                pass
            
            progress_frame = tk.Frame(progress_window, bg="#FFFFFF")
            progress_frame.pack(fill="both", expand=True, padx=20, pady=20)
            
            progress_label = tk.Label(progress_frame, text="매핑 진행 중...", 
                                     font=("맑은 고딕", 11, "bold"), bg="#FFFFFF", fg="#2c3e50")
            progress_label.pack(pady=(0, 10))
            
            progress_bar = ttk.Progressbar(progress_frame, length=460, mode='determinate')
            progress_bar.pack(fill="x", pady=5)
            
            progress_status = tk.Label(progress_frame, text="", 
                                      font=("맑은 고딕", 9), bg="#FFFFFF", fg="#7f8c8d")
            progress_status.pack(pady=5)
            
            results = []  # [(성공여부, 등록솔루션파일, 가공엑셀파일, 패턴, 오류메시지)]
            
            total_pairs = len(matched_pairs)
            for idx, (sol_file, proc_file, pattern, sol_name, proc_name) in enumerate(matched_pairs, 1):
                try:
                    # 프로그레스 바 업데이트
                    progress = (idx - 1) * 100 / total_pairs if total_pairs > 0 else 0
                    progress_bar['value'] = progress
                    progress_status.config(text=f"매핑 중: {idx}/{total_pairs} - {pattern}")
                    progress_window.update_idletasks()
                    
                    self._log("")
                    self._log("=" * 60)
                    self._log(f"📌 매핑 {idx}/{len(matched_pairs)}: {pattern}")
                    self._log("=" * 60)
                    self._log(f"등록 솔루션: {sol_name}")
                    self._log(f"가공 엑셀: {proc_name}")
                    
                    # 파일 로드
                    self.solution_file_path = sol_file
                    self.processed_file_path = proc_file
                    
                    # 이셀러스 솔루션 로드
                    self.solution_df = read_excel_with_fallback(sol_file, sheet_name="기본정보")
                    if len(self.solution_df) > 0:
                        # 2행 설명탭 처리 (3행부터 데이터)
                        self.solution_df = self.solution_df.iloc[2:].reset_index(drop=True)
                    
                    # 가공 엑셀 로드
                    self.processed_df = read_excel_with_fallback(proc_file)
                    
                    # 마켓 감지
                    self.detected_market = self._detect_market_from_filename(proc_file)
                    
                    # 매핑 실행
                    success = self._execute_mapping_internal(skip_popup=True)
                    
                    if success:
                        results.append((True, sol_file, proc_file, pattern, None))
                        self._log(f"✅ 매핑 완료: {pattern}")
                    else:
                        results.append((False, sol_file, proc_file, pattern, "매핑 실행 실패"))
                        self._log(f"❌ 매핑 실패: {pattern}", "ERROR")
                        
                except Exception as e:
                    error_msg = str(e)
                    results.append((False, sol_file, proc_file, pattern, error_msg))
                    self._log(f"❌ 오류 발생: {pattern} - {error_msg}", "ERROR")
                    import traceback
                    self._log(traceback.format_exc(), "ERROR")
            
            # 결과 요약
            self._log("")
            self._log("=" * 60)
            self._log("📊 일괄 매핑 결과 요약")
            self._log("=" * 60)
            
            success_count = sum(1 for r in results if r[0])
            fail_count = len(results) - success_count
            
            self._log(f"✅ 성공: {success_count}개")
            self._log(f"❌ 실패: {fail_count}개")
            self._log("")
            
            if unmatched_solution:
                self._log(f"⚠️ 매칭 안 된 등록 솔루션: {len(unmatched_solution)}개")
                for sol_file, sol_name, pattern in unmatched_solution:
                    self._log(f"   - {sol_name} (패턴: {pattern})")
                self._log("")
            
            if unmatched_processed:
                self._log(f"⚠️ 매칭 안 된 가공 엑셀: {len(unmatched_processed)}개")
                for proc_file, proc_name, pattern in unmatched_processed:
                    self._log(f"   - {proc_name} (패턴: {pattern})")
                self._log("")
            
            # 결과 팝업
            result_msg = f"일괄 매핑 완료!\n\n"
            result_msg += f"✅ 성공: {success_count}개\n"
            result_msg += f"❌ 실패: {fail_count}개\n\n"
            
            if unmatched_solution:
                result_msg += f"⚠️ 매칭 안 된 등록 솔루션: {len(unmatched_solution)}개\n"
            if unmatched_processed:
                result_msg += f"⚠️ 매칭 안 된 가공 엑셀: {len(unmatched_processed)}개\n"
            
            if fail_count > 0:
                result_msg += "\n실패한 파일:\n"
                for success, sol_file, proc_file, pattern, error in results:
                    if not success:
                        result_msg += f"• {pattern}\n"
                        if error:
                            result_msg += f"  오류: {error[:50]}...\n"
            
            if fail_count == 0:
                messagebox.showinfo("완료", result_msg)
            else:
                messagebox.showwarning("완료 (일부 실패)", result_msg)
            
            # 프로그레스 바 완료
            try:
                progress_bar['value'] = 100
                progress_status.config(text="완료!")
                progress_window.update_idletasks()
                progress_window.after(500, progress_window.destroy)  # 0.5초 후 자동 닫기
            except:
                pass
            
            self.status_label.config(text="일괄 매핑 완료", fg="#27ae60")
            
        except Exception as e:
            error_msg = f"일괄 매핑 실행 실패: {str(e)}"
            self._log(error_msg, "ERROR")
            self.status_label.config(text="오류", fg="#e74c3c")
            messagebox.showerror("오류", error_msg)
            import traceback
            self._log(traceback.format_exc(), "ERROR")
    
    def _execute_mapping_internal(self, skip_popup=False, filtered_extension_df=None, 
                                  progress_window=None, progress_bar=None, progress_status=None,
                                  current_file=None, total_files=None):
        """매핑 실행 내부 로직 (팝업 제어 가능)
        
        Args:
            skip_popup: 팝업 표시 여부
            filtered_extension_df: 필터링된 확장정보 시트 (None이면 원본에서 읽기)
            progress_window: 프로그레스 바 창 (선택)
            progress_bar: 프로그레스 바 위젯 (선택)
            progress_status: 프로그레스 상태 레이블 (선택)
            current_file: 현재 파일 번호 (선택)
            total_files: 전체 파일 수 (선택)
        
        Returns:
            bool: 성공 여부
        """
        try:
            if self.solution_df is None or self.processed_df is None or not self.solution_instance:
                return False
            
            # 매핑 정보 수집 (기본 매핑 규칙 사용)
            column_mapping = {}
            default_values = {}
            
            # 기본 매핑 규칙 사용
            default_mapping = self.solution_instance.get_default_mapping()
            for proc_col, sol_col in default_mapping.items():
                column_mapping[proc_col] = sol_col
            
            if not column_mapping:
                return False
            
            # 결과 데이터프레임 생성
            result_df = self.solution_df.copy()
            
            # 설정 가져오기
            config = self.config_manager.get_solution_config(self.solution_name, self.detected_market)
            config["default_values"] = default_values
            
            if self.solution_name == "이셀러스" and self.detected_market:
                config["detected_market"] = self.detected_market
            
            if self.solution_name == "이셀러스" and self.processed_file_path:
                config["processed_file_path"] = self.processed_file_path
            
            if self.solution_name == "이셀러스" and self.solution_file_path:
                config["solution_file_path"] = self.solution_file_path
            
            # 프로그레스 바 업데이트 (매핑 시작)
            if progress_window and progress_bar and progress_status:
                try:
                    if current_file and total_files:
                        base_progress = (current_file - 1) * 100 / total_files
                        progress_bar['value'] = base_progress + 20  # 매핑 시작: 20%
                        progress_status.config(text=f"매핑 중: {current_file}/{total_files} - 컬럼 매핑 중...")
                    else:
                        progress_status.config(text="컬럼 매핑 중...")
                    progress_window.update()
                except:
                    pass
            
            # 매핑 적용
            result_df = self.solution_instance.apply_mapping(
                result_df, self.processed_df, column_mapping, config
            )
            
            # 프로그레스 바 업데이트 (특화 규칙 적용 시작)
            if progress_window and progress_bar and progress_status:
                try:
                    if current_file and total_files:
                        base_progress = (current_file - 1) * 100 / total_files
                        progress_bar['value'] = base_progress + 50  # 특화 규칙: 50%
                        progress_status.config(text=f"매핑 중: {current_file}/{total_files} - 특화 규칙 적용 중...")
                    else:
                        progress_status.config(text="특화 규칙 적용 중...")
                    progress_window.update()
                except:
                    pass
            
            # 솔루션별 특화 규칙 적용
            result_df = self.solution_instance.apply_solution_specific_rules(
                result_df, self.processed_df, config, original_solution_df=self.solution_df
            )
            
            # 프로그레스 바 업데이트 (저장 시작)
            if progress_window and progress_bar and progress_status:
                try:
                    if current_file and total_files:
                        base_progress = (current_file - 1) * 100 / total_files
                        progress_bar['value'] = base_progress + 80  # 저장: 80%
                        progress_status.config(text=f"매핑 중: {current_file}/{total_files} - 파일 저장 중...")
                    else:
                        progress_status.config(text="파일 저장 중...")
                    progress_window.update()
                except:
                    pass
            
            # 결과 저장
            if self.solution_name == "이셀러스":
                if self.processed_file_path and self.solution_file_path:
                    processed_filename = os.path.splitext(os.path.basename(self.processed_file_path))[0]
                    solution_ext = os.path.splitext(self.solution_file_path)[1]
                    output_filename = f"이셀완료{processed_filename}{solution_ext}"
                    output_dir = os.path.dirname(self.processed_file_path)
                    
                    # 솔루션 엑셀 파일명에 '_기본카테고리'가 있는지 확인
                    solution_filename = os.path.basename(self.solution_file_path)
                    has_basic_category = '_기본카테고리' in solution_filename
                    
                    # 날짜 추출 (파일명 앞부분에서 YYYYMMDD 형식 추출)
                    import re
                    date_match = re.match(r'^(\d{8})', processed_filename)
                    if date_match:
                        date_str = date_match.group(1)
                        # 기본카테고리인 경우 별도 폴더 생성
                        if has_basic_category:
                            subfolder_name = f"{date_str}이셀완료_기본카테고리"
                        else:
                            subfolder_name = f"{date_str}이셀완료"
                        output_dir = os.path.join(output_dir, subfolder_name)
                        # 폴더가 없으면 생성
                        if not os.path.exists(output_dir):
                            os.makedirs(output_dir)
                    
                    output_path = os.path.join(output_dir, output_filename)
                    
                    # 확장정보 시트 읽기 (필터링된 것이 있으면 사용, 없으면 원본에서 읽기)
                    extension_df = filtered_extension_df if filtered_extension_df is not None else None
                    if extension_df is None and self.solution_file_path:
                        try:
                            extension_df = read_excel_with_fallback(self.solution_file_path, sheet_name="확장정보")
                            if extension_df is not None and self.detected_market:
                                extension_df = self._process_extension_sheet_by_market(extension_df, self.detected_market)
                        except Exception as e:
                            self._log(f"확장정보 시트 읽기 실패: {str(e)}", "WARNING")
                            extension_df = None
                    elif extension_df is not None and self.detected_market:
                        # 필터링된 확장정보도 마켓별 처리 필요
                        extension_df = self._process_extension_sheet_by_market(extension_df, self.detected_market)
                    
                    # 기본정보 시트 전처리
                    if 'Unnamed: 76' in result_df.columns:
                        result_df['Unnamed: 76'] = ''
                    
                    # 기본정보: 2행 빈 행 추가 (이셀러스 형식: 1행=헤더, 2행=빈행, 3행부터=데이터)
                    empty_row = pd.DataFrame([[''] * len(result_df.columns)], columns=result_df.columns)
                    result_df_with_empty_row = pd.concat([empty_row, result_df], ignore_index=True)
                    
                    # 확장정보: 2행 빈 행 추가 (기본정보와 동일한 형식)
                    extension_df_with_empty_row = None
                    if extension_df is not None and len(extension_df) > 0:
                        extension_empty_row = pd.DataFrame([[''] * len(extension_df.columns)], columns=extension_df.columns)
                        extension_df_with_empty_row = pd.concat([extension_empty_row, extension_df], ignore_index=True)
                    
                    # 저장
                    file_ext = os.path.splitext(output_path)[1].lower()
                    if file_ext == '.xls':
                        output_path_xlsx = output_path.replace('.xls', '.xlsx')
                        with pd.ExcelWriter(output_path_xlsx, engine='openpyxl') as writer:
                            result_df_with_empty_row.to_excel(writer, sheet_name='기본정보', index=False)
                            if extension_df_with_empty_row is not None:
                                extension_df_with_empty_row.to_excel(writer, sheet_name='확장정보', index=False)
                        output_path = output_path_xlsx
                    else:
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            result_df_with_empty_row.to_excel(writer, sheet_name='기본정보', index=False)
                            if extension_df_with_empty_row is not None:
                                extension_df_with_empty_row.to_excel(writer, sheet_name='확장정보', index=False)
                    
                    self._log(f"저장 완료: {output_path}")
                    return True
            
            return False
            
        except Exception as e:
            self._log(f"매핑 실행 오류: {str(e)}", "ERROR")
            return False
    
    def _process_extension_sheet_by_market(self, extension_df: pd.DataFrame, market: str) -> pd.DataFrame:
        """확장정보 시트 마켓별 처리
        
        Args:
            extension_df: 확장정보 시트 DataFrame
            market: 감지된 마켓명
            
        Returns:
            처리된 확장정보 DataFrame
        """
        if extension_df is None or extension_df.empty:
            return extension_df
        
        # 스마트스토어(스스)만 특별 처리
        if market == "스스" or market == "스마트스토어":
            if "마켓 카테고리번호" in extension_df.columns:
                for idx, row in extension_df.iterrows():
                    category_value = row.get("마켓 카테고리번호", "")
                    if pd.notna(category_value) and str(category_value).strip():
                        # 줄바꿈으로 분리
                        lines = str(category_value).strip().split('\n')
                        # '스마트스토어*'로 시작하는 줄만 찾기
                        smartstore_lines = [line.strip() for line in lines if line.strip().startswith('스마트스토어*')]
                        if smartstore_lines:
                            # 첫 번째 매칭되는 줄만 사용
                            extension_df.at[idx, "마켓 카테고리번호"] = smartstore_lines[0]
                        else:
                            # 매칭되는 줄이 없으면 빈 값
                            extension_df.at[idx, "마켓 카테고리번호"] = ""
                self._log(f"스마트스토어: 마켓 카테고리번호에서 '스마트스토어*'로 시작하는 줄만 유지했습니다.")
        # 옥션, 지마켓, 쿠팡, 나머지 마켓은 그대로 유지
        
        return extension_df
    
    def _show_mapping_result_popup(self, result_df: pd.DataFrame, column_mapping: Dict[str, str], 
                                   default_values: Dict[str, str], matched_codes: set = None,
                                   unmatched_result: set = None, unmatched_processed: set = None,
                                   output_path: str = None):
        """매핑 결과 팝업 창 표시"""
        
        # 팝업 창 생성
        popup = tk.Toplevel(self)
        popup.title("📊 매핑 결과")
        
        # 메인 윈도우 위치 가져오기
        main_x = self.winfo_x()
        main_y = self.winfo_y()
        
        # 팝업창 크기 및 위치 설정 (세로 크기 증가, 메인 윈도우와 동일한 위치)
        popup.geometry(f"900x900+{main_x}+{main_y}")
        popup.configure(bg="#FFFFFF")
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(popup, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # 제목
        title_frame = tk.Frame(scrollable_frame, bg="#2C3E50", height=60)
        title_frame.pack(fill="x", padx=10, pady=(0, 10))
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="📊 매핑 결과 요약", 
                font=("맑은 고딕", 16, "bold"), 
                bg="#2C3E50", fg="white").pack(expand=True, pady=15)
        
        # 기본 정보
        info_frame = tk.LabelFrame(scrollable_frame,
                                   text="📋 기본 정보",
                                   font=("맑은 고딕", 11, "bold"),
                                   bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        info_frame.pack(fill="x", padx=10, pady=10)
        
        info_text_frame = tk.Frame(info_frame, bg="#FFFFFF")
        info_text_frame.pack(fill="x", padx=15, pady=15)
        
        tk.Label(info_text_frame, text=f"등록 솔루션 엑셀: {len(self.solution_df)}행",
                font=("맑은 고딕", 10),
                bg="#FFFFFF", fg="#333", anchor="w").pack(fill="x", pady=3)
        tk.Label(info_text_frame, text=f"가공 엑셀: {len(self.processed_df)}행",
                font=("맑은 고딕", 10),
                bg="#FFFFFF", fg="#333", anchor="w").pack(fill="x", pady=3)
        tk.Label(info_text_frame, text=f"결과 파일: {len(result_df)}행, {len(result_df.columns)}컬럼",
                font=("맑은 고딕", 10, "bold"),
                bg="#FFFFFF", fg="#27ae60", anchor="w").pack(fill="x", pady=3)
        tk.Label(info_text_frame, text=f"매핑된 컬럼 수: {len(column_mapping)}개",
                font=("맑은 고딕", 10),
                bg="#FFFFFF", fg="#333", anchor="w").pack(fill="x", pady=3)
        if output_path:
            tk.Label(info_text_frame, text=f"저장 위치: {output_path}",
                    font=("맑은 고딕", 9),
                    bg="#FFFFFF", fg="#666", anchor="w", wraplength=800, justify="left").pack(fill="x", pady=3)
        
        # 상품코드 매칭 통계
        if matched_codes is not None:
            code_frame = tk.LabelFrame(scrollable_frame,
                                       text="🔗 상품코드 매칭 통계",
                                       font=("맑은 고딕", 11, "bold"),
                                       bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
            code_frame.pack(fill="x", padx=10, pady=10)
            
            code_text_frame = tk.Frame(code_frame, bg="#FFFFFF")
            code_text_frame.pack(fill="x", padx=15, pady=15)
            
            tk.Label(code_text_frame, text=f"✅ 매칭된 상품코드: {len(matched_codes)}개",
                    font=("맑은 고딕", 10, "bold"),
                    bg="#FFFFFF", fg="#27ae60", anchor="w").pack(fill="x", pady=3)
            if unmatched_result:
                tk.Label(code_text_frame, text=f"⚠️ 등록 솔루션에만 있는 상품코드: {len(unmatched_result)}개",
                        font=("맑은 고딕", 10),
                        bg="#FFFFFF", fg="#f39c12", anchor="w").pack(fill="x", pady=3)
            if unmatched_processed:
                tk.Label(code_text_frame, text=f"⚠️ 가공 엑셀에만 있는 상품코드: {len(unmatched_processed)}개",
                        font=("맑은 고딕", 10),
                        bg="#FFFFFF", fg="#f39c12", anchor="w").pack(fill="x", pady=3)
        
        # 컬럼별 매핑 결과
        mapping_frame = tk.LabelFrame(scrollable_frame,
                                      text="📈 컬럼별 매핑 결과",
                                      font=("맑은 고딕", 11, "bold"),
                                      bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        mapping_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 테이블 헤더
        header_frame = tk.Frame(mapping_frame, bg="#E8EAF6")
        header_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(header_frame, text="상태", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=5).pack(side="left", padx=5)
        tk.Label(header_frame, text="가공 엑셀 컬럼", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=25).pack(side="left", padx=5)
        tk.Label(header_frame, text="→", font=("맑은 고딕", 10, "bold"),
                bg="#E8EAF6", fg="#546E7A", width=5).pack(side="left")
        tk.Label(header_frame, text="등록 솔루션 컬럼", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=25).pack(side="left", padx=5)
        tk.Label(header_frame, text="매핑 통계", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=30).pack(side="left", padx=5)
        
        # 컬럼별 결과 표시
        mapping_list_frame = tk.Frame(mapping_frame, bg="#FFFFFF")
        mapping_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        for proc_col, sol_col in sorted(column_mapping.items()):
            if sol_col in result_df.columns:
                row_frame = tk.Frame(mapping_list_frame, bg="#FFFFFF", relief="solid", bd=1)
                row_frame.pack(fill="x", padx=5, pady=2)
                
                # 매핑된 값의 개수 (빈 값 제외)
                mapped_count = result_df[sol_col].notna().sum()
                total_count = len(result_df)
                mapped_pct = (mapped_count / total_count * 100) if total_count > 0 else 0
                
                # 기본값으로 채워진 경우 확인
                default_filled_count = 0
                default_val_str = ""
                if sol_col in default_values:
                    default_val = default_values[sol_col]
                    default_filled_count = (result_df[sol_col] == default_val).sum()
                    if default_filled_count > 0:
                        default_val_str = f"\n기본값 사용: {default_filled_count}행"
                
                # 상태 아이콘
                status_icon = "✅" if mapped_count > 0 else "❌"
                status_color = "#27ae60" if mapped_count > 0 else "#e74c3c"
                
                tk.Label(row_frame, text=status_icon, font=("맑은 고딕", 12),
                        bg="#FFFFFF", fg=status_color, width=5).pack(side="left", padx=5)
                
                tk.Label(row_frame, text=proc_col, font=("맑은 고딕", 9),
                        bg="#FFFFFF", fg="#333", width=25, anchor="w").pack(side="left", padx=5)
                
                tk.Label(row_frame, text="→", font=("맑은 고딕", 10),
                        bg="#FFFFFF", fg="#999", width=5).pack(side="left")
                
                tk.Label(row_frame, text=sol_col, font=("맑은 고딕", 9),
                        bg="#FFFFFF", fg="#333", width=25, anchor="w").pack(side="left", padx=5)
                
                stat_text = f"{mapped_count}/{total_count}행 ({mapped_pct:.1f}%){default_val_str}"
                tk.Label(row_frame, text=stat_text, font=("맑은 고딕", 9),
                        bg="#FFFFFF", fg="#666", width=30, anchor="w", justify="left").pack(side="left", padx=5)
        
        # 기본값 정보
        if default_values:
            default_frame = tk.LabelFrame(scrollable_frame,
                                         text="💡 기본값 설정",
                                         font=("맑은 고딕", 11, "bold"),
                                         bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
            default_frame.pack(fill="x", padx=10, pady=10)
            
            default_text_frame = tk.Frame(default_frame, bg="#FFFFFF")
            default_text_frame.pack(fill="x", padx=15, pady=15)
            
            for sol_col, default_val in default_values.items():
                tk.Label(default_text_frame, text=f"• [{sol_col}]: {default_val}",
                        font=("맑은 고딕", 10),
                        bg="#FFFFFF", fg="#333", anchor="w").pack(fill="x", pady=2)
        
        # 닫기 버튼
        btn_frame = tk.Frame(popup, bg="#FFFFFF")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(btn_frame, text="닫기", command=popup.destroy,
                 bg="#3498db", fg="white",
                 font=("맑은 고딕", 10, "bold"),
                 relief="raised", cursor="hand2",
                 padx=30, pady=8).pack(side="right", padx=5)
    
    def _open_market_settings(self):
        """마켓 설정 창 열기 (다팔자 솔루션 전용)"""
        if not self.solution_name or self.solution_name != "다팔자":
            messagebox.showwarning("경고", "다팔자 솔루션에서만 사용할 수 있습니다.")
            return
        
        # 팝업 창 생성
        popup = tk.Toplevel(self)
        popup.title("마켓 설정")
        popup.geometry("900x700")
        popup.configure(bg="#FFFFFF")
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(popup, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # =====================================================================
        # 1. 마켓 접두사 관리
        # =====================================================================
        prefix_frame = tk.LabelFrame(scrollable_frame,
                                     text="📋 마켓 접두사 관리",
                                     font=("맑은 고딕", 11, "bold"),
                                     bg="#FFFFFF", fg="#546E7A", bd=2, relief="groove")
        prefix_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(prefix_frame,
                text="파일명 접두사를 입력하면 해당 마켓으로 자동 감지됩니다.\n예: '스스_상품.xlsx' → 스스 마켓으로 감지",
                font=("맑은 고딕", 9),
                bg="#FFFFFF", fg="#666",
                justify="left").pack(anchor="w", padx=10, pady=10)
        
        # 현재 접두사 목록
        current_prefixes = self.config_manager.get_market_prefixes()
        
        # 접두사 입력 영역
        prefix_list_frame = tk.Frame(prefix_frame, bg="#FFFFFF")
        prefix_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 헤더
        header_frame = tk.Frame(prefix_list_frame, bg="#E8EAF6")
        header_frame.pack(fill="x", padx=5, pady=5)
        tk.Label(header_frame, text="접두사", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=20).pack(side="left", padx=5)
        tk.Label(header_frame, text="마켓명", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=20).pack(side="left", padx=5)
        tk.Label(header_frame, text="작업", font=("맑은 고딕", 9, "bold"),
                bg="#E8EAF6", fg="#333", width=10).pack(side="left", padx=5)
        
        # 접두사 목록 표시
        prefix_vars = {}
        prefix_list_container = tk.Frame(prefix_list_frame, bg="#FFFFFF")
        prefix_list_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        def refresh_prefix_list():
            """접두사 목록 새로고침"""
            for widget in prefix_list_container.winfo_children():
                widget.destroy()
            prefix_vars.clear()
            
            current_prefixes = self.config_manager.get_market_prefixes()
            for prefix, market in current_prefixes.items():
                row_frame = tk.Frame(prefix_list_container, bg="#FFFFFF")
                row_frame.pack(fill="x", padx=5, pady=2)
                
                prefix_var = tk.StringVar(value=prefix)
                market_var = tk.StringVar(value=market)
                
                tk.Entry(row_frame, textvariable=prefix_var, font=("맑은 고딕", 9),
                        width=20).pack(side="left", padx=5)
                tk.Entry(row_frame, textvariable=market_var, font=("맑은 고딕", 9),
                        width=20).pack(side="left", padx=5)
                
                def delete_prefix(p=prefix):
                    if messagebox.askyesno("확인", f"접두사 '{p}'를 삭제하시겠습니까?"):
                        current = self.config_manager.get_market_prefixes()
                        if p in current:
                            del current[p]
                            self.config_manager.save_market_prefixes(current)
                            refresh_prefix_list()
                
                tk.Button(row_frame, text="삭제", command=delete_prefix,
                         bg="#e74c3c", fg="white", font=("맑은 고딕", 8),
                         relief="raised", cursor="hand2", padx=5, pady=2).pack(side="left", padx=5)
                
                prefix_vars[prefix] = {"prefix": prefix_var, "market": market_var}
        
        refresh_prefix_list()
        
        # 새 접두사 추가
        add_frame = tk.Frame(prefix_frame, bg="#FFFFFF")
        add_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(add_frame, text="새 접두사:", font=("맑은 고딕", 9),
                bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        new_prefix_var = tk.StringVar()
        tk.Entry(add_frame, textvariable=new_prefix_var, font=("맑은 고딕", 9),
                width=15).pack(side="left", padx=5)
        
        tk.Label(add_frame, text="마켓명:", font=("맑은 고딕", 9),
                bg="#FFFFFF", fg="#333").pack(side="left", padx=5)
        new_market_var = tk.StringVar()
        tk.Entry(add_frame, textvariable=new_market_var, font=("맑은 고딕", 9),
                width=15).pack(side="left", padx=5)
        
        def add_prefix():
            prefix = new_prefix_var.get().strip()
            market = new_market_var.get().strip()
            if not prefix or not market:
                messagebox.showwarning("경고", "접두사와 마켓명을 모두 입력해주세요.")
                return
            if prefix in prefix_vars:
                messagebox.showwarning("경고", "이미 존재하는 접두사입니다.")
                return
            
            current = self.config_manager.get_market_prefixes()
            current[prefix] = market
            self.config_manager.save_market_prefixes(current)
            refresh_prefix_list()
            new_prefix_var.set("")
            new_market_var.set("")
            self._log(f"마켓 접두사 추가: {prefix} → {market}")
        
        tk.Button(add_frame, text="추가", command=add_prefix,
                 bg="#27ae60", fg="white", font=("맑은 고딕", 9, "bold"),
                 relief="raised", cursor="hand2", padx=10, pady=3).pack(side="left", padx=5)
        
        def save_prefixes():
            """접두사 저장"""
            current = {}
            for prefix, vars_dict in prefix_vars.items():
                new_prefix = vars_dict["prefix"].get().strip()
                new_market = vars_dict["market"].get().strip()
                if new_prefix and new_market:
                    current[new_prefix] = new_market
            
            self.config_manager.save_market_prefixes(current)
            self._log("마켓 접두사 저장 완료")
            messagebox.showinfo("완료", "마켓 접두사가 저장되었습니다.")
            refresh_prefix_list()
        
        # 저장 버튼
        btn_frame = tk.Frame(popup, bg="#FFFFFF")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(btn_frame, text="💾 저장", command=save_prefixes,
                 bg="#3498db", fg="white",
                 font=("맑은 고딕", 10, "bold"),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)
        
        tk.Button(btn_frame, text="취소", command=popup.destroy,
                 bg="#95a5a6", fg="white",
                 font=("맑은 고딕", 10),
                 relief="raised", cursor="hand2",
                 padx=20, pady=5).pack(side="right", padx=5)

# =============================================================================
# 메인 실행
# =============================================================================
if __name__ == "__main__":
    # 모듈 경로 설정
    import sys
    from pathlib import Path
    current_dir = Path(__file__).parent
    parent_dir = current_dir.parent
    if str(parent_dir) not in sys.path:
        sys.path.insert(0, str(parent_dir))
    
    app = UploadMapperGUI()
    app.mainloop()

