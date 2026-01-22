"""
season_filter_manager_gui.py

시즌 필터링 설정 관리 GUI
- Excel 파일 (Season_Filter_Seasons_Keywords.xlsx) 편집
- 시즌, 키워드, 타입 설정 관리
- 추후 data_export.py에 통합 예정
"""

import os
import sys
import json
import warnings
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText


# 환경설정 파일 경로
CONFIG_FILE_NAME = "season_filter_config.json"


def get_config_path() -> str:
    """환경설정 파일 경로 반환"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, CONFIG_FILE_NAME)


def load_default_config() -> Dict:
    """기본 환경설정 로드 (파일이 없으면 기본값 반환)"""
    default_config = {
        "default_sourcing_start_days": 30,
        "default_processing_end_days": 21,
        "default_prep_days": 30,
        "default_grace_days": 21,
        "default_score_min": 1,
        "use_excel_values": True,  # True: Excel 값 우선, False: 환경설정 기본값만 사용
        "common_exclude_keywords": [],  # 공통 제외 키워드 목록
        "description": "시즌 필터링 환경설정\n- Excel에 값이 없을 때 사용되는 기본값\n- use_excel_values가 False이면 Excel 값 무시하고 기본값만 사용"
    }
    
    config_path = get_config_path()
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
                # 기본값과 사용자 설정 병합
                default_config.update(user_config)
                # 공통 제외 키워드가 없으면 빈 리스트로 초기화
                if "common_exclude_keywords" not in default_config:
                    default_config["common_exclude_keywords"] = []
        except Exception as e:
            print(f"[경고] 환경설정 파일 로드 실패, 기본값 사용: {e}")
    
    return default_config


def save_default_config(config: Dict) -> bool:
    """환경설정 저장"""
    try:
        config_path = get_config_path()
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"[오류] 환경설정 파일 저장 실패: {e}")
        return False

# openpyxl Data Validation 경고 필터링
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# tkcalendar 라이브러리 시도 (없으면 간단한 위젯 사용)
try:
    from tkcalendar import DateEntry
    HAS_TKCALENDAR = True
    print("[INFO] tkcalendar 사용 가능 - 달력 위젯 사용")
except ImportError:
    HAS_TKCALENDAR = False
    print("[INFO] tkcalendar 미설치 - 일반 입력 필드 사용 (설치: pip install tkcalendar)")

# 상위 디렉토리에서 모듈 import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


class SeasonFilterManagerGUI(tk.Tk):
    """시즌 필터링 설정 관리 GUI"""
    
    def __init__(self):
        super().__init__()
        self.title("시즌 필터링 설정 관리")
        self.geometry("1400x900")
        
        # 변수 초기화
        self.excel_path = None
        self.excel_data = {}  # {시트명: DataFrame}
        self.original_excel_path = None
        self.sheet_trees = {}  # {시트명: Treeview}
        self.row_count_vars = {}  # {시트명: StringVar}
        self.editing_cell = None  # 현재 편집 중인 셀 정보
        self.has_unsaved_changes = False  # 저장되지 않은 변경사항 추적
        self.auto_backup_enabled = True  # 자동 백업 활성화
        self.auto_backup_interval = 10  # 자동 백업 간격 (분)
        self._auto_backup_job = None  # 자동 백업 스케줄러
        
        # Phase 1: 펼쳐놓고 선택하는 방식
        self.selected_season_id = None  # 현재 선택된 시즌ID
        self.season_list_tree = None  # 시즌 목록 트리뷰
        self.season_detail_frame = None  # 시즌 상세 정보 프레임
        
        # 기본 Excel 파일 경로
        default_excel = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "Season_Filter_Seasons_Keywords.xlsx"
        )
        if os.path.exists(default_excel):
            self.excel_path = default_excel
            self.original_excel_path = default_excel
        
        # UI 구성
        self._configure_styles()
        self._init_ui()
        
        # Excel 파일이 있으면 자동 로드
        if self.excel_path:
            self._load_excel()
        
        # 자동 백업 시작
        self._setup_auto_backup()
        
        # 종료 시 확인
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _configure_styles(self):
        """스타일 설정"""
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except tk.TclError as e:
            self._log(f"테마 설정 실패 (기본 테마 사용): {e}", "WARNING")
        
        bg_color = "#f5f6fa"
        self.configure(background=bg_color)
        
        style.configure("TFrame", background=bg_color)
        style.configure("TLabelframe", background=bg_color, font=("맑은 고딕", 9, "bold"))
        style.configure("TLabelframe.Label", background=bg_color, foreground="#2c3e50")
        style.configure("TLabel", background=bg_color, font=("맑은 고딕", 9))
        style.configure("Action.TButton", font=("맑은 고딕", 10, "bold"), padding=8)
        
        # 단축키 설정
        self._setup_shortcuts()
    
    def _setup_shortcuts(self):
        """단축키 설정"""
        # Ctrl+S: 저장
        self.bind('<Control-s>', lambda e: self._save_excel())
        # Ctrl+O: 파일 열기
        self.bind('<Control-o>', lambda e: self._select_excel_file())
        # F5: 새로고침
        self.bind('<F5>', lambda e: self._reload_excel())
        # Ctrl+Shift+S: 다른 이름으로 저장 (추후 구현 가능)
        # Escape: 편집 취소 (셀 편집 중일 때)
        self.bind('<Escape>', lambda e: self._cancel_editing())
        
        self._log("단축키 설정 완료: Ctrl+S(저장), Ctrl+O(열기), F5(새로고침), Esc(취소)", "INFO")
    
    def _reload_excel(self):
        """Excel 파일 다시 로드"""
        if self.excel_path:
            if self.has_unsaved_changes:
                result = messagebox.askyesno(
                    "새로고침 확인",
                    "저장하지 않은 변경사항이 있습니다.\n\n"
                    "변경사항을 버리고 새로고침하시겠습니까?"
                )
                if not result:
                    return
            
            self.has_unsaved_changes = False
            self.title(self.title().lstrip('*'))
            self._load_excel()
            self._log("Excel 파일 새로고침 완료", "SUCCESS")
    
    def _cancel_editing(self):
        """편집 취소 (셀 편집 중일 때)"""
        if self.editing_cell:
            try:
                tree, item, col_index = self.editing_cell
                # 편집 창이 열려있으면 닫기 (실제로는 편집 창에서 처리)
                # 여기서는 플래그만 초기화
                self.editing_cell = None
            except:
                pass
    
    def _init_ui(self):
        """UI 초기화"""
        main_frame = ttk.Frame(self, padding=15)
        main_frame.pack(fill='both', expand=True)
        
        # 1. 파일 선택 영역
        file_frame = ttk.LabelFrame(main_frame, text="📁 Excel 파일", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))
        
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill='x')
        
        ttk.Label(file_select_frame, text="파일:", width=8).pack(side='left')
        self.excel_path_var = tk.StringVar(value=self.excel_path or "")
        excel_entry = ttk.Entry(file_select_frame, textvariable=self.excel_path_var, state='readonly')
        excel_entry.pack(side='left', fill='x', expand=True, padx=5)
        
        btn_select = ttk.Button(file_select_frame, text="📂 파일 선택", command=self._select_excel_file)
        btn_select.pack(side='right', padx=(5, 0))
        
        btn_reload = ttk.Button(file_select_frame, text="🔄 다시 로드", command=self._load_excel)
        btn_reload.pack(side='right', padx=5)
        
        # 2. 메인 콘텐츠 영역 (좌우 분할 방식)
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Phase 1: 펼쳐놓고 선택하는 방식으로 변경
        # 좌우 분할: 시즌 목록 + 상세 정보
        self._create_season_management_layout(content_frame)
        
        # 기존 탭 방식은 주석 처리 (필요시 활성화 가능)
        # self.notebook = ttk.Notebook(content_frame)
        # self.notebook.pack(fill='both', expand=True)
        # self.sheet_frames = {}
        
        # 3. 하단 버튼 영역
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x')
        
        # Excel 저장 및 JSON 컴파일 통합
        btn_save = ttk.Button(btn_frame, text="💾 저장 및 컴파일", style="Action.TButton", 
                             command=self._save_and_compile)
        btn_save.pack(side='left', padx=(0, 5))
        
        # Excel 열기
        btn_open = ttk.Button(btn_frame, text="📂 Excel 열기", command=self._open_excel)
        btn_open.pack(side='left', padx=5)
        
        btn_config = ttk.Button(btn_frame, text="⚙️ 환경설정", command=self._show_config_dialog)
        btn_config.pack(side='left', padx=5)
        
        
        btn_help = ttk.Button(btn_frame, text="📖 용어 설명", command=self._show_help_dialog)
        btn_help.pack(side='left', padx=5)
        
        # 4. 로그창
        log_frame = ttk.LabelFrame(main_frame, text="📝 로그", padding=10)
        log_frame.pack(fill='both', expand=True)
        
        self.log_widget = ScrolledText(log_frame, height=8, state='disabled', 
                                       font=("Consolas", 9), wrap='word')
        self.log_widget.pack(fill='both', expand=True)
    
    def _log(self, msg: str, level: str = "INFO"):
        """로그 출력"""
        ts = datetime.now().strftime("%H:%M:%S")
        level_symbol = {
            "INFO": "ℹ️",
            "WARNING": "⚠️",
            "ERROR": "❌",
            "SUCCESS": "✅",
            "DEBUG": "🔍"
        }.get(level, "•")
        
        if hasattr(self, 'log_widget') and self.log_widget:
            try:
                self.log_widget.config(state='normal')
                self.log_widget.insert('end', f"[{ts}] {level_symbol} {msg}\n")
                self.log_widget.see('end')
                self.log_widget.config(state='disabled')
            except (tk.TclError, AttributeError) as e:
                # 로그 위젯이 없거나 오류 발생 시 콘솔에 출력
                print(f"[{ts}] {level_symbol} {msg}")
                print(f"로그 위젯 오류: {e}")
    
    def _show_progress_dialog(self, title: str, message: str):
        """진행 상황 다이얼로그 표시"""
        progress_window = tk.Toplevel(self)
        progress_window.title(title)
        progress_window.geometry("400x120")
        progress_window.transient(self)
        progress_window.grab_set()
        
        # 중앙 배치
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (120 // 2)
        progress_window.geometry(f"400x120+{x}+{y}")
        
        frame = ttk.Frame(progress_window, padding=20)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text=message, font=("맑은 고딕", 10)).pack(pady=(0, 15))
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100, length=350, mode='indeterminate')
        progress_bar.pack()
        progress_bar.start(10)  # 애니메이션 시작
        
        progress_window.update()
        
        return progress_window, progress_var, progress_bar
    
    def _close_progress_dialog(self, progress_window):
        """진행 상황 다이얼로그 닫기"""
        try:
            if progress_window and progress_window.winfo_exists():
                progress_window.destroy()
        except (tk.TclError, AttributeError):
            pass
    
    def _show_progress_dialog(self, title: str, message: str):
        """진행 상황 다이얼로그 표시"""
        progress_window = tk.Toplevel(self)
        progress_window.title(title)
        progress_window.geometry("400x120")
        progress_window.transient(self)
        progress_window.grab_set()
        
        # 중앙 배치
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (120 // 2)
        progress_window.geometry(f"400x120+{x}+{y}")
        
        frame = ttk.Frame(progress_window, padding=20)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text=message, font=("맑은 고딕", 10)).pack(pady=(0, 15))
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100, length=350, mode='indeterminate')
        progress_bar.pack()
        progress_bar.start(10)  # 애니메이션 시작
        
        progress_window.update()
        
        return progress_window, progress_var, progress_bar
    
    def _close_progress_dialog(self, progress_window):
        """진행 상황 다이얼로그 닫기"""
        try:
            if progress_window and progress_window.winfo_exists():
                progress_window.destroy()
        except (tk.TclError, AttributeError):
            pass
    
    def _select_excel_file(self):
        """Excel 파일 선택"""
        path = filedialog.askopenfilename(
            title="시즌 필터 Excel 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Season_Filter_Seasons_Keywords.xlsx"
        )
        if path:
            self.excel_path = path
            self.original_excel_path = path
            self.excel_path_var.set(path)
            self._log(f"파일 선택: {os.path.basename(path)}")
            self._load_excel()
    
    def _find_backup_files(self) -> List[str]:
        """백업 파일 검색"""
        if not self.excel_path:
            return []
        
        import glob
        dir_path = os.path.dirname(self.excel_path)
        base_name = os.path.basename(self.excel_path)
        base_name_no_ext = os.path.splitext(base_name)[0]
        
        # 백업 파일 패턴: *_백업_*.xlsx
        backup_pattern = os.path.join(dir_path, f"{base_name_no_ext}*백업*.xlsx")
        backup_files = glob.glob(backup_pattern)
        
        # 파일명으로 정렬 (최신 것이 먼저)
        backup_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        backup_files = [os.path.basename(f) for f in backup_files]
        
        return backup_files
    
    def _load_excel(self):
        """Excel 파일 로드"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            messagebox.showwarning("오류", "Excel 파일을 선택해주세요.")
            return
        
        # zipfile 모듈 import (에러 처리에서도 사용)
        import zipfile
        
        # 진행 상황 다이얼로그 변수
        progress_window = None
        progress_var = None
        
        try:
            # 진행 상황 다이얼로그 표시
            progress_window, progress_var, progress_bar = self._show_progress_dialog(
                "Excel 로드 중", 
                f"Excel 파일을 로드하는 중입니다...\n{os.path.basename(self.excel_path)}"
            )
            
            self._log(f"Excel 파일 로드 중: {os.path.basename(self.excel_path)}")
            
            # 파일 무결성 검사 (zipfile으로 .xlsx 파일 검증)
            try:
                with zipfile.ZipFile(self.excel_path, 'r') as zip_ref:
                    file_list = zip_ref.namelist()
                    if '[Content_Types].xml' not in file_list:
                        raise zipfile.BadZipFile("필수 파일 [Content_Types].xml이 없습니다.")
            except zipfile.BadZipFile as e:
                self._close_progress_dialog(progress_window)
                error_msg = f"Excel 파일이 손상되었습니다.\n\n오류: {e}\n\n해결 방법:\n1. Excel에서 파일을 열어 복구 시도\n2. 백업 파일 사용\n3. 파일을 다시 저장"
                self._log(f"파일 무결성 검사 실패: {e}", "ERROR")
                messagebox.showerror("파일 손상", error_msg)
                return
            except Exception as e:
                self._log(f"파일 검증 중 경고: {e}", "WARNING")
            
            # Excel 파일 읽기 (engine 명시적으로 지정)
            xl = pd.ExcelFile(self.excel_path, engine='openpyxl')
            sheet_names = xl.sheet_names
            
            self._log(f"시트 목록: {', '.join(sheet_names)}")
            
            # 기존 데이터 초기화
            self.excel_data.clear()
            
            # 각 시트별로 데이터 로드
            for sheet_name in sheet_names:
                try:
                    # SEASON_MASTER 시트는 날짜 컬럼을 텍스트로 읽기 위해 dtype 지정
                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                        # 먼저 컬럼명만 읽기
                        df_temp = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl', nrows=0)
                        date_columns = []
                        
                        # 날짜 컬럼 찾기
                        for col in df_temp.columns:
                            col_lower = str(col).lower()
                            if any(keyword in col_lower for keyword in ['시작', 'start', 'start_mmdd', 'start_date']):
                                date_columns.append(col)
                            if any(keyword in col_lower for keyword in ['종료', 'end', 'end_mmdd', 'end_date']):
                                date_columns.append(col)
                        
                        # 날짜 컬럼을 문자열로 지정하여 읽기
                        dtype_dict = {}
                        for col in date_columns:
                            dtype_dict[col] = str
                        
                        if dtype_dict:
                            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl', dtype=dtype_dict)
                        else:
                            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl')
                    else:
                        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl')
                    
                    # 빈 행 제거
                    df = df.dropna(how='all')
                    
                    # SEASON_MASTER 시트의 날짜 컬럼을 문자열로 변환
                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                        start_col = self._find_column(df, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
                        end_col = self._find_column(df, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
                        
                        self._log(f"시트 '{sheet_name}': 시작일 컬럼={start_col}, 종료일 컬럼={end_col}")
                        
                        # datetime 객체를 MM-DD 형식 문자열로 변환
                        if start_col and start_col in df.columns:
                            df[start_col] = df[start_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                        if end_col and end_col in df.columns:
                            df[end_col] = df[end_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                    
                    self.excel_data[sheet_name] = df
                    
                    # 디버깅: 컬럼명 출력
                    self._log(f"시트 '{sheet_name}' 컬럼: {list(df.columns)}")
                    self._log(f"시트 '{sheet_name}' 로드 완료: {len(df)}행, {len(df.columns)}컬럼")
                except Exception as e:
                    self._log(f"시트 '{sheet_name}' 로드 실패: {e}", "ERROR")
                    import traceback
                    self._log(traceback.format_exc(), "ERROR")
                
                # 진행 상황 업데이트
                if progress_var and progress_window:
                    try:
                        progress_var.set((i + 1) / total_sheets * 100)
                        progress_window.update()
                    except:
                        pass
            
            self._log(f"Excel 파일 로드 완료: {len(sheet_names)}개 시트", "SUCCESS")
            
            # 진행 상황 다이얼로그 닫기
            self._close_progress_dialog(progress_window)
            
            # Phase 1: 시즌 목록 트리뷰 업데이트 (새 레이아웃 사용 시)
            if hasattr(self, 'season_list_tree') and self.season_list_tree:
                self._update_season_list_tree()
            
        except KeyError as e:
            self._close_progress_dialog(progress_window)
            error_msg = str(e)
            if "[Content_Types].xml" in error_msg or "Content_Types" in error_msg:
                self._log(f"Excel 파일 손상: {error_msg}", "ERROR")
                # 백업 파일 자동 검색
                backup_files = self._find_backup_files()
                if backup_files:
                    backup_msg = f"Excel 파일이 손상되었습니다.\n\n오류: {error_msg}\n\n다음 백업 파일을 찾았습니다:\n" + "\n".join(backup_files[:3])
                    backup_msg += "\n\n백업 파일 중 하나를 사용하시겠습니까?"
                    if messagebox.askyesno("파일 손상", backup_msg):
                        # 첫 번째 백업 파일 사용
                        backup_path = os.path.join(os.path.dirname(self.excel_path), backup_files[0])
                        if os.path.exists(backup_path):
                            self.excel_path = backup_path
                            self._log(f"백업 파일로 전환: {backup_files[0]}")
                            # 재시도
                            self._load_excel()
                            return
                else:
                    messagebox.showerror(
                        "파일 손상", 
                        f"Excel 파일이 손상되었습니다.\n\n오류: {error_msg}\n\n해결 방법:\n"
                        "1. Excel에서 파일을 열어 '복구' 시도\n"
                        "2. 마지막으로 저장한 버전 확인\n"
                        "3. 파일이 다른 프로그램에서 열려있으면 닫기"
                    )
            else:
                self._log(f"Excel 파일 로드 실패: {error_msg}", "ERROR")
                import traceback
                self._log(traceback.format_exc(), "ERROR")
                messagebox.showerror("오류", f"Excel 파일을 읽는 중 오류가 발생했습니다:\n{error_msg}")
        except zipfile.BadZipFile as e:
            self._close_progress_dialog(progress_window)
            self._log(f"Excel 파일 손상 (BadZipFile): {e}", "ERROR")
            backup_files = self._find_backup_files()
            if backup_files:
                backup_msg = f"Excel 파일이 손상되었습니다.\n\n다음 백업 파일을 찾았습니다:\n" + "\n".join(backup_files[:3])
                backup_msg += "\n\n백업 파일을 사용하시겠습니까?"
                if messagebox.askyesno("파일 손상", backup_msg):
                    backup_path = os.path.join(os.path.dirname(self.excel_path), backup_files[0])
                    if os.path.exists(backup_path):
                        self.excel_path = backup_path
                        self._log(f"백업 파일로 전환: {backup_files[0]}")
                        self._load_excel()
                        return
            messagebox.showerror(
                "파일 손상", 
                f"Excel 파일이 손상되었습니다.\n\n오류: {e}\n\n해결 방법:\n"
                "1. Excel에서 파일을 열어 복구 시도\n"
                "2. 마지막으로 저장한 버전 확인"
            )
        except Exception as e:
            self._close_progress_dialog(progress_window)
            self._log(f"Excel 파일 로드 실패: {e}", "ERROR")
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            error_msg = str(e)
            if "Permission" in error_msg or "denied" in error_msg.lower():
                messagebox.showerror("접근 오류", "Excel 파일이 다른 프로그램에서 열려있거나 접근 권한이 없습니다.\n\n파일을 닫고 다시 시도해주세요.")
            else:
                messagebox.showerror("오류", f"Excel 파일을 읽는 중 오류가 발생했습니다:\n{error_msg}")
    
    def _create_season_management_layout(self, parent_frame: ttk.Frame):
        """Phase 1: 펼쳐놓고 선택하는 방식 - 시즌 관리 레이아웃 생성"""
        # 좌우 분할 프레임
        paned = ttk.PanedWindow(parent_frame, orient='horizontal')
        paned.pack(fill='both', expand=True)
        
        # 왼쪽: 시즌 목록
        left_frame = ttk.LabelFrame(paned, text="📋 시즌 목록", padding=10)
        paned.add(left_frame, weight=1)
        self._create_season_list_panel(left_frame)
        
        # 오른쪽: 시즌 상세 정보
        right_frame = ttk.LabelFrame(paned, text="📝 시즌 상세 정보", padding=10)
        paned.add(right_frame, weight=2)
        self._create_season_detail_panel(right_frame)
    
    def _create_season_list_panel(self, parent_frame: ttk.Frame):
        """시즌 목록 패널 생성 (왼쪽)"""
        # 검색 프레임
        search_frame = ttk.Frame(parent_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 검색:", width=8).pack(side='left', padx=(0, 5))
        self.season_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.season_search_var, width=20)
        search_entry.pack(side='left', fill='x', expand=True, padx=(0, 5))
        search_entry.bind('<KeyRelease>', lambda e: self._filter_season_list())
        
        # 검색 초기화 버튼
        ttk.Button(search_frame, text="✕", width=3, command=self._clear_season_search).pack(side='left')
        
        # 필터 프레임
        filter_frame = ttk.Frame(parent_frame)
        filter_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(filter_frame, text="필터:", width=8).pack(side='left', padx=(0, 5))
        self.season_filter_var = tk.StringVar(value="전체")
        filter_combo = ttk.Combobox(filter_frame, textvariable=self.season_filter_var, 
                                   values=["전체", "활성", "비활성"], state='readonly', width=10)
        filter_combo.pack(side='left')
        filter_combo.bind('<<ComboboxSelected>>', lambda e: self._filter_season_list())
        
        # 상단 버튼
        btn_frame = ttk.Frame(parent_frame)
        btn_frame.pack(fill='x', pady=(0, 10))
        
        btn_add = ttk.Button(btn_frame, text="➕ 시즌 추가", command=self._add_season_dialog)
        btn_add.pack(side='left', padx=(0, 5))
        
        btn_refresh = ttk.Button(btn_frame, text="🔄 새로고침", command=self._refresh_season_list)
        btn_refresh.pack(side='left', padx=5)
        
        # 시즌 목록 트리뷰
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill='both', expand=True)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical')
        
        # 트리뷰 생성
        columns = ("시즌명", "기간", "상태")
        self.season_list_tree = ttk.Treeview(
            tree_frame, 
            columns=columns, 
            show='tree headings',
            yscrollcommand=scrollbar.set,
            selectmode='browse'
        )
        
        # 컬럼 설정
        self.season_list_tree.heading("#0", text="타입 / 시즌ID")
        self.season_list_tree.column("#0", width=200, anchor='w')
        
        self.season_list_tree.heading("시즌명", text="시즌명")
        self.season_list_tree.column("시즌명", width=150, anchor='w')
        
        self.season_list_tree.heading("기간", text="기간")
        self.season_list_tree.column("기간", width=150, anchor='w')
        
        self.season_list_tree.heading("상태", text="상태")
        self.season_list_tree.column("상태", width=100, anchor='center')
        
        # 타입별 색상 태그 설정
        self.season_list_tree.tag_configure("type_event", foreground="#2196F3")
        self.season_list_tree.tag_configure("type_climate", foreground="#4CAF50")
        self.season_list_tree.tag_configure("type_activity", foreground="#FF9800")
        self.season_list_tree.tag_configure("type_lifecycle", foreground="#9C27B0")
        self.season_list_tree.tag_configure("type_other", foreground="#757575")
        
        # 상태별 색상 태그
        self.season_list_tree.tag_configure("status_sourcing", foreground="#2196F3", background="#E3F2FD")
        self.season_list_tree.tag_configure("status_active", foreground="#4CAF50", background="#E8F5E9")
        self.season_list_tree.tag_configure("status_expired", foreground="#F44336", background="#FFEBEE")
        self.season_list_tree.tag_configure("status_inactive", foreground="#9E9E9E", background="#F5F5F5")
        
        # 스크롤바 연결
        scrollbar.config(command=self.season_list_tree.yview)
        
        # 배치
        self.season_list_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 시즌 선택 이벤트
        self.season_list_tree.bind('<<TreeviewSelect>>', self._on_season_select)
        self.season_list_tree.bind('<Double-1>', lambda e: self._edit_season_dialog())
    
    def _create_season_detail_panel(self, parent_frame: ttk.Frame):
        """시즌 상세 정보 패널 생성 (오른쪽) - 노트북으로 섹션 분리"""
        self.season_detail_frame = parent_frame
        
        # 노트북 (탭)으로 섹션 분리
        notebook = ttk.Notebook(parent_frame)
        notebook.pack(fill='both', expand=True)
        
        # 탭 1: 시즌 정보
        season_info_frame = ttk.Frame(notebook, padding=10)
        notebook.add(season_info_frame, text="📝 시즌 정보")
        
        # 스크롤 가능한 프레임 (탭 1 내부)
        canvas = tk.Canvas(season_info_frame)
        scrollbar = ttk.Scrollbar(season_info_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 기본 정보 섹션
        basic_frame = ttk.LabelFrame(scrollable_frame, text="기본 정보", padding=10)
        basic_frame.pack(fill='x', pady=(0, 10))
        self.basic_info_frame = basic_frame
        
        # 키워드 섹션
        keyword_frame = ttk.LabelFrame(scrollable_frame, text="키워드", padding=10)
        keyword_frame.pack(fill='both', expand=True)
        self.keyword_info_frame = keyword_frame
        
        # 액션 버튼
        action_frame = ttk.Frame(scrollable_frame)
        action_frame.pack(fill='x', pady=(10, 0))
        self.action_frame = action_frame
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 탭 2: 예외단어 관리
        exclude_tab_frame = ttk.Frame(notebook, padding=10)
        notebook.add(exclude_tab_frame, text="🚫 예외단어 관리")
        self.exclude_keywords_frame = exclude_tab_frame
        self._create_exclude_keywords_panel(exclude_tab_frame)
        
        self.detail_canvas = canvas
        self.detail_scrollable_frame = scrollable_frame
        
        # 초기 상태: 시즌 미선택
        self._show_no_selection_message()
    
    def _show_no_selection_message(self):
        """시즌 미선택 상태 메시지 표시"""
        # 기존 위젯 제거
        for widget in self.basic_info_frame.winfo_children():
            widget.destroy()
        for widget in self.keyword_info_frame.winfo_children():
            widget.destroy()
        for widget in self.action_frame.winfo_children():
            widget.destroy()
        
        # 메시지 표시
        ttk.Label(
            self.basic_info_frame, 
            text="왼쪽에서 시즌을 선택하거나 새 시즌을 추가하세요.",
            font=("맑은 고딕", 10),
            foreground="#666"
        ).pack(pady=20)
    
    def _refresh_season_list(self):
        """시즌 목록 새로고침"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            return
        
        try:
            # Excel 파일 다시 로드
            self._load_excel()
            # 시즌 목록 업데이트
            self._update_season_list_tree()
        except Exception as e:
            self._log(f"시즌 목록 새로고침 실패: {e}")
    
    def _update_season_list_tree(self):
        """시즌 목록 트리뷰 업데이트 (타입별 그룹화)"""
        if not self.season_list_tree:
            return
        
        # 기존 항목 제거
        for item in self.season_list_tree.get_children():
            self.season_list_tree.delete(item)
        
        # SEASON_MASTER 시트에서 시즌 정보 가져오기
        season_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                season_sheet = sheet_name
                break
        
        if not season_sheet:
            self._log("⚠️ SEASON_MASTER 시트를 찾을 수 없습니다.")
            return
        
        df_seasons = self.excel_data[season_sheet]
        
        # 컬럼 찾기 (괄호 포함 형식 지원)
        season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
        type_col = self._find_column(df_seasons, ["타입", "type", "category", "타입(type: Event/Climate/Activity/Lifecycle)"])
        start_col = self._find_column(df_seasons, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
        end_col = self._find_column(df_seasons, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
        cross_year_col = self._find_column(df_seasons, ["연도교차", "cross_year", "crossyear", "연도교차(Y/N)(cross_year)"])
        use_col = self._find_column(df_seasons, ["사용여부", "use", "enabled", "active", "사용여부(Y/N)(enabled)", "enabled"])
        
        if not season_id_col:
            self._log("⚠️ 시즌ID 컬럼을 찾을 수 없습니다.")
            return
        
        # 시즌 설정 로드 (유효성 체크용)
        season_config = {}
        if hasattr(self, 'excel_path') and self.excel_path:
            try:
                json_path = os.path.join(os.path.dirname(self.excel_path), "season_filters.json")
                season_config = load_season_config(self.excel_path, json_path) or {}
            except Exception as e:
                self._log(f"시즌 설정 로드 실패: {e}", "WARNING")
                season_config = {}
        
        # 현재 날짜
        current_date = datetime.now()
        
        # 타입별로 시즌 그룹화
        type_groups = {}
        type_labels = {
            "Event": "🎉 이벤트",
            "Climate": "🌤️ 기후",
            "Activity": "🏃 활동",
            "Lifecycle": "🔄 생활주기",
        }
        
        # 시즌 데이터 수집
        for idx, row in df_seasons.iterrows():
            season_id = str(row.get(season_id_col, "")).strip()
            if not season_id or season_id in ['nan', 'None', '']:
                continue
            
            # 타입 가져오기
            season_type = str(row.get(type_col, "")).strip() if type_col else "Other"
            if not season_type or season_type in ['nan', 'None', '']:
                season_type = "Other"
            
            # 타입 라벨
            type_label = type_labels.get(season_type, f"📦 {season_type}")
            
            if season_type not in type_groups:
                type_groups[season_type] = {
                    "label": type_label,
                    "seasons": []
                }
            
            # 사용여부 체크
            is_active = True
            if use_col:
                use_val = str(row.get(use_col, "Y")).strip().upper()
                is_active = use_val in ['Y', 'YES', 'TRUE', '1']
            
            # 시즌명
            season_name = str(row.get(season_name_col, season_id)).strip() if season_name_col else season_id
            
            # 기간
            start_val = row.get(start_col, "")
            end_val = row.get(end_col, "")
            start_date_str = self._convert_date_to_mmdd(start_val) if pd.notna(start_val) else ""
            end_date_str = self._convert_date_to_mmdd(end_val) if pd.notna(end_val) else ""
            period = f"{start_date_str} ~ {end_date_str}" if start_date_str and end_date_str else ""
            
            # 연도교차
            cross_year_val = str(row.get(cross_year_col, "N")).strip().upper() if cross_year_col else "N"
            cross_year = cross_year_val in ['Y', 'YES', 'TRUE', '1']
            
            # 시즌 유효성 체크 (활성인 경우만)
            validity_status = None
            if is_active and start_date_str and end_date_str:
                try:
                    # 시즌 정보 구성
                    season_info = {
                        "start_date": start_date_str,
                        "end_date": end_date_str,
                        "cross_year": cross_year
                    }
                    # 유효성 체크
                    validity_status = _check_season_validity(season_info, current_date, season_config)
                except Exception as e:
                    self._log(f"시즌 유효성 체크 실패 ({season_id}): {e}", "WARNING")
                    validity_status = None
            
            # 상태 표시
            if not is_active:
                status_text = "❌ 비활성"
                status_tag = "status_inactive"
            elif validity_status == "SOURCING":
                status_text = "🔵 소싱"
                status_tag = "status_sourcing"
            elif validity_status == "ACTIVE":
                status_text = "✅ 출력 가능"
                status_tag = "status_active"
            elif validity_status == "EXPIRED":
                status_text = "🔴 종료"
                status_tag = "status_expired"
            else:
                status_text = "⚪ 미정"
                status_tag = ""
            
            type_groups[season_type]["seasons"].append({
                "id": season_id,
                "name": season_name,
                "period": period,
                "status": status_text,
                "status_tag": status_tag,
                "validity": validity_status,
                "is_active": is_active
            })
        
        # 타입별로 트리뷰 구성
        type_order = ["Event", "Climate", "Activity", "Lifecycle"]
        for type_key in type_order:
            if type_key not in type_groups:
                continue
            
            type_data = type_groups[type_key]
            type_label = type_data["label"]
            seasons = type_data["seasons"]
            
            # 타입별 상태 카운트
            sourcing_count = sum(1 for s in seasons if s.get("validity") == "SOURCING")
            active_count = sum(1 for s in seasons if s.get("validity") == "ACTIVE")
            expired_count = sum(1 for s in seasons if s.get("validity") == "EXPIRED")
            inactive_count = sum(1 for s in seasons if not s.get("is_active"))
            
            # 타입 노드 생성
            type_tag = f"type_{type_key.lower()}"
            type_summary = f"소싱:{sourcing_count} 출력:{active_count} 종료:{expired_count}"
            if inactive_count > 0:
                type_summary += f" 비활성:{inactive_count}"
            
            type_node = self.season_list_tree.insert(
                '', 'end',
                text=f"{type_label} ({len(seasons)}개)",
                values=("", "", type_summary),
                tags=(type_tag,),
                open=True  # 기본적으로 펼쳐짐
            )
            
            # 시즌들을 타입 노드 아래에 추가
            for season in seasons:
                self.season_list_tree.insert(
                    type_node, 'end',
                    text=season["id"],
                    values=(season["name"], season["period"], season["status"]),
                    tags=(season["id"], season["status_tag"]) if season["status_tag"] else (season["id"],)
                )
        
        # 기타 타입 처리
        for type_key, type_data in type_groups.items():
            if type_key not in type_order:
                type_label = type_data["label"]
                seasons = type_data["seasons"]
                
                sourcing_count = sum(1 for s in seasons if s.get("validity") == "SOURCING")
                active_count = sum(1 for s in seasons if s.get("validity") == "ACTIVE")
                expired_count = sum(1 for s in seasons if s.get("validity") == "EXPIRED")
                inactive_count = sum(1 for s in seasons if not s.get("is_active"))
                
                type_tag = f"type_{type_key.lower()}"
                type_summary = f"소싱:{sourcing_count} 출력:{active_count} 종료:{expired_count}"
                if inactive_count > 0:
                    type_summary += f" 비활성:{inactive_count}"
                
                type_node = self.season_list_tree.insert(
                    '', 'end',
                    text=f"{type_label} ({len(seasons)}개)",
                    values=("", "", type_summary),
                    tags=(type_tag,),
                    open=True
                )
                
                for season in seasons:
                    self.season_list_tree.insert(
                        type_node, 'end',
                        text=season["id"],
                        values=(season["name"], season["period"], season["status"]),
                        tags=(season["id"], season["status_tag"]) if season["status_tag"] else (season["id"],)
                    )
        
        total_count = sum(len(g["seasons"]) for g in type_groups.values())
        self._log(f"시즌 목록 업데이트 완료: {len(type_groups)}개 타입, 총 {total_count}개 시즌", "INFO")
        
        # 검색/필터 적용
        if hasattr(self, 'season_search_var') and hasattr(self, 'season_filter_var'):
            self._filter_season_list()
    
    def _filter_season_list(self):
        """시즌 목록 필터링 (타입별 트리 구조 지원)"""
        if not self.season_list_tree:
            return
        
        search_query = self.season_search_var.get().strip().lower() if hasattr(self, 'season_search_var') else ""
        filter_value = self.season_filter_var.get() if hasattr(self, 'season_filter_var') else "전체"
        
        visible_count = 0
        hidden_count = 0
        
        # 타입 노드 순회
        for type_node in self.season_list_tree.get_children():
            type_values = self.season_list_tree.item(type_node, 'values')
            type_text = self.season_list_tree.item(type_node, 'text')
            type_tags = list(self.season_list_tree.item(type_node, 'tags'))
            
            # 타입 노드가 타입 태그를 가지고 있는지 확인
            is_type_node = any('type_' in tag for tag in type_tags)
            
            if is_type_node:
                # 타입 노드인 경우 - 자식 시즌들을 확인
                visible_seasons = []
                for season_node in self.season_list_tree.get_children(type_node):
                    season_values = self.season_list_tree.item(season_node, 'values')
                    season_text = self.season_list_tree.item(season_node, 'text')
                    season_tags = list(self.season_list_tree.item(season_node, 'tags'))
                    
                    # 검색어 매칭
                    matches_search = True
                    if search_query:
                        search_text = f"{season_text} {' '.join([str(v) for v in season_values])}".lower()
                        matches_search = search_query in search_text
                    
                    # 필터 매칭 (상태 확인)
                    matches_filter = True
                    if filter_value == "활성":
                        status_text = str(season_values[2]) if len(season_values) > 2 else ""
                        matches_filter = ("✅" in status_text or "🔵" in status_text or "출력" in status_text or "소싱" in status_text) and "❌" not in status_text
                    elif filter_value == "비활성":
                        status_text = str(season_values[2]) if len(season_values) > 2 else ""
                        matches_filter = "❌" in status_text or "비활성" in status_text
                    
                    if matches_search and matches_filter:
                        visible_seasons.append(season_node)
                        visible_count += 1
                    else:
                        hidden_count += 1
                
                # 타입 노드 표시/숨김 결정 (자식이 하나라도 보이면 표시)
                if visible_seasons:
                    # 타입 노드 표시
                    self.season_list_tree.item(type_node, open=True)  # 자식이 있으면 펼치기
                    visible_count += 1
                else:
                    # 타입 노드 숨김
                    hidden_count += 1
            else:
                # 일반 노드인 경우 (예전 형식 호환)
                values = type_values
                text = type_text
                
                matches_search = True
                if search_query:
                    search_text = f"{text} {' '.join([str(v) for v in values])}".lower()
                    matches_search = search_query in search_text
                
                matches_filter = True
                if filter_value == "활성":
                    matches_filter = len(values) > 2 and ("활성" in str(values[2]) or "✅" in str(values[2]))
                elif filter_value == "비활성":
                    matches_filter = len(values) > 2 and ("비활성" in str(values[2]) or "❌" in str(values[2]))
                
                if matches_search and matches_filter:
                    visible_count += 1
                else:
                    hidden_count += 1
        
        # 검색 결과 표시
        if search_query or filter_value != "전체":
            self._log(f"검색 결과: {visible_count}개 표시, {hidden_count}개 숨김", "INFO")
    
    def _clear_season_search(self):
        """검색 초기화"""
        if hasattr(self, 'season_search_var'):
            self.season_search_var.set("")
        if hasattr(self, 'season_filter_var'):
            self.season_filter_var.set("전체")
        self._filter_season_list()
    
    def _on_season_select(self, event):
        """시즌 선택 시 상세 정보 업데이트 (타입 노드 선택 방지)"""
        selection = self.season_list_tree.selection()
        if not selection:
            self._show_no_selection_message()
            return
        
        item = selection[0]
        tags = self.season_list_tree.item(item, 'tags')
        
        # 타입 노드인지 확인 (type_ 태그가 있으면 타입 노드)
        is_type_node = any('type_' in str(tag) for tag in tags)
        if is_type_node:
            # 타입 노드는 선택하지 않음
            self.season_list_tree.selection_remove(item)
            self._show_no_selection_message()
            return
        
        # 부모 노드가 있는지 확인 (자식 노드인지)
        parent = self.season_list_tree.parent(item)
        if not parent:
            # 최상위 노드인데 타입 노드가 아니면 선택하지 않음
            self._show_no_selection_message()
            return
        
        season_id = self.season_list_tree.item(item, 'text')
        self.selected_season_id = season_id
        
        # 시즌 상세 정보 표시
        self._update_season_detail(season_id)
    
    def _update_season_detail(self, season_id: str):
        """시즌 상세 정보 업데이트"""
        # 기존 위젯 제거
        for widget in self.basic_info_frame.winfo_children():
            widget.destroy()
        for widget in self.keyword_info_frame.winfo_children():
            widget.destroy()
        for widget in self.action_frame.winfo_children():
            widget.destroy()
        
        # SEASON_MASTER 시트에서 시즌 정보 찾기
        season_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                season_sheet = sheet_name
                break
        
        if not season_sheet:
            ttk.Label(self.basic_info_frame, text="시즌 정보를 찾을 수 없습니다.").pack()
            return
        
        df_seasons = self.excel_data[season_sheet]
        
        # 컬럼 찾기 (괄호 포함 형식 지원)
        season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
        start_col = self._find_column(df_seasons, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
        end_col = self._find_column(df_seasons, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
        cross_col = self._find_column(df_seasons, ["cross_year", "연도넘김", "연도초과", "시즌교차여부", "연도교차(Y/N)(cross_year)"])
        sourcing_start_col = self._find_column(df_seasons, ["소싱시작일수", "sourcing_start_days", "소싱시작"])
        processing_end_col = self._find_column(df_seasons, ["가공완료마감일수", "processing_end_days", "가공완료마감"])
        use_col = self._find_column(df_seasons, ["사용여부", "use", "enabled", "active"])
        
        # 시즌 정보 찾기
        season_row = None
        for idx, row in df_seasons.iterrows():
            if str(row.get(season_id_col, "")).strip() == season_id:
                season_row = row
                break
        
        if season_row is None:
            ttk.Label(self.basic_info_frame, text=f"시즌 '{season_id}'를 찾을 수 없습니다.").pack()
            return
        
        # 기본 정보 표시
        info_grid = ttk.Frame(self.basic_info_frame)
        info_grid.pack(fill='x')
        
        # 시즌ID
        ttk.Label(info_grid, text="시즌ID:", width=15, anchor='e').grid(row=0, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=season_id, font=("맑은 고딕", 9, "bold")).grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # 시즌명
        season_name = str(season_row.get(season_name_col, season_id)).strip() if season_name_col else season_id
        ttk.Label(info_grid, text="시즌명:", width=15, anchor='e').grid(row=1, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=season_name).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        # 시작일 (날짜 형식 변환)
        if start_col:
            start_val = season_row.get(start_col, "")
            self._log(f"시즌 '{season_id}' 시작일 컬럼 '{start_col}' 값: {start_val} (타입: {type(start_val)})")
            if pd.notna(start_val) and str(start_val).strip() and str(start_val).strip() not in ['nan', 'None', '']:
                start_date = self._convert_date_to_mmdd(start_val)
                self._log(f"  → 변환된 시작일: {start_date}")
            else:
                start_date = ""
        else:
            start_date = ""
            self._log(f"⚠️ 시즌 '{season_id}': 시작일 컬럼을 찾을 수 없습니다. 사용 가능한 컬럼: {list(df_seasons.columns)}")
        ttk.Label(info_grid, text="시작일:", width=15, anchor='e').grid(row=2, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=start_date if start_date else "(없음)", 
                 foreground="#666" if not start_date else "#000").grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        # 종료일 (날짜 형식 변환)
        if end_col:
            end_val = season_row.get(end_col, "")
            self._log(f"시즌 '{season_id}' 종료일 컬럼 '{end_col}' 값: {end_val} (타입: {type(end_val)})")
            if pd.notna(end_val) and str(end_val).strip() and str(end_val).strip() not in ['nan', 'None', '']:
                end_date = self._convert_date_to_mmdd(end_val)
                self._log(f"  → 변환된 종료일: {end_date}")
            else:
                end_date = ""
        else:
            end_date = ""
            self._log(f"⚠️ 시즌 '{season_id}': 종료일 컬럼을 찾을 수 없습니다. 사용 가능한 컬럼: {list(df_seasons.columns)}")
        ttk.Label(info_grid, text="종료일:", width=15, anchor='e').grid(row=3, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=end_date if end_date else "(없음)", 
                 foreground="#666" if not end_date else "#000").grid(row=3, column=1, sticky='w', padx=5, pady=5)
        
        # 연도넘김
        cross_year_val = False
        if cross_col:
            cross_val = str(season_row.get(cross_col, "")).strip().upper()
            cross_year_val = cross_val in ['Y', 'YES', 'TRUE', '1']
        ttk.Label(info_grid, text="연도넘김:", width=15, anchor='e').grid(row=4, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text="예" if cross_year_val else "아니오").grid(row=4, column=1, sticky='w', padx=5, pady=5)
        
        # 소싱시작일수
        sourcing_start_days = 30  # 기본값
        if sourcing_start_col:
            try:
                val = season_row.get(sourcing_start_col)
                if pd.notna(val) and str(val).strip():
                    sourcing_start_days = int(val)
            except (ValueError, TypeError) as e:
                self._log(f"소싱시작일수 변환 오류 (시즌={season_id}): {e}", "WARNING")
        ttk.Label(info_grid, text="소싱시작일수:", width=15, anchor='e').grid(row=5, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=f"{sourcing_start_days}일").grid(row=5, column=1, sticky='w', padx=5, pady=5)
        
        # 가공완료마감일수
        processing_end_days = 21  # 기본값
        if processing_end_col:
            try:
                val = season_row.get(processing_end_col)
                if pd.notna(val) and str(val).strip():
                    processing_end_days = int(val)
            except (ValueError, TypeError) as e:
                self._log(f"가공완료마감일수 변환 오류 (시즌={season_id}): {e}", "WARNING")
        ttk.Label(info_grid, text="가공완료마감일수:", width=15, anchor='e').grid(row=6, column=0, sticky='e', padx=5, pady=5)
        ttk.Label(info_grid, text=f"{processing_end_days}일").grid(row=6, column=1, sticky='w', padx=5, pady=5)
        
        # 계산된 기간 표시
        if start_date and end_date:
            try:
                from datetime import datetime, timedelta
                # 날짜 파싱 (MM-DD 형식)
                start_month, start_day = map(int, start_date.split('-'))
                end_month, end_day = map(int, end_date.split('-'))
                current_year = datetime.now().year
                
                start_dt = datetime(current_year, start_month, start_day)
                end_dt = datetime(current_year, end_month, end_day)
                
                # 소싱 시작일 계산
                sourcing_start_dt = start_dt - timedelta(days=sourcing_start_days)
                # 가공 완료 마감일 계산
                processing_end_dt = end_dt - timedelta(days=processing_end_days)
                
                ttk.Separator(info_grid, orient='horizontal').grid(row=7, column=0, columnspan=2, sticky='ew', pady=10)
                
                ttk.Label(info_grid, text="소싱 시작일:", width=15, anchor='e', font=("맑은 고딕", 9, "bold")).grid(row=8, column=0, sticky='e', padx=5, pady=5)
                ttk.Label(info_grid, text=sourcing_start_dt.strftime("%m-%d"), font=("맑은 고딕", 9, "bold"), foreground="#2c3e50").grid(row=8, column=1, sticky='w', padx=5, pady=5)
                
                ttk.Label(info_grid, text="가공 완료 마감:", width=15, anchor='e', font=("맑은 고딕", 9, "bold")).grid(row=9, column=0, sticky='e', padx=5, pady=5)
                ttk.Label(info_grid, text=processing_end_dt.strftime("%m-%d"), font=("맑은 고딕", 9, "bold"), foreground="#2c3e50").grid(row=9, column=1, sticky='w', padx=5, pady=5)
            except Exception as e:
                self._log(f"기간 계산 오류: {e}")
        
        # 키워드 표시
        self._display_keywords(season_id)
        
        # 액션 버튼
        btn_edit = ttk.Button(self.action_frame, text="✏️ 시즌 수정", command=self._edit_season_dialog)
        btn_edit.pack(side='left', padx=5)
        
        btn_delete = ttk.Button(self.action_frame, text="🗑️ 시즌 삭제", command=self._delete_season_dialog)
        btn_delete.pack(side='left', padx=5)
        
        btn_add_keyword = ttk.Button(self.action_frame, text="➕ 키워드 추가", command=self._add_keyword_dialog)
        btn_add_keyword.pack(side='left', padx=5)
    
    def _display_keywords(self, season_id: str):
        """키워드 표시"""
        # KEYWORDS 시트 찾기
        keyword_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                keyword_sheet = sheet_name
                break
        
        if not keyword_sheet:
            ttk.Label(self.keyword_info_frame, text="KEYWORDS 시트를 찾을 수 없습니다.").pack()
            return
        
        df_keywords = self.excel_data[keyword_sheet]
        
        # 컬럼 찾기
        kw_season_id_col = self._find_column(df_keywords, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        keyword_col = self._find_column(df_keywords, ["키워드", "keyword", "단어", "키워드(keyword)"])
        polarity_col = self._find_column(df_keywords, ["polarity", "타입", "종류", "포함여부", "극성", "포함/제외(polarity: include/exclude)"])
        weight_col = self._find_column(df_keywords, ["가중치", "weight", "점수"])
        
        if not kw_season_id_col or not keyword_col:
            ttk.Label(self.keyword_info_frame, text="키워드 컬럼을 찾을 수 없습니다.").pack()
            return
        
        # 해당 시즌의 키워드 필터링
        season_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() == season_id]
        
        if len(season_keywords) == 0:
            ttk.Label(self.keyword_info_frame, text="키워드가 없습니다. 키워드를 추가하세요.").pack()
            return
        
        # 타입별로 그룹화
        include_keywords = []
        exclude_keywords = []
        # allowed_keywords는 더 이상 사용하지 않음 (deprecated)
        
        for idx, row in season_keywords.iterrows():
            keyword = str(row.get(keyword_col, "")).strip()
            if not keyword or keyword in ['nan', 'None', '']:
                continue
            
            polarity_str = str(row.get(polarity_col, "include")).strip().lower() if polarity_col else "include"
            weight = float(row.get(weight_col, 1.0)) if weight_col and pd.notna(row.get(weight_col)) else 1.0
            
            # polarity 변환 (allowed는 더 이상 사용하지 않음)
            if polarity_str in ["include", "포함", "1", "true", "yes"]:
                include_keywords.append((keyword, weight))
            elif polarity_str in ["exclude", "제외", "0", "false", "no"]:
                exclude_keywords.append((keyword, weight))
            # allowed는 deprecated - 하위 호환성을 위해 읽기만 하고 무시함
            elif polarity_str in ["allowed", "예외허용", "allow", "예외"]:
                # allowed 키워드는 더 이상 사용하지 않음 (무시)
                pass
            else:
                include_keywords.append((keyword, weight))  # 기본값
        
        # include 키워드 표시
        if include_keywords:
            ttk.Label(self.keyword_info_frame, text="✅ 포함 키워드 (include):", font=("맑은 고딕", 9, "bold")).pack(anchor='w', pady=(0, 5))
            for keyword, weight in include_keywords:
                frame = ttk.Frame(self.keyword_info_frame)
                frame.pack(fill='x', padx=10, pady=2)
                ttk.Label(frame, text=f"  • {keyword}", foreground="#2c3e50").pack(side='left')
                ttk.Label(frame, text=f"(가중치: {weight})", font=("맑은 고딕", 8), foreground="#666").pack(side='left', padx=(5, 0))
                # 삭제 버튼
                def delete_include_keyword(k=keyword):
                    self._delete_keyword(season_id, k, "include")
                ttk.Button(frame, text="🗑️", command=delete_include_keyword, width=3).pack(side='right')
        
        # exclude 키워드는 공통 섹션에서만 관리하므로 시즌별로는 표시하지 않음
        # (하위 호환성을 위해 읽기는 하지만 표시하지 않음)
        if exclude_keywords:
            ttk.Label(self.keyword_info_frame, 
                     text="※ 제외 키워드는 환경설정의 공통 제외 키워드에서 관리합니다.", 
                     font=("맑은 고딕", 7), foreground="#999", 
                     style="TLabel").pack(anchor='w', padx=10, pady=(10, 5))
        
        # allowed 키워드는 더 이상 사용하지 않음 (하위 호환성을 위해 표시는 하지만 deprecated 표시)
        # if allowed_keywords:
        #     ttk.Label(self.keyword_info_frame, text="⚠️ 예외 허용 키워드 (allowed) - 더 이상 사용하지 않음:", 
        #              font=("맑은 고딕", 9, "bold"), foreground="#999").pack(anchor='w', pady=(10, 5))
        #     for keyword, weight in allowed_keywords:
        #         frame = ttk.Frame(self.keyword_info_frame)
        #         frame.pack(fill='x', padx=10, pady=2)
        #         ttk.Label(frame, text=f"  • {keyword} (deprecated)", foreground="#999").pack(side='left')
    
    def _add_season_dialog(self):
        """시즌 추가 다이얼로그"""
        dialog = tk.Toplevel(self)
        dialog.title("시즌 추가")
        dialog.geometry("600x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # 입력 필드 생성
        self._create_season_input_dialog(dialog, season_id=None, is_new=True)
    
    def _edit_season_dialog(self):
        """시즌 수정 다이얼로그"""
        if not self.selected_season_id:
            messagebox.showwarning("알림", "수정할 시즌을 선택하세요.")
            return
        
        dialog = tk.Toplevel(self)
        dialog.title(f"시즌 수정: {self.selected_season_id}")
        dialog.geometry("600x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # 입력 필드 생성 (기존 데이터 로드)
        self._create_season_input_dialog(dialog, season_id=self.selected_season_id, is_new=False)
    
    def _create_season_input_dialog(self, dialog: tk.Toplevel, season_id: Optional[str] = None, is_new: bool = True):
        """시즌 입력 다이얼로그 생성 (추가/수정 공통)"""
        main_frame = ttk.Frame(dialog, padding=15)
        main_frame.pack(fill='both', expand=True)
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        entry_vars = {}
        
        # 시즌ID
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="시즌ID:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_id = tk.StringVar(value=season_id if season_id else "")
        entry_vars['시즌ID'] = var_id
        if is_new:
            entry = ttk.Entry(row_frame, textvariable=var_id, width=40)
            entry.pack(side='left', fill='x', expand=True)
            entry.bind('<KeyRelease>', lambda e, v=var_id: self._replace_spaces_with_underscore(v))
        else:
            ttk.Label(row_frame, text=season_id, font=("맑은 고딕", 9, "bold")).pack(side='left')
        
        # 시즌명
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="시즌명:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_name = tk.StringVar()
        entry_vars['시즌명'] = var_name
        ttk.Entry(row_frame, textvariable=var_name, width=40).pack(side='left', fill='x', expand=True)
        
        # 시작일
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="시작일 (MM-DD):", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_start = tk.StringVar()
        entry_vars['시작일'] = var_start
        entry_start = ttk.Entry(row_frame, textvariable=var_start, width=15)
        entry_start.pack(side='left', fill='x', expand=True)
        
        # 달력 버튼
        def open_calendar_start():
            if HAS_TKCALENDAR:
                cal_window = tk.Toplevel(dialog)
                cal_window.title("시작일 선택")
                cal_window.transient(dialog)
                cal_window.grab_set()
                
                cal_frame = tk.Frame(cal_window, padx=10, pady=10)
                cal_frame.pack()
                
                cal = DateEntry(cal_frame, width=12, background='darkblue',
                               foreground='white', borderwidth=2,
                               date_pattern='mm/dd/yyyy', year=datetime.now().year)
                cal.pack(pady=10)
                
                def set_date():
                    selected = cal.get_date()
                    var_start.set(selected.strftime("%m-%d"))
                    cal_window.destroy()
                
                ttk.Button(cal_frame, text="선택", command=set_date).pack(pady=5)
        
        ttk.Button(row_frame, text="📅", command=open_calendar_start, width=3).pack(side='left', padx=(5, 0))
        
        # 종료일
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="종료일 (MM-DD):", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_end = tk.StringVar()
        entry_vars['종료일'] = var_end
        entry_end = ttk.Entry(row_frame, textvariable=var_end, width=15)
        entry_end.pack(side='left', fill='x', expand=True)
        
        # 달력 버튼
        def open_calendar_end():
            if HAS_TKCALENDAR:
                cal_window = tk.Toplevel(dialog)
                cal_window.title("종료일 선택")
                cal_window.transient(dialog)
                cal_window.grab_set()
                
                cal_frame = tk.Frame(cal_window, padx=10, pady=10)
                cal_frame.pack()
                
                cal = DateEntry(cal_frame, width=12, background='darkblue',
                               foreground='white', borderwidth=2,
                               date_pattern='mm/dd/yyyy', year=datetime.now().year)
                cal.pack(pady=10)
                
                def set_date():
                    selected = cal.get_date()
                    var_end.set(selected.strftime("%m-%d"))
                    cal_window.destroy()
                
                ttk.Button(cal_frame, text="선택", command=set_date).pack(pady=5)
        
        ttk.Button(row_frame, text="📅", command=open_calendar_end, width=3).pack(side='left', padx=(5, 0))
        
        # 연도넘김
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="연도넘김:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_cross = tk.BooleanVar()
        entry_vars['연도넘김'] = var_cross
        ttk.Checkbutton(row_frame, text="연도 넘김 (예: 12-01 ~ 02-28)", variable=var_cross).pack(side='left')
        
        # 소싱시작일수
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="소싱시작일수:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_sourcing = tk.StringVar(value="30")
        entry_vars['소싱시작일수'] = var_sourcing
        ttk.Entry(row_frame, textvariable=var_sourcing, width=10).pack(side='left')
        ttk.Label(row_frame, text="일 (시즌 시작일 기준)", font=("맑은 고딕", 8), foreground="#666").pack(side='left', padx=(5, 0))
        
        # 가공완료마감일수
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="가공완료마감일수:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_processing = tk.StringVar(value="21")
        entry_vars['가공완료마감일수'] = var_processing
        ttk.Entry(row_frame, textvariable=var_processing, width=10).pack(side='left')
        ttk.Label(row_frame, text="일 (시즌 종료일 기준)", font=("맑은 고딕", 8), foreground="#666").pack(side='left', padx=(5, 0))
        
        # 사용여부
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="사용여부:", width=20, anchor='e').pack(side='left', padx=(0, 10))
        var_use = tk.BooleanVar(value=True)
        entry_vars['사용여부'] = var_use
        ttk.Checkbutton(row_frame, text="사용", variable=var_use).pack(side='left')
        
        # 설명
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="설명:", width=20, anchor='e').pack(side='left', padx=(0, 10), anchor='n')
        var_desc = tk.StringVar()
        entry_vars['설명'] = var_desc
        text_desc = tk.Text(row_frame, width=40, height=3)
        text_desc.pack(side='left', fill='x', expand=True)
        entry_vars['설명_텍스트'] = text_desc
        
        # 기존 데이터 로드 (수정 모드)
        if not is_new and season_id:
            self._load_season_data_to_dialog(season_id, entry_vars)
        
        # 계산된 기간 표시
        calc_frame = ttk.LabelFrame(scrollable_frame, text="계산된 기간", padding=10)
        calc_frame.pack(fill='x', pady=10)
        
        calc_label = ttk.Label(calc_frame, text="시작일과 종료일을 입력하면 자동으로 계산됩니다.", 
                              font=("맑은 고딕", 8), foreground="#666")
        calc_label.pack()
        
        def update_calc():
            try:
                start_str = var_start.get().strip()
                end_str = var_end.get().strip()
                sourcing_days = int(var_sourcing.get() or "30")
                processing_days = int(var_processing.get() or "21")
                
                if start_str and end_str and '-' in start_str and '-' in end_str:
                    from datetime import datetime, timedelta
                    start_month, start_day = map(int, start_str.split('-'))
                    end_month, end_day = map(int, end_str.split('-'))
                    current_year = datetime.now().year
                    
                    start_dt = datetime(current_year, start_month, start_day)
                    end_dt = datetime(current_year, end_month, end_day)
                    
                    sourcing_start = start_dt - timedelta(days=sourcing_days)
                    processing_end = end_dt - timedelta(days=processing_days)
                    
                    calc_text = f"소싱 시작일: {sourcing_start.strftime('%m-%d')} | 가공 완료 마감: {processing_end.strftime('%m-%d')}"
                    calc_label.config(text=calc_text, foreground="#2c3e50")
                else:
                    calc_label.config(text="시작일과 종료일을 입력하세요.", foreground="#666")
            except (ValueError, AttributeError, IndexError) as e:
                calc_label.config(text=f"날짜 형식 오류 (MM-DD 형식으로 입력): {str(e)}", foreground="red")
                self._log(f"날짜 계산 오류: {e}", "WARNING")
        
        var_start.trace('w', lambda *args: update_calc())
        var_end.trace('w', lambda *args: update_calc())
        var_sourcing.trace('w', lambda *args: update_calc())
        var_processing.trace('w', lambda *args: update_calc())
        update_calc()
        
        # 스크롤바 배치
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 하단 버튼
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', padx=15, pady=10)
        
        def save_season():
            try:
                # 데이터 검증
                if is_new:
                    new_season_id = var_id.get().strip()
                    if not new_season_id:
                        messagebox.showerror("오류", "시즌ID를 입력하세요.")
                        return
                    if new_season_id in ['nan', 'None', '']:
                        messagebox.showerror("오류", "유효한 시즌ID를 입력하세요.")
                        return
                else:
                    new_season_id = season_id
                
                season_name = var_name.get().strip()
                start_date = var_start.get().strip()
                end_date = var_end.get().strip()
                cross_year = var_cross.get()
                sourcing_days = var_sourcing.get().strip() or "30"
                processing_days = var_processing.get().strip() or "21"
                use_val = "Y" if var_use.get() else "N"
                desc = entry_vars['설명_텍스트'].get("1.0", "end-1c").strip()
                
                # 날짜 형식 검증 및 변환
                if start_date:
                    # MM-DD 형식으로 변환
                    start_date = self._convert_date_to_mmdd(start_date)
                    if not self._validate_date_format(start_date):
                        messagebox.showerror("오류", f"시작일 형식이 올바르지 않습니다.\n\n입력된 값: {var_start.get()}\n올바른 형식: MM-DD (예: 12-01)")
                        return
                
                if end_date:
                    # MM-DD 형식으로 변환
                    end_date = self._convert_date_to_mmdd(end_date)
                    if not self._validate_date_format(end_date):
                        messagebox.showerror("오류", f"종료일 형식이 올바르지 않습니다.\n\n입력된 값: {var_end.get()}\n올바른 형식: MM-DD (예: 12-31)")
                        return
                
                # 데이터 검증
                season_data = {
                    'id': new_season_id,
                    'name': season_name,
                    'start_date': start_date,
                    'end_date': end_date,
                    'cross_year': cross_year,
                    'sourcing_start_days': sourcing_days,
                    'processing_end_days': processing_days,
                    'is_new': is_new
                }
                
                is_valid, errors, warnings = self._validate_season_data(season_data)
                
                if not is_valid:
                    error_msg = "다음 오류를 수정해주세요:\n\n" + "\n".join(f"• {e}" for e in errors)
                    messagebox.showerror("데이터 검증 실패", error_msg)
                    return
                
                # 경고가 있으면 확인
                if warnings:
                    warning_msg = "다음 경고가 있습니다:\n\n" + "\n".join(f"⚠ {w}" for w in warnings)
                    warning_msg += "\n\n계속하시겠습니까?"
                    if not messagebox.askyesno("데이터 검증 경고", warning_msg):
                        return
                
                # Excel에 저장 (메모리)
                self._save_season_to_excel(
                    season_id=new_season_id,
                    season_name=season_name,
                    start_date=start_date,
                    end_date=end_date,
                    cross_year=cross_year,
                    sourcing_start_days=sourcing_days,
                    processing_end_days=processing_days,
                    use_val=use_val,
                    description=desc,
                    is_new=is_new
                )
                
                # Excel 파일에 실제 저장
                if self.excel_path and os.path.exists(self.excel_path):
                    try:
                        # 백업 생성
                        import shutil
                        backup_path = self.excel_path.replace('.xlsx', f'_백업_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                        shutil.copy2(self.excel_path, backup_path)
                        
                        # Excel 파일로 저장
                        import openpyxl
                        from openpyxl import load_workbook
                        
                        try:
                            wb_original = load_workbook(self.excel_path)
                            original_sheet_names = wb_original.sheetnames
                            original_active_sheet_title = wb_original.active.title if wb_original.active else None
                            original_sheet_states = {ws.title: ws.sheet_state for ws in wb_original.worksheets}
                            wb_original.close()
                        except (FileNotFoundError, PermissionError, Exception) as e:
                            self._log(f"기존 Excel 파일 구조 확인 실패: {e} - 새로 생성합니다.", "WARNING")
                            original_sheet_names = []
                            original_active_sheet_title = None
                            original_sheet_states = {}
                        
                        with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                            for sheet_name in original_sheet_names:
                                if sheet_name in self.excel_data:
                                    df_to_save = self.excel_data[sheet_name].copy()
                                    
                                    # SEASON_MASTER 시트의 날짜 컬럼을 문자열로 변환
                                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                                        start_col = self._find_column(df_to_save, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
                                        end_col = self._find_column(df_to_save, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
                                        
                                        # datetime 객체를 MM-DD 형식 문자열로 변환
                                        if start_col and start_col in df_to_save.columns:
                                            df_to_save[start_col] = df_to_save[start_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                                        if end_col and end_col in df_to_save.columns:
                                            df_to_save[end_col] = df_to_save[end_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                                    
                                    df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                                else:
                                    # 편집되지 않은 시트는 원본 그대로 유지
                                    pass
                            
                            # 새로 추가된 시트 처리
                            for sheet_name, df in self.excel_data.items():
                                if sheet_name not in original_sheet_names:
                                    df_to_save = df.copy()
                                    
                                    # SEASON_MASTER 시트의 날짜 컬럼을 문자열로 변환
                                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                                        start_col = self._find_column(df_to_save, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
                                        end_col = self._find_column(df_to_save, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
                                        
                                        if start_col and start_col in df_to_save.columns:
                                            df_to_save[start_col] = df_to_save[start_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                                        if end_col and end_col in df_to_save.columns:
                                            df_to_save[end_col] = df_to_save[end_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                                    
                                    df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            workbook = writer.book
                            if original_active_sheet_title and original_active_sheet_title in workbook.sheetnames:
                                workbook.active = workbook[original_active_sheet_title]
                        
                        self._log(f"Excel 파일 저장 완료: {os.path.basename(self.excel_path)}", "SUCCESS")
                        self._mark_saved()
                    except Exception as save_error:
                        self._log(f"Excel 파일 저장 실패: {save_error}", "ERROR")
                        messagebox.showwarning(
                            "경고", 
                            f"시즌은 수정되었지만 Excel 파일 저장에 실패했습니다.\n\n"
                            f"오류: {save_error}\n\n"
                            f"메모리에만 저장되었습니다. 'Excel 저장' 버튼을 눌러 수동으로 저장하세요."
                        )
                
                messagebox.showinfo("완료", f"시즌이 {'추가' if is_new else '수정'}되었습니다.\n\nExcel 파일에 저장되었습니다.")
                dialog.destroy()
                
                # 목록 새로고침 (Excel 파일을 다시 읽지 않고 메모리 데이터만 업데이트)
                # _load_excel()을 호출하면 Excel 파일을 다시 읽으면서 날짜 형식이 변환될 수 있음
                # 대신 Treeview만 업데이트
                self._update_season_list_tree()
                
                # 선택된 시즌 업데이트
                if not is_new:
                    self._update_season_detail(new_season_id)
                else:
                    # 새 시즌이면 선택
                    self.selected_season_id = new_season_id
                    self._update_season_detail(new_season_id)
                
            except Exception as e:
                messagebox.showerror("오류", f"시즌 저장 중 오류가 발생했습니다:\n{e}")
                import traceback
                self._log(traceback.format_exc())
        
        ttk.Button(btn_frame, text="💾 저장", command=save_season).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side='right', padx=5)
    
    def _load_season_data_to_dialog(self, season_id: str, entry_vars: Dict):
        """기존 시즌 데이터를 다이얼로그에 로드"""
        season_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                season_sheet = sheet_name
                break
        
        if not season_sheet:
            return
        
        df_seasons = self.excel_data[season_sheet]
        season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season"])
        
        for idx, row in df_seasons.iterrows():
            if str(row.get(season_id_col, "")).strip() == season_id:
                # 데이터 로드
                season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name"])
                # 시작일 컬럼 찾기: "소싱"이 포함된 컬럼은 제외 (소싱시작일수와 혼동 방지)
                start_col = None
                for col_name in ["시작일", "start_date", "시즌시작일", "start_mmdd", "시작(MM-DD)(start_mmdd)"]:
                    found_col = self._find_column(df_seasons, [col_name])
                    if found_col and "소싱" not in str(found_col) and "sourcing" not in str(found_col).lower():
                        start_col = found_col
                        break
                # 위에서 못 찾았으면 일반 검색 (하위 호환성)
                if not start_col:
                    start_col = self._find_column(df_seasons, ["시작일", "start_date", "start", "시즌시작일"])
                end_col = self._find_column(df_seasons, ["종료일", "end_date", "end", "시즌종료일"])
                cross_col = self._find_column(df_seasons, ["cross_year", "연도넘김", "연도초과", "시즌교차여부", "연도교차(Y/N)(cross_year)"])
                sourcing_start_col = self._find_column(df_seasons, ["소싱시작일수", "sourcing_start_days", "소싱시작"])
                processing_end_col = self._find_column(df_seasons, ["가공완료마감일수", "processing_end_days", "가공완료마감"])
                use_col = self._find_column(df_seasons, ["사용여부", "use", "enabled", "active"])
                desc_col = self._find_column(df_seasons, ["설명", "description", "desc"])
                
                if '시즌명' in entry_vars:
                    entry_vars['시즌명'].set(str(row.get(season_name_col, "")).strip() if season_name_col else "")
                if '시작일' in entry_vars:
                    if start_col:
                        start_val = row.get(start_col, "")
                        # datetime 객체인 경우 MM-DD 형식으로 변환
                        if pd.notna(start_val):
                            if isinstance(start_val, datetime):
                                entry_vars['시작일'].set(start_val.strftime("%m-%d"))
                            elif isinstance(start_val, pd.Timestamp):
                                entry_vars['시작일'].set(start_val.strftime("%m-%d"))
                            else:
                                start_str = str(start_val).strip()
                                # YYYY-MM-DD 형식인 경우 MM-DD로 변환
                                if len(start_str) >= 10 and '-' in start_str:
                                    try:
                                        date_obj = pd.to_datetime(start_str)
                                        entry_vars['시작일'].set(date_obj.strftime("%m-%d"))
                                    except:
                                        entry_vars['시작일'].set(start_str)
                                else:
                                    entry_vars['시작일'].set(start_str)
                        else:
                            entry_vars['시작일'].set("")
                    else:
                        entry_vars['시작일'].set("")
                if '종료일' in entry_vars:
                    if end_col:
                        end_val = row.get(end_col, "")
                        # datetime 객체인 경우 MM-DD 형식으로 변환
                        if pd.notna(end_val):
                            if isinstance(end_val, datetime):
                                entry_vars['종료일'].set(end_val.strftime("%m-%d"))
                            elif isinstance(end_val, pd.Timestamp):
                                entry_vars['종료일'].set(end_val.strftime("%m-%d"))
                            else:
                                end_str = str(end_val).strip()
                                # YYYY-MM-DD 형식인 경우 MM-DD로 변환
                                if len(end_str) >= 10 and '-' in end_str:
                                    try:
                                        date_obj = pd.to_datetime(end_str)
                                        entry_vars['종료일'].set(date_obj.strftime("%m-%d"))
                                    except:
                                        entry_vars['종료일'].set(end_str)
                                else:
                                    entry_vars['종료일'].set(end_str)
                        else:
                            entry_vars['종료일'].set("")
                    else:
                        entry_vars['종료일'].set("")
                if '연도넘김' in entry_vars:
                    if cross_col:
                        cross_val = str(row.get(cross_col, "")).strip().upper()
                        entry_vars['연도넘김'].set(cross_val in ['Y', 'YES', 'TRUE', '1'])
                if '소싱시작일수' in entry_vars:
                    val = row.get(sourcing_start_col, 30) if sourcing_start_col else 30
                    entry_vars['소싱시작일수'].set(str(int(val)) if pd.notna(val) else "30")
                if '가공완료마감일수' in entry_vars:
                    val = row.get(processing_end_col, 21) if processing_end_col else 21
                    entry_vars['가공완료마감일수'].set(str(int(val)) if pd.notna(val) else "21")
                if '사용여부' in entry_vars:
                    if use_col:
                        use_val = str(row.get(use_col, "Y")).strip().upper()
                        entry_vars['사용여부'].set(use_val in ['Y', 'YES', 'TRUE', '1'])
                    else:
                        entry_vars['사용여부'].set(True)
                if '설명_텍스트' in entry_vars:
                    desc = str(row.get(desc_col, "")).strip() if desc_col else ""
                    entry_vars['설명_텍스트'].delete("1.0", "end")
                    entry_vars['설명_텍스트'].insert("1.0", desc)
                break
    
    def _save_season_to_excel(self, season_id: str, season_name: str, start_date: str, 
                              end_date: str, cross_year: bool, sourcing_start_days: str,
                              processing_end_days: str, use_val: str, description: str, is_new: bool):
        """시즌 정보를 Excel에 저장"""
        # SEASON_MASTER 시트 찾기
        season_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                season_sheet = sheet_name
                break
        
        if not season_sheet:
            # 시트가 없으면 생성
            season_sheet = "SEASON_MASTER"
            df_new = pd.DataFrame(columns=[
                "시즌ID", "시즌명", "시작일", "종료일", "연도넘김", 
                "소싱시작일수", "가공완료마감일수", "사용여부", "설명"
            ])
            self.excel_data[season_sheet] = df_new
        
        df_seasons = self.excel_data[season_sheet]
        
        # 실제 컬럼명 찾기 (하드코딩 대신 동적으로 찾기)
        season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
        start_col = self._find_column(df_seasons, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
        end_col = self._find_column(df_seasons, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
        cross_col = self._find_column(df_seasons, ["연도넘김", "cross_year", "연도초과", "시즌교차여부", "연도교차(Y/N)(cross_year)"])
        sourcing_start_col = self._find_column(df_seasons, ["소싱시작일수", "sourcing_start_days", "소싱시작", "prep_days"])
        processing_end_col = self._find_column(df_seasons, ["가공완료마감일수", "processing_end_days", "가공완료마감", "grace_days"])
        use_col = self._find_column(df_seasons, ["사용여부", "use", "enabled", "active", "사용여부(Y/N)(enabled)"])
        desc_col = self._find_column(df_seasons, ["설명", "description", "desc"])
        
        # 컬럼이 없으면 생성
        if not season_id_col:
            season_id_col = "시즌ID"
            df_seasons[season_id_col] = ""
        if not season_name_col:
            season_name_col = "시즌명"
            df_seasons[season_name_col] = ""
        if not start_col:
            start_col = "시작일"
            df_seasons[start_col] = ""
        if not end_col:
            end_col = "종료일"
            df_seasons[end_col] = ""
        if not cross_col:
            cross_col = "연도넘김"
            df_seasons[cross_col] = ""
        if not sourcing_start_col:
            sourcing_start_col = "소싱시작일수"
            df_seasons[sourcing_start_col] = ""
        if not processing_end_col:
            processing_end_col = "가공완료마감일수"
            df_seasons[processing_end_col] = ""
        if not use_col:
            use_col = "사용여부"
            df_seasons[use_col] = "Y"
        if not desc_col:
            desc_col = "설명"
            df_seasons[desc_col] = ""
        
        if is_new:
            # 새 시즌 추가
            new_row = {}
            new_row[season_id_col] = str(season_id).strip()
            new_row[season_name_col] = str(season_name).strip()
            # 날짜는 문자열로 저장 (MM-DD 형식 유지, 빈 값 방지)
            new_row[start_col] = str(start_date).strip() if start_date else ""
            new_row[end_col] = str(end_date).strip() if end_date else ""
            new_row[cross_col] = "TRUE" if cross_year else "FALSE"
            new_row[sourcing_start_col] = str(sourcing_start_days).strip()
            new_row[processing_end_col] = str(processing_end_days).strip()
            new_row[use_col] = str(use_val).strip()
            new_row[desc_col] = str(description).strip()
            
            # 모든 컬럼에 대해 빈 값 설정
            for col in df_seasons.columns:
                if col not in new_row:
                    new_row[col] = ""
            
            df_seasons = pd.concat([df_seasons, pd.DataFrame([new_row])], ignore_index=True)
        else:
            # 기존 시즌 수정
            found = False
            for idx in df_seasons.index:
                if str(df_seasons.at[idx, season_id_col]).strip() == season_id:
                    df_seasons.at[idx, season_name_col] = str(season_name).strip()
                    # 날짜는 문자열로 저장 (MM-DD 형식 유지, 빈 값 방지)
                    df_seasons.at[idx, start_col] = str(start_date).strip() if start_date else ""
                    df_seasons.at[idx, end_col] = str(end_date).strip() if end_date else ""
                    df_seasons.at[idx, cross_col] = "TRUE" if cross_year else "FALSE"
                    df_seasons.at[idx, sourcing_start_col] = str(sourcing_start_days).strip()
                    df_seasons.at[idx, processing_end_col] = str(processing_end_days).strip()
                    df_seasons.at[idx, use_col] = str(use_val).strip()
                    df_seasons.at[idx, desc_col] = str(description).strip()
                    found = True
                    self._log(f"✅ 시즌 수정: {season_id} - 시작일={start_date}, 종료일={end_date} (컬럼: {start_col}, {end_col})", "SUCCESS")
                    self._mark_modified()  # 변경사항 표시
                    break
            
            if not found:
                self._log(f"⚠️ 시즌 '{season_id}'를 찾을 수 없습니다. 새로 추가합니다.")
                # 찾지 못했으면 새로 추가
                new_row = {}
                new_row[season_id_col] = str(season_id).strip()
                new_row[season_name_col] = str(season_name).strip()
                new_row[start_col] = str(start_date).strip() if start_date else ""
                new_row[end_col] = str(end_date).strip() if end_date else ""
                new_row[cross_col] = "TRUE" if cross_year else "FALSE"
                new_row[sourcing_start_col] = str(sourcing_start_days).strip()
                new_row[processing_end_col] = str(processing_end_days).strip()
                new_row[use_col] = str(use_val).strip()
                new_row[desc_col] = str(description).strip()
                
                for col in df_seasons.columns:
                    if col not in new_row:
                        new_row[col] = ""
                
                df_seasons = pd.concat([df_seasons, pd.DataFrame([new_row])], ignore_index=True)
        
        self.excel_data[season_sheet] = df_seasons
        self._log(f"✅ 시즌 저장 완료: {season_id} (컬럼: {start_col}={start_date}, {end_col}={end_date})")
    
    def _delete_season_dialog(self):
        """시즌 삭제 확인 다이얼로그"""
        if not self.selected_season_id:
            messagebox.showwarning("알림", "삭제할 시즌을 선택하세요.")
            return
        
        result = messagebox.askyesno(
            "시즌 삭제 확인",
            f"시즌 '{self.selected_season_id}'를 삭제하시겠습니까?\n\n"
            "연결된 키워드도 함께 삭제됩니다."
        )
        
        if result:
            self._delete_season(self.selected_season_id)
    
    def _delete_season(self, season_id: str):
        """시즌 삭제"""
        try:
            # SEASON_MASTER 시트에서 삭제
            season_sheet = None
            for sheet_name in self.excel_data.keys():
                if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                    season_sheet = sheet_name
                    break
            
            if season_sheet:
                df_seasons = self.excel_data[season_sheet]
                season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season"])
                if season_id_col:
                    df_seasons = df_seasons[df_seasons[season_id_col].astype(str).str.strip() != season_id]
                    self.excel_data[season_sheet] = df_seasons
            
            # KEYWORDS 시트에서도 삭제
            keyword_sheet = None
            for sheet_name in self.excel_data.keys():
                if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                    keyword_sheet = sheet_name
                    break
            
            if keyword_sheet:
                df_keywords = self.excel_data[keyword_sheet]
                kw_season_id_col = self._find_column(df_keywords, ["시즌ID", "시즌", "season_id", "season"])
                if kw_season_id_col:
                    df_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() != season_id]
                    self.excel_data[keyword_sheet] = df_keywords
            
            self._log(f"시즌 삭제 완료: {season_id}")
            messagebox.showinfo("완료", f"시즌 '{season_id}'가 삭제되었습니다.")
            
            # 목록 새로고침
            self._update_season_list_tree()
            self._show_no_selection_message()
            self.selected_season_id = None
            
        except Exception as e:
            messagebox.showerror("오류", f"시즌 삭제 중 오류가 발생했습니다:\n{e}")
            import traceback
            self._log(traceback.format_exc())
    
    def _add_keyword_dialog(self):
        """키워드 추가 다이얼로그"""
        if not self.selected_season_id:
            messagebox.showwarning("알림", "키워드를 추가할 시즌을 선택하세요.")
            return
        
        dialog = tk.Toplevel(self)
        dialog.title(f"키워드 추가: {self.selected_season_id}")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=15)
        main_frame.pack(fill='both', expand=True)
        
        # 시즌ID (자동 선택)
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="시즌ID:", width=15, anchor='e').pack(side='left', padx=(0, 10))
        ttk.Label(row_frame, text=self.selected_season_id, font=("맑은 고딕", 9, "bold")).pack(side='left')
        
        # 키워드
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="키워드:", width=15, anchor='e').pack(side='left', padx=(0, 10))
        var_keyword = tk.StringVar()
        ttk.Entry(row_frame, textvariable=var_keyword, width=30).pack(side='left', fill='x', expand=True)
        
        # 타입 (드롭다운)
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="타입:", width=15, anchor='e').pack(side='left', padx=(0, 10))
        var_type = tk.StringVar(value="include")
        # exclude는 공통 섹션에서만 관리하므로 include만 선택 가능
        combo_type = ttk.Combobox(row_frame, textvariable=var_type, 
                                 values=["include"], 
                                 state='readonly', width=27)
        combo_type.pack(side='left', fill='x', expand=True)
        
        # 타입 설명 라벨
        type_desc_frame = ttk.Frame(main_frame)
        type_desc_frame.pack(fill='x', pady=(2, 5))
        ttk.Label(type_desc_frame, text="  • include: 시즌 판단 키워드 (가중치로 점수 추가)", 
                 font=("맑은 고딕", 8), foreground="#2c3e50").pack(anchor='w', padx=(100, 0))
        ttk.Label(type_desc_frame, text="  • exclude: 환경설정의 공통 제외 키워드에서 관리", 
                 font=("맑은 고딕", 8), foreground="#d32f2f").pack(anchor='w', padx=(100, 0))
        
        # 가중치
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="가중치:", width=15, anchor='e').pack(side='left', padx=(0, 10))
        var_weight = tk.StringVar(value="1.0")
        ttk.Entry(row_frame, textvariable=var_weight, width=10).pack(side='left')
        ttk.Label(row_frame, text="(include 키워드만 사용)", font=("맑은 고딕", 8), foreground="#666").pack(side='left', padx=(5, 0))
        
        # 설명
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill='x', pady=5)
        ttk.Label(row_frame, text="설명:", width=15, anchor='e').pack(side='left', padx=(0, 10), anchor='n')
        var_desc = tk.StringVar()
        text_desc = tk.Text(row_frame, width=30, height=3)
        text_desc.pack(side='left', fill='x', expand=True)
        
        # 하단 버튼
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', padx=15, pady=10)
        
        def save_keyword():
            try:
                keyword = var_keyword.get().strip()
                if not keyword:
                    messagebox.showerror("오류", "키워드를 입력하세요.")
                    return
                
                kw_type = var_type.get()
                
                # exclude 타입은 공통 섹션에서만 관리
                if kw_type == "exclude":
                    messagebox.showwarning("알림", 
                        "제외 키워드(exclude)는 환경설정의 공통 제외 키워드에서 관리합니다.\n\n"
                        "환경설정 버튼(⚙️)을 클릭하여 공통 제외 키워드를 추가하세요.")
                    return
                
                weight = float(var_weight.get() or "1.0")
                desc = text_desc.get("1.0", "end-1c").strip()
                
                # Excel에 저장
                self._save_keyword_to_excel(
                    season_id=self.selected_season_id,
                    keyword=keyword,
                    kw_type=kw_type,
                    weight=weight,
                    description=desc
                )
                
                messagebox.showinfo("완료", "키워드가 추가되었습니다.")
                dialog.destroy()
                
                # 상세 정보 업데이트
                self._update_season_detail(self.selected_season_id)
                
            except Exception as e:
                messagebox.showerror("오류", f"키워드 저장 중 오류가 발생했습니다:\n{e}")
                import traceback
                self._log(traceback.format_exc())
        
        ttk.Button(btn_frame, text="💾 저장", command=save_keyword).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side='right', padx=5)
    
    def _save_keyword_to_excel(self, season_id: str, keyword: str, kw_type: str, weight: float, description: str):
        """키워드를 Excel에 저장"""
        # KEYWORDS 시트 찾기 또는 생성
        keyword_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                keyword_sheet = sheet_name
                break
        
        if not keyword_sheet:
            keyword_sheet = "KEYWORDS"
            df_new = pd.DataFrame(columns=["시즌ID", "키워드", "타입", "가중치", "설명"])
            self.excel_data[keyword_sheet] = df_new
        
        df_keywords = self.excel_data[keyword_sheet]
        
        # 컬럼 확인 및 추가
        required_columns = ["시즌ID", "키워드", "타입", "가중치", "설명"]
        for col in required_columns:
            if col not in df_keywords.columns:
                df_keywords[col] = ""
        
        # 새 키워드 추가
        new_row = {
            "시즌ID": season_id,
            "키워드": keyword,
            "타입": kw_type,
            "가중치": weight,
            "설명": description
        }
        df_keywords = pd.concat([df_keywords, pd.DataFrame([new_row])], ignore_index=True)
        self.excel_data[keyword_sheet] = df_keywords
        
        self._log(f"키워드 저장 완료: {season_id} - {keyword}")
    
    def _delete_keyword(self, season_id: str, keyword: str, kw_type: str):
        """키워드 삭제"""
        if not messagebox.askyesno("확인", f"키워드를 삭제하시겠습니까?\n\n시즌ID: {season_id}\n키워드: {keyword}\n타입: {kw_type}"):
            return
        
        # KEYWORDS 시트 찾기
        keyword_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                keyword_sheet = sheet_name
                break
        
        if not keyword_sheet or keyword_sheet not in self.excel_data:
            messagebox.showerror("오류", "KEYWORDS 시트를 찾을 수 없습니다.")
            return
        
        df_keywords = self.excel_data[keyword_sheet].copy()
        
        # 컬럼 찾기
        season_id_col = self._find_column(df_keywords, ["시즌ID", "season_id", "seasonid"])
        keyword_col = self._find_column(df_keywords, ["키워드", "keyword", "kw"])
        polarity_col = self._find_column(df_keywords, ["포함/제외", "polarity", "type", "타입"])
        
        if not season_id_col or not keyword_col:
            messagebox.showerror("오류", "필수 컬럼을 찾을 수 없습니다.")
            return
        
        # 삭제할 행 찾기
        mask = (df_keywords[season_id_col].astype(str).str.strip() == season_id) & \
               (df_keywords[keyword_col].astype(str).str.strip() == keyword)
        
        if polarity_col:
            # 타입도 확인
            polarity_lower = kw_type.lower()
            mask = mask & (df_keywords[polarity_col].astype(str).str.strip().str.lower() == polarity_lower)
        
        rows_to_delete = df_keywords[mask]
        
        if len(rows_to_delete) == 0:
            messagebox.showwarning("알림", "삭제할 키워드를 찾을 수 없습니다.")
            return
        
        # 행 삭제
        df_keywords = df_keywords[~mask]
        self.excel_data[keyword_sheet] = df_keywords
        
        # Excel 파일 저장
        self._save_excel()
        self._log(f"키워드 삭제 완료: {season_id} - {keyword} ({kw_type})")
        
        # 상세 정보 업데이트
        if hasattr(self, 'selected_season_id') and self.selected_season_id == season_id:
            self._update_season_detail(season_id)
        
        messagebox.showinfo("완료", "키워드가 삭제되었습니다.")
    
    def _create_sheet_frame(self, sheet_name: str, df: pd.DataFrame) -> ttk.Frame:
        """시트별 프레임 생성 (편집 가능)"""
        frame = ttk.Frame(self.notebook)
        
        # 상단: 컬럼 정보 및 버튼
        header_frame = ttk.Frame(frame)
        header_frame.pack(fill='x', pady=(0, 10))
        
        info_frame = ttk.Frame(header_frame)
        info_frame.pack(side='left', fill='x', expand=True)
        
        self.row_count_vars = {}  # 시트별 행 수 변수 저장
        row_var = tk.StringVar(value=f"행 수: {len(df):,}개")
        self.row_count_vars[sheet_name] = row_var
        ttk.Label(info_frame, textvariable=row_var, font=("맑은 고딕", 9)).pack(side='left')
        ttk.Label(info_frame, text=f"컬럼: {', '.join(df.columns)}", 
                 font=("맑은 고딕", 9), foreground="#666").pack(side='left', padx=(20, 0))
        
        # 편집 버튼 영역 (READ ME 같은 시트는 제외)
        btn_frame = ttk.Frame(header_frame)
        btn_frame.pack(side='right')
        
        # READ ME, 설명, 안내 등의 시트는 행 추가/삭제 비활성화
        read_only_sheets = ['read me', 'readme', 'read_me', '설명', '안내', 'guide', 'manual']
        is_read_only = any(keyword in sheet_name.lower() for keyword in read_only_sheets)
        
        if not is_read_only:
            btn_add = ttk.Button(btn_frame, text="➕ 행 추가", 
                                command=lambda: self._add_row(sheet_name))
            btn_add.pack(side='left', padx=2)
            
            btn_delete = ttk.Button(btn_frame, text="➖ 행 삭제", 
                                   command=lambda: self._delete_row(sheet_name))
            btn_delete.pack(side='left', padx=2)
        
        # 테이블 뷰 (Treeview 사용)
        table_frame = ttk.Frame(frame)
        table_frame.pack(fill='both', expand=True)
        
        # 스크롤바
        scrollbar_y = ttk.Scrollbar(table_frame, orient='vertical')
        scrollbar_x = ttk.Scrollbar(table_frame, orient='horizontal')
        
        # Treeview 생성
        columns = list(df.columns)
        tree = ttk.Treeview(table_frame, columns=columns, show='headings', 
                          yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set,
                          selectmode='browse')
        
        # 컬럼 설정
        for col in columns:
            tree.heading(col, text=col)
            # 컬럼 너비 자동 조정 (최소 100, 최대 300)
            tree.column(col, width=min(max(100, len(str(col)) * 10), 300), anchor='w')
        
        # 데이터 삽입
        for idx, row in df.iterrows():
            values = [str(val) if pd.notna(val) else "" for val in row]
            tree.insert('', 'end', iid=str(idx), values=values, tags=(str(idx),))
        
        # 스크롤바 연결
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)
        
        # 배치
        tree.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # 편집 기능: 더블클릭으로 셀 편집
        tree.bind('<Double-1>', lambda e, s=sheet_name, t=tree: self._edit_cell(s, t, e))
        
        # 트리뷰를 프레임에 저장 (나중에 접근하기 위해)
        if not hasattr(self, 'sheet_trees'):
            self.sheet_trees = {}
        self.sheet_trees[sheet_name] = tree
        
        return frame
    
    def _add_row(self, sheet_name: str):
        """행 추가 (입력 창 표시)"""
        if sheet_name not in self.sheet_trees:
            return
        
        tree = self.sheet_trees[sheet_name]
        df = self.excel_data[sheet_name]
        columns = list(df.columns)
        
        # 입력 창 열기
        self._open_add_row_dialog(sheet_name, columns)
    
    def _open_add_row_dialog(self, sheet_name: str, columns: List[str]):
        """행 추가 입력 창 (완전히 새로 작성)"""
        dialog = tk.Toplevel(self)
        dialog.title(f"새 행 추가: {sheet_name}")
        dialog.geometry("700x600")
        dialog.transient(self)
        dialog.grab_set()
        
        # 메인 프레임
        main_frame = ttk.Frame(dialog, padding=15)
        main_frame.pack(fill='both', expand=True)
        
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 입력 필드 생성
        entry_vars = {}
        # 시즌교차여부 자동 계산을 위한 참조 변수 (나중에 할당)
        date_vars = {'start': None, 'end': None, 'cross': None}
        
        for col in columns:
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', pady=5)
            
            ttk.Label(row_frame, text=f"{col}:", width=20, anchor='e').pack(side='left', padx=(0, 10))
            
            var = tk.StringVar()
            entry_vars[col] = var
            
            # 시즌ID 컬럼인지 확인 (대소문자 무시)
            if '시즌ID' in col or 'season_id' in col.lower() or 'seasonid' in col.lower():
                # SEASON_MASTER 시트인 경우 중복 체크
                if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                    entry = ttk.Entry(row_frame, textvariable=var, width=40)
                    entry.pack(side='left', fill='x', expand=True)
                    # 띄어쓰기 자동 변환
                    entry.bind('<KeyRelease>', lambda e, v=var: self._replace_spaces_with_underscore(v))
                    # 포커스 시 중복 체크
                    entry.bind('<FocusOut>', lambda e, c=col, v=var: self._check_season_id_duplicate(c, v, sheet_name, dialog))
                else:
                    # KEYWORDS 시트인 경우 시즌ID 드롭다운 + 시즌 정보 표시
                    season_list = self._get_available_seasons_with_full_info()
                    if season_list:
                        # 시즌ID 목록 생성 (ID와 이름 모두 표시)
                        season_options = [f"{s['id']} - {s['name']}" for s in season_list]
                        
                        # 드롭다운 생성 (var와 연결하지 않음 - 별도 관리)
                        combo_var = tk.StringVar()
                        combo = ttk.Combobox(row_frame, textvariable=combo_var, values=season_options, width=30, state='readonly')
                        combo.pack(side='left', fill='x', expand=True)
                        
                        # 시즌 정보 표시 레이블 (드롭다운 옆에 배치)
                        info_label = ttk.Label(row_frame, text="", font=("맑은 고딕", 8), foreground="#666")
                        info_label.pack(side='left', padx=(5, 0))
                        
                        # 시즌 선택 시 정보 업데이트
                        def on_season_select(event=None):
                            try:
                                selected = combo_var.get()
                                if not selected:
                                    info_label.config(text="")
                                    return
                                
                                # "ID - 이름" 형식에서 ID 추출
                                if ' - ' in selected:
                                    season_id = selected.split(' - ')[0].strip()
                                else:
                                    season_id = selected.strip()
                                
                                # 시즌 정보 찾기
                                found = False
                                for season in season_list:
                                    if season['id'] == season_id:
                                        start_date = season.get('start_date', '')
                                        end_date = season.get('end_date', '')
                                        
                                        # 타입 변환 및 정리
                                        if isinstance(start_date, str):
                                            start_date = start_date.strip()
                                        else:
                                            start_date = str(start_date).strip() if start_date else ''
                                        
                                        if isinstance(end_date, str):
                                            end_date = end_date.strip()
                                        else:
                                            end_date = str(end_date).strip() if end_date else ''
                                        
                                        # 빈 값 체크
                                        if not start_date or start_date in ['nan', 'None', 'NaT', '']:
                                            start_date = ''
                                        if not end_date or end_date in ['nan', 'None', 'NaT', '']:
                                            end_date = ''
                                        
                                        if start_date and end_date:
                                            info_text = f"시작: {start_date} | 종료: {end_date}"
                                        elif start_date:
                                            info_text = f"시작: {start_date} | 종료: 없음"
                                        elif end_date:
                                            info_text = f"시작: 없음 | 종료: {end_date}"
                                        else:
                                            info_text = "시작일/종료일 정보 없음"
                                        
                                        info_label.config(text=info_text)
                                        # var에 시즌ID만 저장 (combo는 그대로 유지)
                                        var.set(season_id)
                                        found = True
                                        break
                                
                                if not found:
                                    info_label.config(text="시즌 정보를 찾을 수 없습니다")
                                    var.set("")
                            except Exception as e:
                                import traceback
                                self._log(f"시즌 선택 오류: {e}")
                                self._log(traceback.format_exc())
                                info_label.config(text="오류 발생")
                                var.set("")
                        
                        combo.bind('<<ComboboxSelected>>', on_season_select)
                        
                        # 초기화: 새 행 추가이므로 빈 값으로 설정
                        var.set("")
                        combo_var.set("")
                        info_label.config(text="")
                    else:
                        entry = ttk.Entry(row_frame, textvariable=var, width=40, state='disabled')
                        entry.pack(side='left', fill='x', expand=True)
                        ttk.Label(row_frame, text="⚠️ SEASON_MASTER에 시즌을 먼저 추가하세요", 
                                foreground='red', font=("맑은 고딕", 8)).pack(side='left', padx=(5, 0))
            # 타입ID 컬럼인 경우 (SEASON_MASTER에서만) - '타입'도 포함
            elif (('타입ID' in col or 'type_id' in col.lower() or ('타입' in col and '시즌' not in col)) and 
                  sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']):
                type_list = self._get_available_type_ids()
                if type_list:
                    combo = ttk.Combobox(row_frame, textvariable=var, values=type_list, width=37, state='readonly')
                    combo.pack(side='left', fill='x', expand=True)
                else:
                    # 기본 타입 목록 제공 (Excel에 실제로 사용되는 타입만)
                    default_types = ['Event', 'Climate', 'Activity', 'Lifecycle']
                    combo = ttk.Combobox(row_frame, textvariable=var, values=default_types, width=37, state='readonly')
                    combo.pack(side='left', fill='x', expand=True)
            # polarity 컬럼인 경우 드롭다운 (KEYWORDS에서만)
            elif ('polarity' in col.lower() or '극성' in col) and sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                combo = ttk.Combobox(row_frame, textvariable=var, 
                                    values=["include", "exclude", "포함", "제외"], 
                                    width=37, state='readonly')
                combo.pack(side='left', fill='x', expand=True)
                var.set("include")  # 기본값
            # 날짜 컬럼인 경우 (시작일, 종료일) - SEASON_MASTER에서만 달력 형식
            elif ('시작일' in col or 'start_date' in col.lower()) and sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                # 날짜 입력 프레임
                date_input_frame = ttk.Frame(row_frame)
                date_input_frame.pack(side='left', fill='x', expand=True)
                
                # 입력 필드
                entry = ttk.Entry(date_input_frame, textvariable=var, width=15)
                entry.pack(side='left', fill='x', expand=True)
                
                # 달력 버튼 (항상 표시)
                def open_calendar():
                    if HAS_TKCALENDAR:
                        cal_window = tk.Toplevel(dialog)
                        cal_window.title("날짜 선택")
                        cal_window.transient(dialog)
                        cal_window.grab_set()
                        
                        cal_frame = tk.Frame(cal_window, padx=10, pady=10)
                        cal_frame.pack()
                        
                        cal = DateEntry(cal_frame, width=12, background='darkblue',
                                       foreground='white', borderwidth=2,
                                       date_pattern='MM-dd', year=datetime.now().year)
                        cal.pack(pady=10)
                        
                        def set_date():
                            selected = cal.get_date()
                            var.set(selected.strftime("%m-%d"))
                            if date_vars['start'] and date_vars['end'] and date_vars['cross']:
                                self._calculate_cross_year(date_vars['start'], date_vars['end'], date_vars['cross'])
                            cal_window.destroy()
                        
                        ttk.Button(cal_frame, text="선택", command=set_date).pack(pady=5)
                        ttk.Button(cal_frame, text="취소", command=cal_window.destroy).pack()
                    else:
                        messagebox.showinfo("안내", "tkcalendar가 설치되지 않았습니다.\nMM-DD 형식으로 직접 입력해주세요.\n\n설치: pip install tkcalendar", parent=dialog)
                
                btn_cal = ttk.Button(date_input_frame, text="📅", width=3, command=open_calendar)
                btn_cal.pack(side='left', padx=(5, 0))
                
                hint_label = ttk.Label(date_input_frame, text="(MM-DD)", font=("맑은 고딕", 8), foreground="#999")
                hint_label.pack(side='left', padx=(5, 0))
                
                # 포커스 아웃 시 형식 검증
                def validate_start_date(e):
                    self._validate_and_fix_mmdd(var, dialog)
                    if date_vars['start'] and date_vars['end'] and date_vars['cross']:
                        self._calculate_cross_year(date_vars['start'], date_vars['end'], date_vars['cross'])
                entry.bind('<FocusOut>', validate_start_date)
                entry.bind('<Return>', validate_start_date)  # Enter 키로도 검증
                
                date_vars['start'] = var
                
            elif ('종료일' in col or 'end_date' in col.lower()) and sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                # 날짜 입력 프레임
                date_input_frame = ttk.Frame(row_frame)
                date_input_frame.pack(side='left', fill='x', expand=True)
                
                # 입력 필드
                entry = ttk.Entry(date_input_frame, textvariable=var, width=15)
                entry.pack(side='left', fill='x', expand=True)
                
                # 달력 버튼 (항상 표시)
                def open_calendar():
                    if HAS_TKCALENDAR:
                        cal_window = tk.Toplevel(dialog)
                        cal_window.title("날짜 선택")
                        cal_window.transient(dialog)
                        cal_window.grab_set()
                        
                        cal_frame = tk.Frame(cal_window, padx=10, pady=10)
                        cal_frame.pack()
                        
                        cal = DateEntry(cal_frame, width=12, background='darkblue',
                                       foreground='white', borderwidth=2,
                                       date_pattern='MM-dd', year=datetime.now().year)
                        cal.pack(pady=10)
                        
                        def set_date():
                            selected = cal.get_date()
                            var.set(selected.strftime("%m-%d"))
                            if date_vars['start'] and date_vars['end'] and date_vars['cross']:
                                self._calculate_cross_year(date_vars['start'], date_vars['end'], date_vars['cross'])
                            cal_window.destroy()
                        
                        ttk.Button(cal_frame, text="선택", command=set_date).pack(pady=5)
                        ttk.Button(cal_frame, text="취소", command=cal_window.destroy).pack()
                    else:
                        messagebox.showinfo("안내", "tkcalendar가 설치되지 않았습니다.\nMM-DD 형식으로 직접 입력해주세요.\n\n설치: pip install tkcalendar", parent=dialog)
                
                btn_cal = ttk.Button(date_input_frame, text="📅", width=3, command=open_calendar)
                btn_cal.pack(side='left', padx=(5, 0))
                
                hint_label = ttk.Label(date_input_frame, text="(MM-DD)", font=("맑은 고딕", 8), foreground="#999")
                hint_label.pack(side='left', padx=(5, 0))
                
                # 포커스 아웃 시 형식 검증
                def validate_end_date(e):
                    self._validate_and_fix_mmdd(var, dialog)
                    if date_vars['start'] and date_vars['end'] and date_vars['cross']:
                        self._calculate_cross_year(date_vars['start'], date_vars['end'], date_vars['cross'])
                entry.bind('<FocusOut>', validate_end_date)
                entry.bind('<Return>', validate_end_date)  # Enter 키로도 검증
                
                date_vars['end'] = var
            elif ('시작일' in col or '종료일' in col or 'start_date' in col.lower() or 'end_date' in col.lower()):
                entry = ttk.Entry(row_frame, textvariable=var, width=40)
                entry.pack(side='left', fill='x', expand=True)
                hint_label = ttk.Label(row_frame, text="(YYYY-MM-DD)", font=("맑은 고딕", 8), foreground="#999")
                hint_label.pack(side='left', padx=(5, 0))
            # 시즌교차여부 컬럼인 경우 (자동 계산)
            elif ('cross_year' in col.lower() or '연도넘김' in col or '연도초과' in col or '시즌교차' in col) and sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                entry = ttk.Entry(row_frame, textvariable=var, width=40, state='readonly')
                entry.pack(side='left', fill='x', expand=True)
                ttk.Label(row_frame, text="(자동 계산)", font=("맑은 고딕", 8), foreground="#999").pack(side='left', padx=(5, 0))
                date_vars['cross'] = var
            elif 'cross_year' in col.lower() or '연도넘김' in col or '연도초과' in col or '시즌교차' in col:
                entry = ttk.Entry(row_frame, textvariable=var, width=40)
                entry.pack(side='left', fill='x', expand=True)
            # 시즌명 컬럼인 경우 (SEASON_MASTER에서만 띄어쓰기 자동 변환)
            elif ('시즌명' in col or 'season_name' in col.lower()) and sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                entry = ttk.Entry(row_frame, textvariable=var, width=40)
                entry.pack(side='left', fill='x', expand=True)
                # 띄어쓰기 자동 변환
                entry.bind('<KeyRelease>', lambda e, v=var: self._replace_spaces_with_underscore(v))
            # 사용여부 컬럼인 경우
            elif '사용여부' in col or 'use' in col.lower() or 'enabled' in col.lower() or 'active' in col.lower():
                if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                    combo = ttk.Combobox(row_frame, textvariable=var, values=["Y", "N"], width=37, state='readonly')
                    combo.pack(side='left', fill='x', expand=True)
                    var.set("Y")  # 기본값
                else:
                    entry = ttk.Entry(row_frame, textvariable=var, width=40)
                    entry.pack(side='left', fill='x', expand=True)
            else:
                entry = ttk.Entry(row_frame, textvariable=var, width=40)
                entry.pack(side='left', fill='x', expand=True)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 버튼
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', padx=15, pady=10)
        
        def save_row():
            # 입력값 검증 및 변환
            new_values = []
            for i, col in enumerate(columns):
                value = entry_vars[col].get().strip()
                
                # KEYWORDS 시트에서 시즌ID 컬럼인 경우 이미 ID만 저장되어 있음
                # (드롭다운에서 선택 시 자동으로 ID만 저장되므로 추가 처리 불필요)
                
                # polarity 값 정규화 (포함/제외 → include/exclude)
                if 'polarity' in col.lower() or '극성' in col:
                    if value in ['포함', 'include']:
                        value = 'include'
                    elif value in ['제외', 'exclude']:
                        value = 'exclude'
                
                # SEASON_MASTER에서 날짜 형식 검증 (MM-DD)
                if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                    if ('시작일' in col or '종료일' in col or 'start_date' in col.lower() or 'end_date' in col.lower()) and value:
                        if not self._validate_mmdd_format(value):
                            messagebox.showwarning("날짜 형식 오류", 
                                f"'{col}'는 MM-DD 형식으로 입력해주세요.\n예: 12-01",
                                parent=dialog)
                            return
                    
                    # 시즌ID 형식 검증 (필요시)
                    if ('시즌ID' in col or 'season_id' in col.lower()) and value:
                        if not self._validate_season_id_format(value):
                            messagebox.showwarning("시즌ID 형식 오류",
                                f"시즌ID는 영문, 숫자, 언더스코어만 사용 가능합니다.",
                                parent=dialog)
                            return
                
                # 필수 필드 체크 (시즌ID)
                if ('시즌ID' in col or 'season_id' in col.lower()) and not value:
                    messagebox.showwarning("입력 오류", f"'{col}'는 필수 입력 항목입니다.", parent=dialog)
                    return
                
                new_values.append(value)
            
            # 중복 체크
            if not self._validate_new_row(sheet_name, columns, new_values, dialog):
                return
            
            # Treeview에 추가
            tree = self.sheet_trees[sheet_name]
            existing_ids = [int(iid) for iid in tree.get_children() if iid.isdigit()]
            new_id = str(max(existing_ids) + 1 if existing_ids else 0)
            
            tree.insert('', 'end', iid=new_id, values=new_values)
            
            # 행 수 업데이트
            if sheet_name in self.row_count_vars:
                count = len(tree.get_children())
                self.row_count_vars[sheet_name].set(f"행 수: {count:,}개")
            
            self._log(f"시트 '{sheet_name}': 새 행 추가됨")
            dialog.destroy()
        
        ttk.Button(btn_frame, text="저장", command=save_row).pack(side='right', padx=(5, 0))
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side='right')
    
    def _get_available_season_ids(self) -> List[str]:
        """SEASON_MASTER에서 사용 가능한 시즌ID 목록 가져오기"""
        season_ids = []
        
        # SEASON_MASTER 시트 찾기
        for sheet_name in ['SEASON_MASTER', 'SEASONS', 'Season_Master']:
            if sheet_name in self.excel_data:
                df = self.excel_data[sheet_name]
                # 시즌ID 컬럼 찾기
                season_id_col = None
                for col in df.columns:
                    if '시즌ID' in col or 'season_id' in col.lower() or 'seasonid' in col.lower():
                        season_id_col = col
                        break
                
                if season_id_col:
                    # Treeview에서 현재 값 가져오기
                    if sheet_name in self.sheet_trees:
                        tree = self.sheet_trees[sheet_name]
                        for item_id in tree.get_children():
                            values = tree.item(item_id, 'values')
                            col_index = list(df.columns).index(season_id_col)
                            if col_index < len(values) and values[col_index]:
                                season_id = values[col_index].strip()
                                if season_id and season_id not in season_ids:
                                    season_ids.append(season_id)
                    else:
                        # DataFrame에서 직접 가져오기
                        for val in df[season_id_col].dropna():
                            if val and str(val).strip() not in season_ids:
                                season_ids.append(str(val).strip())
                break
        
        return sorted(season_ids)
    
    def _get_available_seasons_with_names(self) -> List[Dict]:
        """SEASON_MASTER에서 시즌ID와 시즌명 매핑 가져오기"""
        seasons = []
        
        # SEASON_MASTER 시트 찾기
        for sheet_name in ['SEASON_MASTER', 'SEASONS', 'Season_Master']:
            if sheet_name in self.excel_data:
                df = self.excel_data[sheet_name]
                # 시즌ID, 시즌명 컬럼 찾기
                season_id_col = None
                season_name_col = None
                for col in df.columns:
                    if '시즌ID' in col or 'season_id' in col.lower() or 'seasonid' in col.lower():
                        season_id_col = col
                    if '시즌명' in col or 'season_name' in col.lower() or 'seasonname' in col.lower():
                        season_name_col = col
                
                if season_id_col:
                    # Treeview에서 현재 값 가져오기
                    if sheet_name in self.sheet_trees:
                        tree = self.sheet_trees[sheet_name]
                        for item_id in tree.get_children():
                            values = tree.item(item_id, 'values')
                            id_index = list(df.columns).index(season_id_col) if season_id_col else -1
                            name_index = list(df.columns).index(season_name_col) if season_name_col else -1
                            
                            if id_index >= 0 and id_index < len(values) and values[id_index]:
                                season_id = values[id_index].strip()
                                season_name = values[name_index].strip() if name_index >= 0 and name_index < len(values) and values[name_index] else season_id
                                
                                if season_id:
                                    seasons.append({
                                        'id': season_id,
                                        'name': season_name
                                    })
                    else:
                        # DataFrame에서 직접 가져오기
                        for idx, row in df.iterrows():
                            season_id = str(row.get(season_id_col, "")).strip() if season_id_col else ""
                            season_name = str(row.get(season_name_col, season_id)).strip() if season_name_col else season_id
                            
                            if season_id:
                                seasons.append({
                                    'id': season_id,
                                    'name': season_name
                                })
                break
        
        return seasons
    
    def _get_available_seasons_with_full_info(self) -> List[Dict]:
        """SEASON_MASTER에서 시즌 전체 정보 가져오기 (ID, 이름, 시작일, 종료일) - DataFrame 직접 읽기"""
        seasons = []
        
        # SEASON_MASTER 시트 찾기 (대소문자 무시)
        target_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
                target_sheet = sheet_name
                break
        
        if not target_sheet:
            self._log(f"[시즌 정보] SEASON_MASTER 시트를 찾을 수 없습니다. 사용 가능한 시트: {list(self.excel_data.keys())}")
            return seasons
        
        df = self.excel_data[target_sheet]
        
        # 컬럼 찾기
        season_id_col = self._find_column(df, ["시즌ID", "시즌", "season_id", "season"])
        season_name_col = self._find_column(df, ["시즌명", "시즌이름", "season_name", "name"])
        start_date_col = self._find_column(df, ["시작일", "start_date", "start", "시즌시작일"])
        end_date_col = self._find_column(df, ["종료일", "end_date", "end", "시즌종료일"])
        
        if not season_id_col:
            self._log(f"[시즌 정보] 시즌ID 컬럼을 찾을 수 없습니다. 컬럼: {list(df.columns)}")
            return seasons
        
        # DataFrame에서 직접 가져오기
        for idx, row in df.iterrows():
            # 시즌ID
            season_id_val = row.get(season_id_col)
            if pd.isna(season_id_val):
                continue
            season_id = str(season_id_val).strip()
            if not season_id or season_id in ['nan', 'None', 'NaT', '']:
                continue
            
            # 시즌명
            if season_name_col:
                season_name_val = row.get(season_name_col)
                if pd.notna(season_name_val):
                    season_name = str(season_name_val).strip()
                    if season_name in ['nan', 'None', 'NaT', '']:
                        season_name = season_id
                else:
                    season_name = season_id
            else:
                season_name = season_id
            
            # 시작일
            start_date = ""
            if start_date_col:
                start_date_val = row.get(start_date_col)
                if pd.notna(start_date_val):
                    start_date_str = str(start_date_val).strip()
                    # 날짜 형식 정리 (MM-DD 형식으로 변환 시도)
                    if start_date_str and start_date_str not in ['nan', 'None', 'NaT', '']:
                        # YYYY-MM-DD 형식이면 MM-DD로 변환
                        if len(start_date_str) >= 10 and '-' in start_date_str:
                            parts = start_date_str.split('-')
                            if len(parts) >= 2:
                                start_date = f"{parts[-2].zfill(2)}-{parts[-1].zfill(2)}"
                            else:
                                start_date = start_date_str
                        else:
                            start_date = start_date_str
            
            # 종료일
            end_date = ""
            if end_date_col:
                end_date_val = row.get(end_date_col)
                if pd.notna(end_date_val):
                    end_date_str = str(end_date_val).strip()
                    if end_date_str and end_date_str not in ['nan', 'None', 'NaT', '']:
                        # YYYY-MM-DD 형식이면 MM-DD로 변환
                        if len(end_date_str) >= 10 and '-' in end_date_str:
                            parts = end_date_str.split('-')
                            if len(parts) >= 2:
                                end_date = f"{parts[-2].zfill(2)}-{parts[-1].zfill(2)}"
                            else:
                                end_date = end_date_str
                        else:
                            end_date = end_date_str
            
            seasons.append({
                'id': season_id,
                'name': season_name,
                'start_date': start_date,
                'end_date': end_date
            })
        
        # 디버깅용 로그
        if seasons:
            self._log(f"[시즌 정보] {len(seasons)}개 시즌 로드됨")
            for s in seasons[:3]:  # 처음 3개만 로그
                self._log(f"  - {s['id']} ({s['name']}): 시작={s['start_date']}, 종료={s['end_date']}")
        else:
            self._log(f"[시즌 정보] 시즌을 찾을 수 없습니다. 컬럼: {list(df.columns)}")
        
        return seasons
    
    def _get_available_type_ids(self) -> List[str]:
        """TYPE_PRESETS에서 사용 가능한 타입ID 목록 가져오기"""
        type_ids = []
        
        # TYPE_PRESETS 시트 찾기
        for sheet_name in ['TYPE_PRESETS', 'TYPES', 'Type_Presets', 'TYPE']:
            if sheet_name in self.excel_data:
                df = self.excel_data[sheet_name]
                # 타입ID 컬럼 찾기 (더 유연하게)
                type_id_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if ('타입id' in col_lower or 'type_id' in col_lower or 'typeid' in col_lower or 
                        ('타입' in col and 'id' in col_lower) or col_lower == 'type'):
                        type_id_col = col
                        break
                
                if type_id_col:
                    # Treeview에서 현재 값 가져오기
                    if sheet_name in self.sheet_trees:
                        tree = self.sheet_trees[sheet_name]
                        for item_id in tree.get_children():
                            values = tree.item(item_id, 'values')
                            col_index = list(df.columns).index(type_id_col)
                            if col_index < len(values) and values[col_index]:
                                type_id = str(values[col_index]).strip()
                                if type_id and type_id not in type_ids:
                                    type_ids.append(type_id)
                    else:
                        # DataFrame에서 직접 가져오기
                        for val in df[type_id_col].dropna():
                            type_id = str(val).strip()
                            if type_id and type_id not in type_ids:
                                type_ids.append(type_id)
                break
        
        # 타입ID가 없으면 기본값 제공 (Excel에 실제로 사용되는 타입만)
        if not type_ids:
            type_ids = ['Event', 'Climate', 'Activity', 'Lifecycle']
        
        return sorted(type_ids)
    
    def _replace_spaces_with_underscore(self, var: tk.StringVar):
        """띄어쓰기를 언더스코어로 자동 변환"""
        current = var.get()
        if ' ' in current:
            new_value = current.replace(' ', '_')
            var.set(new_value)
    
    def _check_season_id_duplicate(self, col_name: str, var: tk.StringVar, sheet_name: str, dialog: tk.Toplevel):
        """시즌ID 중복 체크"""
        value = var.get().strip()
        if not value:
            return
        
        # 현재 시트의 다른 행에서 중복 확인
        if sheet_name in self.sheet_trees:
            tree = self.sheet_trees[sheet_name]
            df = self.excel_data[sheet_name]
            col_index = list(df.columns).index(col_name)
            
            for item_id in tree.get_children():
                values = tree.item(item_id, 'values')
                if col_index < len(values) and values[col_index].strip() == value:
                    # 현재 편집 중인 행이 아니면 중복
                    messagebox.showwarning("중복 오류", 
                        f"'{value}'는 이미 사용 중인 시즌ID입니다.\n다른 값을 입력해주세요.",
                        parent=dialog)
                    var.set("")
                    return
    
    def _validate_new_row(self, sheet_name: str, columns: List[str], new_values: List[str], dialog: tk.Toplevel) -> bool:
        """새 행 검증 (중복 체크 등)"""
        # SEASON_MASTER 시트인 경우 시즌ID, 시즌명 중복 체크
        if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS']:
            season_id_col_index = None
            season_name_col_index = None
            
            for i, col in enumerate(columns):
                if '시즌ID' in col or 'season_id' in col.lower():
                    season_id_col_index = i
                if '시즌명' in col or 'season_name' in col.lower():
                    season_name_col_index = i
            
            # 시즌ID 중복 체크
            if season_id_col_index is not None and season_id_col_index < len(new_values):
                season_id = new_values[season_id_col_index].strip()
                if season_id:
                    if sheet_name in self.sheet_trees:
                        tree = self.sheet_trees[sheet_name]
                        for item_id in tree.get_children():
                            values = tree.item(item_id, 'values')
                            if season_id_col_index < len(values) and values[season_id_col_index].strip() == season_id:
                                messagebox.showwarning("중복 오류", 
                                    f"시즌ID '{season_id}'는 이미 존재합니다.\n다른 값을 입력해주세요.",
                                    parent=dialog)
                                return False
            
            # 시즌명 중복 체크
            if season_name_col_index is not None and season_name_col_index < len(new_values):
                season_name = new_values[season_name_col_index].strip()
                if season_name:
                    if sheet_name in self.sheet_trees:
                        tree = self.sheet_trees[sheet_name]
                        for item_id in tree.get_children():
                            values = tree.item(item_id, 'values')
                            if season_name_col_index < len(values) and values[season_name_col_index].strip() == season_name:
                                messagebox.showwarning("중복 오류", 
                                    f"시즌명 '{season_name}'는 이미 존재합니다.\n다른 값을 입력해주세요.",
                                    parent=dialog)
                                return False
        
        # KEYWORDS 시트인 경우 SEASON_MASTER에 해당 시즌ID가 있는지 확인 및 키워드 중복 체크
        if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
            season_id_col_index = None
            keyword_col_index = None
            
            for i, col in enumerate(columns):
                if '시즌ID' in col or 'season_id' in col.lower():
                    season_id_col_index = i
                if '키워드' in col or 'keyword' in col.lower():
                    keyword_col_index = i
            
            if season_id_col_index is not None and season_id_col_index < len(new_values):
                season_id = new_values[season_id_col_index].strip()
                if season_id:
                    available_ids = self._get_available_season_ids()
                    if season_id not in available_ids:
                        messagebox.showerror("오류", 
                            f"시즌ID '{season_id}'는 SEASON_MASTER에 존재하지 않습니다.\n"
                            f"먼저 SEASON_MASTER에 시즌을 추가해주세요.",
                            parent=dialog)
                        return False
                
                # 키워드 중복 체크 (같은 시즌ID + 같은 키워드)
                if keyword_col_index is not None and keyword_col_index < len(new_values):
                    keyword = new_values[keyword_col_index].strip()
                    if keyword:
                        if sheet_name in self.sheet_trees:
                            tree = self.sheet_trees[sheet_name]
                            for item_id in tree.get_children():
                                values = tree.item(item_id, 'values')
                                if (season_id_col_index < len(values) and 
                                    keyword_col_index < len(values) and
                                    values[season_id_col_index].strip() == season_id and
                                    values[keyword_col_index].strip() == keyword):
                                    messagebox.showwarning("중복 오류",
                                        f"시즌 '{season_id}'에 키워드 '{keyword}'는 이미 존재합니다.\n다른 키워드를 입력해주세요.",
                                        parent=dialog)
                                    return False
        
        return True
    
    def _validate_date_format(self, date_str: str) -> bool:
        """날짜 형식 검증 (YYYY-MM-DD)"""
        import re
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        if not re.match(pattern, date_str):
            return False
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except:
            return False
    
    def _validate_mmdd_format(self, date_str: str) -> bool:
        """날짜 형식 검증 (MM-DD)"""
        import re
        pattern = r'^\d{2}-\d{2}$'
        if not re.match(pattern, date_str):
            return False
        try:
            month, day = date_str.split('-')
            month_int = int(month)
            day_int = int(day)
            if month_int < 1 or month_int > 12:
                return False
            if day_int < 1 or day_int > 31:
                return False
            # 간단한 날짜 유효성 체크 (2월 30일 같은 경우는 허용하되, 실제 사용 시 검증)
            return True
        except:
            return False
    
    def _validate_and_fix_mmdd(self, var: tk.StringVar, parent=None):
        """MM-DD 형식 검증 및 자동 수정"""
        value = var.get().strip()
        if not value:
            return
        
        # YYYY-MM-DD 형식인 경우 MM-DD로 변환
        import re
        if re.match(r'^\d{4}-\d{2}-\d{2}$', value):
            parts = value.split('-')
            var.set(f"{parts[1]}-{parts[2]}")
            return
        
        # MM/DD 형식을 MM-DD로 변환
        if re.match(r'^\d{2}/\d{2}$', value):
            parts = value.split('/')
            var.set(f"{parts[0]}-{parts[1]}")
            return
        
        # MM-DD 형식 검증
        if not self._validate_mmdd_format(value):
            if parent:
                messagebox.showwarning("날짜 형식 오류", 
                    f"날짜는 MM-DD 형식으로 입력해주세요.\n예: 12-01\n\n입력된 값: {value}",
                    parent=parent)
            else:
                messagebox.showwarning("날짜 형식 오류", 
                    f"날짜는 MM-DD 형식으로 입력해주세요.\n예: 12-01\n\n입력된 값: {value}")
            var.set("")
    
    def _calculate_cross_year(self, start_var: tk.StringVar, end_var: tk.StringVar, cross_var: tk.StringVar):
        """시즌교차여부 자동 계산"""
        if not start_var or not end_var or not cross_var:
            return
        
        start_str = start_var.get().strip()
        end_str = end_var.get().strip()
        
        if not start_str or not end_str:
            cross_var.set("")
            return
        
        try:
            # MM-DD 형식 파싱
            start_month, start_day = map(int, start_str.split('-'))
            end_month, end_day = map(int, end_str.split('-'))
            
            # 종료일의 월이 시작일의 월보다 낮으면 연도 넘김 (Y)
            if end_month < start_month:
                cross_var.set("Y")
            else:
                cross_var.set("N")
        except (ValueError, AttributeError) as e:
            self._log(f"연도넘김 계산 오류: {e} (시작일={start_str}, 종료일={end_str})", "WARNING")
            cross_var.set("")
    
    def _validate_season_id_format(self, season_id: str) -> bool:
        """시즌ID 형식 검증 (영문, 숫자, 언더스코어만)"""
        import re
        pattern = r'^[a-zA-Z0-9_]+$'
        return bool(re.match(pattern, season_id))
    
    def _validate_season_data(self, season_data: Dict) -> tuple[bool, List[str]]:
        """시즌 데이터 검증 후 오류 목록 반환"""
        errors = []
        warnings = []
        
        # 필수 필드 확인
        season_id = season_data.get('id', '').strip()
        if not season_id:
            errors.append("시즌ID가 필요합니다")
        elif not self._validate_season_id_format(season_id):
            errors.append(f"시즌ID 형식 오류: '{season_id}' (영문, 숫자, 언더스코어만 허용)")
        
        # 중복 시즌ID 확인
        if season_id and self._is_duplicate_season_id(season_id, season_data.get('is_new', False)):
            errors.append(f"중복된 시즌ID: '{season_id}'")
        
        # 날짜 형식 검증
        start_date = season_data.get('start_date', '').strip()
        end_date = season_data.get('end_date', '').strip()
        
        if start_date:
            if not self._validate_date_format(start_date):
                errors.append(f"시작일 형식 오류: '{start_date}' (예상 형식: MM-DD)")
            else:
                # 날짜 범위 검증
                start_valid, start_msg = self._validate_date_range(start_date)
                if not start_valid:
                    errors.append(f"시작일 범위 오류: {start_msg}")
        
        if end_date:
            if not self._validate_date_format(end_date):
                errors.append(f"종료일 형식 오류: '{end_date}' (예상 형식: MM-DD)")
            else:
                # 날짜 범위 검증
                end_valid, end_msg = self._validate_date_range(end_date)
                if not end_valid:
                    errors.append(f"종료일 범위 오류: {end_msg}")
        
        # 날짜 논리 검증
        if start_date and end_date and self._validate_date_format(start_date) and self._validate_date_format(end_date):
            logic_valid, logic_msg = self._validate_date_logic(start_date, end_date, season_data.get('cross_year', False))
            if not logic_valid:
                warnings.append(f"날짜 논리 경고: {logic_msg}")
        
        # 숫자 필드 검증
        sourcing_days = season_data.get('sourcing_start_days', '')
        if sourcing_days:
            try:
                days = int(sourcing_days)
                if days < 0 or days > 365:
                    warnings.append(f"소싱시작일수가 비정상적입니다: {days}일 (권장: 0-365일)")
            except (ValueError, TypeError):
                errors.append(f"소싱시작일수 형식 오류: '{sourcing_days}' (숫자여야 함)")
        
        processing_days = season_data.get('processing_end_days', '')
        if processing_days:
            try:
                days = int(processing_days)
                if days < 0 or days > 365:
                    warnings.append(f"가공완료마감일수가 비정상적입니다: {days}일 (권장: 0-365일)")
            except (ValueError, TypeError):
                errors.append(f"가공완료마감일수 형식 오류: '{processing_days}' (숫자여야 함)")
        
        return len(errors) == 0, errors, warnings
    
    def _validate_date_format(self, date_str: str) -> bool:
        """날짜 형식 검증 (MM-DD)"""
        import re
        if not date_str or not isinstance(date_str, str):
            return False
        pattern = r'^\d{2}-\d{2}$'
        if not re.match(pattern, date_str):
            return False
        
        try:
            month, day = map(int, date_str.split('-'))
            return 1 <= month <= 12 and 1 <= day <= 31
        except (ValueError, AttributeError):
            return False
    
    def _validate_date_range(self, date_str: str) -> tuple[bool, str]:
        """날짜 범위 검증"""
        try:
            month, day = map(int, date_str.split('-'))
            if not (1 <= month <= 12):
                return False, f"월이 1-12 범위를 벗어남: {month}"
            if not (1 <= day <= 31):
                return False, f"일이 1-31 범위를 벗어남: {day}"
            
            # 각 월의 최대 일수 확인
            days_in_month = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            if day > days_in_month[month - 1]:
                return False, f"{month}월은 최대 {days_in_month[month - 1]}일까지 가능: {day}일"
            
            return True, ""
        except (ValueError, IndexError) as e:
            return False, f"날짜 파싱 오류: {e}"
    
    def _validate_date_logic(self, start_date: str, end_date: str, cross_year: bool) -> tuple[bool, str]:
        """날짜 논리 검증 (시작일 <= 종료일)"""
        try:
            start_month, start_day = map(int, start_date.split('-'))
            end_month, end_day = map(int, end_date.split('-'))
            
            if cross_year:
                # 연도 넘김인 경우: 시작일이 종료일보다 늦어야 함 (예: 12-01 ~ 02-28)
                if start_month < end_month:
                    return False, f"연도넘김 시즌인데 시작월({start_month})이 종료월({end_month})보다 작습니다"
                elif start_month == end_month and start_day <= end_day:
                    return False, f"연도넘김 시즌인데 같은 월에서 시작일({start_day})이 종료일({end_day})보다 작거나 같습니다"
            else:
                # 일반 시즌: 시작일이 종료일보다 빨라야 함
                if start_month > end_month:
                    return False, f"시작월({start_month})이 종료월({end_month})보다 큽니다"
                elif start_month == end_month and start_day > end_day:
                    return False, f"같은 월에서 시작일({start_day})이 종료일({end_day})보다 큽니다"
            
            return True, ""
        except (ValueError, AttributeError) as e:
            return False, f"날짜 비교 오류: {e}"
    
    def _is_duplicate_season_id(self, season_id: str, is_new: bool = False) -> bool:
        """중복 시즌ID 확인"""
        if not season_id:
            return False
        
        # SEASON_MASTER 시트에서 확인
        season_sheet = None
        for sheet_name in self.excel_data.keys():
            if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                season_sheet = sheet_name
                break
        
        if not season_sheet:
            return False
        
        df_seasons = self.excel_data[season_sheet]
        season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        
        if not season_id_col:
            return False
        
        # 중복 확인
        for idx, row in df_seasons.iterrows():
            existing_id = str(row.get(season_id_col, "")).strip()
            if existing_id == season_id:
                # 새로 추가하는 경우에만 중복으로 간주
                if is_new:
                    return True
                # 수정하는 경우는 자기 자신이므로 중복 아님
                return False
        
        return False
    
    def _convert_date_to_mmdd(self, date_val) -> str:
        """날짜 값을 MM-DD 형식 문자열로 변환"""
        if pd.isna(date_val) or date_val == "" or date_val is None:
            return ""
        
        # 숫자(Excel serial number)인 경우 변환
        if isinstance(date_val, (int, float)):
            try:
                # Excel serial number (1900-01-01을 기준으로 한 일수)
                # 예: 1 = 1900-01-01, 44927 = 2023-01-01
                # 하지만 MM-DD 형식으로 저장된 텍스트가 숫자로 잘못 읽힌 경우도 있음
                # 작은 숫자(1-365)는 연도의 일수로 해석될 수 있음
                if 1 <= date_val <= 365:
                    # 1-365 범위면 Excel serial number가 아닐 수 있음
                    # 문자열로 변환 후 다시 시도
                    date_str = str(int(date_val))
                    # 50 같은 경우 특별 처리
                    if len(date_str) <= 3:
                        # 작은 숫자는 그대로 반환하지 않고, 다른 방식으로 처리
                        # 실제로는 날짜가 아니라 다른 값일 수 있음
                        return ""
                else:
                    # Excel serial number로 변환
                    from openpyxl.utils import datetime as openpyxl_datetime
                    if hasattr(openpyxl_datetime, 'from_excel'):
                        date_obj = openpyxl_datetime.from_excel(date_val)
                        return date_obj.strftime("%m-%d")
                    else:
                        # openpyxl 버전에 따라 다른 방식 사용
                        base_date = datetime(1899, 12, 30)  # Excel의 기준 날짜
                        date_obj = base_date + pd.Timedelta(days=int(date_val))
                        return date_obj.strftime("%m-%d")
            except (ValueError, OverflowError, AttributeError) as e:
                # 숫자 변환 실패 시 빈 문자열 반환
                self._log(f"Excel serial number 변환 실패: {e} (값={date_val})", "DEBUG")
                return ""
        
        # datetime 객체인 경우
        if isinstance(date_val, datetime):
            return date_val.strftime("%m-%d")
        
        # pd.Timestamp인 경우
        if isinstance(date_val, pd.Timestamp):
            return date_val.strftime("%m-%d")
        
        # 문자열인 경우
        date_str = str(date_val).strip()
        
        # 이미 MM-DD 형식이면 그대로 반환
        import re
        if re.match(r'^\d{2}-\d{2}$', date_str):
            return date_str
        
        # MM/DD 형식인 경우 MM-DD로 변환
        if re.match(r'^\d{2}/\d{2}$', date_str):
            return date_str.replace('/', '-')
        
        # YYYY-MM-DD 형식인 경우 MM-DD로 변환
        if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
            try:
                parts = date_str.split('-')
                if len(parts) >= 3:
                    return f"{parts[1]}-{parts[2]}"
            except (IndexError, ValueError) as e:
                self._log(f"날짜 파싱 오류 (YYYY-MM-DD): {e} (입력={date_str})", "DEBUG")
        
        # YYYY-MM-DD HH:MM:SS 형식인 경우
        if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', date_str):
            try:
                date_obj = pd.to_datetime(date_str)
                return date_obj.strftime("%m-%d")
            except (ValueError, TypeError) as e:
                self._log(f"날짜 파싱 오류 (YYYY-MM-DD HH:MM:SS): {e} (입력={date_str})", "DEBUG")
        
        # 숫자만 있는 문자열 (Excel serial number)
        if re.match(r'^\d+\.?\d*$', date_str):
            try:
                num_val = float(date_str)
                if 1 <= num_val <= 365:
                    # 1-365 범위는 Excel serial number가 아닐 수 있음
                    return ""
                else:
                    # Excel serial number로 변환
                    base_date = datetime(1899, 12, 30)
                    date_obj = base_date + pd.Timedelta(days=int(num_val))
                    return date_obj.strftime("%m-%d")
            except (ValueError, OverflowError) as e:
                self._log(f"숫자 날짜 변환 오류: {e} (입력={date_str})", "DEBUG")
        
        # pandas가 날짜로 인식한 경우
        try:
            date_obj = pd.to_datetime(date_str)
            return date_obj.strftime("%m-%d")
        except (ValueError, TypeError) as e:
            self._log(f"pandas 날짜 변환 오류: {e} (입력={date_str})", "DEBUG")
        
        # 변환 실패 시 원본 반환
        return date_str
    
    def _delete_row(self, sheet_name: str):
        """행 삭제"""
        if sheet_name not in self.sheet_trees:
            return
        
        tree = self.sheet_trees[sheet_name]
        selected = tree.selection()
        
        if not selected:
            messagebox.showwarning("선택 필요", "삭제할 행을 선택해주세요.")
            return
        
        # 확인 메시지
        if not messagebox.askyesno("확인", f"{len(selected)}개 행을 삭제하시겠습니까?"):
            return
        
        # 선택된 행 삭제
        for item_id in selected:
            tree.delete(item_id)
        
        # 행 수 업데이트
        if sheet_name in self.row_count_vars:
            count = len(tree.get_children())
            self.row_count_vars[sheet_name].set(f"행 수: {count:,}개")
        
        self._log(f"시트 '{sheet_name}': {len(selected)}개 행 삭제됨")
    
    def _edit_cell(self, sheet_name: str, tree: ttk.Treeview, event):
        """셀 편집 (더블클릭)"""
        # 편집 중이면 무시
        if self.editing_cell:
            return
        
        # 클릭한 위치 확인
        region = tree.identify_region(event.x, event.y)
        if region == "heading":
            return  # 헤더 클릭은 무시
        
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        
        if not item or not column:
            return
        
        # 컬럼 인덱스 (column은 '#1', '#2' 형식)
        col_index = int(column.replace('#', '')) - 1  # '#0'은 첫 번째 컬럼이지만 보통 숨김
        
        df = self.excel_data[sheet_name]
        columns = list(df.columns)
        
        if col_index >= len(columns):
            return
        
        col_name = columns[col_index]
        
        # 현재 값 가져오기
        values = list(tree.item(item, 'values'))
        current_value = values[col_index] if col_index < len(values) else ""
        
        # 편집 창 생성
        self._open_cell_editor(tree, item, col_index, col_name, current_value, sheet_name)
    
    def _open_cell_editor(self, tree: ttk.Treeview, item: str, col_index: int, 
                         col_name: str, current_value: str, sheet_name: str):
        """셀 편집 창 열기"""
        # 편집 창
        edit_window = tk.Toplevel(self)
        edit_window.title(f"셀 편집: {col_name}")
        edit_window.geometry("500x150")
        edit_window.transient(self)
        edit_window.grab_set()
        
        # 현재 위치 계산
        bbox = tree.bbox(item, column=f"#{col_index+1}")
        if bbox:
            x, y, width, height = bbox
            edit_window.geometry(f"500x150+{x+100}+{y+100}")
        
        frame = ttk.Frame(edit_window, padding=15)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text=f"컬럼: {col_name}", font=("맑은 고딕", 10, "bold")).pack(anchor='w', pady=(0, 10))
        
        # 입력 필드
        value_var = tk.StringVar(value=str(current_value))
        entry = ttk.Entry(frame, textvariable=value_var, font=("맑은 고딕", 10))
        entry.pack(fill='x', pady=(0, 15))
        entry.select_range(0, tk.END)
        entry.focus()
        
        # 버튼
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x')
        
        def save_and_close():
            new_value = value_var.get().strip()
            
            # Treeview 값 업데이트
            values = list(tree.item(item, 'values'))
            while len(values) <= col_index:
                values.append("")
            values[col_index] = new_value
            tree.item(item, values=values)
            
            self._log(f"시트 '{sheet_name}': {col_name} 셀 수정됨")
            self._mark_modified()  # 변경사항 표시
            edit_window.destroy()
            self.editing_cell = None
        
        def cancel():
            edit_window.destroy()
            self.editing_cell = None
        
        ttk.Button(btn_frame, text="저장", command=save_and_close).pack(side='right', padx=(5, 0))
        ttk.Button(btn_frame, text="취소", command=cancel).pack(side='right')
        
        # Enter 키로 저장
        entry.bind('<Return>', lambda e: save_and_close())
        entry.bind('<Escape>', lambda e: cancel())
        
        self.editing_cell = (tree, item, col_index)
    
    def _mark_modified(self):
        """변경사항이 있다고 표시"""
        if not self.has_unsaved_changes:
            self.has_unsaved_changes = True
            current_title = self.title()
            if not current_title.startswith('*'):
                self.title('*' + current_title)
    
    def _mark_saved(self):
        """저장 완료 표시"""
        self.has_unsaved_changes = False
        current_title = self.title()
        if current_title.startswith('*'):
            self.title(current_title.lstrip('*'))
    
    def _on_closing(self):
        """창 닫기 시 확인"""
        if self.has_unsaved_changes:
            result = messagebox.askyesnocancel(
                "저장 확인",
                "저장하지 않은 변경사항이 있습니다.\n\n"
                "저장하고 종료하시겠습니까?\n\n"
                "예: 저장 후 종료\n"
                "아니오: 저장하지 않고 종료\n"
                "취소: 종료 취소"
            )
            if result is None:  # 취소
                return
            elif result:  # 예 - 저장 후 종료
                if self._save_excel():
                    self.destroy()
                else:
                    return  # 저장 실패 시 종료 취소
            # 아니오 - 저장하지 않고 종료
        self.destroy()
    
    def _setup_auto_backup(self):
        """자동 백업 설정"""
        if not self.auto_backup_enabled:
            return
        
        def auto_backup():
            if self.excel_path and self.has_unsaved_changes and os.path.exists(self.excel_path):
                try:
                    backup_dir = os.path.join(os.path.dirname(self.excel_path), "자동백업")
                    os.makedirs(backup_dir, exist_ok=True)
                    
                    backup_filename = os.path.basename(self.excel_path).replace(
                        '.xlsx', 
                        f'_자동백업_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                    )
                    backup_path = os.path.join(backup_dir, backup_filename)
                    
                    import shutil
                    shutil.copy2(self.excel_path, backup_path)
                    self._log(f"자동 백업 생성: {backup_filename}", "INFO")
                    
                    # 오래된 백업 파일 정리 (최대 20개만 유지)
                    self._cleanup_old_backups(backup_dir)
                except Exception as e:
                    self._log(f"자동 백업 실패: {e}", "ERROR")
            
            # 다음 자동 백업 스케줄링
            if self.auto_backup_enabled:
                self._auto_backup_job = self.after(
                    self.auto_backup_interval * 60 * 1000, 
                    auto_backup
                )
        
        # 첫 자동 백업 스케줄링
        self._auto_backup_job = self.after(
            self.auto_backup_interval * 60 * 1000, 
            auto_backup
        )
    
    def _cleanup_old_backups(self, backup_dir: str, max_backups: int = 20):
        """오래된 백업 파일 정리"""
        try:
            backup_files = [
                os.path.join(backup_dir, f)
                for f in os.listdir(backup_dir)
                if f.endswith('.xlsx') and '자동백업' in f
            ]
            
            # 수정 시간 기준으로 정렬 (오래된 것부터)
            backup_files.sort(key=lambda x: os.path.getmtime(x))
            
            # 최대 개수 초과 시 오래된 파일 삭제
            if len(backup_files) > max_backups:
                files_to_delete = backup_files[:-max_backups]
                for file_path in files_to_delete:
                    try:
                        os.remove(file_path)
                        self._log(f"오래된 백업 파일 삭제: {os.path.basename(file_path)}", "DEBUG")
                    except Exception as e:
                        self._log(f"백업 파일 삭제 실패: {e}", "WARNING")
        except Exception as e:
            self._log(f"백업 정리 실패: {e}", "WARNING")
    
    def _save_excel(self):
        """Excel 파일 저장 (편집된 내용 반영, 기존 시트 구조 보존)"""
        if not self.excel_path:
            messagebox.showwarning("오류", "저장할 Excel 파일이 없습니다.")
            return False
        
        try:
            self._log("Excel 파일 저장 중...")
            
            # 백업 생성
            if os.path.exists(self.excel_path):
                backup_path = self.excel_path.replace('.xlsx', f'_백업_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                import shutil
                shutil.copy2(self.excel_path, backup_path)
                self._log(f"백업 생성: {os.path.basename(backup_path)}")
            
            # 기존 Excel 파일의 모든 시트 구조 보존을 위해 원본 파일 로드
            import openpyxl
            from openpyxl import load_workbook
            
            try:
                # 기존 파일을 openpyxl로 열어서 구조 보존
                wb_original = load_workbook(self.excel_path)
                original_sheet_names = wb_original.sheetnames
                original_active = wb_original.active
                wb_original.close()
                self._log(f"기존 시트 목록 확인: {', '.join(original_sheet_names)}")
            except Exception as e:
                self._log(f"⚠️ 기존 파일 구조 확인 실패: {e} - 새로 생성합니다.")
                original_sheet_names = []
                original_active = None
            
            # Treeview에서 DataFrame으로 데이터 복원 (편집된 시트만)
            updated_data = {}
            for sheet_name, tree in self.sheet_trees.items():
                # Treeview의 모든 행 읽기
                rows = []
                columns = list(self.excel_data[sheet_name].columns)
                
                for item_id in tree.get_children():
                    values = tree.item(item_id, 'values')
                    row_dict = {}
                    for i, col in enumerate(columns):
                        if i < len(values):
                            val = values[i]
                            # 빈 문자열은 NaN으로 변환
                            row_dict[col] = val if val else None
                        else:
                            row_dict[col] = None
                    rows.append(row_dict)
                
                # DataFrame 생성
                updated_df = pd.DataFrame(rows)
                # 원본 컬럼 순서 유지
                updated_df = updated_df[columns]
                updated_data[sheet_name] = updated_df
                self._log(f"시트 '{sheet_name}': {len(updated_df)}행 저장 준비")
            
            # 편집되지 않은 시트는 원본 데이터 유지
            for sheet_name in original_sheet_names:
                if sheet_name not in updated_data and sheet_name in self.excel_data:
                    updated_data[sheet_name] = self.excel_data[sheet_name]
                    self._log(f"시트 '{sheet_name}': 원본 데이터 유지 ({len(self.excel_data[sheet_name])}행)")
            
            # Excel 파일로 저장
            # mode='w'로 새로 작성하되, 모든 시트를 포함
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='w') as writer:
                # 시트 순서 유지 (원본 순서 우선, 새 시트는 뒤에 추가)
                sheet_order = []
                
                # 1. 원본 순서대로 추가
                for orig_sheet in original_sheet_names:
                    if orig_sheet in updated_data:
                        sheet_order.append(orig_sheet)
                
                # 2. 새로운 시트 추가
                for new_sheet in updated_data.keys():
                    if new_sheet not in sheet_order:
                        sheet_order.append(new_sheet)
                
                # 시트 저장
                for sheet_name in sheet_order:
                    df = updated_data[sheet_name].copy()
                    
                    # SEASON_MASTER 시트의 날짜 컬럼을 문자열로 변환
                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                        start_col = self._find_column(df, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
                        end_col = self._find_column(df, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
                        
                        # datetime 객체를 MM-DD 형식 문자열로 변환
                        if start_col and start_col in df.columns:
                            df[start_col] = df[start_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                        if end_col and end_col in df.columns:
                            df[end_col] = df[end_col].apply(lambda x: self._convert_date_to_mmdd(x) if pd.notna(x) else "")
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
                # workbook 가져오기 및 시트 상태 설정
                workbook = writer.book
                
                # 시트 상태 확인 및 설정
                visible_sheets = []
                for ws in workbook.worksheets:
                    if ws.sheet_state != 'hidden' and ws.sheet_state != 'veryHidden':
                        visible_sheets.append(ws)
                
                # 최소 하나의 시트는 visible 상태 유지
                if not visible_sheets and workbook.worksheets:
                    first_sheet = workbook.worksheets[0]
                    first_sheet.sheet_state = 'visible'
                    visible_sheets.append(first_sheet)
                    self._log(f"⚠️ 모든 시트가 숨겨져 있어 첫 번째 시트 '{first_sheet.title}'를 visible로 설정했습니다.")
                
                # Active 시트 설정 (원본 active 시트 우선)
                if original_active and original_active.title in [ws.title for ws in workbook.worksheets]:
                    try:
                        workbook.active = workbook[original_active.title]
                    except:
                        if workbook.worksheets:
                            workbook.active = workbook.worksheets[0]
                elif workbook.worksheets:
                    workbook.active = workbook.worksheets[0]
            
            # 메모리 데이터도 업데이트 (편집된 시트만)
            for sheet_name in updated_data:
                self.excel_data[sheet_name] = updated_data[sheet_name]
            
            self._log(f"✅ Excel 파일 저장 완료: {os.path.basename(self.excel_path)}", "SUCCESS")
            self._mark_saved()
            messagebox.showinfo("완료", f"Excel 파일이 저장되었습니다.\n\n{os.path.basename(self.excel_path)}")
            return True
            
        except PermissionError:
            self._log("❌ Excel 파일이 열려있어 저장할 수 없습니다.")
            messagebox.showerror("저장 실패", 
                "Excel 파일이 다른 프로그램에서 열려있습니다.\n\n"
                "파일을 닫고 다시 시도해주세요.")
        except Exception as e:
            self._log(f"❌ Excel 저장 실패: {e}")
            import traceback
            self._log(traceback.format_exc())
            messagebox.showerror("오류", f"Excel 파일 저장 중 오류가 발생했습니다:\n{e}")
    
    def _save_and_compile(self):
        """Excel 저장 및 JSON 컴파일 통합"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            messagebox.showwarning("오류", "Excel 파일을 먼저 로드해주세요.")
            return
        
        try:
            # 1. Excel 저장
            self._log("Excel 저장 중...")
            if self._save_excel():
                self._log("✅ Excel 저장 완료")
            else:
                messagebox.showerror("오류", "Excel 저장에 실패했습니다.")
                return
            
            # 2. JSON 컴파일
            self._log("JSON 컴파일 시작...")
            
            # JSON 파일 경로
            json_path = os.path.join(
                os.path.dirname(self.excel_path),
                "season_filters.json"
            )
            
            # Excel → JSON 변환
            config = self._parse_excel_to_config()
            
            # JSON 저장
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            self._log(f"✅ JSON 컴파일 완료: {os.path.basename(json_path)}")
            messagebox.showinfo("완료", f"Excel 저장 및 JSON 컴파일이 완료되었습니다.\n\nJSON 파일: {json_path}")
            
        except Exception as e:
            self._log(f"❌ 저장 및 컴파일 실패: {e}", "ERROR")
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            messagebox.showerror("오류", f"저장 및 컴파일 중 오류가 발생했습니다:\n{e}")
    
    def _open_excel(self):
        """Excel 파일 열기"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            messagebox.showwarning("오류", "Excel 파일을 먼저 로드해주세요.")
            return
        
        try:
            import subprocess
            import platform
            
            if platform.system() == 'Windows':
                os.startfile(self.excel_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', self.excel_path])
            else:  # Linux
                subprocess.call(['xdg-open', self.excel_path])
            
            self._log(f"✅ Excel 파일 열기: {os.path.basename(self.excel_path)}")
        except Exception as e:
            self._log(f"❌ Excel 파일 열기 실패: {e}", "ERROR")
            messagebox.showerror("오류", f"Excel 파일을 열 수 없습니다:\n{e}")
    
    def _compile_to_json(self):
        """Excel을 JSON으로 컴파일 (하위 호환성용, _save_and_compile 사용 권장)"""
        self._save_and_compile()
    
    def _parse_excel_to_config(self) -> Dict:
        """Excel을 JSON 구조로 변환"""
        config = {
            "version": "2.0",
            "settings": {
                "filter_mode": "exclude_expired",
                "case_sensitive": False
            },
            "types": {},
            "seasons": []
        }
        
        # 시트 이름 찾기 (대소문자 무시)
        def find_sheet(name_variants):
            for variant in name_variants:
                for sheet_name in self.excel_data.keys():
                    if sheet_name.upper() == variant.upper():
                        return sheet_name
            return None
        
        # TYPE_PRESETS 시트 처리
        type_sheet = find_sheet(["TYPE_PRESETS", "TYPES", "Type_Presets"])
        if type_sheet:
            df_types = self.excel_data[type_sheet]
            self._log(f"[JSON 컴파일] TYPE_PRESETS 시트 발견: {type_sheet}, 컬럼: {list(df_types.columns)}")
            
            type_id_col = self._find_column(df_types, ["타입ID", "타입", "type_id", "type"])
            prep_col = self._find_column(df_types, ["prep_days", "prep", "소싱기간"])
            grace_col = self._find_column(df_types, ["grace_days", "grace", "유예기간"])
            score_col = self._find_column(df_types, ["score_min", "score", "점수최소값"])
            
            self._log(f"[JSON 컴파일] TYPE_PRESETS 컬럼 매핑: type_id={type_id_col}, prep={prep_col}, grace={grace_col}, score={score_col}")
            
            for idx, row in df_types.iterrows():
                if type_id_col:
                    type_id = str(row.get(type_id_col, "")).strip()
                    if type_id and type_id not in ['nan', 'None', '']:
                        try:
                            prep_val = row.get(prep_col, 30) if prep_col else 30
                            grace_val = row.get(grace_col, 7) if grace_col else 7
                            score_val = row.get(score_col, 1) if score_col else 1
                            
                            config["types"][type_id] = {
                                "prep_days": int(prep_val) if pd.notna(prep_val) and str(prep_val).strip() else 30,
                                "grace_days": int(grace_val) if pd.notna(grace_val) and str(grace_val).strip() else 7,
                                "score_min": int(score_val) if pd.notna(score_val) and str(score_val).strip() else 1
                            }
                            self._log(f"[JSON 컴파일] 타입 추가: {type_id}")
                        except Exception as e:
                            self._log(f"[JSON 컴파일] 타입 처리 오류 (행 {idx}): {e}")
        else:
            self._log("[JSON 컴파일] ⚠️ TYPE_PRESETS 시트를 찾을 수 없습니다.")
        
        # SEASON_MASTER 시트 처리
        season_sheet = find_sheet(["SEASON_MASTER", "SEASONS", "Season_Master"])
        if season_sheet:
            df_seasons = self.excel_data[season_sheet]
            self._log(f"[JSON 컴파일] SEASON_MASTER 시트 발견: {season_sheet}, 컬럼: {list(df_seasons.columns)}")
            
            season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
            season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
            type_id_col = self._find_column(df_seasons, ["타입ID", "타입", "type_id", "type", "타입(type: Event/Climate/Activity/Lifecycle)"])
            start_col = self._find_column(df_seasons, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
            end_col = self._find_column(df_seasons, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
            cross_col = self._find_column(df_seasons, ["cross_year", "연도넘김", "연도초과", "시즌교차여부", "연도교차(Y/N)(cross_year)"])
            # Phase 1: 가공 기간 관련 컬럼 추가
            sourcing_start_col = self._find_column(df_seasons, ["소싱시작일수", "sourcing_start_days", "소싱시작", "prep_days"])
            processing_end_col = self._find_column(df_seasons, ["가공완료마감일수", "processing_end_days", "가공완료마감", "grace_days"])
            # 하위 호환성: 기존 컬럼명도 지원
            prep_override_col = self._find_column(df_seasons, ["prep_override", "prep_override_days"])
            grace_override_col = self._find_column(df_seasons, ["grace_override", "grace_override_days"])
            priority_col = self._find_column(df_seasons, ["priority", "우선순위", "우선도"])
            use_col = self._find_column(df_seasons, ["사용여부", "use", "enabled", "active"])
            
            self._log(f"[JSON 컴파일] SEASON_MASTER 컬럼 매핑: season_id={season_id_col}, name={season_name_col}, type={type_id_col}, start={start_col}, end={end_col}")
            
            for idx, row in df_seasons.iterrows():
                if not season_id_col:
                    continue
                    
                season_id = str(row.get(season_id_col, "")).strip()
                if not season_id or season_id in ['nan', 'None', '']:
                    continue
                
                # 사용여부 체크
                if use_col:
                    use_val = str(row.get(use_col, "Y")).strip().upper()
                    if use_val not in ['Y', 'YES', 'TRUE', '1']:
                        continue
                
                type_id = str(row.get(type_id_col, "")).strip() if type_id_col else ""
            
            # 환경설정에서 기본값 로드
            default_config = load_default_config()
            default_sourcing_start_days = default_config.get("default_sourcing_start_days", 30)
            default_processing_end_days = default_config.get("default_processing_end_days", 21)
            use_excel_values = default_config.get("use_excel_values", True)
            force_default_values = default_config.get("force_default_values", False)
            
            # 기본값 강제 적용 옵션이 켜져있으면 Excel 값 무시
            if force_default_values:
                use_excel_values = False
            
            # 타입별 설정 우선 적용 (환경설정의 type_configs)
            type_configs = default_config.get("type_configs", {})
            type_specific_config = type_configs.get(type_id, {})
            
            type_defaults = config["types"].get(type_id, {
                "prep_days": type_specific_config.get("sourcing_start_days", default_config.get("default_prep_days", 30)), 
                "grace_days": type_specific_config.get("processing_end_days", default_config.get("default_grace_days", 21)), 
                "score_min": default_config.get("default_score_min", 1)
            })
            
            # Phase 1: 가공 기간 관련 값 가져오기 (새 컬럼 우선, 기존 컬럼은 하위 호환)
            sourcing_start_days = None
            if use_excel_values:
                if sourcing_start_col:
                    sourcing_start_days = row.get(sourcing_start_col, None)
                if (sourcing_start_days is None or pd.isna(sourcing_start_days) or str(sourcing_start_days).strip() == "") and prep_override_col:
                    sourcing_start_days = row.get(prep_override_col, None)
            
            processing_end_days = None
            if use_excel_values:
                if processing_end_col:
                    processing_end_days = row.get(processing_end_col, None)
                if (processing_end_days is None or pd.isna(processing_end_days) or str(processing_end_days).strip() == "") and grace_override_col:
                    processing_end_days = row.get(grace_override_col, None)
                
            # 기본값 적용 (force_default_values가 True이면 항상 환경설정 기본값 사용)
            if force_default_values:
                # 타입별 설정이 있으면 타입별 설정 사용, 없으면 환경설정 기본값 사용
                if type_specific_config:
                    sourcing_start_days = type_specific_config.get("sourcing_start_days", default_sourcing_start_days)
                    processing_end_days = type_specific_config.get("processing_end_days", default_processing_end_days)
                else:
                    sourcing_start_days = default_sourcing_start_days
                    processing_end_days = default_processing_end_days
            else:
                # 기존 로직: Excel 값이 없으면 타입별 설정 → 타입 기본값 → 환경설정 기본값 순으로 사용
                if sourcing_start_days is None or pd.isna(sourcing_start_days) or str(sourcing_start_days).strip() == "":
                    # 타입별 설정 우선, 없으면 타입 기본값, 마지막으로 환경설정 기본값
                    sourcing_start_days = (type_specific_config.get("sourcing_start_days") 
                                         or type_defaults.get("prep_days") 
                                         or default_sourcing_start_days)
                if processing_end_days is None or pd.isna(processing_end_days) or str(processing_end_days).strip() == "":
                    processing_end_days = (type_specific_config.get("processing_end_days")
                                         or type_defaults.get("grace_days")
                                         or default_processing_end_days)
                
                # cross_year 처리
                cross_year_val = False
                if cross_col:
                    cross_val = str(row.get(cross_col, "")).strip().upper()
                    cross_year_val = cross_val in ['Y', 'YES', 'TRUE', '1', 'TRUE']
                
                season = {
                    "id": season_id,
                    "name": str(row.get(season_name_col, season_id)).strip() if season_name_col else season_id,
                    "type": type_id if type_id else "default",
                    "start_date": str(row.get(start_col, "")).strip() if start_col else "",
                    "end_date": str(row.get(end_col, "")).strip() if end_col else "",
                    "cross_year": cross_year_val,
                    # Phase 1: 가공 기간 관련 필드 (새 필드명 사용)
                    "sourcing_start_days": int(sourcing_start_days) if pd.notna(sourcing_start_days) else default_sourcing_start_days,
                    "processing_end_days": int(processing_end_days) if pd.notna(processing_end_days) else default_processing_end_days,
                    # 하위 호환성: 기존 필드명도 유지
                    "prep_days": int(sourcing_start_days) if pd.notna(sourcing_start_days) else default_sourcing_start_days,
                    "grace_days": int(processing_end_days) if pd.notna(processing_end_days) else default_processing_end_days,
                    "priority": int(row.get(priority_col, 1)) if priority_col and pd.notna(row.get(priority_col)) else 1,
                    "keywords": {
                        "include": [],
                        "exclude": [],
                        "allowed": []  # Phase 1: allowed 타입 추가
                    }
                }
                
                # KEYWORDS 시트에서 해당 시즌 키워드 찾기
                keyword_sheet = find_sheet(["KEYWORDS", "KEYWORD"])
                if keyword_sheet:
                    df_keywords = self.excel_data[keyword_sheet]
                    kw_season_id_col = self._find_column(df_keywords, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                    keyword_col = self._find_column(df_keywords, ["키워드", "keyword", "단어", "키워드(keyword)"])
                    polarity_col = self._find_column(df_keywords, ["polarity", "타입", "종류", "포함여부", "극성", "포함/제외(polarity: include/exclude)"])
                    weight_col = self._find_column(df_keywords, ["가중치", "weight", "점수"])
                    
                    if kw_season_id_col and keyword_col:
                        season_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() == season_id]
                        for _, kw_row in season_keywords.iterrows():
                            keyword = str(kw_row.get(keyword_col, "")).strip()
                            if not keyword or keyword in ['nan', 'None', '']:
                                continue
                                
                            # Phase 1: 타입 컬럼도 지원 (polarity와 동일)
                            type_col = self._find_column(df_keywords, ["타입", "type"])
                            polarity_str = None
                            if polarity_col:
                                polarity_str = str(kw_row.get(polarity_col, "")).strip().lower()
                            if (not polarity_str or polarity_str in ['nan', 'none', '']) and type_col:
                                polarity_str = str(kw_row.get(type_col, "")).strip().lower()
                            if not polarity_str or polarity_str in ['nan', 'none', '']:
                                polarity_str = "include"
                            
                            weight = float(kw_row.get(weight_col, 1.0)) if weight_col and pd.notna(kw_row.get(weight_col)) else 1.0
                            
                            # polarity/타입 변환 (Phase 1: allowed 추가)
                            if polarity_str in ["include", "포함", "1", "true", "yes"]:
                                polarity = "include"
                            elif polarity_str in ["exclude", "제외", "0", "false", "no"]:
                                polarity = "exclude"
                            elif polarity_str in ["allowed", "예외허용", "allow", "예외", "allowed"]:
                                polarity = "allowed"
                            else:
                                polarity = "include"  # 기본값
            
                            season["keywords"][polarity].append({
                                "keyword": keyword,
                                "weight": weight
                            })
                
                config["seasons"].append(season)
                self._log(f"[JSON 컴파일] 시즌 추가: {season_id} ({season['name']})")
        else:
            self._log("[JSON 컴파일] ⚠️ SEASON_MASTER 시트를 찾을 수 없습니다.")
        
        self._log(f"[JSON 컴파일] 완료: 타입 {len(config['types'])}개, 시즌 {len(config['seasons'])}개")
        return config
    
    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """DataFrame에서 가능한 컬럼명 중 하나를 찾기 (괄호 포함 형식도 지원)"""
        # 1. 정확한 매칭 시도
        for name in possible_names:
            if name in df.columns:
                return name
        
        # 2. 괄호 포함 형식 매칭 (예: "시즌ID(season_id)" -> "시즌ID" 또는 "season_id" 매칭)
        #    또는 "시작(MM-DD)(start_mmdd)" -> "시작", "start", "start_mmdd" 매칭
        for name in possible_names:
            for col in df.columns:
                col_str = str(col)
                # 컬럼명에 괄호가 있으면 괄호 앞부분 또는 괄호 안 내용 매칭
                if '(' in col_str and ')' in col_str:
                    # 괄호 앞부분 추출 (예: "시작(MM-DD)(start_mmdd)" -> "시작(MM-DD)")
                    # 여러 괄호가 있을 수 있으므로 첫 번째 괄호 전까지
                    before_paren = col_str.split('(')[0].strip()
                    # 모든 괄호 안 내용 추출
                    import re
                    inside_parens = re.findall(r'\(([^)]+)\)', col_str)
                    
                    # 괄호 앞부분과 매칭
                    if name.lower() == before_paren.lower():
                        return col
                    
                    # 괄호 안 내용 중 하나와 매칭
                    for inside in inside_parens:
                        inside_clean = inside.strip()
                        # 부분 매칭 (예: "start_mmdd"에서 "start" 매칭)
                        if name.lower() == inside_clean.lower() or name.lower() in inside_clean.lower():
                            return col
                        
                        # 언더스코어로 분리된 부분 매칭
                        for part in inside_clean.split('_'):
                            if name.lower() == part.lower():
                                return col
                
                # 괄호 없이도 부분 매칭 시도
                if name.lower() in col_str.lower() or col_str.lower() in name.lower():
                    return col
        
        return None
    
    # _preview_json 함수는 더 이상 사용하지 않음 (제거됨)
    
    def _show_config_dialog(self):
        """환경설정 다이얼로그"""
        config_window = tk.Toplevel(self)
        config_window.title("⚙️ 시즌 필터링 환경설정")
        config_window.geometry("700x750")
        config_window.transient(self)
        config_window.grab_set()
        
        # 현재 설정 로드
        current_config = load_default_config()
        
        # 메인 컨테이너 프레임
        main_container = ttk.Frame(config_window)
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 스크롤 가능한 프레임 생성
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        main_frame = scrollable_frame
        main_frame.configure(padding=20)
        
        # 설명
        desc_label = ttk.Label(main_frame, 
                              text="시즌 필터링 기본값 설정\nExcel에 값이 없을 때 사용되는 기본값입니다.",
                              font=("맑은 고딕", 9))
        desc_label.pack(pady=(0, 20))
        
        # 용어 설명 프레임
        explanation_frame = ttk.LabelFrame(main_frame, text="📖 용어 설명", padding=15)
        explanation_frame.pack(fill='x', pady=(0, 15))
        
        explanation_text = (
            "• 소싱 시작일수: 시즌 시작일 기준으로 몇 일 전부터 상품 소싱을 시작할지 설정합니다.\n"
            "  예) 시즌 시작일이 01-15이고 소싱 시작일수가 30일이면, 12-16부터 소싱을 시작합니다.\n\n"
            "• 가공완료마감일수: 시즌 종료일 기준으로 몇 일 전까지 가공을 완료해야 하는지 설정합니다.\n"
            "  예) 시즌 종료일이 03-31이고 가공완료마감일수가 21일이면, 03-10까지 가공을 완료해야 합니다."
        )
        ttk.Label(explanation_frame, text=explanation_text,
                 font=("맑은 고딕", 8), foreground="#333", justify='left').pack(anchor='w')
        
        # 설정 값 입력 프레임
        config_frame = ttk.LabelFrame(main_frame, text="전역 기본값 설정", padding=15)
        config_frame.pack(fill='x', pady=(0, 15))
        
        entries = {}
        
        # 소싱 시작일수
        row1 = ttk.Frame(config_frame)
        row1.pack(fill='x', pady=5)
        ttk.Label(row1, text="소싱 시작일수 (기본값):", width=25, anchor='e').pack(side='left', padx=(0, 10))
        var1 = tk.StringVar(value=str(current_config.get("default_sourcing_start_days", 30)))
        entry1 = ttk.Entry(row1, textvariable=var1, width=10)
        entry1.pack(side='left')
        ttk.Label(row1, text="일", width=5).pack(side='left', padx=(5, 0))
        entries["default_sourcing_start_days"] = (var1, entry1)
        
        # 가공완료마감일수
        row2 = ttk.Frame(config_frame)
        row2.pack(fill='x', pady=5)
        ttk.Label(row2, text="가공완료마감일수 (기본값):", width=25, anchor='e').pack(side='left', padx=(0, 10))
        var2 = tk.StringVar(value=str(current_config.get("default_processing_end_days", 21)))
        entry2 = ttk.Entry(row2, textvariable=var2, width=10)
        entry2.pack(side='left')
        ttk.Label(row2, text="일", width=5).pack(side='left', padx=(5, 0))
        entries["default_processing_end_days"] = (var2, entry2)
        
        # Excel 값 사용 여부
        row3 = ttk.Frame(config_frame)
        row3.pack(fill='x', pady=5)
        use_excel_var = tk.BooleanVar(value=current_config.get("use_excel_values", True))
        ttk.Checkbutton(row3, text="Excel 값 우선 사용", 
                       variable=use_excel_var).pack(side='left')
        
        # 기본값 강제 적용 옵션
        row4 = ttk.Frame(config_frame)
        row4.pack(fill='x', pady=5)
        force_default_var = tk.BooleanVar(value=current_config.get("force_default_values", False))
        ttk.Checkbutton(row4, text="환경설정 기본값을 모든 시즌에 강제 적용 (Excel 값 무시)", 
                       variable=force_default_var).pack(side='left')
        
        # 설명
        info_label = ttk.Label(config_frame, 
                              text="※ Excel 값 우선 사용: Excel에 값이 있으면 Excel 값 사용, 없으면 환경설정 기본값 사용\n"
                                   "※ 기본값 강제 적용: 체크하면 모든 시즌에 환경설정 기본값 적용 (Excel 값 무시)\n"
                                   "※ 둘 다 체크 해제: Excel 값이 없을 때만 환경설정 기본값 사용",
                              font=("맑은 고딕", 8),
                              foreground="#666")
        info_label.pack(pady=(10, 0), anchor='w')
        
        # 타입별 설정 프레임
        type_config_frame = ttk.LabelFrame(main_frame, text="타입별 기본값 설정", padding=15)
        type_config_frame.pack(fill='both', expand=True, pady=(0, 15))
        
        # 타입 목록 가져오기
        type_ids = self._get_available_type_ids() if hasattr(self, '_get_available_type_ids') else []
        if not type_ids and hasattr(self, 'excel_data'):
            # Excel에서 직접 가져오기
            for sheet_name in ['TYPE_PRESETS', 'TYPES', 'Type_Presets']:
                if sheet_name in self.excel_data:
                    df = self.excel_data[sheet_name]
                    type_id_col = self._find_column(df, ["타입ID", "타입", "type_id", "type"])
                    if type_id_col:
                        type_ids = [str(val).strip() for val in df[type_id_col].dropna() 
                                  if str(val).strip() and str(val).strip().lower() not in ['nan', 'none', '']]
                        break
        
        type_entries = {}  # {type_id: {"sourcing": (var, entry), "processing": (var, entry)}}
        
        if type_ids:
            # 스크롤 가능한 타입 리스트
            type_canvas = tk.Canvas(type_config_frame, height=200)
            type_scrollbar = ttk.Scrollbar(type_config_frame, orient="vertical", command=type_canvas.yview)
            type_scrollable_frame = ttk.Frame(type_canvas)
            
            type_scrollable_frame.bind(
                "<Configure>",
                lambda e: type_canvas.configure(scrollregion=type_canvas.bbox("all"))
            )
            
            type_canvas.create_window((0, 0), window=type_scrollable_frame, anchor="nw")
            type_canvas.configure(yscrollcommand=type_scrollbar.set)
            
            # 타입별 설정 로드
            type_configs = current_config.get("type_configs", {})
            
            for type_id in sorted(type_ids):
                type_row_frame = ttk.Frame(type_scrollable_frame)
                type_row_frame.pack(fill='x', pady=3)
                
                # 타입 ID 레이블
                ttk.Label(type_row_frame, text=f"• {type_id}:", width=15, anchor='e').pack(side='left', padx=(0, 10))
                
                # 소싱 시작일수
                ttk.Label(type_row_frame, text="소싱", width=5).pack(side='left')
                type_sourcing_var = tk.StringVar(value=str(type_configs.get(type_id, {}).get("sourcing_start_days", 
                                current_config.get("default_sourcing_start_days", 30))))
                type_sourcing_entry = ttk.Entry(type_row_frame, textvariable=type_sourcing_var, width=8)
                type_sourcing_entry.pack(side='left', padx=(0, 5))
                ttk.Label(type_row_frame, text="일", width=3).pack(side='left', padx=(0, 10))
                
                # 가공완료마감일수
                ttk.Label(type_row_frame, text="가공", width=5).pack(side='left')
                type_processing_var = tk.StringVar(value=str(type_configs.get(type_id, {}).get("processing_end_days",
                                current_config.get("default_processing_end_days", 21))))
                type_processing_entry = ttk.Entry(type_row_frame, textvariable=type_processing_var, width=8)
                type_processing_entry.pack(side='left', padx=(0, 5))
                ttk.Label(type_row_frame, text="일", width=3).pack(side='left')
                
                type_entries[type_id] = {
                    "sourcing": (type_sourcing_var, type_sourcing_entry),
                    "processing": (type_processing_var, type_processing_entry)
                }
            
            type_canvas.pack(side='left', fill='both', expand=True)
            type_scrollbar.pack(side='right', fill='y')
            
            type_info_label = ttk.Label(type_config_frame,
                                      text="※ 타입별 설정은 Excel의 TYPE_PRESETS 시트에서 가져옵니다.\n"
                                           "※ 타입별 값이 설정되면 해당 타입의 시즌에 타입별 값이 우선 적용됩니다.",
                                      font=("맑은 고딕", 8),
                                      foreground="#666")
            type_info_label.pack(pady=(10, 0), anchor='w')
        else:
            ttk.Label(type_config_frame, 
                     text="타입 정보를 찾을 수 없습니다. Excel 파일에 TYPE_PRESETS 시트가 있는지 확인하세요.",
                     font=("맑은 고딕", 8),
                     foreground="#999").pack(pady=20)
        
        # Canvas와 Scrollbar 배치 (메인 컨테이너 내부)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 버튼 프레임 (config_window에 직접 배치)
        btn_frame = ttk.Frame(config_window)
        btn_frame.pack(fill='x', padx=20, pady=10)
        
        def save_config():
            try:
                # 값 검증 - 전역 기본값
                sourcing_days = int(entries["default_sourcing_start_days"][0].get())
                processing_days = int(entries["default_processing_end_days"][0].get())
                
                if sourcing_days < 0 or processing_days < 0:
                    messagebox.showerror("오류", "일수는 0 이상이어야 합니다.")
                    return
                
                # 타입별 설정 검증 및 수집
                type_configs = {}
                for type_id, type_entry_dict in type_entries.items():
                    try:
                        type_sourcing = int(type_entry_dict["sourcing"][0].get())
                        type_processing = int(type_entry_dict["processing"][0].get())
                        
                        if type_sourcing < 0 or type_processing < 0:
                            messagebox.showerror("오류", f"타입 '{type_id}'의 일수는 0 이상이어야 합니다.")
                            return
                        
                        type_configs[type_id] = {
                            "sourcing_start_days": type_sourcing,
                            "processing_end_days": type_processing
                        }
                    except ValueError:
                        messagebox.showerror("오류", f"타입 '{type_id}'의 일수는 숫자여야 합니다.")
                        return
                
                # 설정 저장
                new_config = {
                    "default_sourcing_start_days": sourcing_days,
                    "default_processing_end_days": processing_days,
                    "default_prep_days": sourcing_days,  # 하위 호환성
                    "default_grace_days": processing_days,  # 하위 호환성
                    "default_score_min": current_config.get("default_score_min", 1),
                    "use_excel_values": use_excel_var.get(),
                    "force_default_values": force_default_var.get(),
                    "type_configs": type_configs,  # 타입별 설정 추가
                    "common_exclude_keywords": current_config.get("common_exclude_keywords", []),
                    "description": current_config.get("description", "")
                }
                
                # force_default_values가 True이면 use_excel_values를 무효화
                if force_default_var.get():
                    new_config["use_excel_values"] = False
                
                if save_default_config(new_config):
                    # force_default_values가 True이면 Excel 파일에도 값을 업데이트
                    if force_default_var.get():
                        apply_result = self._apply_default_values_to_excel(sourcing_days, processing_days)
                        if apply_result:
                            type_msg = f"\n\n타입별 설정 {len(type_configs)}개 저장됨" if type_configs else ""
                            msg = f"환경설정이 저장되었습니다.\n\n모든 시즌에 기본값이 적용되었습니다:\n- 소싱 시작일수: {sourcing_days}일\n- 가공완료마감일수: {processing_days}일{type_msg}\n\nExcel 파일도 업데이트되었습니다."
                        else:
                            msg = "환경설정이 저장되었습니다.\n\n하지만 Excel 파일 업데이트에 실패했습니다.\nExcel 파일을 수동으로 확인해주세요."
                    else:
                        type_msg = f" (타입별 설정 {len(type_configs)}개 포함)" if type_configs else ""
                        msg = f"환경설정이 저장되었습니다.{type_msg}\n\n변경사항을 적용하려면 JSON을 다시 컴파일하세요."
                    
                    messagebox.showinfo("완료", msg)
                    log_msg = f"✅ 환경설정 저장: 소싱 시작일수={sourcing_days}일, 가공완료마감일수={processing_days}일"
                    if type_configs:
                        log_msg += f", 타입별 설정 {len(type_configs)}개"
                    self._log(log_msg)
                    config_window.destroy()
                    
                    # Excel 파일이 열려있으면 다시 로드
                    if self.excel_path and os.path.exists(self.excel_path):
                        self._load_excel()
                else:
                    messagebox.showerror("오류", "환경설정 저장에 실패했습니다.")
            except ValueError:
                messagebox.showerror("오류", "일수는 숫자여야 합니다.")
            except Exception as e:
                messagebox.showerror("오류", f"환경설정 저장 중 오류 발생:\n{e}")
        
        ttk.Button(btn_frame, text="💾 저장", command=save_config).pack(side='right', padx=(5, 0))
        ttk.Button(btn_frame, text="취소", command=config_window.destroy).pack(side='right')
    
    def _apply_default_values_to_excel(self, sourcing_days: int, processing_days: int) -> bool:
        """모든 시즌에 기본값을 Excel 파일에 적용"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            return False
        
        if 'SEASON_MASTER' not in self.excel_data:
            self._log("SEASON_MASTER 시트를 찾을 수 없습니다.", "WARNING")
            return False
        
        try:
            df = self.excel_data['SEASON_MASTER'].copy()
            
            # 컬럼 찾기
            sourcing_col = self._find_column(df, ["소싱시작일수", "sourcing_start_days", "prep_days", "prep_override"])
            processing_col = self._find_column(df, ["가공완료마감일수", "processing_end_days", "grace_days", "grace_override"])
            
            # 컬럼이 없으면 생성
            if not sourcing_col:
                sourcing_col = "소싱시작일수"
                if sourcing_col not in df.columns:
                    df[sourcing_col] = None
            if not processing_col:
                processing_col = "가공완료마감일수"
                if processing_col not in df.columns:
                    df[processing_col] = None
            
            # 모든 시즌에 값 적용
            df[sourcing_col] = sourcing_days
            df[processing_col] = processing_days
            
            # 메모리 데이터 업데이트
            self.excel_data['SEASON_MASTER'] = df
            
            # Excel 파일에 저장
            result = self._save_excel()
            
            if result:
                self._log(f"✅ Excel 파일의 모든 시즌에 기본값 적용 완료: 소싱={sourcing_days}일, 가공완료={processing_days}일", "SUCCESS")
            return result
            
        except Exception as e:
            self._log(f"Excel 파일에 기본값 적용 실패: {e}", "ERROR")
            import traceback
            self._log(traceback.format_exc(), "ERROR")
            return False
    
    def _create_exclude_keywords_panel(self, parent_frame: ttk.Frame):
        """예외단어 관리 패널 생성 (메인 GUI 내부)"""
        # 설명
        desc_frame = ttk.Frame(parent_frame)
        desc_frame.pack(fill='x', pady=(0, 10))
        
        # 중요 안내
        important_label = ttk.Label(desc_frame, 
                              text="⚠️ 중요: 예외단어는 Event (이벤트) 타입 시즌에만 적용됩니다!",
                              font=("맑은 고딕", 9, "bold"),
                              foreground="#F44336")
        important_label.pack(anchor='w', pady=(0, 5))
        
        # 상세 설명
        desc_label = ttk.Label(desc_frame, 
                              text="적용 정책:\n"
                                   "  ✅ 🎉 Event (이벤트): 예외단어 적용 (크리스마스, 할로윈 등)\n"
                                   "  ❌ 🏃 Activity (활동): 예외단어 미적용 (스키, 캠핑 등)\n"
                                   "  ❌ 🌤️ Climate (기후): 예외단어 미적용 (여름, 겨울 등)\n"
                                   "  ❌ 🔄 Lifecycle (생활주기): 예외단어 미적용 (이사철, 입학 등)\n\n"
                                   "※ 예외단어 추가 시 Event 타입 시즌의 키워드만 고려하세요\n"
                                   "※ 띄어쓰기, 특수문자 사용 불가 | 중복 등록 불가",
                              font=("맑은 고딕", 8),
                              foreground="#666",
                              justify='left')
        desc_label.pack(anchor='w', padx=(0, 10))
        
        # 필터 및 검색 프레임
        filter_frame = ttk.Frame(parent_frame)
        filter_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(filter_frame, text="필터:", font=("맑은 고딕", 9)).pack(side='left', padx=(0, 5))
        filter_var = tk.StringVar(value="all")
        search_var = tk.StringVar()
        
        ttk.Radiobutton(filter_frame, text="전체", variable=filter_var, value="all").pack(side='left', padx=5)
        ttk.Radiobutton(filter_frame, text="A-Z", variable=filter_var, value="az").pack(side='left', padx=5)
        ttk.Radiobutton(filter_frame, text="가-하", variable=filter_var, value="korean").pack(side='left', padx=5)
        
        ttk.Label(filter_frame, text="검색:", font=("맑은 고딕", 9)).pack(side='left', padx=(20, 5))
        search_entry = ttk.Entry(filter_frame, textvariable=search_var, width=20)
        search_entry.pack(side='left', padx=5)
        
        # 키워드 목록 프레임
        list_frame = ttk.Frame(parent_frame)
        list_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Treeview로 변경 (정렬 및 필터링 용이)
        columns = ('keyword',)
        keyword_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)
        keyword_tree.heading('keyword', text='예외단어')
        keyword_tree.column('keyword', width=400)
        
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=keyword_tree.yview)
        keyword_tree.configure(yscrollcommand=scrollbar.set)
        
        keyword_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        
        # 현재 공통 제외 키워드 로드
        current_config = load_default_config()
        current_keywords = current_config.get("common_exclude_keywords", [])
        keyword_data = []
        for kw in current_keywords:
            if isinstance(kw, dict):
                keyword_data.append(kw.get("keyword", kw.get("key", "")))
            else:
                keyword_data.append(str(kw))
        
        # 중복 제거 및 빈 값 제거
        keyword_data = list(set([k for k in keyword_data if k and k.strip()]))
        
        # 초기 데이터 저장 (필터링용)
        initial_keywords = keyword_data.copy()
        
        # Treeview에 데이터 추가 (정렬은 _sort_exclude_keywords에서 처리)
        for kw in keyword_data:
            keyword_tree.insert('', 'end', values=(kw,))
        
        # 초기 정렬
        self._sort_exclude_keywords(keyword_tree)
        
        # 필터 함수 수정 (초기 데이터 사용)
        def filter_keywords():
            filter_type = filter_var.get()
            search_text = search_var.get().lower()
            
            # 모든 키워드 가져오기 (초기 데이터 사용)
            all_keywords = initial_keywords.copy()
            
            # 필터링
            filtered = []
            for kw in all_keywords:
                kw_lower = kw.lower()
                
                # 검색 필터
                if search_text and search_text not in kw_lower:
                    continue
                
                # 타입 필터
                if filter_type == "all":
                    filtered.append(kw)
                elif filter_type == "az":
                    # 영문/숫자만
                    if kw and (kw[0].isalpha() or kw[0].isdigit()) and (not kw or ord(kw[0]) < 0xAC00):
                        filtered.append(kw)
                elif filter_type == "korean":
                    # 한글만
                    if kw and len(kw) > 0 and ord(kw[0]) >= 0xAC00 and ord(kw[0]) <= 0xD7A3:
                        filtered.append(kw)
            
            # Treeview 재구성
            for item in keyword_tree.get_children():
                keyword_tree.delete(item)
            
            # 필터링된 항목 추가
            for kw in filtered:
                keyword_tree.insert('', 'end', values=(kw,))
            
            # 정렬
            self._sort_exclude_keywords(keyword_tree)
        
        # 필터 함수 업데이트
        filter_var.trace('w', lambda *args: filter_keywords())
        search_var.trace('w', lambda *args: filter_keywords())
        
        # 키워드 추가/삭제 프레임
        control_frame = ttk.Frame(parent_frame)
        control_frame.pack(fill='x', pady=(0, 10))
        
        # 단일 키워드 추가
        single_add_frame = ttk.Frame(control_frame)
        single_add_frame.pack(fill='x', pady=(0, 5))
        
        ttk.Label(single_add_frame, text="새 키워드:", font=("맑은 고딕", 9)).pack(side='left', padx=(0, 5))
        var_new_keyword = tk.StringVar()
        entry_new_keyword = ttk.Entry(single_add_frame, textvariable=var_new_keyword, width=25)
        entry_new_keyword.pack(side='left', padx=5)
        
        def validate_and_add_keyword(keyword_str=None):
            if keyword_str is None:
                keyword_str = var_new_keyword.get().strip()
            else:
                keyword_str = keyword_str.strip()
            
            if not keyword_str:
                return False, "키워드를 입력하세요."
            
            # 띄어쓰기 체크
            if ' ' in keyword_str:
                return False, "띄어쓰기는 사용할 수 없습니다."
            
            # 특수문자 체크 (한글, 영문, 숫자만 허용)
            import re
            if not re.match(r'^[가-힣a-zA-Z0-9]+$', keyword_str):
                return False, "특수문자는 사용할 수 없습니다.\n한글, 영문, 숫자만 사용 가능합니다."
            
            # 중복 체크 (초기 데이터 기준)
            if keyword_str in initial_keywords:
                return False, "이미 등록된 키워드입니다."
            
            # 초기 데이터에 추가
            initial_keywords.append(keyword_str)
            filter_keywords()  # 필터 재적용 (자동으로 추가됨)
            return True, f"'{keyword_str}' 추가 완료"
        
        def add_single_keyword():
            keyword = var_new_keyword.get().strip()
            success, msg = validate_and_add_keyword(keyword)
            if success:
                var_new_keyword.set("")
                self._log(f"✅ 예외단어 추가: {keyword}", "SUCCESS")
            else:
                messagebox.showerror("오류", msg)
        
        def add_batch_keywords():
            """예외단어 일괄 추가 다이얼로그"""
            batch_dialog = tk.Toplevel(parent_frame)
            batch_dialog.title("예외단어 일괄 추가")
            batch_dialog.geometry("500x400")
            batch_dialog.transient(self)
            batch_dialog.grab_set()
            
            main_frame = ttk.Frame(batch_dialog, padding=15)
            main_frame.pack(fill='both', expand=True)
            
            ttk.Label(main_frame, 
                     text="키워드를 쉼표(,) 또는 줄바꿈으로 구분하여 입력하세요:",
                     font=("맑은 고딕", 9)).pack(anchor='w', pady=(0, 5))
            
            # 텍스트 입력 영역
            text_frame = ttk.Frame(main_frame)
            text_frame.pack(fill='both', expand=True, pady=(0, 10))
            
            text_widget = ScrolledText(text_frame, height=15, wrap='word', font=("맑은 고딕", 9))
            text_widget.pack(fill='both', expand=True)
            
            # 하단 버튼
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(fill='x')
            
            def process_batch():
                text = text_widget.get("1.0", "end-1c").strip()
                if not text:
                    messagebox.showwarning("알림", "키워드를 입력하세요.")
                    return
                
                # 쉼표와 줄바꿈으로 분리
                keywords = []
                # 먼저 줄바꿈으로 분리
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    # 쉼표로도 분리
                    comma_separated = [k.strip() for k in line.split(',')]
                    keywords.extend([k for k in comma_separated if k])
                
                if not keywords:
                    messagebox.showwarning("알림", "추가할 키워드가 없습니다.")
                    return
                
                # 중복 제거
                keywords = list(set(keywords))
                
                # 각 키워드 검증 및 추가
                success_count = 0
                error_count = 0
                error_messages = []
                success_keywords = []
                
                for kw in keywords:
                    success, msg = validate_and_add_keyword(kw)
                    if success:
                        success_count += 1
                        success_keywords.append(kw)
                    else:
                        error_count += 1
                        error_messages.append(f"'{kw}': {msg}")
                
                # 결과 메시지
                result_msg = f"추가 완료: {success_count}개\n"
                if error_count > 0:
                    result_msg += f"실패: {error_count}개"
                
                if error_messages:
                    result_msg += "\n\n실패한 키워드:\n" + "\n".join(error_messages[:10])
                    if len(error_messages) > 10:
                        result_msg += f"\n... 외 {len(error_messages) - 10}개"
                
                messagebox.showinfo("일괄 추가 완료", result_msg)
                batch_dialog.destroy()
                
                if success_count > 0:
                    self._log(f"✅ 예외단어 일괄 추가: {success_count}개 추가됨", "SUCCESS")
            
            ttk.Button(btn_frame, text="일괄 추가", command=process_batch).pack(side='right', padx=(5, 0))
            ttk.Button(btn_frame, text="취소", command=batch_dialog.destroy).pack(side='right')
            
            text_widget.focus()
        
        def remove_keyword():
            selection = keyword_tree.selection()
            if not selection:
                messagebox.showwarning("알림", "삭제할 키워드를 선택하세요.")
                return
            
            keyword = keyword_tree.item(selection[0], 'values')[0]
            keyword_tree.delete(selection[0])
            if keyword in initial_keywords:
                initial_keywords.remove(keyword)
            filter_keywords()  # 필터 재적용
            self._log(f"✅ 예외단어 삭제: {keyword}", "SUCCESS")
        
        # 버튼 프레임
        btn_add_frame = ttk.Frame(control_frame)
        btn_add_frame.pack(fill='x')
        
        ttk.Button(single_add_frame, text="추가", command=add_single_keyword).pack(side='left', padx=2)
        ttk.Button(btn_add_frame, text="📋 일괄 추가 (쉼표/줄바꿈)", command=add_batch_keywords).pack(side='left', padx=(0, 5))
        ttk.Button(btn_add_frame, text="삭제", command=remove_keyword).pack(side='left', padx=2)
        
        # Enter 키로 추가
        entry_new_keyword.bind('<Return>', lambda e: add_single_keyword())
        
        # Event 타입 시즌 키워드 참고 섹션
        reference_frame = ttk.LabelFrame(parent_frame, text="📋 Event 타입 시즌 키워드 참고", padding=10)
        reference_frame.pack(fill='both', expand=True, pady=(10, 0))
        
        # 새로고침 버튼
        refresh_btn_frame = ttk.Frame(reference_frame)
        refresh_btn_frame.pack(fill='x', pady=(0, 5))
        ttk.Button(refresh_btn_frame, text="🔄 새로고침", 
                  command=lambda: self._refresh_event_keywords_reference(reference_frame, var_new_keyword, entry_new_keyword),
                  width=15).pack(side='right')
        
        # Event 타입 시즌 키워드 로드
        event_keywords = []
        error_msg = None
        
        # Excel 파일 로드 상태 확인
        if not hasattr(self, 'excel_data') or not self.excel_data:
            error_msg = "Excel 파일이 로드되지 않았습니다. 파일을 먼저 로드해주세요."
        elif not hasattr(self, 'excel_path') or not self.excel_path:
            error_msg = "Excel 파일 경로가 설정되지 않았습니다."
        else:
            try:
                # SEASON_MASTER 시트 찾기 (_update_season_list_tree와 동일한 로직)
                season_sheet = None
                for sheet_name in self.excel_data.keys():
                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                        season_sheet = sheet_name
                        break
                
                if not season_sheet:
                    error_msg = "SEASON_MASTER 시트를 찾을 수 없습니다."
                else:
                    df_seasons = self.excel_data[season_sheet]
                    
                    # 컬럼 찾기 (_update_season_list_tree와 동일한 로직)
                    type_col = self._find_column(df_seasons, ["타입", "type", "category", "타입(type: Event/Climate/Activity/Lifecycle)"])
                    season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                    season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
                    
                    if not type_col:
                        error_msg = "타입 컬럼을 찾을 수 없습니다."
                    elif not season_id_col:
                        error_msg = "시즌ID 컬럼을 찾을 수 없습니다."
                    else:
                        # Event 타입 시즌 찾기
                        event_seasons = df_seasons[df_seasons[type_col].astype(str).str.strip().str.upper() == 'EVENT']
                        
                        if len(event_seasons) == 0:
                            error_msg = "Event 타입 시즌을 찾을 수 없습니다."
                        else:
                            # KEYWORDS 시트 찾기
                            keyword_sheet = None
                            for sheet_name in self.excel_data.keys():
                                if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                                    keyword_sheet = sheet_name
                                    break
                            
                            if not keyword_sheet:
                                error_msg = "KEYWORDS 시트를 찾을 수 없습니다."
                            else:
                                df_keywords = self.excel_data[keyword_sheet]
                                
                                # 키워드 컬럼 찾기
                                kw_season_id_col = self._find_column(df_keywords, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                                keyword_col = self._find_column(df_keywords, ["키워드", "keyword", "단어", "키워드(keyword)"])
                                polarity_col = self._find_column(df_keywords, ["polarity", "타입", "종류", "포함여부", "극성", "포함/제외(polarity: include/exclude)"])
                                
                                if not kw_season_id_col or not keyword_col:
                                    error_msg = "KEYWORDS 시트의 컬럼을 찾을 수 없습니다."
                                else:
                                    # Event 타입 시즌별 키워드 수집
                                    for _, season_row in event_seasons.iterrows():
                                        season_id = str(season_row.get(season_id_col, "")).strip()
                                        season_name = str(season_row.get(season_name_col, season_id)).strip() if season_name_col else season_id
                                        
                                        if not season_id or season_id in ['nan', 'None', '']:
                                            continue
                                        
                                        # 해당 시즌의 include 키워드 찾기
                                        season_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() == season_id]
                                        include_keywords = []
                                        
                                        for _, kw_row in season_keywords.iterrows():
                                            keyword = str(kw_row.get(keyword_col, "")).strip()
                                            if not keyword or keyword in ['nan', 'None', '']:
                                                continue
                                            
                                            # include 키워드만 (polarity가 없거나 include인 경우)
                                            polarity_str = ""
                                            if polarity_col:
                                                polarity_str = str(kw_row.get(polarity_col, "")).strip().lower()
                                            
                                            # include이거나 polarity가 없는 경우 (기본값은 include)
                                            if not polarity_str or polarity_str in ["include", "포함", "1", "true", "yes", ""]:
                                                include_keywords.append(keyword)
                                        
                                        if include_keywords:
                                            event_keywords.append({
                                                "season_id": season_id,
                                                "season_name": season_name,
                                                "keywords": include_keywords
                                            })
                                    
                                    if len(event_keywords) == 0:
                                        error_msg = "Event 타입 시즌에 include 키워드가 없습니다."
            except Exception as e:
                error_msg = f"Event 타입 시즌 키워드 로드 실패: {str(e)}"
                import traceback
                self._log(f"Event 타입 시즌 키워드 로드 오류: {traceback.format_exc()}", "WARNING")
        
        # Event 키워드 표시
        if error_msg:
            ttk.Label(reference_frame, 
                     text=f"⚠️ {error_msg}",
                     font=("맑은 고딕", 8),
                     foreground="#F44336",
                     wraplength=600).pack(pady=10, padx=10, anchor='w')
        elif event_keywords:
            # 스크롤 가능한 프레임
            ref_canvas = tk.Canvas(reference_frame, height=150)
            ref_scrollbar = ttk.Scrollbar(reference_frame, orient='vertical', command=ref_canvas.yview)
            ref_scrollable = ttk.Frame(ref_canvas)
            
            ref_scrollable.bind(
                "<Configure>",
                lambda e: ref_canvas.configure(scrollregion=ref_canvas.bbox("all"))
            )
            
            ref_canvas.create_window((0, 0), window=ref_scrollable, anchor="nw")
            ref_canvas.configure(yscrollcommand=ref_scrollbar.set)
            
            ttk.Label(ref_scrollable, 
                     text="Event 타입 시즌의 include 키워드 (클릭하면 예외단어 입력 필드에 추가):",
                     font=("맑은 고딕", 8, "bold"),
                     foreground="#2196F3").pack(anchor='w', pady=(0, 5))
            
            for event_data in event_keywords:
                season_name = event_data["season_name"]
                keywords = event_data["keywords"]
                
                # 시즌명 표시
                season_label = ttk.Label(ref_scrollable, 
                                        text=f"  🎉 {season_name}:",
                                        font=("맑은 고딕", 8, "bold"),
                                        foreground="#2196F3")
                season_label.pack(anchor='w', padx=(5, 0), pady=(5, 2))
                
                # 키워드 버튼들
                kw_frame = ttk.Frame(ref_scrollable)
                kw_frame.pack(anchor='w', padx=(20, 0), pady=(0, 5), fill='x')
                
                for kw in keywords[:10]:  # 최대 10개만 표시
                    def add_to_entry(k=kw):
                        var_new_keyword.set(k)
                        entry_new_keyword.focus()
                    
                    btn = ttk.Button(kw_frame, text=kw, width=len(kw)+2, 
                                   command=add_to_entry)
                    btn.pack(side='left', padx=2, pady=2)
                
                if len(keywords) > 10:
                    ttk.Label(kw_frame, text=f"... 외 {len(keywords)-10}개", 
                             font=("맑은 고딕", 7),
                             foreground="#999").pack(side='left', padx=5)
            
            ref_canvas.pack(side='left', fill='both', expand=True)
            ref_scrollbar.pack(side='right', fill='y')
        else:
            ttk.Label(reference_frame, 
                     text="Event 타입 시즌을 찾을 수 없거나 Excel 파일이 로드되지 않았습니다.",
                     font=("맑은 고딕", 8),
                     foreground="#999").pack(pady=10)
        
        # 저장 버튼
        btn_frame = ttk.Frame(parent_frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        
        def save_exclude_keywords():
            # 초기 데이터에서 모든 키워드 가져오기 (필터링과 무관하게 전체 목록)
            all_keywords = initial_keywords.copy()
            
            # 현재 설정 로드
            current_config = load_default_config()
            current_config["common_exclude_keywords"] = all_keywords
            
            if save_default_config(current_config):
                self._log(f"✅ 예외단어 {len(all_keywords)}개가 저장되었습니다. 변경사항을 적용하려면 '저장 및 컴파일'을 실행하세요.", "SUCCESS")
                messagebox.showinfo("완료", f"예외단어 {len(all_keywords)}개가 저장되었습니다.\n\n변경사항을 적용하려면 '저장 및 컴파일'을 실행하세요.")
            else:
                self._log("❌ 예외단어 저장에 실패했습니다.", "ERROR")
                messagebox.showerror("오류", "예외단어 저장에 실패했습니다.")
        
        ttk.Button(btn_frame, text="💾 저장", command=save_exclude_keywords).pack(side='left', padx=(0, 5))
        
        # 필터 함수에 trace 추가
        def on_filter_change(*args):
            filter_keywords()
        
        filter_var.trace('w', on_filter_change)
        search_var.trace('w', on_filter_change)
        
        # 인스턴스 변수로 저장 (나중에 업데이트할 수 있도록)
        self.exclude_filter_var = filter_var
        self.exclude_search_var = search_var
        self.exclude_keyword_tree = keyword_tree
        self.exclude_initial_keywords = initial_keywords
        
        # 필터 함수를 위해 keyword_tree 참조 저장
        keyword_listbox = keyword_tree  # 호환성을 위해
    
    def _refresh_event_keywords_reference(self, reference_frame, var_new_keyword, entry_new_keyword):
        """Event 타입 시즌 키워드 참고 섹션 새로고침"""
        # 기존 내용 모두 제거 (새로고침 버튼 프레임 제외)
        children_to_keep = []
        for widget in reference_frame.winfo_children():
            # 새로고침 버튼이 있는 프레임은 유지
            if isinstance(widget, ttk.Frame):
                has_refresh_btn = any(
                    isinstance(child, ttk.Button) and '🔄 새로고침' in str(child.cget('text'))
                    for child in widget.winfo_children() if hasattr(child, 'winfo_children')
                )
                if not has_refresh_btn:
                    widget.destroy()
                else:
                    children_to_keep.append(widget)
            else:
                widget.destroy()
        
        # Event 키워드 다시 로드 (기존 로직 재사용)
        event_keywords = []
        error_msg = None
        
        if not hasattr(self, 'excel_data') or not self.excel_data:
            error_msg = "Excel 파일이 로드되지 않았습니다. 파일을 먼저 로드해주세요."
        elif not hasattr(self, 'excel_path') or not self.excel_path:
            error_msg = "Excel 파일 경로가 설정되지 않았습니다."
        else:
            try:
                season_sheet = None
                for sheet_name in self.excel_data.keys():
                    if sheet_name.upper() in ['SEASON_MASTER', 'SEASONS', 'SEASON_MASTER']:
                        season_sheet = sheet_name
                        break
                
                if not season_sheet:
                    error_msg = "SEASON_MASTER 시트를 찾을 수 없습니다."
                else:
                    df_seasons = self.excel_data[season_sheet]
                    type_col = self._find_column(df_seasons, ["타입", "type", "category", "타입(type: Event/Climate/Activity/Lifecycle)"])
                    season_id_col = self._find_column(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                    season_name_col = self._find_column(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
                    
                    if not type_col:
                        error_msg = "타입 컬럼을 찾을 수 없습니다."
                    elif not season_id_col:
                        error_msg = "시즌ID 컬럼을 찾을 수 없습니다."
                    else:
                        event_seasons = df_seasons[df_seasons[type_col].astype(str).str.strip().str.upper() == 'EVENT']
                        
                        if len(event_seasons) == 0:
                            error_msg = "Event 타입 시즌을 찾을 수 없습니다."
                        else:
                            keyword_sheet = None
                            for sheet_name in self.excel_data.keys():
                                if sheet_name.upper() in ['KEYWORDS', 'KEYWORD']:
                                    keyword_sheet = sheet_name
                                    break
                            
                            if not keyword_sheet:
                                error_msg = "KEYWORDS 시트를 찾을 수 없습니다."
                            else:
                                df_keywords = self.excel_data[keyword_sheet]
                                kw_season_id_col = self._find_column(df_keywords, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                                keyword_col = self._find_column(df_keywords, ["키워드", "keyword", "단어", "키워드(keyword)"])
                                polarity_col = self._find_column(df_keywords, ["polarity", "타입", "종류", "포함여부", "극성", "포함/제외(polarity: include/exclude)"])
                                
                                if not kw_season_id_col or not keyword_col:
                                    error_msg = "KEYWORDS 시트의 컬럼을 찾을 수 없습니다."
                                else:
                                    for _, season_row in event_seasons.iterrows():
                                        season_id = str(season_row.get(season_id_col, "")).strip()
                                        season_name = str(season_row.get(season_name_col, season_id)).strip() if season_name_col else season_id
                                        
                                        if not season_id or season_id in ['nan', 'None', '']:
                                            continue
                                        
                                        season_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() == season_id]
                                        include_keywords = []
                                        
                                        for _, kw_row in season_keywords.iterrows():
                                            keyword = str(kw_row.get(keyword_col, "")).strip()
                                            if not keyword or keyword in ['nan', 'None', '']:
                                                continue
                                            
                                            polarity_str = ""
                                            if polarity_col:
                                                polarity_str = str(kw_row.get(polarity_col, "")).strip().lower()
                                            
                                            if not polarity_str or polarity_str in ["include", "포함", "1", "true", "yes", ""]:
                                                include_keywords.append(keyword)
                                        
                                        if include_keywords:
                                            event_keywords.append({
                                                "season_id": season_id,
                                                "season_name": season_name,
                                                "keywords": include_keywords
                                            })
                                    
                                    if len(event_keywords) == 0:
                                        error_msg = "Event 타입 시즌에 include 키워드가 없습니다."
                                        
            except Exception as e:
                error_msg = f"Event 타입 시즌 키워드 로드 실패: {str(e)}"
                import traceback
                self._log(f"Event 타입 시즌 키워드 새로고침 오류: {traceback.format_exc()}", "WARNING")
        
        # 내용 다시 그리기
        if error_msg:
            ttk.Label(reference_frame, 
                     text=f"⚠️ {error_msg}",
                     font=("맑은 고딕", 8),
                     foreground="#F44336",
                     wraplength=600).pack(pady=10, padx=10, anchor='w')
        elif event_keywords:
            ref_canvas = tk.Canvas(reference_frame, height=150)
            ref_scrollbar = ttk.Scrollbar(reference_frame, orient='vertical', command=ref_canvas.yview)
            ref_scrollable = ttk.Frame(ref_canvas)
            
            ref_scrollable.bind(
                "<Configure>",
                lambda e: ref_canvas.configure(scrollregion=ref_canvas.bbox("all"))
            )
            
            ref_canvas.create_window((0, 0), window=ref_scrollable, anchor="nw")
            ref_canvas.configure(yscrollcommand=ref_scrollbar.set)
            
            ttk.Label(ref_scrollable, 
                     text="Event 타입 시즌의 include 키워드 (클릭하면 예외단어 입력 필드에 추가):",
                     font=("맑은 고딕", 8, "bold"),
                     foreground="#2196F3").pack(anchor='w', pady=(0, 5))
            
            for event_data in event_keywords:
                season_name = event_data["season_name"]
                keywords = event_data["keywords"]
                
                season_label = ttk.Label(ref_scrollable, 
                                        text=f"  🎉 {season_name}:",
                                        font=("맑은 고딕", 8, "bold"),
                                        foreground="#2196F3")
                season_label.pack(anchor='w', padx=(5, 0), pady=(5, 2))
                
                kw_frame = ttk.Frame(ref_scrollable)
                kw_frame.pack(anchor='w', padx=(20, 0), pady=(0, 5), fill='x')
                
                for kw in keywords[:10]:
                    def add_to_entry(k=kw):
                        var_new_keyword.set(k)
                        entry_new_keyword.focus()
                    
                    btn = ttk.Button(kw_frame, text=kw, width=len(kw)+2, 
                                   command=add_to_entry)
                    btn.pack(side='left', padx=2, pady=2)
                
                if len(keywords) > 10:
                    ttk.Label(kw_frame, text=f"... 외 {len(keywords)-10}개", 
                             font=("맑은 고딕", 7),
                             foreground="#999").pack(side='left', padx=5)
            
            ref_canvas.pack(side='left', fill='both', expand=True)
            ref_scrollbar.pack(side='right', fill='y')
        else:
            ttk.Label(reference_frame, 
                     text="Event 타입 시즌을 찾을 수 없거나 Excel 파일이 로드되지 않았습니다.",
                     font=("맑은 고딕", 8),
                     foreground="#999").pack(pady=10)
        
        self._log(f"Event 타입 시즌 키워드 참고 섹션 새로고침 완료: {len(event_keywords)}개 시즌", "INFO")
    
    def _sort_exclude_keywords(self, tree):
        """예외단어 목록 정렬 (한글 우선, 영문 다음)"""
        items = [(tree.item(item, 'values')[0], item) for item in tree.get_children()]
        
        def sort_key(item):
            keyword = item[0]
            # 한글인지 영문인지 판단
            if keyword and ord(keyword[0]) >= 0xAC00 and ord(keyword[0]) <= 0xD7A3:
                return (0, keyword)  # 한글 우선
            else:
                return (1, keyword.lower())  # 영문은 소문자로 정렬
        
        items.sort(key=sort_key)
        
        # Treeview 재구성
        for item in tree.get_children():
            tree.delete(item)
        
        for keyword, _ in items:
            tree.insert('', 'end', values=(keyword,))
    
    def _filter_exclude_keywords(self, tree, filter_var, search_var):
        """예외단어 필터링"""
        filter_type = filter_var.get()
        search_text = search_var.get().lower()
        
        # 모든 항목 가져오기
        all_items = []
        for item in tree.get_children():
            keyword = tree.item(item, 'values')[0]
            keyword_lower = keyword.lower()
            
            # 검색 필터
            if search_text and search_text not in keyword_lower:
                continue
            
            # 타입 필터
            if filter_type == "all":
                all_items.append(keyword)
            elif filter_type == "az":
                # 영문만 (숫자 포함)
                if keyword and (keyword[0].isalpha() or keyword[0].isdigit()) and (not keyword or ord(keyword[0]) < 0xAC00):
                    all_items.append(keyword)
            elif filter_type == "korean":
                # 한글만
                if keyword and len(keyword) > 0 and ord(keyword[0]) >= 0xAC00 and ord(keyword[0]) <= 0xD7A3:
                    all_items.append(keyword)
        
        # Treeview 재구성
        for item in tree.get_children():
            tree.delete(item)
        
        # 필터링된 항목 추가
        for kw in all_items:
            tree.insert('', 'end', values=(kw,))
        
        # 정렬
        self._sort_exclude_keywords(tree)
    
    def _show_help_dialog(self):
        """용어 설명 도움말 창"""
        help_window = tk.Toplevel(self)
        help_window.title("📖 용어 설명")
        help_window.geometry("900x700")
        help_window.transient(self)
        
        # 메인 프레임
        main_frame = ttk.Frame(help_window, padding=15)
        main_frame.pack(fill='both', expand=True)
        
        # 제목
        title_label = ttk.Label(main_frame, text="시즌 필터링 도구 - 용어 설명", 
                               font=("맑은 고딕", 14, "bold"))
        title_label.pack(pady=(0, 15))
        
        # 스크롤 가능한 텍스트 영역
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side='right', fill='y')
        
        help_text = ScrolledText(text_frame, wrap='word', font=("맑은 고딕", 10),
                                yscrollcommand=scrollbar.set, state='normal')
        help_text.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=help_text.yview)
        
        # 용어 설명 내용
        help_content = """📖 시즌 필터링 도구 - 용어 설명

기본 용어, 시즌 설정, 키워드 설정, 가공 기간 등에 대한 자세한 설명은
시즌필터링_사용자_용어설명.md 파일을 참고하세요."""

        help_text.insert('1.0', help_content)
        help_text.config(state='disabled')
    
    # _show_excel_structure 함수는 더 이상 사용하지 않음 (제거됨)


# ============================================================================
# 필터링 로직 함수 (data_export.py에서 사용)
# ============================================================================

def load_season_config(excel_path: str, json_path: str) -> Optional[Dict]:
    """
    시즌 설정 로딩 (자동 캐시 갱신)
    - Excel과 JSON의 mtime 비교
    - Excel이 더 최신이면 자동 재생성
    
    Args:
        excel_path: Excel 설정 파일 경로
        json_path: JSON 캐시 파일 경로
    
    Returns:
        시즌 설정 딕셔너리 또는 None
    """
    if not os.path.exists(excel_path):
        return None
    
    # mtime 비교
    excel_mtime = os.path.getmtime(excel_path)
    json_exists = os.path.exists(json_path)
    json_mtime = os.path.getmtime(json_path) if json_exists else 0
    
    # Excel이 더 최신이거나 JSON이 없으면 재생성
    if not json_exists or excel_mtime > json_mtime:
        try:
            # Excel 읽기 (engine 명시적으로 지정)
            xl = pd.ExcelFile(excel_path, engine='openpyxl')
            config = _parse_excel_to_config_static(xl)
            
            # JSON 저장
            os.makedirs(os.path.dirname(json_path), exist_ok=True)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            return config
        except Exception as e:
            # JSON이 있으면 이전 캐시 사용
            if json_exists:
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except:
                    pass
            return None
    else:
        # JSON 사용
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return None


def _parse_excel_to_config_static(xl: pd.ExcelFile) -> Dict:
    """Excel을 JSON 구조로 변환 (정적 함수)"""
    # 환경설정에서 공통 제외 키워드 가져오기
    default_config = load_default_config()
    common_exclude_keywords = default_config.get("common_exclude_keywords", [])
    
    config = {
        "version": "2.0",
        "settings": {
            "filter_mode": "exclude_expired",
            "case_sensitive": False,
            "common_exclude_keywords": common_exclude_keywords  # 공통 제외 키워드 추가
        },
        "types": {},
        "seasons": [],
        "keywords": {}  # 시즌별 키워드 저장 (빠른 접근용)
    }
    
    # 시트 이름 찾기 (대소문자 무시)
    def find_sheet(name_variants):
        for variant in name_variants:
            for sheet_name in xl.sheet_names:
                if sheet_name.upper() == variant.upper():
                    return sheet_name
        return None
    
    # TYPE_PRESETS 시트 처리
    type_sheet = find_sheet(["TYPE_PRESETS", "TYPES", "Type_Presets"])
    if type_sheet:
        df_types = pd.read_excel(xl, sheet_name=type_sheet)
        df_types = df_types.dropna(how='all')  # 빈 행 제거
        
        type_id_col = _find_column_static(df_types, ["타입ID", "타입", "type_id", "type"])
        prep_col = _find_column_static(df_types, ["prep_days", "prep", "소싱기간"])
        grace_col = _find_column_static(df_types, ["grace_days", "grace", "유예기간"])
        score_col = _find_column_static(df_types, ["score_min", "score", "점수최소값"])
        
        for _, row in df_types.iterrows():
            if type_id_col:
                type_id = str(row.get(type_id_col, "")).strip()
                if type_id and type_id not in ['nan', 'None', '']:
                    try:
                        prep_val = row.get(prep_col, 30) if prep_col else 30
                        grace_val = row.get(grace_col, 7) if grace_col else 7
                        score_val = row.get(score_col, 1) if score_col else 1
                        
                        config["types"][type_id] = {
                            "prep_days": int(prep_val) if pd.notna(prep_val) and str(prep_val).strip() else 30,
                            "grace_days": int(grace_val) if pd.notna(grace_val) and str(grace_val).strip() else 7,
                            "score_min": int(score_val) if pd.notna(score_val) and str(score_val).strip() else 1
                        }
                    except:
                        pass
    
    # SEASON_MASTER 시트 처리
    season_sheet = find_sheet(["SEASON_MASTER", "SEASONS", "Season_Master"])
    if season_sheet:
        df_seasons = pd.read_excel(xl, sheet_name=season_sheet)
        df_seasons = df_seasons.dropna(how='all')  # 빈 행 제거
        
        season_id_col = _find_column_static(df_seasons, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
        season_name_col = _find_column_static(df_seasons, ["시즌명", "시즌이름", "season_name", "name", "시즌명(season_name)"])
        type_id_col = _find_column_static(df_seasons, ["타입ID", "타입", "type_id", "type", "타입(type: Event/Climate/Activity/Lifecycle)"])
        start_col = _find_column_static(df_seasons, ["시작일", "start_date", "start", "시즌시작일", "시작(MM-DD)(start_mmdd)", "start_mmdd"])
        end_col = _find_column_static(df_seasons, ["종료일", "end_date", "end", "시즌종료일", "종료(MM-DD)(end_mmdd)", "end_mmdd"])
        cross_col = _find_column_static(df_seasons, ["cross_year", "연도넘김", "연도초과", "시즌교차여부", "연도교차(Y/N)(cross_year)"])
        # Phase 1: 가공 기간 관련 컬럼 추가
        sourcing_start_col = _find_column_static(df_seasons, ["소싱시작일수", "sourcing_start_days", "소싱시작", "prep_days"])
        processing_end_col = _find_column_static(df_seasons, ["가공완료마감일수", "processing_end_days", "가공완료마감", "grace_days"])
        # 하위 호환성: 기존 컬럼명도 지원
        prep_override_col = _find_column_static(df_seasons, ["prep_override", "prep_override_days"])
        grace_override_col = _find_column_static(df_seasons, ["grace_override", "grace_override_days"])
        priority_col = _find_column_static(df_seasons, ["priority", "우선순위", "우선도"])
        use_col = _find_column_static(df_seasons, ["사용여부", "use", "enabled", "active", "사용여부(Y/N)(enabled)"])
        
        for _, row in df_seasons.iterrows():
            if not season_id_col:
                continue
                
            season_id = str(row.get(season_id_col, "")).strip()
            if not season_id or season_id in ['nan', 'None', '']:
                continue
            
            # 사용여부 체크
            if use_col:
                use_val = str(row.get(use_col, "Y")).strip().upper()
                if use_val not in ['Y', 'YES', 'TRUE', '1']:
                    continue
            
            type_id = str(row.get(type_id_col, "")).strip() if type_id_col else ""
            
            # 환경설정에서 기본값 로드
            default_config = load_default_config()
            default_sourcing_start_days = default_config.get("default_sourcing_start_days", 30)
            default_processing_end_days = default_config.get("default_processing_end_days", 21)
            use_excel_values = default_config.get("use_excel_values", True)
            
            # 타입별 설정 우선 적용 (환경설정의 type_configs)
            type_configs = default_config.get("type_configs", {})
            type_specific_config = type_configs.get(type_id, {})
            
            type_defaults = config["types"].get(type_id, {
                "prep_days": type_specific_config.get("sourcing_start_days", default_config.get("default_prep_days", 30)), 
                "grace_days": type_specific_config.get("processing_end_days", default_config.get("default_grace_days", 21)), 
                "score_min": default_config.get("default_score_min", 1)
            })
            
            # Phase 1: 가공 기간 관련 값 가져오기 (새 컬럼 우선, 기존 컬럼은 하위 호환)
            sourcing_start_days = None
            if use_excel_values:
                if sourcing_start_col:
                    sourcing_start_days = row.get(sourcing_start_col, None)
                if (sourcing_start_days is None or pd.isna(sourcing_start_days) or str(sourcing_start_days).strip() == "") and prep_override_col:
                    sourcing_start_days = row.get(prep_override_col, None)
            
            processing_end_days = None
            if use_excel_values:
                if processing_end_col:
                    processing_end_days = row.get(processing_end_col, None)
                if (processing_end_days is None or pd.isna(processing_end_days) or str(processing_end_days).strip() == "") and grace_override_col:
                    processing_end_days = row.get(grace_override_col, None)
            
            # 기본값 적용 (환경설정 우선, 타입 기본값 다음, 마지막으로 하드코딩 값)
            if sourcing_start_days is None or pd.isna(sourcing_start_days) or str(sourcing_start_days).strip() == "":
                # 타입별 설정 우선, 없으면 타입 기본값, 마지막으로 환경설정 기본값
                sourcing_start_days = (type_specific_config.get("sourcing_start_days") 
                                     or type_defaults.get("prep_days") 
                                     or default_sourcing_start_days)
            if processing_end_days is None or pd.isna(processing_end_days) or str(processing_end_days).strip() == "":
                processing_end_days = (type_specific_config.get("processing_end_days")
                                     or type_defaults.get("grace_days")
                                     or default_processing_end_days)
            
            # cross_year 처리
            cross_year_val = False
            if cross_col:
                cross_val = str(row.get(cross_col, "")).strip().upper()
                cross_year_val = cross_val in ['Y', 'YES', 'TRUE', '1', 'TRUE']
            
            season = {
                "id": season_id,
                "name": str(row.get(season_name_col, season_id)).strip() if season_name_col else season_id,
                "type": type_id if type_id else "default",
                "start_date": str(row.get(start_col, "")).strip() if start_col else "",
                "end_date": str(row.get(end_col, "")).strip() if end_col else "",
                "cross_year": cross_year_val,
                # Phase 1: 가공 기간 관련 필드 (새 필드명 사용)
                "sourcing_start_days": int(sourcing_start_days) if pd.notna(sourcing_start_days) else default_sourcing_start_days,
                "processing_end_days": int(processing_end_days) if pd.notna(processing_end_days) else default_processing_end_days,
                # 하위 호환성: 기존 필드명도 유지
                "prep_days": int(sourcing_start_days) if pd.notna(sourcing_start_days) else default_sourcing_start_days,
                "grace_days": int(processing_end_days) if pd.notna(processing_end_days) else default_processing_end_days,
                "priority": int(row.get(priority_col, 1)) if priority_col and pd.notna(row.get(priority_col)) else 1,
                "keywords": {
                    "include": [],
                    "exclude": [],
                    "allowed": []  # Phase 1: allowed 타입 추가
                }
            }
            
            # KEYWORDS 시트에서 해당 시즌 키워드 찾기
            keyword_sheet = find_sheet(["KEYWORDS", "KEYWORD"])
            if keyword_sheet:
                df_keywords = pd.read_excel(xl, sheet_name=keyword_sheet)
                df_keywords = df_keywords.dropna(how='all')  # 빈 행 제거
                
                kw_season_id_col = _find_column_static(df_keywords, ["시즌ID", "시즌", "season_id", "season", "시즌ID(season_id)"])
                keyword_col = _find_column_static(df_keywords, ["키워드", "keyword", "단어", "키워드(keyword)"])
                polarity_col = _find_column_static(df_keywords, ["polarity", "타입", "종류", "포함여부", "극성", "포함/제외(polarity: include/exclude)"])
                type_col = _find_column_static(df_keywords, ["타입", "type"])
                weight_col = _find_column_static(df_keywords, ["가중치", "weight", "점수"])
                
                if kw_season_id_col and keyword_col:
                    season_keywords = df_keywords[df_keywords[kw_season_id_col].astype(str).str.strip() == season_id]
                    for _, kw_row in season_keywords.iterrows():
                        keyword = str(kw_row.get(keyword_col, "")).strip()
                        if not keyword or keyword in ['nan', 'None', '']:
                            continue
                        
                        # Phase 1: 타입 컬럼도 지원 (polarity와 동일)
                        polarity_str = None
                        if polarity_col:
                            polarity_str = str(kw_row.get(polarity_col, "")).strip().lower()
                        if (not polarity_str or polarity_str in ['nan', 'none', '']) and type_col:
                            polarity_str = str(kw_row.get(type_col, "")).strip().lower()
                        if not polarity_str or polarity_str in ['nan', 'none', '']:
                            polarity_str = "include"
                        
                        weight = float(kw_row.get(weight_col, 1.0)) if weight_col and pd.notna(kw_row.get(weight_col)) else 1.0
                        
                        # polarity/타입 변환 (Phase 1: allowed 추가)
                        if polarity_str in ["include", "포함", "1", "true", "yes"]:
                            polarity = "include"
                        elif polarity_str in ["exclude", "제외", "0", "false", "no"]:
                            polarity = "exclude"
                        elif polarity_str in ["allowed", "예외허용", "allow", "예외", "allowed"]:
                            polarity = "allowed"
                        else:
                            polarity = "include"  # 기본값
                        
                        kw_data = {
                            "keyword": keyword,
                            "weight": weight
                        }
                        
                        season["keywords"][polarity].append(kw_data)
                        
                        # config["keywords"]에도 저장 (빠른 접근용)
                        if season_id not in config["keywords"]:
                            config["keywords"][season_id] = {
                                "include": [],
                                "exclude": [],
                                "allowed": []
                            }
                        config["keywords"][season_id][polarity].append(kw_data)
            
            config["seasons"].append(season)
    
    return config


def _find_column_static(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """컬럼명 찾기 (정적 함수, 괄호 포함 형식도 지원)"""
    # 1. 정확한 매칭 시도
    for name in possible_names:
        if name in df.columns:
            return name
    
    # 2. 괄호 포함 형식 매칭 (예: "시즌ID(season_id)" -> "시즌ID" 또는 "season_id" 매칭)
    #    또는 "시작(MM-DD)(start_mmdd)" -> "시작", "start", "start_mmdd" 매칭
    for name in possible_names:
        for col in df.columns:
            col_str = str(col)
            # 컬럼명에 괄호가 있으면 괄호 앞부분 또는 괄호 안 내용 매칭
            if '(' in col_str and ')' in col_str:
                # 괄호 앞부분 추출 (예: "시작(MM-DD)(start_mmdd)" -> "시작(MM-DD)")
                before_paren = col_str.split('(')[0].strip()
                # 모든 괄호 안 내용 추출
                import re
                inside_parens = re.findall(r'\(([^)]+)\)', col_str)
                
                # 괄호 앞부분과 매칭
                if name.lower() == before_paren.lower():
                    return col
                
                # 괄호 안 내용 중 하나와 매칭
                for inside in inside_parens:
                    inside_clean = inside.strip()
                    # 부분 매칭 (예: "start_mmdd"에서 "start" 매칭)
                    if name.lower() == inside_clean.lower() or name.lower() in inside_clean.lower():
                        return col
                    
                    # 언더스코어로 분리된 부분 매칭
                    for part in inside_clean.split('_'):
                        if name.lower() == part.lower():
                            return col
            
            # 괄호 없이도 부분 매칭 시도
            if name.lower() in col_str.lower() or col_str.lower() in name.lower():
                return col
    
    return None


def filter_products_by_season(products: List[Dict], season_config: Dict, 
                              current_date: Optional[datetime] = None) -> tuple:
    """
    상품 목록을 시즌 필터로 필터링
    
    Args:
        products: 상품 딕셔너리 리스트 (각 상품은 '상품명' 또는 'product_name' 키 포함)
        season_config: 시즌 설정 딕셔너리
        current_date: 현재 날짜 (기본값: 오늘)
    
    Returns:
        (filtered_products, excluded_count, excluded_seasons, included_seasons, season_stats)
        - filtered_products: 필터링된 상품 리스트
        - excluded_count: 제외된 상품 수
        - excluded_seasons: 제외된 시즌 정보 {season_id: {'count': int, 'reason': str}}
        - included_seasons: 포함된 시즌 정보 {season_id: {'count': int, 'name': str}}
        - season_stats: 통계 정보 {'non_season': int, 'season_valid': int, 'season_invalid': int}
    """
    if not season_config or not products:
        return products, 0, {}, {}, {'non_season': len(products), 'season_valid': 0, 'season_invalid': 0}
    
    if current_date is None:
        current_date = datetime.now()
    
    excluded_count = 0
    excluded_seasons = {}  # {season_id: {'count': int, 'reason': str, 'name': str}}
    included_seasons = {}  # {season_id: {'count': int, 'name': str}}
    filtered_products = []
    season_stats = {'non_season': 0, 'season_valid': 0, 'season_invalid': 0}
    
    # 공통 예외단어 로드 (최우선순위: 예외허용단어 > 타입 > 시즌 > 단어)
    common_exclude_keywords = season_config.get("settings", {}).get("common_exclude_keywords", [])
    
    # 각 상품에 대해 시즌 감지 및 필터링
    for product in products:
        # 상품명 가져오기
        product_name = str(product.get('상품명', product.get('product_name', ''))).lower()
        if not product_name:
            filtered_products.append(product)
            season_stats['non_season'] += 1
            continue
        
        category_name = product.get("카테고리명", "")
        
        # 검색 텍스트 (예외단어 체크용)
        search_text_for_exclude = product_name
        if category_name:
            search_text_for_exclude = f"{product_name} {category_name}"
        search_text_for_exclude = search_text_for_exclude.lower()
        
        # 예외단어 체크 (시즌 감지 전에 확인하지만, 시즌 감지는 항상 수행)
        has_exclude_keyword = False
        for kw in common_exclude_keywords:
            if isinstance(kw, dict):
                kw_str = kw.get("keyword", kw.get("key", "")).lower()
            else:
                kw_str = str(kw).lower()
            
            if kw_str and kw_str in search_text_for_exclude:
                has_exclude_keyword = True
                break
        
        # 우선순위 2: 시즌 감지 (예외단어가 있어도 수행 - 시즌 키워드가 있으면 시즌으로 분류)
        detected_seasons = _detect_seasons_from_product(product_name, season_config, category_name)
        
        if not detected_seasons:
            # 시즌이 감지되지 않으면 통과 (일반 상품)
            filtered_products.append(product)
            season_stats['non_season'] += 1
            continue
        
        # 우선순위 3: 시즌 유효성 확인 (타입, 시즌, 단어 순)
        is_valid = False
        valid_season_id = None
        # 각 시즌별 제외 사유를 저장 (시즌ID -> 사유)
        season_reasons = {}
        
        for season_id, score in detected_seasons:
            season_info = next((s for s in season_config.get("seasons", []) if s["id"] == season_id), None)
            if not season_info:
                continue
            
            season_name = season_info.get("name", season_id)
            validity = _check_season_validity(season_info, current_date, season_config)
            season_type = season_info.get("type", "").strip().upper()
            
            if validity == "ACTIVE":
                # 우선순위 3-1: 예외단어 체크 (ACTIVE 시즌에서만 예외단어 적용)
                # 예외단어가 있으면 ACTIVE Event 타입에서 제외하지만, 일반 상품으로 처리
                if has_exclude_keyword and season_type == "EVENT":
                    # 예외단어가 있고 Event 타입이면 ACTIVE 시즌에서 제외
                    season_reasons[season_id] = f"{season_name}(공통 제외 키워드 매칭 - Event 타입)"
                    continue  # 이 시즌은 제외하지만, 다른 시즌 확인 계속
                
                # 우선순위 3-2: 시즌 점수 확인 (단어 매칭 점수)
                score_min = season_info.get("score_min", 1)
                if score >= score_min:
                    # 점수를 만족하면 포함 (예외단어가 없거나, 예외단어가 있어도 Event가 아닌 타입)
                    is_valid = True
                    valid_season_id = season_id
                    break  # 첫 번째 유효한 시즌을 찾으면 중단
                else:
                    # 점수 부족으로 제외 (하지만 다른 시즌 확인 계속)
                    season_reasons[season_id] = f"{season_name}(점수 부족: {score}/{score_min})"
            elif validity == "SOURCING":
                # 소싱 기간: Event 타입만 제외, 다른 타입은 포함
                # 중요 정책: 예외단어와 무관하게 시즌 키워드가 있으면 시즌으로 분류
                if season_type == "EVENT":
                    # Event 타입만 소싱 기간 제외
                    season_reasons[season_id] = f"{season_name}(소싱 기간 - 이미 가공 완료)"
                else:
                    # 다른 타입은 소싱 기간이어도 시즌 키워드가 있으면 시즌으로 분류 (예외단어 무관)
                    # ACTIVE 시즌이 없을 때 포함 처리
                    score_min = season_info.get("score_min", 1)
                    if score >= score_min:
                        is_valid = True
                        valid_season_id = season_id
                        break
            else:
                # EXPIRED인 경우: Event 타입만 제외, 다른 타입은 포함
                # 중요 정책: 예외단어와 무관하게 시즌 키워드가 있으면 시즌으로 분류
                if season_type == "EVENT":
                    # Event 타입만 시즌 종료 시 제외
                    season_reasons[season_id] = f"{season_name}(시즌 종료됨 - Event 타입)"
                else:
                    # 다른 타입은 시즌 종료되어도 시즌 키워드가 있으면 시즌으로 분류 (예외단어 무관)
                    # ACTIVE 시즌이 없을 때 포함 처리
                    score_min = season_info.get("score_min", 1)
                    if score >= score_min:
                        is_valid = True
                        valid_season_id = season_id
                    break
        
        if is_valid:
            # 유효한 시즌이 있으면 포함
            filtered_products.append(product)
            season_stats['season_valid'] += 1
            
            # 포함된 시즌 기록
            if valid_season_id:
                season_info = next((s for s in season_config.get("seasons", []) if s["id"] == valid_season_id), None)
                season_name = season_info.get("name", valid_season_id) if season_info else valid_season_id
                validity_status = _check_season_validity(season_info, current_date, season_config) if season_info else "ACTIVE"
                
                if valid_season_id not in included_seasons:
                    included_seasons[valid_season_id] = {'count': 0, 'name': season_name, 'status': validity_status}
                included_seasons[valid_season_id]['count'] += 1
        else:
            # 유효한 시즌이 없으면
            # 예외단어가 있고 모든 시즌이 제외된 경우, 일반 상품으로 처리
            if has_exclude_keyword:
                # 예외단어가 있으면 일반 상품으로 처리 (시즌 제외)
                filtered_products.append(product)
                season_stats['non_season'] += 1
            else:
                # 예외단어가 없으면 제외
                excluded_count += 1
                season_stats['season_invalid'] += 1
            
            # 제외된 시즌 기록 (각 시즌별로 정확한 사유 기록)
            for season_id, score in detected_seasons:
                season_info = next((s for s in season_config.get("seasons", []) if s["id"] == season_id), None)
                season_name = season_info.get("name", season_id) if season_info else season_id
                
                if season_id not in excluded_seasons:
                    excluded_seasons[season_id] = {'count': 0, 'reason': '', 'name': season_name}
                excluded_seasons[season_id]['count'] += 1
                
                # 제외 사유 설정 (해당 시즌의 정확한 사유)
                if not excluded_seasons[season_id]['reason']:
                    if season_id in season_reasons:
                        excluded_seasons[season_id]['reason'] = season_reasons[season_id]
                    else:
                        excluded_seasons[season_id]['reason'] = f"{season_name}(시즌 기간 외 또는 점수 부족)"
    
    return filtered_products, excluded_count, excluded_seasons, included_seasons, season_stats


def _detect_seasons_from_product(product_name: str, season_config: Dict, category_name: str = "") -> List[tuple]:
    """
    상품명과 카테고리에서 시즌 감지 (점수 기반)
    
    로직:
    1. include 키워드만 사용 → 가중치만큼 score 추가 (시즌 판단)
    2. exclude 키워드는 사용하지 않음 (시즌별 exclude는 미사용)
    3. 제외 키워드는 공통 제외 키워드만 사용 (filter_products_by_season에서 처리)
       - 공통 제외 키워드는 Event 타입 ACTIVE 기간에만 적용
    
    Args:
        product_name: 상품명
        season_config: 시즌 설정 딕셔너리
        category_name: 카테고리명 (선택적)
    
    Returns:
        [(season_id, score), ...] 리스트
    """
    detected = []
    case_sensitive = season_config.get("settings", {}).get("case_sensitive", False)
    
    # 검색 대상 텍스트: 상품명 + 카테고리
    search_text = product_name
    if category_name:
        search_text = f"{product_name} {category_name}"
    
    if not case_sensitive:
        search_text = search_text.lower()
    
    for season in season_config.get("seasons", []):
        season_id = season.get("id")
        if not season_id:
            continue
        
        # 키워드는 season 객체에서 직접 가져오기
        keywords = season.get("keywords", {})
        if not keywords:
            # config["keywords"]에서도 시도 (하위 호환성)
            keywords = season_config.get("keywords", {}).get(season_id, {})
        
        score = 0
        
        # 중요: 시즌별 exclude 키워드는 사용하지 않음
        # - 시즌별 키워드는 모두 include만 사용
        # - 제외는 공통 제외 키워드만 사용 (Event 타입 ACTIVE 기간에만 적용)
        
        # include 키워드만 매칭하여 점수 계산 (시즌 판단 키워드)
        for kw_info in keywords.get("include", []):
            kw = kw_info.get("keyword", "")
            if not case_sensitive:
                kw = kw.lower()
            if kw in search_text:
                weight = kw_info.get("weight", 1)
                score += weight
        
        if score > 0:
            detected.append((season_id, score))
    
    # 점수 순으로 정렬 (가장 높은 점수의 시즌 우선 확인)
    # 주의: 현재는 첫 번째 유효한 시즌만 사용하지만, 향후 점수 차이를 활용할 수 있도록 정렬 유지
    detected.sort(key=lambda x: x[1], reverse=True)
    return detected


def _parse_date_string(date_str: str, default_year: Optional[int] = None) -> Optional[datetime]:
    """
    다양한 날짜 형식 파싱
    - "%Y-%m-%d": 2024-12-01
    - "%Y/%m/%d": 2024/12/01
    - "%m-%d": 12-01 (현재 연도 사용, default_year 지정 시 해당 연도)
    - "%m/%d": 12/01 (현재 연도 사용, default_year 지정 시 해당 연도)
    """
    if not date_str or pd.isna(date_str):
        return None
    
    date_str = str(date_str).strip()
    if not date_str or date_str in ['nan', 'None', '']:
        return None
    
    # 기본 연도 설정
    if default_year is None:
        default_year = datetime.now().year
    
    # 1. "%Y-%m-%d" 형식
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except:
        pass
    
    # 2. "%Y/%m/%d" 형식
    try:
        return datetime.strptime(date_str, "%Y/%m/%d")
    except:
        pass
    
    # 3. "%m-%d" 형식 (현재 연도 사용)
    try:
        date_obj = datetime.strptime(date_str, "%m-%d")
        return date_obj.replace(year=default_year)
    except:
        pass
    
    # 4. "%m/%d" 형식 (현재 연도 사용)
    try:
        date_obj = datetime.strptime(date_str, "%m/%d")
        return date_obj.replace(year=default_year)
    except:
        pass
    
    # 5. pandas를 이용한 자동 파싱
    try:
        return pd.to_datetime(date_str).to_pydatetime()
    except:
        pass
    
    return None


def _check_season_validity(season_info: Dict, current_date: datetime, season_config: Dict) -> str:
    """
    시즌 유효성 확인 (3단계 구분)
    
    기간 구분:
    1. SOURCING: 소싱 기간 (시즌시작일 - 소싱시작일수 ~ 시즌시작일 - 1일) - 미리 가공해야 하는 기간
    2. ACTIVE: 출력 가능 기간 (시즌시작일 ~ 시즌종료일 - 가공완료마감일수) - 지금 출력 가능한 기간
    3. EXPIRED: 제외 기간 (그 외) - 삭제해야 하는 기간
    
    Returns:
        "SOURCING", "ACTIVE", "EXPIRED" 중 하나
    """
    try:
        start_date_str = season_info.get("start_date", "")
        end_date_str = season_info.get("end_date", "")
        
        if not start_date_str or not end_date_str:
            return "EXCLUDE"
        
        # 날짜 파싱
        start_date = _parse_date_string(start_date_str)
        end_date = _parse_date_string(end_date_str)
        
        if start_date is None or end_date is None:
            return "EXCLUDE"
        
        cross_year = season_info.get("cross_year", False)
        
        # 연도넘김 처리 (개선된 로직)
        if cross_year:
            # 시작일과 종료일의 월 비교
            start_month = start_date.month
            end_month = end_date.month
            
            # 종료일이 시작일보다 앞서 있으면 연도를 넘김 (예: 12월 ~ 2월)
            if end_month < start_month:
                # 현재 날짜가 시작일 전반기 (예: 11월 이전)면 이전 연도
                if current_date.month < start_month:
                    start_date = start_date.replace(year=current_date.year - 1)
                    end_date = end_date.replace(year=current_date.year)  # 종료일은 현재 연도
                # 현재 날짜가 종료일 후반기 (예: 3월 이후)면 다음 연도
                elif current_date.month > end_month:
                    start_date = start_date.replace(year=current_date.year)
                    end_date = end_date.replace(year=current_date.year + 1)  # 종료일은 다음 연도
                # 현재 날짜가 시작일 ~ 종료일 사이 (예: 12월 ~ 2월)
                else:
                    # 현재 날짜가 시작일보다 크면 현재 연도, 작으면 이전 연도
                    if current_date.month >= start_month:
                        start_date = start_date.replace(year=current_date.year)
                        end_date = end_date.replace(year=current_date.year + 1)
                    else:
                        start_date = start_date.replace(year=current_date.year - 1)
                        end_date = end_date.replace(year=current_date.year)
            else:
                # 일반적인 경우 (같은 연도 내)
                start_date = start_date.replace(year=current_date.year)
                end_date = end_date.replace(year=current_date.year)
        else:
            # 연도넘김 없으면 현재 연도 사용
            start_date = start_date.replace(year=current_date.year)
            end_date = end_date.replace(year=current_date.year)
        
        # Phase 1: 하드코딩 제거, Excel 설정값 사용
        sourcing_start_days = season_info.get("sourcing_start_days", season_info.get("prep_days", 30))
        processing_end_days = season_info.get("processing_end_days", season_info.get("grace_days", 21))
        
        # 3단계 기간 계산
        # 1. 소싱 시작: 시즌시작일 - 소싱시작일수
        sourcing_start = start_date - pd.Timedelta(days=int(sourcing_start_days))
        # 2. 출력 시작: 시즌시작일
        active_start = start_date
        # 3. 출력 종료: 시즌종료일 - 가공완료마감일수
        active_end = end_date - pd.Timedelta(days=int(processing_end_days))
        
        # 3단계 구분
        if sourcing_start <= current_date < active_start:
            return "SOURCING"  # 소싱 기간 (미리 가공해야 하는)
        elif active_start <= current_date <= active_end:
            return "ACTIVE"  # 출력 가능 기간 (지금 출력 가능한)
        else:
            return "EXPIRED"  # 제외 기간 (삭제해야 하는)
    
    except Exception as e:
        # 오류 발생 시 제외 처리
        return "EXCLUDE"


if __name__ == "__main__":
    app = SeasonFilterManagerGUI()
    app.mainloop()

